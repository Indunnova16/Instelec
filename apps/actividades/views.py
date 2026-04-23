"""
Views for activity management.
"""
from typing import Any
from uuid import UUID
from datetime import date, timedelta

from django.db.models import QuerySet, Count, Q
from django.views.generic import ListView, DetailView, TemplateView, FormView, View
from django.contrib.auth.mixins import LoginRequiredMixin
from django.utils import timezone
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from apps.core.mixins import HTMXMixin, RoleRequiredMixin
from .models import Actividad, ProgramacionMensual, TipoActividad, HistorialIntervencion


class ActividadListView(LoginRequiredMixin, HTMXMixin, ListView):
    """List activities with filters."""
    model = Actividad
    template_name = 'actividades/lista.html'
    partial_template_name = 'actividades/partials/lista_actividades.html'
    context_object_name = 'actividades'
    paginate_by = 20

    def get_queryset(self) -> QuerySet[Actividad]:
        qs = super().get_queryset().select_related(
            'linea', 'torre', 'tipo_actividad', 'cuadrilla'
        ).defer(
            'observaciones_programacion', 'motivo_reprogramacion',
            'motivo_cancelacion',
        )

        # Search by aviso SAP
        buscar_aviso = self.request.GET.get('buscar_aviso', '').strip()
        if buscar_aviso:
            qs = qs.filter(aviso_sap__icontains=buscar_aviso)

        # Filters
        estado = self.request.GET.get('estado')
        if estado and estado in dict(Actividad.Estado.choices):
            qs = qs.filter(estado=estado)

        linea = self.request.GET.get('linea')
        if linea:
            try:
                UUID(linea)
                qs = qs.filter(linea_id=linea)
            except ValueError:
                pass

        cuadrilla = self.request.GET.get('cuadrilla')
        if cuadrilla:
            try:
                UUID(cuadrilla)
                qs = qs.filter(cuadrilla_id=cuadrilla)
            except ValueError:
                pass

        tipo_actividad = self.request.GET.get('tipo_actividad')
        if tipo_actividad:
            try:
                UUID(tipo_actividad)
                qs = qs.filter(tipo_actividad_id=tipo_actividad)
            except ValueError:
                pass

        # Month/year filter
        mes = self.request.GET.get('mes')
        if mes:
            try:
                qs = qs.filter(fecha_programada__month=int(mes))
            except (ValueError, TypeError):
                pass

        anio = self.request.GET.get('anio')
        if anio:
            try:
                qs = qs.filter(fecha_programada__year=int(anio))
            except (ValueError, TypeError):
                pass

        # Store for stats calculation
        self._filtered_qs = qs
        return qs

    def get_context_data(self, **kwargs: Any) -> dict[str, Any]:
        context = super().get_context_data(**kwargs)

        from apps.lineas.models import Linea
        from apps.cuadrillas.models import Cuadrilla

        context['estados'] = Actividad.Estado.choices
        context['tipos'] = TipoActividad.objects.filter(activo=True)
        context['lineas'] = Linea.objects.filter(activa=True)
        context['cuadrillas'] = Cuadrilla.objects.filter(activa=True)

        # Month/year selector options
        context['meses'] = [
            (1, 'Enero'), (2, 'Febrero'), (3, 'Marzo'), (4, 'Abril'),
            (5, 'Mayo'), (6, 'Junio'), (7, 'Julio'), (8, 'Agosto'),
            (9, 'Septiembre'), (10, 'Octubre'), (11, 'Noviembre'), (12, 'Diciembre')
        ]
        context['anios'] = range(date.today().year - 1, date.today().year + 2)

        # Progress stats from filtered queryset (before pagination)
        qs = getattr(self, '_filtered_qs', self.get_queryset())
        stats = qs.aggregate(
            total=Count('id'),
            ejecutadas=Count('id', filter=Q(estado='COMPLETADA')),
            pendientes=Count('id', filter=Q(estado__in=['PENDIENTE', 'PROGRAMADA'])),
            en_curso=Count('id', filter=Q(estado='EN_CURSO')),
            canceladas=Count('id', filter=Q(estado__in=['CANCELADA', 'REPROGRAMADA'])),
        )
        context['stats'] = stats
        total = stats['total'] or 0
        context['porcentaje_ejecucion'] = (
            round((stats['ejecutadas'] / total) * 100, 1) if total > 0 else 0
        )

        return context


class ActividadDetailView(LoginRequiredMixin, HTMXMixin, DetailView):
    """Detail view for an activity."""
    model = Actividad
    template_name = 'actividades/detalle.html'
    context_object_name = 'actividad'

    def get_queryset(self) -> QuerySet[Actividad]:
        return super().get_queryset().select_related(
            'linea', 'torre', 'tipo_actividad', 'cuadrilla'
        ).prefetch_related(
            'registros_campo__usuario',
            'registros_campo__evidencias'
        )

    def get_context_data(self, **kwargs: Any) -> dict[str, Any]:
        context = super().get_context_data(**kwargs)
        # Get field records for this activity (already prefetched)
        context['registros'] = self.object.registros_campo.all()
        return context


class ActividadDetailPartialView(ActividadDetailView):
    """Partial view for HTMX loading."""
    template_name = 'actividades/partials/detalle_actividad.html'


class CalendarioView(LoginRequiredMixin, TemplateView):
    """Calendar view for activity scheduling."""
    template_name = 'actividades/calendario.html'

    def get_context_data(self, **kwargs: Any) -> dict[str, Any]:
        context = super().get_context_data(**kwargs)

        # Get current month activities
        hoy = timezone.now().date()
        try:
            mes = int(self.request.GET.get('mes', hoy.month))
        except (ValueError, TypeError):
            mes = hoy.month
        try:
            anio = int(self.request.GET.get('anio', hoy.year))
        except (ValueError, TypeError):
            anio = hoy.year

        actividades = Actividad.objects.filter(
            fecha_programada__year=anio,
            fecha_programada__month=mes
        ).select_related('linea', 'torre', 'tipo_actividad', 'cuadrilla')

        # Group by date
        from collections import defaultdict
        actividades_por_fecha = defaultdict(list)
        for act in actividades:
            actividades_por_fecha[act.fecha_programada.day].append(act)

        context['actividades_por_fecha'] = dict(actividades_por_fecha)
        context['mes'] = mes
        context['anio'] = anio

        # Generate calendar data
        import calendar
        cal = calendar.Calendar(firstweekday=0)
        context['semanas'] = cal.monthdayscalendar(anio, mes)
        context['nombre_mes'] = calendar.month_name[mes]

        return context


class ProgramacionListView(LoginRequiredMixin, RoleRequiredMixin, ListView):
    """List monthly programming with activities filtered by month."""
    model = Actividad
    template_name = 'actividades/programacion.html'
    context_object_name = 'actividades'
    paginate_by = 50
    allowed_roles = ['admin', 'director', 'coordinador', 'ing_residente']

    def get_queryset(self) -> QuerySet[Actividad]:
        qs = Actividad.objects.select_related(
            'linea', 'torre', 'tipo_actividad', 'cuadrilla'
        ).order_by('linea__codigo', 'tipo_actividad__nombre', 'torre__numero')

        # Default to current month/year
        hoy = date.today()
        try:
            self.selected_mes = int(self.request.GET.get('mes', hoy.month))
        except (ValueError, TypeError):
            self.selected_mes = hoy.month
        try:
            self.selected_anio = int(self.request.GET.get('anio', hoy.year))
        except (ValueError, TypeError):
            self.selected_anio = hoy.year

        qs = qs.filter(
            fecha_programada__month=self.selected_mes,
            fecha_programada__year=self.selected_anio,
        )

        # Filter by linea(s) - multiple selection
        linea_ids = self.request.GET.getlist('linea')
        valid_linea_ids = []
        for lid in linea_ids:
            try:
                UUID(lid)
                valid_linea_ids.append(lid)
            except ValueError:
                pass
        if valid_linea_ids:
            qs = qs.filter(linea_id__in=valid_linea_ids)

        # Filter by tipo_actividad
        tipo_id = self.request.GET.get('tipo_actividad')
        if tipo_id:
            try:
                UUID(tipo_id)
                qs = qs.filter(tipo_actividad_id=tipo_id)
            except ValueError:
                pass

        # Filter by estado
        estado = self.request.GET.get('estado')
        if estado and estado in dict(Actividad.Estado.choices):
            qs = qs.filter(estado=estado)

        # Filter by cuadrilla(s) - multiple selection
        cuadrilla_ids = self.request.GET.getlist('cuadrilla')
        valid_cuadrilla_ids = []
        for cid in cuadrilla_ids:
            try:
                UUID(cid)
                valid_cuadrilla_ids.append(cid)
            except ValueError:
                pass
        if valid_cuadrilla_ids:
            qs = qs.filter(cuadrilla_id__in=valid_cuadrilla_ids)

        # Search by aviso SAP
        buscar_aviso = self.request.GET.get('buscar_aviso', '').strip()
        if buscar_aviso:
            qs = qs.filter(aviso_sap__icontains=buscar_aviso)

        self._filtered_qs = qs
        return qs

    def get_context_data(self, **kwargs: Any) -> dict[str, Any]:
        context = super().get_context_data(**kwargs)
        from apps.lineas.models import Linea
        from apps.cuadrillas.models import Cuadrilla

        context['lineas'] = Linea.objects.filter(activa=True)
        context['tipos'] = TipoActividad.objects.filter(activo=True)
        context['estados'] = Actividad.Estado.choices
        context['cuadrillas'] = Cuadrilla.objects.filter(activa=True)
        context['selected_cuadrillas'] = self.request.GET.getlist('cuadrilla')
        context['selected_lineas'] = self.request.GET.getlist('linea')
        context['meses'] = [
            (1, 'Enero'), (2, 'Febrero'), (3, 'Marzo'), (4, 'Abril'),
            (5, 'Mayo'), (6, 'Junio'), (7, 'Julio'), (8, 'Agosto'),
            (9, 'Septiembre'), (10, 'Octubre'), (11, 'Noviembre'), (12, 'Diciembre')
        ]
        context['anios'] = range(date.today().year - 1, date.today().year + 2)
        context['selected_mes'] = self.selected_mes
        context['selected_anio'] = self.selected_anio

        # Get month name
        nombres_mes = {
            1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril',
            5: 'Mayo', 6: 'Junio', 7: 'Julio', 8: 'Agosto',
            9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'
        }
        context['nombre_mes'] = nombres_mes.get(self.selected_mes, '')

        # Stats
        qs = getattr(self, '_filtered_qs', self.get_queryset())
        stats = qs.aggregate(
            total=Count('id'),
            ejecutadas=Count('id', filter=Q(estado='COMPLETADA')),
            pendientes=Count('id', filter=Q(estado__in=['PENDIENTE', 'PROGRAMADA'])),
            en_curso=Count('id', filter=Q(estado='EN_CURSO')),
        )
        context['stats'] = stats
        total = stats['total'] or 0
        context['porcentaje_ejecucion'] = (
            round((stats['ejecutadas'] / total) * 100, 1) if total > 0 else 0
        )

        # Programaciones for this month
        context['programaciones'] = ProgramacionMensual.objects.filter(
            mes=self.selected_mes, anio=self.selected_anio
        ).select_related('linea')

        return context


class ImportarProgramacionView(LoginRequiredMixin, RoleRequiredMixin, TemplateView):
    """View for importing programming from Excel."""
    template_name = 'actividades/importar.html'
    allowed_roles = ['admin', 'director', 'coordinador']

    def get_context_data(self, **kwargs: Any) -> dict[str, Any]:
        context = super().get_context_data(**kwargs)
        from apps.lineas.models import Linea
        context['lineas'] = Linea.objects.filter(activa=True)
        context['anios'] = range(date.today().year - 1, date.today().year + 2)
        context['meses'] = [
            (1, 'Enero'), (2, 'Febrero'), (3, 'Marzo'), (4, 'Abril'),
            (5, 'Mayo'), (6, 'Junio'), (7, 'Julio'), (8, 'Agosto'),
            (9, 'Septiembre'), (10, 'Octubre'), (11, 'Noviembre'), (12, 'Diciembre')
        ]
        return context

    def post(self, request, *args, **kwargs):
        """Handle Excel file upload and import."""
        from .importers import ProgramaTranselcaImporter, AvisosTranselcaImporter
        from apps.lineas.models import Linea

        archivo = request.FILES.get('archivo')
        if not archivo:
            messages.error(request, 'Debe seleccionar un archivo Excel')
            return redirect('actividades:importar')

        # Validar extensión
        if not archivo.name.endswith(('.xlsx', '.xls')):
            messages.error(request, 'El archivo debe ser un Excel (.xlsx o .xls)')
            return redirect('actividades:importar')

        # Obtener parámetros
        linea_id = request.POST.get('linea')
        anio_str = request.POST.get('anio', '').strip()
        mes_str = request.POST.get('mes', '').strip()
        actualizar_existentes = request.POST.get('actualizar_existentes') == 'on'

        # Convertir a int, permitir None para detección automática
        anio = int(anio_str) if anio_str else None
        mes = int(mes_str) if mes_str else None

        # Si no hay línea, intentar importar como avisos (detecta línea del Excel)
        if not linea_id:
            importer = AvisosTranselcaImporter()
            resultado = importer.importar(
                archivo,
                anio=anio,
                mes=mes,
                opciones={'actualizar_existentes': actualizar_existentes}
            )

            if resultado['exito']:
                mensaje = (
                    f"Importación exitosa: {resultado['actividades_creadas']} actividades creadas, "
                    f"{resultado['actividades_actualizadas']} actualizadas."
                )
                if resultado.get('advertencias'):
                    mensaje += f" {len(resultado['advertencias'])} advertencias."
                messages.success(request, mensaje)
            else:
                messages.error(request, f"Error en importación: {resultado.get('error', 'Error desconocido')}")

            return redirect('actividades:programacion')

        try:
            linea = Linea.objects.get(id=linea_id)
        except Linea.DoesNotExist as e:
            messages.error(request, f'Línea no encontrada: {e}')
            return redirect('actividades:importar')

        # Si año o mes no se especifican, detectarlos del Excel
        if not anio or not mes:
            importer_detect = ProgramaTranselcaImporter()
            from openpyxl import load_workbook
            try:
                wb = load_workbook(archivo, read_only=True, data_only=True)
                rows = list(wb.active.iter_rows(values_only=True))
                importer_detect._detectar_columnas(rows[0])
                anio_excel, mes_excel = importer_detect._detectar_fecha_excel(rows[1:])
                anio = anio or anio_excel
                mes = mes or mes_excel
            except Exception:
                pass

        # Usar valores por defecto si no se encuentran
        if not anio or not mes:
            from datetime import date
            hoy = date.today()
            anio = anio or hoy.year
            mes = mes or hoy.month

        # Crear o obtener programación mensual
        programacion, created = ProgramacionMensual.objects.get_or_create(
            anio=anio,
            mes=mes,
            linea=linea,
            defaults={
                'archivo_origen': archivo,
            }
        )

        if not created:
            programacion.archivo_origen = archivo
            programacion.save(update_fields=['archivo_origen', 'updated_at'])

        # Importar
        importer = ProgramaTranselcaImporter()
        resultado = importer.importar(
            archivo,
            programacion,
            opciones={'actualizar_existentes': actualizar_existentes}
        )

        if resultado['exito']:
            mensaje = (
                f"Importación exitosa: {resultado['actividades_creadas']} actividades creadas, "
                f"{resultado['actividades_actualizadas']} actualizadas, "
                f"{resultado['filas_omitidas']} filas omitidas."
            )
            if resultado['advertencias']:
                mensaje += f" {len(resultado['advertencias'])} advertencias."
            messages.success(request, mensaje)

            # Guardar datos importados en la programación
            programacion.datos_importados = {
                'resultado': resultado,
                'fecha_importacion': timezone.now().isoformat(),
                'usuario': request.user.get_full_name(),
            }
            programacion.save(update_fields=['datos_importados', 'updated_at'])
        else:
            messages.error(request, f"Error en importación: {resultado.get('error', 'Error desconocido')}")

        return redirect('actividades:programacion')


class ImportarAvancesView(LoginRequiredMixin, RoleRequiredMixin, TemplateView):
    """View for importing activity progress/avances from Excel."""
    template_name = 'actividades/importar_avances.html'
    allowed_roles = ['admin', 'director', 'coordinador', 'ing_residente']

    def get_context_data(self, **kwargs: Any) -> dict[str, Any]:
        context = super().get_context_data(**kwargs)
        context['columnas_soportadas'] = {
            'Aviso SAP (requerido)': ['aviso sap', 'aviso', 'nro aviso', 'sap'],
            'Porcentaje de Avance': ['avance', '% avance', 'porcentaje', 'progreso'],
            'Estado': ['estado', 'status'],
            'Observaciones': ['observaciones', 'notas', 'comentarios'],
        }
        return context

    def post(self, request, *args, **kwargs):
        from .importers import AvancesImporter

        archivo = request.FILES.get('archivo')
        if not archivo or not archivo.name.endswith(('.xlsx', '.xls')):
            messages.error(request, 'Por favor suba un archivo Excel (.xlsx o .xls)')
            return redirect('actividades:importar_avances')

        importer = AvancesImporter()
        resultado = importer.importar(archivo)

        if resultado['exito']:
            parts = [f"✓ {resultado['actividades_actualizadas']} actividades actualizadas"]
            if resultado['filas_omitidas']:
                parts.append(f"{resultado['filas_omitidas']} filas omitidas")
            if resultado['advertencias']:
                parts.append(f"{len(resultado['advertencias'])} advertencias")
            if resultado['errores']:
                parts.append(f"{len(resultado['errores'])} errores")
            messages.success(request, ' | '.join(parts))
        else:
            messages.error(request, f"Error: {resultado.get('error', 'Error desconocido')}")

        return redirect('actividades:programacion')


class TorresParaLineaView(LoginRequiredMixin, View):
    """JSON endpoint to get torres for a specific linea (HTMX dynamic select)."""

    def get(self, request, linea_id, *args, **kwargs):
        from apps.lineas.models import Torre
        try:
            UUID(str(linea_id))
        except ValueError:
            return JsonResponse([], safe=False)

        torres = Torre.objects.filter(linea_id=linea_id).order_by('numero').values('id', 'numero')
        return JsonResponse(list(torres), safe=False)


class ActividadCreateView(LoginRequiredMixin, RoleRequiredMixin, HTMXMixin, TemplateView):
    """View for creating a new activity."""
    template_name = 'actividades/crear.html'
    partial_template_name = 'actividades/partials/form_actividad.html'
    allowed_roles = ['admin', 'director', 'coordinador', 'ing_residente']

    def get_context_data(self, **kwargs: Any) -> dict[str, Any]:
        context = super().get_context_data(**kwargs)
        context['tipos'] = TipoActividad.objects.filter(activo=True)
        from apps.lineas.models import Linea
        from apps.cuadrillas.models import Cuadrilla
        context['lineas'] = Linea.objects.filter(activa=True)
        context['cuadrillas'] = Cuadrilla.objects.filter(activa=True)
        return context

    def post(self, request, *args, **kwargs):
        """Handle activity creation."""
        from apps.lineas.models import Linea, Torre
        from apps.cuadrillas.models import Cuadrilla

        tipo_actividad_id = request.POST.get('tipo_actividad')
        linea_id = request.POST.get('linea')
        torre_id = request.POST.get('torre')
        cuadrilla_id = request.POST.get('cuadrilla') or None
        fecha_programada = request.POST.get('fecha_programada')
        aviso_sap = request.POST.get('aviso_sap', '').strip()
        orden_sap = request.POST.get('orden_sap', '').strip()
        observaciones = request.POST.get('observaciones_programacion', '').strip()

        # If aviso_sap provided but missing other fields, try to fill from existing activity
        tipo_actividad = None
        linea = None
        torre = None

        if aviso_sap and not all([tipo_actividad_id, linea_id, torre_id]):
            existing = Actividad.objects.filter(aviso_sap__iexact=aviso_sap).select_related(
                'linea', 'torre', 'tipo_actividad'
            ).first()
            if existing:
                if not tipo_actividad_id:
                    tipo_actividad = existing.tipo_actividad
                if not linea_id:
                    linea = existing.linea
                if not torre_id:
                    torre = existing.torre
                if not orden_sap:
                    orden_sap = existing.orden_sap or ''

        # Resolve from IDs if not already set from lookup
        if not tipo_actividad and tipo_actividad_id:
            try:
                tipo_actividad = TipoActividad.objects.get(id=tipo_actividad_id)
            except TipoActividad.DoesNotExist:
                pass

        if not linea and linea_id:
            try:
                linea = Linea.objects.get(id=linea_id)
            except Linea.DoesNotExist:
                pass

        if not torre and torre_id:
            try:
                torre = Torre.objects.get(id=torre_id)
            except Torre.DoesNotExist:
                pass

        # Validation - only tipo_actividad and linea are strictly required
        if not tipo_actividad:
            messages.error(request, 'Debe seleccionar un tipo de actividad.')
            return self.get(request, *args, **kwargs)

        if not linea:
            messages.error(request, 'Debe seleccionar una linea.')
            return self.get(request, *args, **kwargs)

        # If no torre, use the first torre of the linea
        if not torre:
            torre = linea.torres.first()
            if not torre:
                messages.error(request, 'No se encontro una torre para la linea seleccionada.')
                return self.get(request, *args, **kwargs)

        # Validate torre belongs to linea
        if torre.linea_id != linea.id:
            messages.error(request, 'La torre seleccionada no pertenece a la linea.')
            return self.get(request, *args, **kwargs)

        # Default fecha_programada to today if not provided
        if not fecha_programada:
            from datetime import date as date_cls
            fecha_programada = date_cls.today()

        cuadrilla = None
        if cuadrilla_id:
            try:
                cuadrilla = Cuadrilla.objects.get(id=cuadrilla_id)
            except Cuadrilla.DoesNotExist:
                pass

        # Fix 1 abril 2026: Permitir crear actividades sin aviso_sap en emergencias
        # Si no hay aviso_sap y el tipo de actividad contiene "EMERGENCIA", generar código automático
        if not aviso_sap:
            if tipo_actividad and 'EMERGENCIA' in tipo_actividad.nombre.upper():
                from datetime import datetime
                aviso_sap = f"EMG-{datetime.now().strftime('%Y%m%d%H%M%S')}"
            else:
                # Para otros casos, aviso_sap es opcional
                aviso_sap = ''

        try:
            actividad = Actividad.objects.create(
                tipo_actividad=tipo_actividad,
                linea=linea,
                torre=torre,
                cuadrilla=cuadrilla,
                fecha_programada=fecha_programada,
                aviso_sap=aviso_sap,
                orden_sap=orden_sap,
                observaciones_programacion=observaciones,
                estado=Actividad.Estado.PENDIENTE,
                prioridad=Actividad.Prioridad.NORMAL,
            )
            messages.success(request, f'Actividad creada exitosamente.')
            return redirect('actividades:detalle', pk=actividad.pk)
        except Exception as e:
            messages.error(request, f'Error al guardar: {str(e)}')
            return self.get(request, *args, **kwargs)


class ActividadEditView(LoginRequiredMixin, RoleRequiredMixin, HTMXMixin, DetailView):
    """View for editing an activity."""
    model = Actividad
    template_name = 'actividades/editar.html'
    context_object_name = 'actividad'
    allowed_roles = ['admin', 'director', 'coordinador', 'ing_residente']

    def get_queryset(self):
        return super().get_queryset().prefetch_related('cuadrillas')

    def get_context_data(self, **kwargs: Any) -> dict[str, Any]:
        context = super().get_context_data(**kwargs)
        context['tipos'] = TipoActividad.objects.filter(activo=True)
        context['estados'] = Actividad.Estado.choices
        context['prioridades'] = Actividad.Prioridad.choices
        from apps.lineas.models import Linea
        from apps.cuadrillas.models import Cuadrilla
        context['lineas'] = Linea.objects.filter(activa=True)
        context['cuadrillas'] = Cuadrilla.objects.filter(activa=True)
        return context

    def post(self, request, *args, **kwargs):
        """Handle activity update."""
        actividad = self.get_object()

        tipo_actividad_id = request.POST.get('tipo_actividad')
        estado = request.POST.get('estado', '').strip()
        prioridad = request.POST.get('prioridad', '').strip()
        cuadrilla_id = request.POST.get('cuadrilla') or None
        fecha_programada = request.POST.get('fecha_programada')
        aviso_sap = request.POST.get('aviso_sap', '').strip()
        observaciones = request.POST.get('observaciones_programacion', '').strip()

        if not fecha_programada:
            messages.error(request, 'La fecha programada es obligatoria.')
            return self.get(request, *args, **kwargs)

        try:
            if tipo_actividad_id:
                actividad.tipo_actividad = TipoActividad.objects.get(id=tipo_actividad_id)

            if estado and estado in dict(Actividad.Estado.choices):
                actividad.estado = estado

            if prioridad and prioridad in dict(Actividad.Prioridad.choices):
                actividad.prioridad = prioridad

            # Handle cuadrilla assignment (supports multiple via M2M)
            from apps.cuadrillas.models import Cuadrilla
            cuadrilla_ids = request.POST.getlist('cuadrilla')
            # Filter empty values
            cuadrilla_ids = [cid for cid in cuadrilla_ids if cid]

            if cuadrilla_ids:
                try:
                    # Set the FK to the first selected cuadrilla (backward compat)
                    actividad.cuadrilla = Cuadrilla.objects.get(id=cuadrilla_ids[0])
                except Cuadrilla.DoesNotExist:
                    actividad.cuadrilla = None
            else:
                actividad.cuadrilla = None

            actividad.fecha_programada = fecha_programada
            actividad.aviso_sap = aviso_sap
            actividad.observaciones_programacion = observaciones
            actividad.comentarios_restricciones = request.POST.get('comentarios_restricciones', '').strip()

            actividad.save()

            # Sync M2M cuadrillas
            if cuadrilla_ids:
                cuadrillas_qs = Cuadrilla.objects.filter(id__in=cuadrilla_ids)
                actividad.cuadrillas.set(cuadrillas_qs)
            else:
                actividad.cuadrillas.clear()
            messages.success(request, 'Actividad actualizada exitosamente.')
            return redirect('actividades:detalle', pk=actividad.pk)
        except Exception as e:
            messages.error(request, f'Error al guardar: {str(e)}')
            return self.get(request, *args, **kwargs)


class CambiarEstadoView(LoginRequiredMixin, RoleRequiredMixin, DetailView):
    """View for changing activity status via HTMX."""
    model = Actividad
    allowed_roles = ['admin', 'director', 'coordinador', 'ing_residente', 'supervisor']

    def post(self, request, *args, **kwargs):
        actividad = self.get_object()
        nuevo_estado = request.POST.get('estado')
        if nuevo_estado in dict(Actividad.Estado.choices):
            actividad.estado = nuevo_estado
            actividad.save(update_fields=['estado', 'updated_at'])
            return JsonResponse({'success': True, 'estado': nuevo_estado})
        return JsonResponse({'success': False, 'error': 'Estado inválido'}, status=400)


class BulkAsignarCuadrillaView(LoginRequiredMixin, RoleRequiredMixin, View):
    """Bulk assign cuadrillas (one or multiple) to activities via HTMX POST."""
    allowed_roles = ['admin', 'director', 'coordinador', 'ing_residente']

    def post(self, request, *args, **kwargs):
        import json
        from apps.cuadrillas.models import Cuadrilla

        actividad_ids = request.POST.getlist('actividad_ids')
        cuadrilla_ids = request.POST.getlist('cuadrilla_id')  # Handle multiple selections

        if not actividad_ids:
            return JsonResponse({'success': False, 'error': 'No se seleccionaron avisos'}, status=400)
        if not cuadrilla_ids:
            return JsonResponse({'success': False, 'error': 'Debe seleccionar al menos una cuadrilla'}, status=400)

        try:
            cuadrillas = Cuadrilla.objects.filter(id__in=cuadrilla_ids)
            if not cuadrillas.exists():
                return JsonResponse({'success': False, 'error': 'Cuadrillas no encontradas'}, status=404)
        except Exception as e:
            return JsonResponse({'success': False, 'error': f'Error al obtener cuadrillas: {str(e)}'}, status=400)

        updated = 0
        # Asignar todas las cuadrillas seleccionadas a todas las actividades
        for actividad in Actividad.objects.filter(id__in=actividad_ids):
            # Si es una sola cuadrilla, asignarla como cuadrilla principal
            if len(cuadrilla_ids) == 1:
                actividad.cuadrilla = cuadrillas.first()
                actividad.save(update_fields=['cuadrilla', 'updated_at'])

            # Agregar todas las cuadrillas a la relación M2M
            for cuadrilla in cuadrillas:
                actividad.cuadrillas.add(cuadrilla)
            updated += 1

        cuadrilla_nombres = ', '.join([f'{c.codigo}' for c in cuadrillas])

        if request.headers.get('HX-Request'):
            response = HttpResponse()
            response['HX-Trigger'] = json.dumps({
                'showToast': {
                    'message': f'{updated} aviso(s) asignados a cuadrilla(s): {cuadrilla_nombres}',
                    'type': 'success',
                },
                'refreshTable': True,
            })
            return response

        return JsonResponse({'success': True, 'updated': updated})


class BulkCambiarEstadoView(LoginRequiredMixin, RoleRequiredMixin, View):
    """Bulk change status of multiple activities via HTMX POST."""
    allowed_roles = ['admin', 'director', 'coordinador', 'ing_residente', 'supervisor']

    def post(self, request, *args, **kwargs):
        import json

        actividad_ids = request.POST.getlist('actividad_ids')
        nuevo_estado = request.POST.get('estado')

        if not actividad_ids:
            return JsonResponse({'success': False, 'error': 'No se seleccionaron avisos'}, status=400)
        if not nuevo_estado or nuevo_estado not in dict(Actividad.Estado.choices):
            return JsonResponse({'success': False, 'error': 'Estado invalido'}, status=400)

        updated = Actividad.objects.filter(id__in=actividad_ids).update(estado=nuevo_estado)

        if request.headers.get('HX-Request'):
            response = HttpResponse()
            response['HX-Trigger'] = json.dumps({
                'showToast': {
                    'message': f'{updated} aviso(s) cambiados a {dict(Actividad.Estado.choices)[nuevo_estado]}',
                    'type': 'success',
                },
                'refreshTable': True,
            })
            return response

        return JsonResponse({'success': True, 'updated': updated})


class ExportarProgramacionView(LoginRequiredMixin, RoleRequiredMixin, View):
    """View for exporting weekly programming to Excel."""
    allowed_roles = ['admin', 'director', 'coordinador', 'ing_residente', 'supervisor']

    def get(self, request, *args, **kwargs):
        """Generate and download weekly programming Excel."""
        from .exporters import ProgramacionSemanalExporter

        # Obtener parámetros de fecha
        fecha_inicio_str = request.GET.get('fecha_inicio')
        fecha_fin_str = request.GET.get('fecha_fin')
        linea_id = request.GET.get('linea')
        cuadrilla_id = request.GET.get('cuadrilla')

        # Calcular fechas (default: semana actual)
        hoy = date.today()
        if fecha_inicio_str:
            try:
                fecha_inicio = date.fromisoformat(fecha_inicio_str)
            except ValueError:
                fecha_inicio = hoy - timedelta(days=hoy.weekday())  # Lunes de esta semana
        else:
            fecha_inicio = hoy - timedelta(days=hoy.weekday())

        if fecha_fin_str:
            try:
                fecha_fin = date.fromisoformat(fecha_fin_str)
            except ValueError:
                fecha_fin = fecha_inicio + timedelta(days=6)
        else:
            fecha_fin = fecha_inicio + timedelta(days=6)

        # Generar Excel
        exporter = ProgramacionSemanalExporter()
        excel_content = exporter.generar_excel(
            semana_inicio=fecha_inicio,
            semana_fin=fecha_fin,
            linea_id=linea_id,
            cuadrilla_id=cuadrilla_id
        )

        # Preparar respuesta
        filename = f"programacion_semanal_{fecha_inicio.strftime('%Y%m%d')}_{fecha_fin.strftime('%Y%m%d')}.xlsx"
        response = HttpResponse(
            excel_content.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response


class ExportarAvanceView(LoginRequiredMixin, RoleRequiredMixin, View):
    """View for exporting advance report to Excel."""
    allowed_roles = ['admin', 'director', 'coordinador', 'ing_residente']

    def get(self, request, *args, **kwargs):
        """Generate and download advance report Excel."""
        from .exporters import ReporteAvanceExporter

        linea_id = request.GET.get('linea')
        fecha_corte_str = request.GET.get('fecha_corte')

        if not linea_id:
            return JsonResponse({'error': 'Debe especificar una línea'}, status=400)

        fecha_corte = None
        if fecha_corte_str:
            try:
                fecha_corte = date.fromisoformat(fecha_corte_str)
            except ValueError:
                pass

        try:
            exporter = ReporteAvanceExporter()
            excel_content = exporter.generar_excel(linea_id, fecha_corte)
        except ValueError as e:
            return JsonResponse({'error': str(e)}, status=404)

        # Preparar respuesta
        filename = f"reporte_avance_{date.today().strftime('%Y%m%d')}.xlsx"
        response = HttpResponse(
            excel_content.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response


class EventosAPIView(LoginRequiredMixin, View):
    """API endpoint for FullCalendar events."""

    def get(self, request, *args, **kwargs):
        """Return events in FullCalendar format."""
        from datetime import datetime

        # Parse date range from FullCalendar
        start_str = request.GET.get('start', '')
        end_str = request.GET.get('end', '')
        linea_id = request.GET.get('linea')

        # Build queryset
        qs = Actividad.objects.select_related(
            'linea', 'torre', 'tipo_actividad', 'cuadrilla'
        )

        # Filter by date range
        if start_str:
            try:
                start_date = datetime.fromisoformat(start_str.replace('Z', '+00:00')).date()
                qs = qs.filter(fecha_programada__gte=start_date)
            except ValueError:
                pass

        if end_str:
            try:
                end_date = datetime.fromisoformat(end_str.replace('Z', '+00:00')).date()
                qs = qs.filter(fecha_programada__lte=end_date)
            except ValueError:
                pass

        # Filter by linea
        if linea_id:
            from uuid import UUID
            try:
                UUID(linea_id)
                qs = qs.filter(linea_id=linea_id)
            except ValueError:
                pass

        # Build events list for FullCalendar
        events = []
        for actividad in qs:
            # Determine color based on status and priority
            if actividad.prioridad == 'URGENTE':
                color = '#EF4444'  # red-500
            elif actividad.estado == 'COMPLETADA':
                color = '#22C55E'  # green-500
            elif actividad.estado == 'EN_CURSO':
                color = '#EAB308'  # yellow-500
            elif actividad.estado == 'CANCELADA':
                color = '#6B7280'  # gray-500
            else:
                color = '#9CA3AF'  # gray-400

            events.append({
                'id': str(actividad.id),
                'title': f"T{actividad.torre.numero} - {actividad.linea.codigo}",
                'start': actividad.fecha_programada.isoformat(),
                'backgroundColor': color,
                'borderColor': color,
                'extendedProps': {
                    'tipo': actividad.tipo_actividad.nombre,
                    'cuadrilla': actividad.cuadrilla.nombre if actividad.cuadrilla else None,
                    'estado': actividad.get_estado_display(),
                    'prioridad': actividad.get_prioridad_display(),
                }
            })

        return JsonResponse(events, safe=False)


class BuscarAvisoSAPView(LoginRequiredMixin, View):
    """API endpoint to search activities by Aviso SAP number."""

    def get(self, request, *args, **kwargs):
        aviso = request.GET.get('aviso', '').strip()
        if not aviso:
            return JsonResponse({'found': False})

        actividad = Actividad.objects.filter(
            aviso_sap__iexact=aviso
        ).select_related('linea', 'torre', 'tipo_actividad', 'cuadrilla').first()

        if not actividad:
            # Try partial match
            actividad = Actividad.objects.filter(
                aviso_sap__icontains=aviso
            ).select_related('linea', 'torre', 'tipo_actividad', 'cuadrilla').first()

        if not actividad:
            return JsonResponse({'found': False})

        data = {
            'found': True,
            'tipo_actividad_id': str(actividad.tipo_actividad_id),
            'tipo_actividad_nombre': actividad.tipo_actividad.nombre,
            'linea_id': str(actividad.linea_id),
            'linea_nombre': f"{actividad.linea.codigo} - {actividad.linea.nombre}",
            'torre_id': str(actividad.torre_id) if actividad.torre_id else '',
            'torre_numero': actividad.torre.numero if actividad.torre else '',
            'cuadrilla_id': str(actividad.cuadrilla_id) if actividad.cuadrilla_id else '',
            'orden_sap': actividad.orden_sap or '',
            'pt_sap': actividad.pt_sap or '',
            'observaciones': actividad.observaciones_programacion or '',
            'descripcion': str(actividad),
        }
        return JsonResponse(data)


class ActividadDetalleModalView(LoginRequiredMixin, DetailView):
    """Partial view for activity detail in modal."""
    model = Actividad
    template_name = 'actividades/partials/detalle_modal.html'
    context_object_name = 'actividad'

    def get_queryset(self):
        return super().get_queryset().select_related(
            'linea', 'torre', 'tipo_actividad', 'cuadrilla'
        )


class EditarRestriccionesView(LoginRequiredMixin, RoleRequiredMixin, DetailView):
    """Quick edit modal for activity restrictions/comments."""
    model = Actividad
    template_name = 'actividades/partials/modal_restricciones.html'
    context_object_name = 'actividad'
    allowed_roles = ['admin', 'director', 'coordinador', 'ing_residente', 'supervisor']

    def post(self, request, *args, **kwargs):
        import json
        actividad = self.get_object()
        actividad.comentarios_restricciones = request.POST.get('comentarios_restricciones', '')
        actividad.save(update_fields=['comentarios_restricciones', 'updated_at'])

        if request.headers.get('HX-Request'):
            response = HttpResponse()
            response['HX-Trigger'] = json.dumps({
                'showToast': {'message': 'Restricciones actualizadas', 'type': 'success'},
                'closeModal': True,
            })
            return response

        messages.success(request, 'Restricciones actualizadas exitosamente.')
        return redirect('actividades:detalle', pk=actividad.pk)


class ListaOperativaView(LoginRequiredMixin, RoleRequiredMixin, HTMXMixin, ListView):
    """
    Vista de histórico de intervenciones en líneas.
    Agregado: 1 abril 2026

    Muestra todas las intervenciones realizadas con filtros por:
    - Línea
    - Fecha desde/hasta
    - Cuadrilla
    - Tipo de intervención
    """
    model = HistorialIntervencion
    template_name = 'actividades/lista_operativa.html'
    partial_template_name = 'actividades/partials/lista_operativa.html'
    context_object_name = 'intervenciones'
    paginate_by = 50
    allowed_roles = ['admin', 'director', 'coordinador', 'ing_residente', 'supervisor']

    def get_queryset(self):
        qs = super().get_queryset().select_related(
            'linea', 'actividad', 'cuadrilla', 'usuario',
            'torre_inicio', 'torre_fin', 'actividad__tipo_actividad'
        )

        # Filtro por línea
        linea_id = self.request.GET.get('linea')
        if linea_id:
            qs = qs.filter(linea_id=linea_id)

        # Filtro por fecha desde
        fecha_desde = self.request.GET.get('fecha_desde')
        if fecha_desde:
            qs = qs.filter(fecha_intervencion__gte=fecha_desde)

        # Filtro por fecha hasta
        fecha_hasta = self.request.GET.get('fecha_hasta')
        if fecha_hasta:
            from datetime import datetime, time
            # Incluir todo el día hasta las 23:59:59
            fecha_hasta_dt = datetime.strptime(fecha_hasta, '%Y-%m-%d')
            fecha_hasta_dt = datetime.combine(fecha_hasta_dt.date(), time(23, 59, 59))
            qs = qs.filter(fecha_intervencion__lte=fecha_hasta_dt)

        # Filtro por cuadrilla
        cuadrilla_id = self.request.GET.get('cuadrilla')
        if cuadrilla_id:
            qs = qs.filter(cuadrilla_id=cuadrilla_id)

        # Filtro por tipo de intervención
        tipo = self.request.GET.get('tipo')
        if tipo:
            qs = qs.filter(tipo_intervencion__icontains=tipo)

        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Datos para filtros
        from apps.lineas.models import Linea
        from apps.cuadrillas.models import Cuadrilla

        context['lineas'] = Linea.objects.filter(activa=True).order_by('codigo')
        context['cuadrillas'] = Cuadrilla.objects.filter(activa=True).order_by('codigo')

        # Obtener tipos de intervención únicos
        tipos = HistorialIntervencion.objects.values_list(
            'tipo_intervencion', flat=True
        ).distinct().order_by('tipo_intervencion')
        context['tipos_intervencion'] = list(tipos)

        # Valores actuales de filtros
        context['filtros'] = {
            'linea': self.request.GET.get('linea', ''),
            'fecha_desde': self.request.GET.get('fecha_desde', ''),
            'fecha_hasta': self.request.GET.get('fecha_hasta', ''),
            'cuadrilla': self.request.GET.get('cuadrilla', ''),
            'tipo': self.request.GET.get('tipo', ''),
        }

        return context


class DescargarPlantillaProgramacionView(LoginRequiredMixin, RoleRequiredMixin, View):
    """
    Descarga plantilla Excel para carga masiva de actividades.

    Agregado: 13 abril 2026
    """
    allowed_roles = ['admin', 'director', 'coordinador']

    def get(self, request):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from datetime import date

        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Programación"

        # Estilos
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Headers
        headers = [
            'AvísoSAP',                 # A - Aviso SAP (primer campo)
            'Línea',                    # B
            'Torre',                    # C
            'TipoActividad',            # D
            'Fecha',                    # E
            'Cuadrilla',                # F
            'Prioridad',                # G
            'Descripción',              # H
        ]

        # Escribir headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Fila de ejemplo
        ejemplo = [
            '4500001234',                           # AvísoSAP
            'L-838',                                # Línea
            '25',                                   # Torre
            'PODA',                                 # TipoActividad
            date.today().strftime('%Y-%m-%d'),    # Fecha
            'Cuadrilla 01',                        # Cuadrilla
            'ALTA',                                 # Prioridad
            'Poda de vegetación en torre T-25',   # Descripción
        ]

        ws.append(ejemplo)

        # Aplicar estilos a fila de ejemplo
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=2, column=col_num)
            cell.border = thin_border

        # Ajustar anchos de columna
        column_widths = {
            'A': 14,    # AvísoSAP
            'B': 12,    # Línea
            'C': 8,     # Torre
            'D': 18,    # TipoActividad
            'E': 12,    # Fecha
            'F': 16,    # Cuadrilla
            'G': 12,    # Prioridad
            'H': 35,    # Descripción
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Agregar hoja de instrucciones
        ws_instrucciones = wb.create_sheet("Instrucciones")
        ws_instrucciones.column_dimensions['A'].width = 100

        instrucciones = [
            "INSTRUCCIONES PARA CARGA MASIVA DE ACTIVIDADES PROGRAMADAS",
            "",
            "1. FORMATO DE COLUMNAS (en la hoja 'Programación'):",
            "",
            "   A. AvísoSAP (Requerido):",
            "      - Número de aviso en el sistema SAP de Transelca",
            "      - Ejemplo: 4500001234",
            "",
            "   B. Línea (Requerido):",
            "      - Código de la línea de transmisión",
            "      - Debe existir en el sistema",
            "      - Ejemplo: L-838, L-811",
            "",
            "   C. Torre (Requerido):",
            "      - Número de la torre",
            "      - Ejemplo: 25, 30, T-25",
            "",
            "   D. TipoActividad (Requerido):",
            "      - Código del tipo de actividad",
            "      - Ejemplo: PODA, INSPECCION, HERRAJES, AISLADORES",
            "      - Tipos válidos:",
            "        * PODA - Poda de Vegetación",
            "        * HERRAJES - Cambio de Herrajes",
            "        * AISLADORES - Cambio de Aisladores",
            "        * INSPECCION - Inspección General",
            "        * LIMPIEZA - Limpieza",
            "        * SEÑALIZACION - Señalización",
            "        * MEDICION - Medición",
            "        * LAVADO - Lavado Tradicional",
            "        * Y otros tipos disponibles en el sistema",
            "",
            "   E. Fecha (Requerido):",
            "      - Formato: YYYY-MM-DD",
            "      - Ejemplo: 2026-04-15",
            "",
            "   F. Cuadrilla (Opcional):",
            "      - Código de la cuadrilla asignada",
            "      - Ejemplo: Cuadrilla 01, CUA-001",
            "",
            "   G. Prioridad (Opcional):",
            "      - Valores: BAJA, NORMAL, ALTA, URGENTE",
            "      - Default: NORMAL",
            "",
            "   H. Descripción (Opcional):",
            "      - Descripción detallada de la actividad",
            "",
            "2. REQUISITOS:",
            "   - La línea DEBE existir en el sistema",
            "   - La torre DEBE existir en la línea especificada",
            "   - El tipo de actividad DEBE existir en el sistema",
            "   - Si especifica cuadrilla, DEBE existir y estar activa",
            "",
            "3. EJEMPLO DE FILA COMPLETA:",
            "   4500001234 | L-838 | 25 | PODA | 2026-04-15 | Cuadrilla 01 | ALTA | Poda de vegetación",
            "",
            "4. NOTAS IMPORTANTES:",
            "   - No modifique los nombres de las columnas",
            "   - Use el formato de fecha especificado (YYYY-MM-DD)",
            "   - El AvísoSAP es obligatorio en cada fila",
            "   - Las líneas vacías serán ignoradas",
            "   - Si hay errores, se mostrarán advertencias pero el proceso continuará",
            "",
            "5. DESPUÉS DE CARGAR:",
            "   - Verifique que las actividades se crearon correctamente",
            "   - Si hay advertencias, revise los datos",
            "   - Las actividades aparecerán en la programación mensual",
        ]

        for row_num, instruccion in enumerate(instrucciones, 1):
            cell = ws_instrucciones.cell(row=row_num, column=1)
            cell.value = instruccion
            cell.alignment = Alignment(wrap_text=True, vertical="top")

            if row_num == 1:
                cell.font = Font(bold=True, size=14, color="FFFFFF")
                cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            elif instruccion.startswith(('1.', '2.', '3.', '4.', '5.')):
                cell.font = Font(bold=True, size=11)
            elif instruccion.startswith('   -') or instruccion.startswith('   *') or instruccion.startswith('      '):
                cell.font = Font(size=10)

        # Preparar respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'plantilla_programacion_{date.today().strftime("%Y%m%d")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename={filename}'

        wb.save(response)
        return response
