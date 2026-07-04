"""
Importers for activity programming from Excel files.
"""
import logging
from datetime import date
from decimal import Decimal

from django.db import transaction
from openpyxl import load_workbook

logger = logging.getLogger(__name__)


class ProgramaTranselcaImporter:
    """
    Importa Excel de Transelca con columnas:
    - Aviso SAP
    - Línea
    - Tipo Actividad
    - Mes programado
    - Ejecutor (OUTSOURCING)
    - Tramo (opcional)
    - Torre inicio / Torre fin (opcional)
    - Valor facturación (opcional)

    El formato puede variar, por lo que se intenta detectar las columnas
    por sus nombres en la primera fila.
    """

    # Mapeo de nombres de columnas posibles a campos internos
    COLUMN_MAPPINGS = {
        'aviso_sap': ['aviso sap', 'aviso', 'nro aviso', 'numero aviso', 'sap', 'no. aviso',
                      'avisosap', 'avísosap'],  # defensivo: plantillas viejas sin espacio o con tilde
        'linea': ['línea', 'linea', 'line', 'codigo linea', 'código línea'],
        'tipo_actividad': ['tipo actividad', 'actividad', 'tipo', 'tipo de actividad', 'descripcion actividad'],
        'mes': ['mes', 'mes programado', 'fecha programada', 'mes ejecucion'],
        'anio': ['año', 'ano', 'year'],
        'ejecutor': ['ejecutor', 'contratista', 'outsourcing', 'empresa'],
        'tramo': ['tramo', 'sector', 'seccion'],
        'torre_inicio': ['torre inicio', 'torre ini', 'desde torre', 'torre desde'],
        'torre_fin': ['torre fin', 'torre final', 'hasta torre', 'torre hasta'],
        'valor_facturacion': ['valor', 'valor facturacion', 'facturacion', 'precio', 'monto'],
        'observaciones': ['observaciones', 'notas', 'comentarios', 'obs'],
    }

    def __init__(self):
        self.errores = []
        self.advertencias = []
        self.actividades_creadas = []
        self.actividades_actualizadas = []
        self.filas_omitidas = []
        self.column_indices = {}

    def importar(self, archivo_excel, programacion_mensual, opciones=None):
        """
        Importa actividades desde un archivo Excel de Transelca.

        Args:
            archivo_excel: File object or path to Excel file
            programacion_mensual: ProgramacionMensual instance to associate activities
            opciones: Dict with import options (actualizar_existentes, etc.)

        Returns:
            Dict with import summary
        """

        from .models import Actividad

        opciones = opciones or {}
        actualizar_existentes = opciones.get('actualizar_existentes', False)

        try:
            workbook = load_workbook(archivo_excel, read_only=True, data_only=True)
            sheet = workbook.active
        except Exception as e:
            logger.error(f"Error loading Excel file: {e}")
            return {
                'exito': False,
                'error': f'Error al cargar archivo Excel: {str(e)}',
                'actividades_creadas': 0,
                'actividades_actualizadas': 0,
            }

        # Detectar columnas en la primera fila
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return {
                'exito': False,
                'error': 'El archivo está vacío',
                'actividades_creadas': 0,
                'actividades_actualizadas': 0,
            }

        header_row = rows[0]
        self._detectar_columnas(header_row)

        if 'linea' not in self.column_indices:
            return {
                'exito': False,
                'error': 'No se encontró la columna de Línea en el archivo',
                'actividades_creadas': 0,
                'actividades_actualizadas': 0,
            }

        if 'tipo_actividad' not in self.column_indices:
            return {
                'exito': False,
                'error': 'No se encontró la columna de Tipo de Actividad en el archivo',
                'actividades_creadas': 0,
                'actividades_actualizadas': 0,
            }

        # Procesar filas de datos
        linea_asociada = programacion_mensual.linea

        with transaction.atomic():
            for row_num, row in enumerate(rows[1:], start=2):
                try:
                    resultado = self._procesar_fila(
                        row, row_num, programacion_mensual, linea_asociada, actualizar_existentes
                    )
                    if resultado == 'creada':
                        self.actividades_creadas.append(row_num)
                    elif resultado == 'actualizada':
                        self.actividades_actualizadas.append(row_num)
                    elif resultado == 'omitida':
                        self.filas_omitidas.append(row_num)
                except Exception as e:
                    logger.warning(f"Error processing row {row_num}: {e}")
                    self.errores.append({
                        'fila': row_num,
                        'error': str(e)
                    })

        # Actualizar programación mensual
        programacion_mensual.total_actividades = Actividad.objects.filter(
            programacion=programacion_mensual
        ).count()
        programacion_mensual.save(update_fields=['total_actividades', 'updated_at'])

        return {
            'exito': True,
            'actividades_creadas': len(self.actividades_creadas),
            'actividades_actualizadas': len(self.actividades_actualizadas),
            'filas_omitidas': len(self.filas_omitidas),
            'errores': self.errores,
            'advertencias': self.advertencias,
            'columnas_detectadas': list(self.column_indices.keys()),
        }

    def _detectar_columnas(self, header_row):
        """Detecta las columnas en la fila de encabezado."""
        for col_idx, cell_value in enumerate(header_row):
            if cell_value is None:
                continue
            cell_lower = str(cell_value).lower().strip()

            for field_name, posibles in self.COLUMN_MAPPINGS.items():
                if cell_lower in posibles:
                    self.column_indices[field_name] = col_idx
                    break

        logger.info(f"Detected columns: {self.column_indices}")

    def _get_cell_value(self, row, field_name):
        """Obtiene el valor de una celda por nombre de campo."""
        if field_name not in self.column_indices:
            return None
        idx = self.column_indices[field_name]
        if idx < len(row):
            return row[idx]
        return None

    def _detectar_fecha_excel(self, data_rows):
        """Detecta año y mes del Excel inspeccionando datos."""
        anio = None
        mes = None

        for row in data_rows[:5]:
            if 'anio' in self.column_indices:
                val = self._get_cell_value(row, 'anio')
                if val:
                    try:
                        anio = int(val)
                        if 2000 < anio < 2100:
                            break
                    except (ValueError, TypeError):
                        pass

            if 'mes' in self.column_indices:
                val = self._get_cell_value(row, 'mes')
                if val:
                    try:
                        mes = int(val)
                        if 1 <= mes <= 12:
                            break
                    except (ValueError, TypeError):
                        pass

        return anio, mes

    def _procesar_fila(self, row, row_num, programacion_mensual, linea_asociada, actualizar_existentes):
        """Procesa una fila del Excel y crea/actualiza la actividad."""
        from apps.lineas.models import Linea, Torre, Tramo

        from .models import Actividad, TipoActividad

        # Obtener valores de la fila
        aviso_sap = self._get_cell_value(row, 'aviso_sap')
        linea_codigo = self._get_cell_value(row, 'linea')
        tipo_actividad_nombre = self._get_cell_value(row, 'tipo_actividad')
        tramo_codigo = self._get_cell_value(row, 'tramo')
        torre_inicio_num = self._get_cell_value(row, 'torre_inicio')
        _torre_fin_num = self._get_cell_value(row, 'torre_fin')  # noqa: F841
        valor_facturacion = self._get_cell_value(row, 'valor_facturacion')
        observaciones = self._get_cell_value(row, 'observaciones')

        # Validaciones básicas
        if not linea_codigo and not linea_asociada:
            self.advertencias.append({
                'fila': row_num,
                'mensaje': 'No se especificó línea y no hay línea asociada a la programación'
            })
            return 'omitida'

        if not tipo_actividad_nombre:
            self.advertencias.append({
                'fila': row_num,
                'mensaje': 'No se especificó tipo de actividad'
            })
            return 'omitida'

        # Buscar línea
        linea = linea_asociada
        if linea_codigo:
            try:
                linea = Linea.objects.get(codigo__iexact=str(linea_codigo).strip())
            except Linea.DoesNotExist:
                self.advertencias.append({
                    'fila': row_num,
                    'mensaje': f'Línea no encontrada: {linea_codigo}'
                })
                if not linea_asociada:
                    return 'omitida'

        # Buscar tipo de actividad
        try:
            tipo_actividad = TipoActividad.objects.get(
                nombre__iexact=str(tipo_actividad_nombre).strip()
            )
        except TipoActividad.DoesNotExist:
            # Intentar búsqueda parcial
            tipos = TipoActividad.objects.filter(
                nombre__icontains=str(tipo_actividad_nombre).strip()
            )
            if tipos.exists():
                tipo_actividad = tipos.first()
                self.advertencias.append({
                    'fila': row_num,
                    'mensaje': f'Tipo de actividad "{tipo_actividad_nombre}" mapeado a "{tipo_actividad.nombre}"'
                })
            else:
                self.advertencias.append({
                    'fila': row_num,
                    'mensaje': f'Tipo de actividad no encontrado: {tipo_actividad_nombre}'
                })
                return 'omitida'

        # Buscar tramo (opcional)
        tramo = None
        if tramo_codigo:
            try:
                tramo = Tramo.objects.get(codigo__iexact=str(tramo_codigo).strip())
            except Tramo.DoesNotExist:
                self.advertencias.append({
                    'fila': row_num,
                    'mensaje': f'Tramo no encontrado: {tramo_codigo}'
                })

        # Buscar torre (opcional, usa la torre de inicio del tramo si hay tramo)
        torre = None
        if tramo:
            torre = tramo.torre_inicio
        elif torre_inicio_num:
            try:
                torre = Torre.objects.get(
                    linea=linea,
                    numero=str(torre_inicio_num).strip()
                )
            except Torre.DoesNotExist:
                self.advertencias.append({
                    'fila': row_num,
                    'mensaje': f'Torre no encontrada: {torre_inicio_num}'
                })

        # Preparar valor de facturación
        valor_fact = Decimal('0')
        if valor_facturacion:
            try:
                valor_fact = Decimal(str(valor_facturacion).replace(',', '.').replace('$', '').strip())
            except (ValueError, TypeError):
                self.advertencias.append({
                    'fila': row_num,
                    'mensaje': f'Valor de facturación inválido: {valor_facturacion}'
                })

        # Verificar si ya existe (por aviso SAP)
        actividad_existente = None
        if aviso_sap:
            aviso_sap_str = str(aviso_sap).strip()
            try:
                actividad_existente = Actividad.objects.get(aviso_sap=aviso_sap_str)
            except Actividad.DoesNotExist:
                pass

        if actividad_existente:
            if actualizar_existentes:
                # Actualizar actividad existente
                actividad_existente.linea = linea
                actividad_existente.tipo_actividad = tipo_actividad
                actividad_existente.torre = torre
                actividad_existente.tramo = tramo
                actividad_existente.programacion = programacion_mensual
                if valor_fact > 0:
                    actividad_existente.valor_facturacion = valor_fact
                if observaciones:
                    actividad_existente.observaciones_programacion = str(observaciones)
                actividad_existente.save()
                return 'actualizada'
            else:
                self.advertencias.append({
                    'fila': row_num,
                    'mensaje': f'Actividad con Aviso SAP {aviso_sap} ya existe, omitiendo'
                })
                return 'omitida'

        # Crear nueva actividad
        Actividad.objects.create(
            linea=linea,
            torre=torre,
            tipo_actividad=tipo_actividad,
            programacion=programacion_mensual,
            tramo=tramo,
            aviso_sap=str(aviso_sap).strip() if aviso_sap else '',
            fecha_programada=date(programacion_mensual.anio, programacion_mensual.mes, 1),
            estado=Actividad.Estado.PENDIENTE,
            prioridad=Actividad.Prioridad.NORMAL,
            valor_facturacion=valor_fact,
            observaciones_programacion=str(observaciones) if observaciones else '',
        )

        return 'creada'


class AvisosTranselcaImporter:
    """
    Importador especializado para el archivo de Avisos Abiertos de Transelca.

    Formato esperado del Excel:
    - LINEA: Código de línea (L-838, L-5156, etc.)
    - CIRCUITO: Número de circuito
    - TIPO: Categoría de actividad
    - AVISO: Número de aviso SAP
    - PT SAP: Puesto de trabajo SAP
    - CENTRO EMPL.: Centro de emplazamiento
    - Columnas de equipos: cantidad de elementos por tipo de actividad
    """

    # Mapeo de categorías del Excel a TipoActividad
    CATEGORIA_MAPPING = {
        'poda': 'PODA',
        'lavado tradicional': 'LAVADO',
        'lavado': 'LAVADO',
        'servidumbre': 'SERVIDUMBRE',
        'corredor eléctrico': 'CORREDOR',
        'corredor electrico': 'CORREDOR',
        'gestionar permiso': 'PERMISO',
        'permiso': 'PERMISO',
        'inspección pedestre': 'INSPECCION_PED',
        'inspeccion pedestre': 'INSPECCION_PED',
        'termografía': 'TERMOGRAFIA',
        'termografia': 'TERMOGRAFIA',
        'descargas parciales': 'DESCARGAS',
        'mtto electromecánico': 'ELECTROMEC',
        'mtto electromecanico': 'ELECTROMEC',
        'medida puesta tierra': 'MEDICION_PT',
        'medición': 'MEDICION',
        'medicion': 'MEDICION',
        'limpieza': 'LIMPIEZA',
        'cambio herrajes': 'HERRAJES',
        'herrajes': 'HERRAJES',
        'cambio aisladores': 'AISLADORES',
        'aisladores': 'AISLADORES',
        'señalización': 'SEÑALIZACION',
        'señalizacion': 'SEÑALIZACION',
    }

    # Columnas esperadas
    COLUMN_MAPPINGS = {
        'linea': ['linea', 'línea', 'line', 'codigo_linea'],
        'circuito': ['circuito', 'cto', 'circuit'],
        'tipo': ['tipo', 'tipo actividad', 'actividad', 'categoria'],
        'aviso': ['aviso', 'aviso sap', 'nro aviso', 'no. aviso'],
        'pt_sap': ['pt sap', 'pt', 'puesto trabajo', 'puesto de trabajo'],
        'centro_empl': ['centro empl', 'centro empl.', 'centro emplazamiento', 'ce'],
        'contratista': ['contratista', 'ejecutor', 'outsourcing', 'empresa'],
        'torre_inicio': ['torre inicio', 'desde', 'torre desde'],
        'torre_fin': ['torre fin', 'hasta', 'torre hasta'],
        'descripcion': ['descripcion', 'descripción', 'descripcion actividad'],
        'anio': ['año', 'año fin prioridad', 'ano', 'year'],
        'mes': ['mes', 'mes ejecucion', 'mes\nejecu', 'month'],
    }

    def __init__(self):
        self.errores = []
        self.advertencias = []
        self.actividades_creadas = []
        self.actividades_actualizadas = []
        self.actividades_omitidas = []
        self.column_indices = {}

    def importar(self, archivo_excel, anio: int = None, mes: int = None, opciones=None):
        """
        Importa avisos desde el archivo Excel de Transelca.

        Args:
            archivo_excel: File object or path to Excel file
            anio: Año de la programación (opcional, se detecta del Excel si no se proporciona)
            mes: Mes de la programación (opcional, se detecta del Excel si no se proporciona)
            opciones: Dict with import options

        Returns:
            Dict with import summary
        """
        from apps.lineas.models import Linea
        from datetime import date

        from .models import TipoActividad

        opciones = opciones or {}
        actualizar_existentes = opciones.get('actualizar_existentes', False)
        crear_lineas = opciones.get('crear_lineas', False)

        try:
            workbook = load_workbook(archivo_excel, read_only=True, data_only=True)
            sheet = workbook.active
        except Exception as e:
            logger.error(f"Error loading Excel file: {e}")
            return {
                'exito': False,
                'error': f'Error al cargar archivo Excel: {str(e)}',
                'actividades_creadas': 0,
            }

        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return {'exito': False, 'error': 'El archivo está vacío'}

        # Detectar columnas
        header_row = rows[0]
        self._detectar_columnas(header_row)

        if 'aviso' not in self.column_indices:
            return {
                'exito': False,
                'error': 'No se encontró la columna de Aviso SAP',
            }

        # Detectar año y mes del Excel si no se proporcionan
        if not anio or not mes:
            anio_excel, mes_excel = self._detectar_fecha_excel(rows[1:])
            if anio_excel and mes_excel:
                anio = anio or anio_excel
                mes = mes or mes_excel

        # Usar valores por defecto si no se encuentran
        if not anio or not mes:
            hoy = date.today()
            anio = anio or hoy.year
            mes = mes or hoy.month

        # Cache de líneas y tipos de actividad
        lineas_cache = {linea_obj.codigo: linea_obj for linea_obj in Linea.objects.all()}
        tipos_cache = {t.categoria: t for t in TipoActividad.objects.filter(activo=True)}

        # También crear cache por nombre
        tipos_nombre_cache = {t.nombre.lower(): t for t in TipoActividad.objects.filter(activo=True)}

        with transaction.atomic():
            for row_num, row in enumerate(rows[1:], start=2):
                try:
                    resultado = self._procesar_fila(
                        row, row_num, anio, mes,
                        lineas_cache, tipos_cache, tipos_nombre_cache,
                        actualizar_existentes, crear_lineas
                    )
                    if resultado == 'creada':
                        self.actividades_creadas.append(row_num)
                    elif resultado == 'actualizada':
                        self.actividades_actualizadas.append(row_num)
                    elif resultado == 'omitida':
                        self.actividades_omitidas.append(row_num)
                except Exception as e:
                    logger.warning(f"Error processing row {row_num}: {e}")
                    self.errores.append({'fila': row_num, 'error': str(e)})

        return {
            'exito': True,
            'actividades_creadas': len(self.actividades_creadas),
            'actividades_actualizadas': len(self.actividades_actualizadas),
            'actividades_omitidas': len(self.actividades_omitidas),
            'errores': self.errores,
            'advertencias': self.advertencias,
            'columnas_detectadas': list(self.column_indices.keys()),
        }

    def _detectar_columnas(self, header_row):
        """Detecta las columnas en la fila de encabezado."""
        for col_idx, cell_value in enumerate(header_row):
            if cell_value is None:
                continue
            cell_lower = str(cell_value).lower().strip()

            for field_name, posibles in self.COLUMN_MAPPINGS.items():
                if cell_lower in posibles:
                    self.column_indices[field_name] = col_idx
                    break

        logger.info(f"Avisos importer detected columns: {self.column_indices}")

    def _detectar_fecha_excel(self, data_rows):
        """Detecta año y mes del Excel inspeccionando datos."""
        anio = None
        mes = None

        for row in data_rows[:5]:  # Revisar primeras 5 filas
            if 'anio' in self.column_indices:
                val = self._get_cell(row, 'anio')
                if val:
                    try:
                        anio = int(val)
                        if 2000 < anio < 2100:
                            break
                    except (ValueError, TypeError):
                        pass

            if 'mes' in self.column_indices:
                val = self._get_cell(row, 'mes')
                if val:
                    try:
                        mes = int(val)
                        if 1 <= mes <= 12:
                            break
                    except (ValueError, TypeError):
                        pass

        return anio, mes

    def _get_cell(self, row, field_name):
        """Obtiene el valor de una celda."""
        if field_name not in self.column_indices:
            return None
        idx = self.column_indices[field_name]
        if idx < len(row):
            return row[idx]
        return None

    def _mapear_categoria(self, tipo_str):
        """Mapea el tipo del Excel a la categoría del modelo."""
        if not tipo_str:
            return None
        tipo_lower = str(tipo_str).lower().strip()
        return self.CATEGORIA_MAPPING.get(tipo_lower)

    def _procesar_fila(self, row, row_num, anio, mes, lineas_cache, tipos_cache,
                       tipos_nombre_cache, actualizar_existentes, crear_lineas):
        """Procesa una fila del Excel."""

        from .models import Actividad, ProgramacionMensual

        aviso_sap = self._get_cell(row, 'aviso')
        if not aviso_sap:
            return 'omitida'

        aviso_sap = str(aviso_sap).strip()

        linea_codigo = self._get_cell(row, 'linea')
        tipo_str = self._get_cell(row, 'tipo')
        pt_sap = self._get_cell(row, 'pt_sap')
        # centro_empl and circuito read for future use
        _centro_empl = self._get_cell(row, 'centro_empl')  # noqa: F841
        _circuito = self._get_cell(row, 'circuito')  # noqa: F841
        descripcion = self._get_cell(row, 'descripcion')

        # Buscar línea
        linea = None
        if linea_codigo:
            linea_codigo_str = str(linea_codigo).strip()
            linea = lineas_cache.get(linea_codigo_str)
            if not linea:
                # Intentar búsqueda flexible
                for codigo, linea_obj in lineas_cache.items():
                    if linea_codigo_str in codigo or codigo in linea_codigo_str:
                        linea = linea_obj
                        break

            if not linea:
                self.advertencias.append({
                    'fila': row_num,
                    'mensaje': f'Línea no encontrada: {linea_codigo_str}'
                })
                return 'omitida'

        # Buscar tipo de actividad
        tipo_actividad = None
        if tipo_str:
            categoria = self._mapear_categoria(tipo_str)
            if categoria and categoria in tipos_cache:
                tipo_actividad = tipos_cache[categoria]
            else:
                # Buscar por nombre
                tipo_lower = str(tipo_str).lower().strip()
                for nombre, t in tipos_nombre_cache.items():
                    if tipo_lower in nombre or nombre in tipo_lower:
                        tipo_actividad = t
                        break

        if not tipo_actividad:
            # Usar tipo genérico
            tipo_actividad = tipos_cache.get('OTRO')
            if tipo_actividad:
                self.advertencias.append({
                    'fila': row_num,
                    'mensaje': f'Tipo "{tipo_str}" mapeado a "Otro"'
                })

        if not linea or not tipo_actividad:
            self.advertencias.append({
                'fila': row_num,
                'mensaje': 'Faltan datos requeridos (línea o tipo)'
            })
            return 'omitida'

        # Buscar o crear programación mensual
        programacion, _ = ProgramacionMensual.objects.get_or_create(
            anio=anio,
            mes=mes,
            linea=linea,
            defaults={'total_actividades': 0}
        )

        # Buscar torre por defecto (primera de la línea)
        torre = linea.torres.first()

        # Verificar si existe
        try:
            actividad = Actividad.objects.get(aviso_sap=aviso_sap)
            if actualizar_existentes:
                actividad.linea = linea
                actividad.tipo_actividad = tipo_actividad
                actividad.pt_sap = str(pt_sap).strip() if pt_sap else ''
                actividad.programacion = programacion
                if descripcion:
                    actividad.observaciones_programacion = str(descripcion)
                actividad.save()
                return 'actualizada'
            return 'omitida'
        except Actividad.DoesNotExist:
            pass

        # Crear actividad
        Actividad.objects.create(
            linea=linea,
            torre=torre,
            tipo_actividad=tipo_actividad,
            programacion=programacion,
            aviso_sap=aviso_sap,
            pt_sap=str(pt_sap).strip() if pt_sap else '',
            fecha_programada=date(anio, mes, 1),
            estado=Actividad.Estado.PENDIENTE,
            prioridad=Actividad.Prioridad.NORMAL,
            observaciones_programacion=str(descripcion) if descripcion else '',
        )

        return 'creada'


class ImportadorExcelGenerico:
    """
    Importador genérico para diferentes formatos de Excel.
    Útil para importar datos de otras fuentes.
    """

    def __init__(self, mapping_columnas=None):
        self.mapping_columnas = mapping_columnas or {}
        self.errores = []
        self.registros_procesados = 0

    def leer_excel(self, archivo_excel, hoja=None):
        """
        Lee un archivo Excel y retorna los datos como lista de diccionarios.

        Args:
            archivo_excel: File object or path
            hoja: Nombre de la hoja (None = hoja activa)

        Returns:
            List of dicts with row data
        """
        try:
            workbook = load_workbook(archivo_excel, read_only=True, data_only=True)
            if hoja:
                sheet = workbook[hoja]
            else:
                sheet = workbook.active
        except Exception as e:
            logger.error(f"Error loading Excel: {e}")
            return []

        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []

        headers = [str(h).strip() if h else f'col_{i}' for i, h in enumerate(rows[0])]
        data = []

        for row in rows[1:]:
            row_dict = {}
            for i, value in enumerate(row):
                if i < len(headers):
                    col_name = headers[i]
                    # Aplicar mapping si existe
                    if col_name.lower() in self.mapping_columnas:
                        col_name = self.mapping_columnas[col_name.lower()]
                    row_dict[col_name] = value
            data.append(row_dict)

        return data


class AvancesImporter:
    """
    Importa avances de actividades desde Excel con detección flexible de columnas.
    Busca actividades por aviso_sap y actualiza porcentaje_avance, estado, observaciones.
    """

    COLUMN_MAPPINGS = {
        'aviso_sap': ['aviso sap', 'aviso', 'nro aviso', 'numero aviso', 'sap', 'no. aviso', 'aviso_sap', 'id aviso'],
        'porcentaje_avance': ['avance', 'porcentaje avance', '% avance', 'avance %', 'porcentaje', 'progreso', '% progreso', 'porcentaje_avance'],
        'estado': ['estado', 'estado actividad', 'status'],
        'observaciones': ['observaciones', 'notas', 'comentarios', 'obs', 'nota'],
    }

    ESTADO_MAPPING = {
        'pendiente': 'PENDIENTE',
        'programada': 'PROGRAMADA',
        'en curso': 'EN_CURSO',
        'en_curso': 'EN_CURSO',
        'completada': 'COMPLETADA',
        'cancelada': 'CANCELADA',
        'reprogramada': 'REPROGRAMADA',
    }

    def __init__(self):
        self.errores = []
        self.advertencias = []
        self.actividades_actualizadas = []
        self.filas_omitidas = []
        self.column_indices = {}

    def importar(self, archivo_excel):
        """Importa avances desde un archivo Excel."""
        try:
            workbook = load_workbook(archivo_excel, read_only=True, data_only=True)
            sheet = workbook.active
        except Exception as e:
            logger.error(f"Error loading Excel file: {e}")
            return {'exito': False, 'error': f'Error al cargar archivo Excel: {str(e)}'}

        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return {'exito': False, 'error': 'El archivo está vacío'}

        self._detectar_columnas(rows[0])

        if 'aviso_sap' not in self.column_indices:
            columnas_recibidas = [str(h) for h in rows[0] if h]
            return {
                'exito': False,
                'error': f'No se encontró columna de Aviso SAP. Columnas detectadas: {", ".join(columnas_recibidas)}',
                'columnas_recibidas': columnas_recibidas,
            }

        with transaction.atomic():
            for row_num, row in enumerate(rows[1:], start=2):
                try:
                    self._procesar_fila(row, row_num)
                except Exception as e:
                    logger.warning(f"Error processing row {row_num}: {e}")
                    self.errores.append({'fila': row_num, 'error': str(e)})

        return {
            'exito': True,
            'actividades_actualizadas': len(self.actividades_actualizadas),
            'filas_omitidas': len(self.filas_omitidas),
            'errores': self.errores,
            'advertencias': self.advertencias,
            'columnas_detectadas': list(self.column_indices.keys()),
        }

    def _detectar_columnas(self, header_row):
        """Detecta las columnas en la fila de encabezado."""
        for col_idx, cell_value in enumerate(header_row):
            if cell_value is None:
                continue
            cell_lower = str(cell_value).lower().strip()
            for field_name, posibles in self.COLUMN_MAPPINGS.items():
                if cell_lower in posibles and field_name not in self.column_indices:
                    self.column_indices[field_name] = col_idx
                    break

        logger.info(f"Avances importer detected columns: {self.column_indices}")

    def _get_cell(self, row, field_name):
        """Obtiene el valor de una celda por nombre de campo."""
        if field_name not in self.column_indices:
            return None
        idx = self.column_indices[field_name]
        return row[idx] if idx < len(row) else None

    def _procesar_fila(self, row, row_num):
        """Procesa una fila del Excel y actualiza la actividad."""
        from .models import Actividad

        aviso_sap = self._get_cell(row, 'aviso_sap')
        if not aviso_sap:
            self.filas_omitidas.append(row_num)
            return

        aviso_sap = str(aviso_sap).strip()
        try:
            actividad = Actividad.objects.get(aviso_sap=aviso_sap)
        except Actividad.DoesNotExist:
            self.advertencias.append({
                'fila': row_num,
                'mensaje': f'Aviso SAP no encontrado: {aviso_sap}'
            })
            self.filas_omitidas.append(row_num)
            return

        actualizado = False

        porcentaje = self._get_cell(row, 'porcentaje_avance')
        if porcentaje is not None:
            try:
                pct = Decimal(str(porcentaje).replace('%', '').replace(',', '.').strip())
                actividad.actualizar_avance(pct)
                actualizado = True
            except (ValueError, TypeError):
                self.advertencias.append({
                    'fila': row_num,
                    'mensaje': f'Porcentaje inválido para SAP {aviso_sap}: {porcentaje}'
                })

        estado_raw = self._get_cell(row, 'estado')
        if estado_raw:
            estado_mapped = self.ESTADO_MAPPING.get(str(estado_raw).lower().strip())
            if estado_mapped and estado_mapped in dict(Actividad.Estado.choices):
                actividad.estado = estado_mapped
                actividad.save(update_fields=['estado', 'updated_at'])
                actualizado = True

        obs = self._get_cell(row, 'observaciones')
        if obs:
            actividad.observaciones_programacion = str(obs)
            actividad.save(update_fields=['observaciones_programacion', 'updated_at'])
            actualizado = True

        if actualizado:
            self.actividades_actualizadas.append(aviso_sap)


class ProgramacionSemanalImporter:
    """
    Importa Excel de programación semanal con el formato real de Instelec.

    Formato esperado:
    - Una pestaña por semana (nombres `02`, `03`, ... `S18`); pestañas no
      numéricas se ignoran (ej. `vc`, `Hoja1`).
    - Row 0: banner con 'Fecha de envio:' en algún punto y una fecha.
    - Row 1: encabezados ('#', 'ACTIVIDAD', 'LINEA', 'TRAMO', 'INICIO',
      'FIN', 'PERSONAL', 'CEDULA', 'CELULAR', 'CARGO', 'ROL', 'PLACA',
      'AVISOS', 'ORDEN', 'PT SAP', 'Comentarios').
    - Row 2+: filas de datos. Una fila con `#` no vacío es ACTIVIDAD; las
      filas siguientes con `#` vacío son MIEMBROS de la cuadrilla.
    - AVISOS/ORDEN/LINEA pueden contener múltiples valores separados por
      saltos de línea.

    Genera una `Actividad` por cada aviso × línea expandida. Asigna las
    cuadrillas detectadas vía cédulas de los miembros listados.
    """

    COLUMN_MAPPINGS = {
        'numero':         ['#', 'no', 'no.', 'item'],
        'tipo_actividad': ['actividad', 'tipo actividad', 'tipo de actividad'],
        'linea':          ['linea', 'línea', 'lineas', 'líneas'],
        'tramo':          ['tramo', 'tramos', 'sector', 'sección'],
        'fecha_inicio':   ['inicio', 'fecha inicio', 'fecha de inicio'],
        'fecha_fin':      ['fin', 'fecha fin', 'fecha de fin'],
        'personal':       ['personal', 'nombre', 'nombres'],
        'cedula':         ['cedula', 'cédula', 'documento', 'identificacion'],
        'cargo':          ['cargo', 'rol cuadrilla'],
        'rol':            ['rol'],
        'avisos':         ['avisos', 'aviso', 'aviso sap'],
        'orden':          ['orden', 'orden sap', 'ot'],
        'pt_sap':         ['pt sap', 'pt', 'puesto trabajo', 'puesto de trabajo'],
        'observaciones':  ['comentarios', 'observaciones', 'obs', 'notas'],
    }

    SHEETS_EXCLUIR = {'vc', 'hoja1', 'sheet1', 'resumen', 'instrucciones'}

    def __init__(self):
        self.errores = []
        self.advertencias = []
        self.actividades_creadas = []
        self.actividades_actualizadas = []
        self.programaciones_tocadas = set()  # set[(anio, mes, linea_id)]
        self.resumen_por_hoja = {}
        self.column_indices = {}
        # Issue #178 (A2): filas de la sección NOVEDADES, independientes de
        # cualquier actividad — ver `_guardar_novedad`.
        self.novedades_pendientes = []
        self.novedades_creadas = 0

    def importar(self, archivo_excel, opciones=None):
        """
        Recorre todas las pestañas válidas del Excel y crea actividades.

        Args:
            archivo_excel: file-like or path
            opciones: dict — actualizar_existentes (bool)

        Returns:
            dict con resumen.
        """
        from apps.lineas.models import Linea  # noqa: F401  (lookup arriba)

        opciones = opciones or {}

        try:
            # Issue #178 (A1): NO usar read_only=True — openpyxl no expone
            # `Worksheet.merged_cells` en modo read_only (AttributeError), y el
            # parser de bloques necesita el RANGO real de la celda combinada de
            # columna A/D para detectar el fin de bloque de forma 100% confiable
            # (antes se dependía implícitamente de que la celda "#" viniera en
            # blanco por el merge, sin verificar el merge en sí).
            workbook = load_workbook(archivo_excel, data_only=True)
        except Exception as e:
            logger.error(f"Error loading Excel file: {e}")
            return self._resultado_error(f'Error al cargar archivo Excel: {e}')

        sheets_procesadas = []
        for sheet_name in workbook.sheetnames:
            if not self._es_hoja_semanal(sheet_name):
                logger.info(f"Hoja '{sheet_name}' omitida (no es semana válida)")
                continue
            try:
                semana = self._numero_semana(sheet_name)
                resumen_hoja = self._procesar_hoja(workbook[sheet_name], opciones, semana)
                self.resumen_por_hoja[sheet_name] = resumen_hoja
                sheets_procesadas.append(sheet_name)
            except Exception as e:
                logger.exception(f"Error procesando hoja {sheet_name!r}")
                self.errores.append({'hoja': sheet_name, 'error': str(e)})

        # Refresco contadores en cada ProgramacionMensual tocada
        from .models import Actividad, ProgramacionMensual
        for (anio, mes, linea_id) in self.programaciones_tocadas:
            ProgramacionMensual.objects.filter(
                anio=anio, mes=mes, linea_id=linea_id
            ).update(
                total_actividades=Actividad.objects.filter(
                    programacion__anio=anio,
                    programacion__mes=mes,
                    programacion__linea_id=linea_id,
                ).count()
            )

        return {
            'exito': True,
            'sheets_procesadas': sheets_procesadas,
            'actividades_creadas': len(self.actividades_creadas),
            'actividades_actualizadas': len(self.actividades_actualizadas),
            'novedades_creadas': self.novedades_creadas,
            'errores': self.errores,
            'advertencias': self.advertencias,
            'resumen_por_hoja': self.resumen_por_hoja,
        }

    # -- helpers ---------------------------------------------------------

    @staticmethod
    def _es_hoja_semanal(sheet_name):
        import re
        nombre = sheet_name.strip().lower()
        if nombre in ProgramacionSemanalImporter.SHEETS_EXCLUIR:
            return False
        # Aceptar '02', '18', 'S18', 'Semana 18', 'semana_05', '12 (2)', '18 (1)'
        # B5 fix: las copias de hojas que Excel nombra '12 (2)' SON
        # válidas (mismo formato de programación). Antes el regex las
        # excluía silenciosamente y se perdían filas de avisos.
        return bool(re.fullmatch(r's?(emana)?[\s_]*\d+(\s*\(\d+\))?', nombre))

    @staticmethod
    def _mapa_bloques_por_merge(sheet, col_numero, col_alt=None):
        """Mapa ``{fila_excel_1based: 'inicio'|'continuacion'}`` derivado del
        RANGO de la celda combinada en la columna ``numero`` (o ``col_alt``,
        p.ej. ``tramo``/columna D, como respaldo) de ``sheet`` (issue #178, A1).

        Requiere que ``sheet`` se haya cargado SIN ``read_only=True`` —
        openpyxl no expone ``merged_cells`` en modo read_only.

        Devuelve ``{}`` si no se encontró ningún merge de ≥2 filas en esas
        columnas (hoja "plana", sin formato de bloques) — el llamador debe
        entonces usar el heurístico de respaldo (columna '#' en blanco).
        """
        mapa = {}
        columnas = [c for c in (col_numero, col_alt) if c is not None]
        merges = getattr(sheet, 'merged_cells', None)
        if merges is None:
            return mapa
        for col_idx in columnas:
            col_excel = col_idx + 1  # openpyxl es 1-based
            encontrados = False
            for rango in merges.ranges:
                if (
                    rango.min_col == col_excel
                    and rango.max_col == col_excel
                    and rango.max_row > rango.min_row
                ):
                    encontrados = True
                    for fila in range(rango.min_row, rango.max_row + 1):
                        mapa[fila] = 'inicio' if fila == rango.min_row else 'continuacion'
            if encontrados:
                break
        return mapa

    @staticmethod
    def _es_fila_continuacion(row_idx, numero_str, mapa_bloques, usa_fallback_legado):
        """¿La fila `row_idx` es continuación del bloque anterior (miembro
        adicional) o el encabezado de una actividad nueva? (issue #178, A1)

        Primario: RANGO de la celda combinada (`mapa_bloques`). Respaldo
        (`usa_fallback_legado=True`, o fila fuera de cualquier merge conocido
        — p.ej. bloque de 1 sola fila que no genera merge): heurístico legado
        columna '#' en blanco = continuación.
        """
        if not usa_fallback_legado:
            estado = mapa_bloques.get(row_idx)
            if estado == 'continuacion':
                return True
            if estado == 'inicio':
                return False
        return numero_str == ''

    def _resultado_error(self, mensaje):
        return {
            'exito': False,
            'error': mensaje,
            'actividades_creadas': 0,
            'actividades_actualizadas': 0,
            'novedades_creadas': 0,
            'errores': self.errores,
            'advertencias': self.advertencias,
        }

    @staticmethod
    def _numero_semana(sheet_name):
        """Extrae el número de semana del nombre de la hoja (issue #178, A2)
        — usado para persistir NovedadPersonalSemana con el contexto
        correcto. Devuelve 0 si el nombre no trae dígitos (p.ej. 'vc')."""
        import re
        m = re.search(r'\d+', sheet_name)
        return int(m.group()) if m else 0

    def _detectar_columnas(self, header_row):
        self.column_indices = {}
        for col_idx, value in enumerate(header_row):
            if value is None:
                continue
            cell_lower = str(value).lower().strip()
            for field_name, posibles in self.COLUMN_MAPPINGS.items():
                if cell_lower in posibles and field_name not in self.column_indices:
                    self.column_indices[field_name] = col_idx
                    break

    @staticmethod
    def _localizar_header(rows):
        """B5 fix: encuentra la fila índice cuyas celdas contienen el header
        ('#' + 'AVISOS' + 'ACTIVIDAD'). Busca en las primeras 6 filas.

        Devuelve el índice (0-based) o None si no encuentra.
        Antes el código asumía rows[1] siempre; falla cuando el archivo trae
        2 filas de banner o si las plantillas evolucionan.
        """
        for idx in range(min(6, len(rows))):
            row = rows[idx]
            celdas = {str(c).lower().strip() for c in row if c is not None}
            if '#' in celdas and 'avisos' in celdas and 'actividad' in celdas:
                return idx
        return None

    @staticmethod
    def _es_numero_actividad(numero_str):
        """B5 fix: una fila es actividad real si '#' es un entero positivo.
        Filtra ruido como '-' (separador), 'NOVEDADES' (sección final),
        notas libres, etc. — los archivos reales de planta tienen estas
        filas mezcladas con las actividades.
        """
        s = str(numero_str).strip()
        if not s:
            return False
        return s.isdigit() and int(s) > 0

    def _crear_torre_placeholder(self, linea):
        """B5 fix: cuando una línea no tiene torres registradas todavía,
        crear una torre placeholder 'T-AUTO' para no perder la actividad.
        Coordenadas (0,0) marcadas para que mantenimiento las corrija.
        """
        from apps.lineas.models import Torre
        torre, _ = Torre.objects.get_or_create(
            linea=linea,
            numero='T-AUTO',
            defaults={
                'tipo': Torre.TipoTorre.SUSPENSION,
                'estado': Torre.EstadoTorre.BUENO,
                'latitud': 0,
                'longitud': 0,
                'observaciones': 'Torre placeholder creada por importer de programación semanal — corregir coordenadas',
            },
        )
        return torre

    def _get_cell(self, row, field_name):
        idx = self.column_indices.get(field_name)
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    def _split_multi(self, value):
        """Devuelve lista de strings limpios separados por \\n, /, ',' o ';'."""
        if value is None:
            return []
        texto = str(value).strip()
        if not texto or texto == '-':
            return []
        import re
        partes = [p.strip() for p in re.split(r'[\n/;,]+', texto) if p.strip()]
        return partes or []

    def _procesar_hoja(self, sheet, opciones, semana=0):
        from .models import Actividad, ProgramacionMensual, TipoActividad
        from apps.lineas.models import Linea

        actualizar_existentes = opciones.get('actualizar_existentes', False)

        rows = list(sheet.iter_rows(values_only=True))
        if len(rows) < 3:
            return {'creadas': 0, 'actualizadas': 0, 'omitidas': 0, 'nota': 'hoja vacía'}

        # B5 fix: detectar la fila del header dinámicamente (puede no ser
        # row index 1). Algunos archivos tienen el banner extendido a 2-3
        # filas y el header termina en row index 2 o 3.
        header_row_idx = self._localizar_header(rows)
        if header_row_idx is None:
            self.advertencias.append({
                'hoja': sheet.title,
                'mensaje': 'no se pudo localizar fila de encabezados; hoja omitida',
            })
            return {'creadas': 0, 'actualizadas': 0, 'omitidas': 0, 'nota': 'sin header'}

        self._detectar_columnas(rows[header_row_idx])
        requeridos = {'numero', 'tipo_actividad', 'linea', 'avisos'}
        faltantes = requeridos - set(self.column_indices)
        if faltantes:
            self.advertencias.append({
                'hoja': sheet.title,
                'mensaje': f'columnas faltantes: {sorted(faltantes)}; hoja omitida',
            })
            return {'creadas': 0, 'actualizadas': 0, 'omitidas': 0, 'nota': 'columnas faltantes'}

        # Issue #178 (A1): límites de bloque por celda combinada de columna A
        # (numero) o D (tramo) — 100% confiable (verificado contra 106 bloques
        # reales, vs 97% del heurístico '#' en blanco que dependía de que el
        # merge dejara la celda vacía sin verificar el merge en sí). Si la
        # hoja no trae merges en ninguna columna (caso raro/manual) se cae al
        # heurístico legado con advertencia explícita.
        mapa_bloques = self._mapa_bloques_por_merge(
            sheet, self.column_indices.get('numero'), self.column_indices.get('tramo')
        )
        usa_fallback_legado = not mapa_bloques
        if usa_fallback_legado:
            self.advertencias.append({
                'hoja': sheet.title,
                'mensaje': 'no se detectaron celdas combinadas en columna '
                           'numero/tramo; se usa el heurístico de respaldo '
                           '(columna "#" en blanco = continuación de bloque, '
                           'menos confiable)',
            })

        creadas = 0
        actualizadas = 0
        omitidas = 0
        actividades_cuadrilla_pendiente = []  # [(actividad_obj, [cedulas])]
        current_actividad_idx = None  # idx en actividades_cuadrilla_pendiente
        # Issue #178 (A2): al entrar a NOVEDADES se cierra la actividad activa
        # y las filas de personal que siguen se persisten como registro
        # INDEPENDIENTE (no como miembro/cédula de la última actividad).
        en_novedades = False
        ultimo_anio = None
        novedades_hoja = 0

        for row_idx, row in enumerate(rows[header_row_idx + 1:], start=header_row_idx + 2):
            numero = self._get_cell(row, 'numero')
            cedula = self._get_cell(row, 'cedula')

            numero_str = '' if numero is None else str(numero).strip()

            if numero_str.strip().upper() == 'NOVEDADES':
                current_actividad_idx = None
                en_novedades = True
                continue

            # B5 fix: filas con '#' no-numérico ('-', notas sueltas) NO son
            # actividades ni miembros; ignorarlas silenciosamente sin alterar
            # el modo NOVEDADES ni current_actividad_idx.
            if numero_str and not self._es_numero_actividad(numero_str):
                continue

            es_continuacion = self._es_fila_continuacion(
                row_idx, numero_str, mapa_bloques, usa_fallback_legado
            )

            if es_continuacion:
                if en_novedades:
                    novedad = self._parsear_novedad(row, row_idx, semana, ultimo_anio, sheet.title)
                    if novedad:
                        self._guardar_novedad(novedad)
                        novedades_hoja += 1
                elif current_actividad_idx is not None and cedula:
                    # Fila de personal — agregar cédula a actividad anterior
                    actividades_cuadrilla_pendiente[current_actividad_idx][1].append(cedula)
                continue

            # Es fila de ACTIVIDAD → sale de NOVEDADES (si estaba).
            en_novedades = False
            tipo_actividad_raw = self._get_cell(row, 'tipo_actividad')
            avisos_raw = self._get_cell(row, 'avisos')
            lineas_raw = self._get_cell(row, 'linea')
            fecha_inicio = self._get_cell(row, 'fecha_inicio')
            fecha_fin = self._get_cell(row, 'fecha_fin')

            avisos = self._split_multi(avisos_raw)
            if not avisos:
                # Sin aviso explícito: skip pero registrar advertencia
                self.advertencias.append({
                    'hoja': sheet.title, 'fila': row_idx,
                    'mensaje': 'fila sin avisos — omitida',
                })
                omitidas += 1
                continue

            codigos_linea = self._split_multi(lineas_raw)
            if not codigos_linea:
                self.advertencias.append({
                    'hoja': sheet.title, 'fila': row_idx,
                    'mensaje': 'fila sin línea — omitida',
                })
                omitidas += 1
                continue

            # Resolver línea (toma la primera que matchea)
            linea_obj = self._buscar_linea(codigos_linea)
            if linea_obj is None:
                self.advertencias.append({
                    'hoja': sheet.title, 'fila': row_idx,
                    'mensaje': f'línea no encontrada para {codigos_linea}',
                })
                omitidas += 1
                continue

            # Resolver torre — primera disponible de la línea.
            # B5 fix: si la línea no tiene torres, NO descartar la actividad
            # (estaba perdiéndolas en silencio). Crear placeholder T-AUTO.
            torre = linea_obj.torres.order_by('numero').first()
            if torre is None:
                torre = self._crear_torre_placeholder(linea_obj)
                self.advertencias.append({
                    'hoja': sheet.title, 'fila': row_idx,
                    'mensaje': f'línea {linea_obj.codigo} sin torres — '
                               f'creada torre placeholder {torre.numero!r}',
                })

            # Resolver tipo_actividad
            tipo_obj = self._resolver_tipo_actividad(tipo_actividad_raw)

            # Fecha inicio para calcular mes
            anio_act, mes_act = self._extraer_anio_mes(fecha_inicio)
            ultimo_anio = anio_act  # issue #178 (A2): contexto de año para NOVEDADES
            programacion = self._get_or_create_programacion(linea_obj, anio_act, mes_act)
            self.programaciones_tocadas.add((anio_act, mes_act, linea_obj.id))

            fecha_inicio_norm = self._normalizar_fecha(fecha_inicio)
            fecha_fin_norm = self._normalizar_fecha(fecha_fin)
            observaciones = self._get_cell(row, 'observaciones')
            tramo_raw = self._get_cell(row, 'tramo')
            orden_raw = self._get_cell(row, 'orden')
            pt_raw = self._get_cell(row, 'pt_sap')

            obs_consolidadas = self._tramo_obs(tramo_raw, observaciones, fecha_fin_norm)
            fecha_prog = fecha_inicio_norm or date.today()

            with transaction.atomic():
                for aviso in avisos:
                    actividad, created = Actividad.objects.get_or_create(
                        aviso_sap=str(aviso).strip(),
                        defaults={
                            'linea': linea_obj,
                            'torre': torre,
                            'tipo_actividad': tipo_obj,
                            'programacion': programacion,
                            'fecha_programada': fecha_prog,
                            'observaciones_programacion': obs_consolidadas,
                            'orden_sap': self._first_token(orden_raw),
                            'pt_sap': self._first_token(pt_raw),
                            'estado': Actividad.Estado.PROGRAMADA,
                        },
                    )
                    if created:
                        creadas += 1
                        self.actividades_creadas.append(actividad.aviso_sap)
                    elif actualizar_existentes:
                        actividad.linea = linea_obj
                        actividad.torre = torre
                        actividad.tipo_actividad = tipo_obj
                        actividad.programacion = programacion
                        actividad.fecha_programada = fecha_prog
                        actividad.observaciones_programacion = obs_consolidadas
                        actividad.orden_sap = self._first_token(orden_raw)
                        actividad.pt_sap = self._first_token(pt_raw)
                        actividad.save()
                        actualizadas += 1
                        self.actividades_actualizadas.append(actividad.aviso_sap)
                    else:
                        omitidas += 1

                    # Guardar para asignar cuadrillas tras leer miembros
                    actividades_cuadrilla_pendiente.append((actividad, []))
                    current_actividad_idx = len(actividades_cuadrilla_pendiente) - 1

        # Resolver cuadrillas via cédulas acumuladas
        self._asignar_cuadrillas(actividades_cuadrilla_pendiente)

        return {
            'creadas': creadas,
            'actualizadas': actualizadas,
            'omitidas': omitidas,
            'novedades': novedades_hoja,
        }

    def _parsear_novedad(self, row, row_idx, semana, anio, hoja_origen):
        """Fila de personal dentro de la sección NOVEDADES (issue #178, A2).

        La nota (p.ej. "Vacaciones", "Incapacidad") viene en la columna
        AVISOS de la fila, NO en una columna dedicada (confirmado contra el
        Excel real del cliente).
        """
        nombre = self._get_cell(row, 'personal')
        cedula = self._get_cell(row, 'cedula')
        nombre = str(nombre).strip() if nombre else ''
        cedula = str(cedula).strip() if cedula else ''
        if not nombre and not cedula:
            return None
        return {
            'nombre': nombre,
            'cedula': cedula,
            'cargo': self._get_cell(row, 'cargo') or '',
            'nota': ', '.join(self._split_multi(self._get_cell(row, 'avisos'))),
            'semana': semana,
            'anio': anio or date.today().year,
            'hoja_origen': hoja_origen,
        }

    def _guardar_novedad(self, novedad):
        """Persiste una fila de NOVEDADES como registro independiente
        (issue #178, A2) — sin FK a Actividad/Cuadrilla. get_or_create sobre
        (cedula, semana, anio, nota) para que un re-import no duplique."""
        from apps.cuadrillas.models import NovedadPersonalSemana

        _, creado = NovedadPersonalSemana.objects.get_or_create(
            cedula=novedad['cedula'],
            semana=novedad['semana'],
            anio=novedad['anio'],
            nota=novedad['nota'],
            defaults={
                'nombre': novedad['nombre'],
                'cargo': str(novedad['cargo']).strip(),
                'hoja_origen': novedad['hoja_origen'],
            },
        )
        if creado:
            self.novedades_creadas += 1

    def _buscar_linea(self, codigos):
        from apps.lineas.models import Linea
        for codigo in codigos:
            codigo = str(codigo).strip()
            qs = Linea.objects.filter(codigo__iexact=codigo)
            if qs.exists():
                return qs.first()
            qs = Linea.objects.filter(codigo_transelca__iexact=codigo)
            if qs.exists():
                return qs.first()
            qs = Linea.objects.filter(codigo__icontains=codigo)
            if qs.exists():
                return qs.first()
        return None

    def _resolver_tipo_actividad(self, raw):
        from .models import TipoActividad
        nombre = str(raw or '').strip()
        if not nombre:
            tipo, _ = TipoActividad.objects.get_or_create(
                codigo='SIN_TIPO',
                defaults={'nombre': 'Sin tipo', 'categoria': TipoActividad.Categoria.OTRO},
            )
            return tipo
        qs = TipoActividad.objects.filter(nombre__iexact=nombre)
        if qs.exists():
            return qs.first()
        qs = TipoActividad.objects.filter(codigo__iexact=nombre[:20])
        if qs.exists():
            return qs.first()
        codigo = self._slug_codigo(nombre)
        categoria = self._inferir_categoria(nombre)
        tipo, _ = TipoActividad.objects.get_or_create(
            codigo=codigo,
            defaults={'nombre': nombre[:100], 'categoria': categoria},
        )
        return tipo

    @staticmethod
    def _slug_codigo(nombre):
        import re
        s = re.sub(r'[^A-Z0-9]', '_', nombre.upper())[:20]
        return s.strip('_') or 'OTRO'

    @staticmethod
    def _inferir_categoria(nombre):
        from .models import TipoActividad
        n = nombre.lower()
        mapping = [
            ('servidumbre', TipoActividad.Categoria.SERVIDUMBRE),
            ('poda', TipoActividad.Categoria.PODA),
            ('lavado', TipoActividad.Categoria.LAVADO),
            ('aislador', TipoActividad.Categoria.AISLADORES),
            ('herraje', TipoActividad.Categoria.HERRAJES),
            ('inspecci', TipoActividad.Categoria.INSPECCION),
            ('termograf', TipoActividad.Categoria.TERMOGRAFIA),
            ('descarga', TipoActividad.Categoria.DESCARGAS),
            ('electromec', TipoActividad.Categoria.ELECTROMEC),
            ('puesta tierra', TipoActividad.Categoria.MEDICION_PT),
            ('permiso', TipoActividad.Categoria.PERMISO),
        ]
        for keyword, categoria in mapping:
            if keyword in n:
                return categoria
        return TipoActividad.Categoria.OTRO

    @staticmethod
    def _normalizar_fecha(valor):
        if valor is None:
            return None
        if isinstance(valor, date):
            return valor
        try:
            from datetime import datetime
            if isinstance(valor, datetime):
                return valor.date()
        except Exception:
            pass
        try:
            from datetime import datetime
            return datetime.strptime(str(valor)[:10], '%Y-%m-%d').date()
        except Exception:
            return None

    def _extraer_anio_mes(self, fecha):
        fecha_norm = self._normalizar_fecha(fecha)
        if fecha_norm:
            return fecha_norm.year, fecha_norm.month
        hoy = date.today()
        return hoy.year, hoy.month

    def _get_or_create_programacion(self, linea, anio, mes):
        from .models import ProgramacionMensual
        prog, _ = ProgramacionMensual.objects.get_or_create(
            anio=anio, mes=mes, linea=linea,
        )
        return prog

    @staticmethod
    def _first_token(valor):
        if valor is None:
            return ''
        s = str(valor).strip()
        if not s or s == '-':
            return ''
        return s.split('\n')[0].strip()[:20]

    @staticmethod
    def _tramo_obs(tramo_raw, obs, fecha_fin=None):
        partes = []
        if tramo_raw:
            partes.append(f'Tramo: {str(tramo_raw).strip()}')
        if fecha_fin:
            partes.append(f'Fecha fin: {fecha_fin.isoformat()}')
        if obs:
            partes.append(str(obs).strip())
        return ' | '.join(partes)[:500]

    def _asignar_cuadrillas(self, lista_pendiente):
        """Resuelve cuadrillas a partir de cédulas acumuladas y las asigna."""
        from apps.cuadrillas.models import CuadrillaMiembro
        for actividad, cedulas in lista_pendiente:
            if not cedulas:
                continue
            cuadrillas_ids = set()
            for cedula in cedulas:
                cedula_str = str(cedula).strip()
                miembro = CuadrillaMiembro.objects.filter(
                    usuario__documento=cedula_str,
                    activo=True,
                ).first()
                if miembro:
                    cuadrillas_ids.add(miembro.cuadrilla_id)
                else:
                    self.advertencias.append({
                        'mensaje': f'cédula {cedula_str} no vinculada a ninguna cuadrilla — actividad {actividad.aviso_sap}',
                    })
            if cuadrillas_ids:
                actividad.cuadrillas.add(*cuadrillas_ids)
                if not actividad.cuadrilla_id:
                    actividad.cuadrilla_id = next(iter(cuadrillas_ids))
                    actividad.save(update_fields=['cuadrilla', 'updated_at'])
