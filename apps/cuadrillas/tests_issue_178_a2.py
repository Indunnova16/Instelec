"""Tests #178 — Sprint (2026-07-25), sub-item A2: exportador Excel horizontal
exacto del formato que Alcides ya usa (pedido A).

Issue: Indunnova16/Instelec#178

A2 agrega ``ProgramacionSemanalHorizontalExporter``
(``apps/actividades/exporters.py``) + el helper ``_parsear_avisos_orden_ptsap``
(reverso de ``ProgramacionS18CuadrillaImporter._construir_observaciones``) +
el endpoint ``ProgramacionSemanalExportarHorizontalView``
(GET /cuadrillas/semanal/<anio>/<semana>/exportar/, name
``cuadrillas:semanal_exportar_horizontal``).

🔴 Riesgo crítico verificado en este archivo (ver PLAN): el dict
``_bloque_a_dict`` tiene las claves ``cargo``/``rol`` INVERTIDAS respecto a
como se llaman las columnas Excel CARGO/ROL del cliente — Excel CARGO debe
salir de ``miembro['rol']`` (puesto real, ej. "Liniero I") y Excel ROL de
``miembro['cargo']`` (JT/CTA). El test ``test_happy_...`` compara CADA fila
del Excel contra el dict fuente (``_bloque_a_dict``) campo a campo, no contra
un valor "parece razonable".

Ejecutar:
  DJANGO_SETTINGS_MODULE=config.settings.dev_lite \
    venv/bin/python -m pytest apps/cuadrillas/tests_issue_178_a2.py -v \
    -o python_files="tests_*.py test_*.py"
"""

from datetime import date
from io import BytesIO

from django.contrib.auth import get_user_model
from django.test import TestCase
from django.urls import reverse
from openpyxl import load_workbook

from apps.actividades.exporters import _parsear_avisos_orden_ptsap
from apps.cuadrillas.models import Cuadrilla, CuadrillaMiembro
from apps.cuadrillas.views_semanal import _bloque_a_dict

Usuario = get_user_model()

HEADERS_ESPERADOS = [
    "#",
    "ACTIVIDAD",
    "LINEA",
    "TRAMO",
    "INICIO",
    "FIN",
    "PERSONAL",
    "CEDULA",
    "CELULAR",
    "CARGO",
    "ROL",
    "PLACA",
    "AVISOS",
    "ORDEN",
    "PT SAP",
    "Comentarios",
]


def _crear_admin():
    # Rol admin (NO superuser) para ejercitar el gate real de RoleRequiredMixin.
    return Usuario.objects.create_user(
        email="admin_178a2@test.com",
        password="testpass123!",
        first_name="Admin",
        last_name="A2",
        rol="admin",
        is_staff=True,
    )


def _crear_usuario(documento, nombre):
    partes = nombre.split(maxsplit=1)
    return Usuario.objects.create(
        email=f"{documento}@test.local",
        documento=documento,
        first_name=partes[0],
        last_name=partes[1] if len(partes) > 1 else "",
        rol="liniero",
        is_active=True,
    )


def _crear_bloque(codigo, fecha, fecha_fin=None, observaciones="", nombre="MANTENIMIENTO - 809"):
    return Cuadrilla.objects.create(
        codigo=codigo,
        nombre=nombre,
        activa=True,
        fecha=fecha,
        fecha_fin=fecha_fin,
        observaciones=observaciones,
    )


def _agregar_miembro(cuadrilla, usuario, rol_cuadrilla_id, cargo, placa=""):
    return CuadrillaMiembro.objects.create(
        cuadrilla=cuadrilla,
        usuario=usuario,
        rol_cuadrilla_id=rol_cuadrilla_id,
        cargo=cargo,
        costo_dia=0,
        fecha_inicio=cuadrilla.fecha or date.today(),
        activo=True,
        placa_vehiculo=placa,
    )


class TestParsearAvisosOrdenPtsap(TestCase):
    """Unit: reverso de ``_construir_observaciones`` — sin BD."""

    def test_los_3_prefijos_presentes(self):
        obs = "Avisos: 5720754, 5720792 | Orden: 1, 2 | PT SAP: 123 | Poda urgente"
        avisos, orden, pt_sap, comentarios = _parsear_avisos_orden_ptsap(obs)
        self.assertEqual(avisos, "5720754, 5720792")
        self.assertEqual(orden, "1, 2")
        self.assertEqual(pt_sap, "123")
        self.assertEqual(comentarios, "Poda urgente")

    def test_subconjunto_de_prefijos(self):
        """El importer solo antepone los prefijos que vienen no vacíos —
        el parser no debe asumir que estén los 3."""
        obs = "Avisos: 5720754 | Sin observaciones adicionales"
        avisos, orden, pt_sap, comentarios = _parsear_avisos_orden_ptsap(obs)
        self.assertEqual(avisos, "5720754")
        self.assertEqual(orden, "")
        self.assertEqual(pt_sap, "")
        self.assertEqual(comentarios, "Sin observaciones adicionales")

    def test_sin_prefijo_todo_a_comentarios(self):
        """Bloques creados desde el grid editable (#188) no tienen prefijo —
        degrada limpio: todo va a Comentarios."""
        avisos, orden, pt_sap, comentarios = _parsear_avisos_orden_ptsap("Poda sector norte")
        self.assertEqual((avisos, orden, pt_sap), ("", "", ""))
        self.assertEqual(comentarios, "Poda sector norte")

    def test_vacio_no_rompe(self):
        self.assertEqual(_parsear_avisos_orden_ptsap(""), ("", "", "", ""))
        self.assertEqual(_parsear_avisos_orden_ptsap(None), ("", "", "", ""))


class TestA2ExportHorizontal(TestCase):
    def setUp(self):
        self.admin = _crear_admin()
        self.client.force_login(self.admin)

    def test_happy_columnas_exactas_y_cargo_rol_no_invertidos(self):
        """happy: semana con 2 bloques QA_E2E_ (>=2 personas c/u) -> export
        200, columnas EXACTAS, CARGO/ROL NO invertidos comparando fila a fila
        contra el dict fuente (_bloque_a_dict)."""
        u1 = _crear_usuario("178A2-001", "QA_E2E_ JHON JIMENEZ")
        u2 = _crear_usuario("178A2-002", "QA_E2E_ ANA TORRES")
        u3 = _crear_usuario("178A2-003", "QA_E2E_ LUIS PENA")
        u4 = _crear_usuario("178A2-004", "QA_E2E_ MARIA LOPEZ")

        c1 = _crear_bloque(
            "45-2026-0001-MAN",
            date(2026, 11, 2),
            date(2026, 11, 6),
            observaciones="Avisos: 5720754 | Orden: 1 | PT SAP: 999 | Poda urgente",
            nombre="Servidumbre Completa - 817",
        )
        _agregar_miembro(c1, u1, "LINIERO_I", CuadrillaMiembro.CargoJerarquico.JT_CTA)
        _agregar_miembro(c1, u2, "AYUDANTE", CuadrillaMiembro.CargoJerarquico.MIEMBRO)

        c2 = _crear_bloque(
            "45-2026-0002-POD",
            date(2026, 11, 3),
            None,
            observaciones="Poda sin prefijo del cliente",
            nombre="Mantenimiento Preventivo",
        )
        _agregar_miembro(c2, u3, "LINIERO_II", CuadrillaMiembro.CargoJerarquico.MIEMBRO)
        _agregar_miembro(
            c2, u4, "CONDUCTOR", CuadrillaMiembro.CargoJerarquico.MIEMBRO, placa="ABC123"
        )

        resp = self.client.get(reverse("cuadrillas:semanal_exportar_horizontal", args=[2026, 45]))
        self.assertEqual(resp.status_code, 200)
        self.assertIn("spreadsheetml.sheet", resp["Content-Type"])

        wb = load_workbook(BytesIO(resp.content))
        ws = wb.active
        headers = [c.value for c in ws[1]]
        self.assertEqual(headers, HEADERS_ESPERADOS)

        filas = list(ws.iter_rows(min_row=2, values_only=True))
        # 1 fila por persona, sin merges: 2 bloques con 2 miembros c/u = 4 filas.
        self.assertEqual(len(filas), 4)

        b1 = _bloque_a_dict(Cuadrilla.objects.get(pk=c1.pk))
        b2 = _bloque_a_dict(Cuadrilla.objects.get(pk=c2.pk))

        # Fila 1: primer miembro de c1 -- trae actividad + AVISOS/ORDEN/PT SAP.
        fila = filas[0]
        m = b1["miembros"][0]
        self.assertEqual(fila[0], 1)
        self.assertEqual(fila[1], "Servidumbre Completa - 817")
        self.assertEqual(fila[4], "02/11/2026")
        self.assertEqual(fila[5], "06/11/2026")
        self.assertEqual(fila[6], m["nombre"])
        self.assertEqual(fila[7], m["documento"])
        # 🔴 Mapeo obligatorio NO invertido.
        self.assertEqual(fila[9], m["rol"])  # Excel CARGO <- dict['rol'] (puesto real)
        self.assertEqual(fila[10], m["cargo"])  # Excel ROL  <- dict['cargo'] (JT/CTA)
        self.assertEqual(fila[12], "5720754")
        self.assertEqual(fila[13], "1")
        self.assertEqual(fila[14], "999")
        self.assertEqual(fila[15], "Poda urgente")

        # Fila 2: segundo miembro de c1 -- MISMO bloque, actividad/avisos en
        # blanco (openpyxl retorna None al releer una celda vacía).
        fila2 = filas[1]
        m2 = b1["miembros"][1]
        self.assertIsNone(fila2[0])
        self.assertIsNone(fila2[1])
        self.assertIsNone(fila2[12])
        self.assertEqual(fila2[6], m2["nombre"])
        self.assertEqual(fila2[9], m2["rol"])
        self.assertEqual(fila2[10], m2["cargo"])

        # Bloque 2: primer miembro trae la actividad; el conductor trae su placa.
        fila3, fila4 = filas[2], filas[3]
        self.assertEqual(fila3[0], 2)
        self.assertEqual(fila3[1], "Mantenimiento Preventivo")
        for f, m in zip((fila3, fila4), b2["miembros"], strict=True):
            self.assertEqual(f[9], m["rol"])
            self.assertEqual(f[10], m["cargo"])
            # openpyxl retorna None (no "") al releer una celda escrita vacía.
            self.assertEqual(f[11] or "", m["placa_vehiculo"])
        conductor_row = fila3 if fila3[11] == "ABC123" else fila4
        self.assertEqual(conductor_row[11], "ABC123")

    def test_edge_observaciones_sin_prefijo_avisos_orden_ptsap_vacios(self):
        """edge 1: bloque con observaciones SIN prefijo Avisos/Orden/PT SAP
        (creado desde la UI del grid editable, #188) -> esas 3 columnas
        vacías, Comentarios = texto completo."""
        u = _crear_usuario("178A2-010", "QA_E2E_ CARLOS RUIZ")
        _agregar_miembro(
            _crear_bloque(
                "45-2026-0003-CAR",
                date(2026, 11, 2),
                observaciones="Revisión de poste torcido reportada por vecino",
            ),
            u,
            "LINIERO_I",
            CuadrillaMiembro.CargoJerarquico.MIEMBRO,
        )

        resp = self.client.get(reverse("cuadrillas:semanal_exportar_horizontal", args=[2026, 45]))
        wb = load_workbook(BytesIO(resp.content))
        fila = next(wb.active.iter_rows(min_row=2, values_only=True))
        # openpyxl retorna None al releer una celda a la que se escribió "".
        self.assertIsNone(fila[12])  # AVISOS
        self.assertIsNone(fila[13])  # ORDEN
        self.assertIsNone(fila[14])  # PT SAP
        self.assertEqual(fila[15], "Revisión de poste torcido reportada por vecino")

    def test_edge_fecha_fin_none_legacy_columna_fin_vacia(self):
        """edge 2: bloque con fecha_fin=None (legacy pre-A1, o cargado antes
        de que existiera la columna) -> columna FIN vacía, no rompe."""
        u = _crear_usuario("178A2-020", "QA_E2E_ SANDRA PARDO")
        _agregar_miembro(
            _crear_bloque("45-2026-0004-SAN", date(2026, 11, 2), fecha_fin=None),
            u,
            "LINIERO_I",
            CuadrillaMiembro.CargoJerarquico.MIEMBRO,
        )

        resp = self.client.get(reverse("cuadrillas:semanal_exportar_horizontal", args=[2026, 45]))
        self.assertEqual(resp.status_code, 200)
        wb = load_workbook(BytesIO(resp.content))
        fila = next(wb.active.iter_rows(min_row=2, values_only=True))
        self.assertEqual(fila[4], "02/11/2026")  # INICIO sigue presente
        self.assertIsNone(fila[5])  # FIN vacío (None al releer), no rompe

    def test_semana_sin_bloques_export_vacio_no_rompe(self):
        """Semana sin ninguna programación -- el export sigue siendo 200 con
        solo la fila de encabezados (no hay bloques que iterar)."""
        resp = self.client.get(reverse("cuadrillas:semanal_exportar_horizontal", args=[2099, 1]))
        self.assertEqual(resp.status_code, 200)
        wb = load_workbook(BytesIO(resp.content))
        self.assertEqual(wb.active.max_row, 1)
