"""
Tests issue #176 (bounce 3) — Salario enlazado a Cargo + Import/Export masivo.

Archivo NUEVO (A8 del plan `SPRINTS/PLAN_2026-07-19_176_salario_cargo_import_export.md`),
consolidado por sub-item conforme se fue implementando el sprint completo
(A1-A7) en el mismo worktree/branch. Convención de nombres de clase:
`Test<subitem><Tema>`.

HALLAZGO CRÍTICO documentado en el plan (y que este archivo verifica
explícitamente en `TestA4RegresionFinanciero*`): `financiero/reports.py` y
`apps/cuadrillas/views_semanal.py` **NO leen** `Cargo.salario_base` ni
`PersonalCuadrilla.salario_base` hoy — usan fuentes de datos completamente
distintas (un dict hardcoded con bug de casing preexistente, y el campo
persistido `CuadrillaMiembro.costo_dia`, respectivamente). El pedido
original del cliente ("el salario permite calcular costo/día en reportes")
NO queda resuelto por el alcance de este issue — es una decisión de scope
explícita, documentada acá y en el comentario de cierre (A10, F6).

Issue: Indunnova16/Instelec#176

Ejecutar con:
    DJANGO_SETTINGS_MODULE=config.settings.dev_lite \
        venv/bin/python -m pytest apps/cuadrillas/tests_issue_176_salario.py -v
"""

from datetime import date
from decimal import Decimal
from io import BytesIO

from django.contrib.auth import get_user_model
from django.test import Client, TestCase
from django.urls import reverse

from apps.cuadrillas.models import Cargo, Cuadrilla, CuadrillaMiembro, PersonalCuadrilla
from tests.factories import (
    ActividadCompletadaFactory,
    CuadrillaMiembroFactory,
)

Usuario = get_user_model()


def _crear_admin(sufijo="176sal"):
    return Usuario.objects.create_user(
        email=f"admin_{sufijo}@test.com",
        password="testpass123!",
        first_name="Admin",
        last_name="Salario176",
        rol="admin",
        is_staff=True,
        is_superuser=True,
    )


def _crear_usuario(documento, nombre):
    partes = nombre.split(maxsplit=1)
    return Usuario.objects.create(
        email=f"{documento}@test.local",
        documento=documento,
        first_name=partes[0],
        last_name=partes[1] if len(partes) > 1 else "",
        rol="liniero",
        is_active=True,
    )


# ---------------------------------------------------------------------------
# A4 — Test de regresión financiero (premisa corregida vs. supuesto original)
# ---------------------------------------------------------------------------
class TestA4RegresionFinancieroReportsPy(TestCase):
    """`financiero.reports.CuadroCostosGenerator._calcular_costos_personal`
    sigue usando su propio dict interno `valores_dia` (con el bug de casing
    preexistente ya documentado, fuera de scope) -- NO lee `Cargo.salario_base`
    ni `PersonalCuadrilla.salario_base`, aunque A1-A3 ya existan."""

    def test_calcular_costos_personal_ignora_cargo_salario_base(self):
        cargo = Cargo.objects.filter(codigo="LINIERO_I").first()
        if cargo is None:
            cargo = Cargo.objects.create(codigo="LINIERO_I", nombre="Liniero I")
        # Valor absurdo/inconfundible: si el código llegara a leer esto, el
        # test lo detecta inmediatamente (no matchea el 80000 del fallback).
        cargo.salario_base = Decimal("999999999")
        cargo.save(update_fields=["salario_base"])

        actividad = ActividadCompletadaFactory()
        CuadrillaMiembroFactory(
            cuadrilla=actividad.cuadrilla,
            rol_cuadrilla_id="LINIERO_I",
            fecha_inicio=actividad.fecha_programada,
        )

        from apps.actividades.models import Actividad
        from apps.financiero.reports import CuadroCostosGenerator

        generator = CuadroCostosGenerator(
            anio=actividad.fecha_programada.year, mes=actividad.fecha_programada.month
        )
        actividades_qs = Actividad.objects.filter(pk=actividad.pk)
        resultado = generator._calcular_costos_personal(actividades_qs)

        self.assertEqual(len(resultado["detalle"]), 1)
        valor_dia = resultado["detalle"][0]["valor_dia"]
        # Bug preexistente (fuera de scope): 'LINIERO_I' (mayúscula, código
        # real) nunca matchea las keys lowercase de valores_dia -> cae al
        # default 80000. Lo que este test PROTEGE es que A1-A3 no lo cambien
        # a leer Cargo.salario_base (999999999) por accidente.
        self.assertEqual(valor_dia, Decimal("80000"))
        self.assertNotEqual(valor_dia, cargo.salario_base)

    def test_calcular_costos_personal_ignora_personal_cuadrilla_salario_base(self):
        """Aunque el PersonalCuadrilla del miembro tenga salario_base editado
        vía el CRUD de Colaboradores (A1 bounce 1), _calcular_costos_personal
        no lo consulta -- trabaja sobre CuadrillaMiembro, no PersonalCuadrilla."""
        usuario = _crear_usuario("176sal-fin-001", "Regresion Financiera")
        PersonalCuadrilla.objects.filter(documento="176sal-fin-001").delete()
        PersonalCuadrilla.objects.create(
            nombre="Regresion Financiera",
            documento="176sal-fin-001",
            rol_cuadrilla_id="AYUDANTE",
            salario_base=Decimal("777777777"),
        )
        actividad = ActividadCompletadaFactory()
        CuadrillaMiembro.objects.create(
            cuadrilla=actividad.cuadrilla,
            usuario=usuario,
            rol_cuadrilla_id="AYUDANTE",
            fecha_inicio=actividad.fecha_programada,
            activo=True,
        )

        from apps.actividades.models import Actividad
        from apps.financiero.reports import CuadroCostosGenerator

        generator = CuadroCostosGenerator(
            anio=actividad.fecha_programada.year, mes=actividad.fecha_programada.month
        )
        resultado = generator._calcular_costos_personal(
            Actividad.objects.filter(pk=actividad.pk)
        )
        valor_dia = resultado["detalle"][0]["valor_dia"]
        self.assertNotEqual(valor_dia, Decimal("777777777"))


class TestA4RegresionFinancieroViewsSemanal(TestCase):
    """`views_semanal.ProgramacionSemanalDuplicarView` sigue copiando el
    campo PERSISTIDO `CuadrillaMiembro.costo_dia` (fijado una vez al crear
    el miembro) -- no recalcula desde `Cargo.salario_base` en cada
    duplicado, aunque el Cargo tenga un salario distinto asignado después."""

    def setUp(self):
        self.admin = _crear_admin("176sal-sem")
        self.client = Client()
        self.client.force_login(self.admin)

    def test_duplicar_semana_copia_costo_dia_persistido_no_salario_base_actual(self):
        cargo = Cargo.objects.filter(codigo="AYUDANTE").first()
        if cargo is None:
            cargo = Cargo.objects.create(codigo="AYUDANTE", nombre="Ayudante")
        usuario = _crear_usuario("176sal-sem-001", "Duplicado Semanal")

        origen = Cuadrilla.objects.create(
            codigo="20-2026-9001-MAN",
            nombre="MANTENIMIENTO - 176SAL",
            activa=True,
            fecha=date(2026, 5, 11),
        )
        CuadrillaMiembro.objects.create(
            cuadrilla=origen,
            usuario=usuario,
            rol_cuadrilla_id="AYUDANTE",
            cargo=CuadrillaMiembro.CargoJerarquico.MIEMBRO,
            costo_dia=Decimal("1750905"),
            fecha_inicio=date(2026, 5, 11),
            activo=True,
        )

        # Después de crear el miembro (costo_dia ya fijado), el Cargo cambia
        # de salario -- si el duplicado leyera Cargo.salario_base en vez del
        # costo_dia persistido, el valor copiado sería distinto.
        cargo.salario_base = Decimal("555555555")
        cargo.save(update_fields=["salario_base"])

        resp = self.client.post(reverse("cuadrillas:semanal_duplicar", args=[2026, 21]))
        self.assertEqual(resp.status_code, 302)

        copia = Cuadrilla.objects.get(codigo="21-2026-9001-MAN")
        miembro_copia = copia.miembros.get(usuario=usuario)
        self.assertEqual(miembro_copia.costo_dia, Decimal("1750905"))
        self.assertNotEqual(miembro_copia.costo_dia, cargo.salario_base)


class TestA4SmokeNominaRenderOnly(TestCase):
    """Smoke E2E render-only: /financiero/nomina/ sigue respondiendo 200
    después de A1-A3 (no regresión de import/carga de módulos)."""

    def setUp(self):
        self.admin = _crear_admin("176sal-nomina")
        self.client = Client()
        self.client.force_login(self.admin)

    def test_nomina_renderiza_sin_error(self):
        resp = self.client.get(reverse("financiero:nomina"))
        self.assertEqual(resp.status_code, 200)


# ---------------------------------------------------------------------------
# A5 — Import/export masivo de Cargo (upsert por código)
# ---------------------------------------------------------------------------
class TestA5ImportExportCargo(TestCase):
    """CargoUploadView (upsert por código) + CargoExportView (xlsx)."""

    def setUp(self):
        self.admin = _crear_admin("176sal-a5")
        self.client = Client()
        self.client.force_login(self.admin)

    @staticmethod
    def _build_workbook(rows):
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Código", "Nombre", "Salario Base"])
        for row in rows:
            ws.append(row)
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        buf.name = "cargos.xlsx"
        return buf

    def test_import_crea_codigo_nuevo(self):
        archivo = self._build_workbook([["QA_A5_NUEVO", "Cargo A5 Nuevo", 2500000]])
        url = reverse("cuadrillas:cargos_upload")
        resp = self.client.post(url, {"archivo": archivo})
        self.assertIn(resp.status_code, (200, 302))
        cargo = Cargo.objects.get(codigo="QA_A5_NUEVO")
        self.assertEqual(cargo.nombre, "Cargo A5 Nuevo")
        self.assertEqual(cargo.salario_base, Decimal("2500000"))
        self.assertTrue(cargo.activo)

    def test_import_actualiza_nombre_y_salario_de_codigo_existente(self):
        Cargo.objects.create(codigo="QA_A5_EXISTE", nombre="Nombre Viejo", salario_base=Decimal("1"))
        archivo = self._build_workbook([["QA_A5_EXISTE", "Nombre Actualizado", 3300000]])
        url = reverse("cuadrillas:cargos_upload")
        resp = self.client.post(url, {"archivo": archivo})
        self.assertIn(resp.status_code, (200, 302))
        cargo = Cargo.objects.get(codigo="QA_A5_EXISTE")
        self.assertEqual(cargo.nombre, "Nombre Actualizado")
        self.assertEqual(cargo.salario_base, Decimal("3300000"))
        # Sigue siendo UN solo registro (upsert, no duplicado).
        self.assertEqual(Cargo.objects.filter(codigo="QA_A5_EXISTE").count(), 1)

    def test_import_dos_filas_una_nueva_una_existente_ambas_persisten(self):
        """Test contra dato legacy real: LINIERO_I (sembrado por el conftest
        autouse) se actualiza junto con un código nuevo en la MISMA carga --
        no se aborta el lote por mezclar creación + actualización."""
        cargo_legacy = Cargo.objects.get(codigo="LINIERO_I")
        salario_original = cargo_legacy.salario_base
        archivo = self._build_workbook(
            [
                ["LINIERO_I", "Liniero I", 3176095],
                ["QA_A5_LOTE_NUEVO", "Cargo Lote Nuevo", 1900000],
            ]
        )
        url = reverse("cuadrillas:cargos_upload")
        resp = self.client.post(url, {"archivo": archivo})
        self.assertIn(resp.status_code, (200, 302))
        cargo_legacy.refresh_from_db()
        self.assertEqual(cargo_legacy.salario_base, Decimal("3176095"))
        self.assertNotEqual(cargo_legacy.salario_base, salario_original)
        self.assertTrue(Cargo.objects.filter(codigo="QA_A5_LOTE_NUEVO").exists())

    def test_import_fila_con_salario_invalido_no_aborta_lote_completo(self):
        """Edge case: una fila con salario no numérico cae a 0 (no revienta
        el proceso) y la fila siguiente válida igual se procesa."""
        archivo = self._build_workbook(
            [
                ["QA_A5_SALARIO_MALO", "Cargo Salario Malo", "no-es-un-numero"],
                ["QA_A5_SALARIO_OK", "Cargo Salario OK", 1000000],
            ]
        )
        url = reverse("cuadrillas:cargos_upload")
        resp = self.client.post(url, {"archivo": archivo})
        self.assertIn(resp.status_code, (200, 302))
        malo = Cargo.objects.get(codigo="QA_A5_SALARIO_MALO")
        self.assertEqual(malo.salario_base, Decimal("0"))
        ok = Cargo.objects.get(codigo="QA_A5_SALARIO_OK")
        self.assertEqual(ok.salario_base, Decimal("1000000"))

    def test_export_descarga_xlsx_con_tres_columnas_correctas(self):
        Cargo.objects.create(codigo="QA_A5_EXPORT", nombre="Cargo Exportable", salario_base=Decimal("4200000"))
        url = reverse("cuadrillas:cargos_export")
        resp = self.client.get(url)
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(
            resp["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        import openpyxl

        wb = openpyxl.load_workbook(BytesIO(resp.content))
        ws = wb.active
        header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        self.assertEqual(header, ["Código", "Nombre", "Salario Base"])
        codigos = [row[0] for row in ws.iter_rows(min_row=2, values_only=True)]
        self.assertIn("QA_A5_EXPORT", codigos)

    def test_import_sin_archivo_muestra_error_no_500(self):
        url = reverse("cuadrillas:cargos_upload")
        resp = self.client.post(url, {})
        self.assertIn(resp.status_code, (200, 302))


# ---------------------------------------------------------------------------
# A6 — Exponer import de Colaboradores en su pantalla natural + export
# ---------------------------------------------------------------------------
class TestA6ImportColaboradoresExpuestoYExport(TestCase):
    """El import (PersonalCuadrillaUploadView) YA FUNCIONA -- A6 solo agrega
    una segunda entrada visual en /cuadrillas/colaboradores/ apuntando al
    MISMO endpoint (sin tocar su lógica) + el export nuevo."""

    def setUp(self):
        self.admin = _crear_admin("176sal-a6")
        self.client = Client()
        self.client.force_login(self.admin)

    def test_modal_importar_visible_en_colaboradores_lista(self):
        resp = self.client.get(reverse("cuadrillas:colaboradores_lista"))
        self.assertEqual(resp.status_code, 200)
        content = resp.content.decode()
        self.assertIn("modal-importar-colaboradores", content)
        self.assertIn(reverse("cuadrillas:personal_upload"), content)

    def test_modal_nuevo_postea_al_mismo_endpoint_ya_validado(self):
        """El modal nuevo de colaboradores_lista.html debe postear al MISMO
        personal_upload que ya funciona (validado en vivo por el cliente
        2026-07-17) -- no se duplica ni se reimplementa la lógica."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Nombre", "Documento", "Cargo", "Salario Base", "Fecha Ingreso", "Fecha Salida"])
        ws.append(["Via Modal Nuevo", "176sal-a6-001", "AYUDANTE", 1750905, "2025-02-01", ""])
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        buf.name = "colaboradores.xlsx"

        url = reverse("cuadrillas:personal_upload")
        resp = self.client.post(url, {"archivo": buf})
        self.assertIn(resp.status_code, (200, 302))
        persona = PersonalCuadrilla.objects.get(documento="176sal-a6-001")
        self.assertEqual(persona.nombre, "Via Modal Nuevo")
        self.assertEqual(persona.salario_base, Decimal("1750905"))

    def test_export_descarga_xlsx_con_seis_columnas_incluye_registro_legacy(self):
        """Test contra dato legacy real: Andrea (documento 43482087,
        LINIERO_I, salario_base=15000.00) debe aparecer en el export."""
        cargo = Cargo.objects.filter(codigo="LINIERO_I").first()
        if cargo is None:
            cargo = Cargo.objects.create(codigo="LINIERO_I", nombre="Liniero I")
        PersonalCuadrilla.objects.filter(documento="43482087").delete()
        PersonalCuadrilla.objects.create(
            nombre="Andrea",
            documento="43482087",
            rol_cuadrilla_id="LINIERO_I",
            salario_base=Decimal("15000.00"),
            fecha_ingreso=date(2025, 1, 1),
        )

        url = reverse("cuadrillas:colaboradores_export")
        resp = self.client.get(url)
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(
            resp["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        import openpyxl

        wb = openpyxl.load_workbook(BytesIO(resp.content))
        ws = wb.active
        header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        self.assertEqual(
            header,
            ["Documento", "Nombre", "Cargo", "Salario Base", "Fecha Ingreso", "Fecha Salida"],
        )
        filas = {row[0]: row for row in ws.iter_rows(min_row=2, values_only=True)}
        self.assertIn("43482087", filas)
        fila_andrea = filas["43482087"]
        self.assertEqual(fila_andrea[1], "Andrea")
        self.assertEqual(fila_andrea[2], "Liniero I")
        self.assertEqual(fila_andrea[3], 15000.0)

    def test_export_colaborador_sin_fecha_salida_exporta_celda_vacia(self):
        """Edge case: colaborador activo (sin fecha_salida) no debe romper
        el export -- la celda queda en blanco (openpyxl round-tripea '' como
        None al releer, comportamiento normal de Excel para celda vacía),
        no lanza excepción ni escribe un string 'None' literal."""
        PersonalCuadrilla.objects.filter(documento="176sal-a6-002").delete()
        PersonalCuadrilla.objects.create(
            nombre="Sin Fecha Salida",
            documento="176sal-a6-002",
            rol_cuadrilla_id="AYUDANTE",
            activo=True,
        )
        url = reverse("cuadrillas:colaboradores_export")
        resp = self.client.get(url)
        self.assertEqual(resp.status_code, 200)
        import openpyxl

        wb = openpyxl.load_workbook(BytesIO(resp.content))
        ws = wb.active
        filas = {row[0]: row for row in ws.iter_rows(min_row=2, values_only=True)}
        self.assertIn("176sal-a6-002", filas)
        self.assertIn(filas["176sal-a6-002"][5], (None, ""))


# ---------------------------------------------------------------------------
# A7 — Copy de orden de dependencia Cargo -> Colaboradores
# ---------------------------------------------------------------------------
class TestA7CopyDependenciaCargoColaboradores(TestCase):
    """Texto de orden de dependencia visible en el modal de import de
    Colaboradores (mismo archivo que A6)."""

    def setUp(self):
        self.admin = _crear_admin("176sal-a7")
        self.client = Client()
        self.client.force_login(self.admin)

    def test_copy_orden_dependencia_visible_en_modal(self):
        resp = self.client.get(reverse("cuadrillas:colaboradores_lista"))
        self.assertEqual(resp.status_code, 200)
        content = resp.content.decode()
        self.assertIn("deben existir primero", content)
        self.assertIn(reverse("cuadrillas:cargos_lista"), content)

    def test_copy_queda_dentro_del_modal_importar_no_en_otro_lugar(self):
        """Edge case: el copy debe estar asociado al modal de import, no
        aparecer suelto en cualquier parte de la página (regresión de
        ubicación, similar al bounce=1 de #176 donde un boton quedo mezclado
        con el card equivocado)."""
        resp = self.client.get(reverse("cuadrillas:colaboradores_lista"))
        content = resp.content.decode()
        idx_modal = content.index('id="modal-importar-colaboradores"')
        idx_copy = content.index("deben existir primero")
        idx_cierre_modal = content.index("</form>", idx_modal)
        self.assertLess(idx_modal, idx_copy)
        self.assertLess(idx_copy, idx_cierre_modal)


# ---------------------------------------------------------------------------
# A8 (backfill A1) — CRUD Cargo con salario_base vía UI
# ---------------------------------------------------------------------------
class TestA1CargoCRUDSalarioBase(TestCase):
    """A1: Cargo.salario_base editable desde el CRUD /cuadrillas/cargos/
    (la migración 0022 y el modelo se verificaron en el commit de A1 contra
    la suite pre-existente tests_issue_176.py; este archivo consolida la
    cobertura formal específica de salario_base pedida por el plan)."""

    def setUp(self):
        self.admin = _crear_admin("176sal-a1")
        self.client = Client()
        self.client.force_login(self.admin)

    def test_crear_cargo_con_salario_persiste(self):
        url = reverse("cuadrillas:cargos_crear")
        resp = self.client.post(
            url, {"codigo": "QA_A1_SALARIO", "nombre": "Cargo Con Salario", "salario_base": "2100000", "activo": "on"}
        )
        self.assertIn(resp.status_code, (200, 302))
        cargo = Cargo.objects.get(codigo="QA_A1_SALARIO")
        self.assertEqual(cargo.salario_base, Decimal("2100000"))

    def test_crear_cargo_sin_salario_cae_a_default_cero(self):
        """Edge case: el campo es opcional (blank=True) -- si se omite del
        POST, cae al default del modelo (0), no revienta con IntegrityError
        por intentar guardar NULL."""
        url = reverse("cuadrillas:cargos_crear")
        resp = self.client.post(url, {"codigo": "QA_A1_SIN_SALARIO", "nombre": "Sin Salario", "activo": "on"})
        self.assertIn(resp.status_code, (200, 302))
        cargo = Cargo.objects.get(codigo="QA_A1_SIN_SALARIO")
        self.assertEqual(cargo.salario_base, Decimal("0"))

    def test_salario_negativo_rechazado_con_mensaje_dominio(self):
        """Edge case: el widget pone min=0 (solo HTML) -- el backend debe
        rechazar un POST directo con salario negativo, no lanzar 500 ni
        persistir un valor inválido."""
        url = reverse("cuadrillas:cargos_crear")
        resp = self.client.post(
            url, {"codigo": "QA_A1_NEGATIVO", "nombre": "Salario Negativo", "salario_base": "-500", "activo": "on"}
        )
        self.assertEqual(resp.status_code, 200)
        self.assertFalse(Cargo.objects.filter(codigo="QA_A1_NEGATIVO").exists())
        form = resp.context["form"]
        self.assertFalse(form.is_valid())
        self.assertIn("salario_base", form.errors)

    def test_editar_cargo_actualiza_salario_base(self):
        cargo = Cargo.objects.create(codigo="QA_A1_EDITAR", nombre="Editar Salario", salario_base=Decimal("100"))
        url = reverse("cuadrillas:cargos_editar", args=[cargo.pk])
        resp = self.client.post(
            url, {"codigo": "QA_A1_EDITAR", "nombre": "Editar Salario", "salario_base": "9999999", "activo": "on"}
        )
        self.assertIn(resp.status_code, (200, 302))
        cargo.refresh_from_db()
        self.assertEqual(cargo.salario_base, Decimal("9999999"))

    def test_columna_salario_base_visible_en_lista_con_separador_de_miles(self):
        Cargo.objects.create(codigo="QA_A1_LISTA", nombre="Cargo Lista", salario_base=Decimal("3500000"))
        resp = self.client.get(reverse("cuadrillas:cargos_lista"))
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "3.500.000")


# ---------------------------------------------------------------------------
# A8 (backfill A2) — CostoRolAPIView retrofit, cobertura formal
# ---------------------------------------------------------------------------
class TestA2CostoRolAPIRetrofit(TestCase):
    """A2: costo_rol_api lee Cargo.salario_base (retrofit del dict hardcoded
    huérfano). Verificado manualmente en el commit de A2 vía test Client
    directo; este archivo agrega la cobertura formal pytest."""

    def setUp(self):
        self.admin = _crear_admin("176sal-a2")
        self.client = Client()
        self.client.force_login(self.admin)

    def test_rol_existente_devuelve_salario_base_correcto(self):
        Cargo.objects.create(codigo="QA_A2_ROL", nombre="Rol QA A2", salario_base=Decimal("1234567.89"), activo=True)
        url = reverse("cuadrillas:costo_rol_api")
        resp = self.client.get(url, {"rol": "QA_A2_ROL"})
        self.assertEqual(resp.status_code, 200)
        data = resp.json()
        self.assertEqual(data["costo_dia"], 1234567.89)
        self.assertIsInstance(data["costo_dia"], float)

    def test_rol_inexistente_devuelve_cero_sin_500(self):
        url = reverse("cuadrillas:costo_rol_api")
        resp = self.client.get(url, {"rol": "QA_A2_NO_EXISTE"})
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.json()["costo_dia"], 0)

    def test_rol_vacio_devuelve_cero_sin_500(self):
        url = reverse("cuadrillas:costo_rol_api")
        resp = self.client.get(url, {"rol": ""})
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.json()["costo_dia"], 0)

    def test_cargo_inactivo_no_se_usa_para_autocompletar(self):
        """Edge case: un cargo INACTIVO no debe autocompletar salario (mismo
        criterio que el resto del maestro: solo activos son elegibles)."""
        Cargo.objects.create(
            codigo="QA_A2_INACTIVO", nombre="Rol Inactivo", salario_base=Decimal("5000000"), activo=False
        )
        url = reverse("cuadrillas:costo_rol_api")
        resp = self.client.get(url, {"rol": "QA_A2_INACTIVO"})
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.json()["costo_dia"], 0)

    def test_conductor_diferencia_interno_externo(self):
        # "CONDUCTOR" ya viene sembrado por el conftest autouse (14 códigos) --
        # actualizar en vez de crear, para no chocar con el unique de codigo.
        Cargo.objects.filter(codigo="CONDUCTOR").update(salario_base=Decimal("480000"), activo=True)
        url = reverse("cuadrillas:costo_rol_api")
        resp = self.client.get(url, {"rol": "CONDUCTOR", "conductor_interno": "false"})
        data = resp.json()
        self.assertTrue(data["es_conductor"])
        self.assertFalse(data["conductor_interno"])


# ---------------------------------------------------------------------------
# A8 — Integración: round-trip export -> import de Cargo
# ---------------------------------------------------------------------------
class TestA8IntegracionRoundTripCargo(TestCase):
    """Cross-cutting: exportar el catálogo Cargo y reimportarlo debe dejar
    los datos exactamente iguales (upsert idempotente sobre el propio
    export) -- integra A5 (import/export) sin duplicar ni perder datos."""

    def setUp(self):
        self.admin = _crear_admin("176sal-a8")
        self.client = Client()
        self.client.force_login(self.admin)

    def test_exportar_y_reimportar_cargo_es_idempotente(self):
        Cargo.objects.create(codigo="QA_A8_ROUNDTRIP", nombre="Round Trip", salario_base=Decimal("2222222"))
        total_antes = Cargo.objects.count()

        export_resp = self.client.get(reverse("cuadrillas:cargos_export"))
        self.assertEqual(export_resp.status_code, 200)

        archivo = BytesIO(export_resp.content)
        archivo.name = "cargos_export.xlsx"
        import_resp = self.client.post(reverse("cuadrillas:cargos_upload"), {"archivo": archivo})
        self.assertIn(import_resp.status_code, (200, 302))

        # Ningún cargo nuevo se creó (upsert por código, no duplica) y el
        # que ya existía conserva su salario.
        self.assertEqual(Cargo.objects.count(), total_antes)
        cargo = Cargo.objects.get(codigo="QA_A8_ROUNDTRIP")
        self.assertEqual(cargo.salario_base, Decimal("2222222"))

    def test_suite_completa_176_salario_no_deja_estado_cruzado(self):
        """Meta-test: confirma que el catálogo de Cargo sembrado por el
        conftest autouse (14 códigos) sigue intacto en cantidad mínima --
        anti falso-verde de que algún test anterior no lo haya vaciado por
        accidente (p.ej. un import mal armado que reemplace en vez de
        upsertear)."""
        codigos_base = {
            "SUPERVISOR", "LINIERO_I", "LINIERO_II", "AYUDANTE", "CONDUCTOR",
            "ADMINISTRADOR_OBRA", "PROFESIONAL_SST", "ING_RESIDENTE",
            "SERVICIO_GENERAL", "ALMACENISTA", "SUPERVISOR_FOREST",
            "ASISTENTE_FOREST", "MALACATERO", "COORDINADOR_HSQ",
        }
        existentes = set(Cargo.objects.values_list("codigo", flat=True))
        self.assertTrue(codigos_base.issubset(existentes))
