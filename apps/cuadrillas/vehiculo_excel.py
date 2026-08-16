"""Importación y exportación segura del catálogo de vehículos (#226)."""
from decimal import Decimal, InvalidOperation

from django.db import transaction
from openpyxl import load_workbook

from .models import Vehiculo


VEHICULO_HEADERS = [
    "PLACA", "MARCA", "TIPO", "DESCRIPCION", "ESTADO", "MODELO", "ANO",
    "CAPACIDAD_PERSONAS", "COSTO_DIA", "OBSERVACIONES",
]


def normalizar_encabezado(value):
    return str(value or "").strip().upper().replace("Ó", "O").replace("Ñ", "N")


class VehiculoExcelImporter:
    """Valida todas las filas antes de persistir cualquier vehículo."""

    def importar(self, archivo):
        try:
            workbook = load_workbook(archivo, read_only=True, data_only=True)
            rows = list(workbook.active.iter_rows(values_only=True))
            workbook.close()
        except Exception as exc:
            return {"exito": False, "creados": 0, "errores": [{"fila": 0, "mensaje": f"No se pudo leer el Excel: {exc}"}]}

        if not rows:
            return {"exito": False, "creados": 0, "errores": [{"fila": 1, "mensaje": "El archivo está vacío."}]}

        indices = {normalizar_encabezado(value): index for index, value in enumerate(rows[0]) if value}
        faltantes = [header for header in ("PLACA", "MARCA", "TIPO", "ESTADO") if header not in indices]
        if faltantes:
            return {"exito": False, "creados": 0, "errores": [{"fila": 1, "mensaje": "Faltan cabeceras requeridas: " + ", ".join(faltantes)}]}

        errores, preparados, placas_archivo = [], [], set()
        for numero_fila, row in enumerate(rows[1:], start=2):
            if not any(value not in (None, "") for value in row):
                continue
            data = {header: row[index] if index < len(row) else None for header, index in indices.items()}
            fila_errores, preparado = self._validar_fila(numero_fila, data, placas_archivo)
            errores.extend(fila_errores)
            if preparado:
                preparados.append(preparado)

        placas_existentes = set(Vehiculo.objects.filter(placa__in=[item["placa"] for item in preparados]).values_list("placa", flat=True))
        for item in preparados:
            if item["placa"] in placas_existentes:
                errores.append({"fila": item["_fila"], "mensaje": f'La placa "{item["placa"]}" ya existe.'})
        if errores:
            return {"exito": False, "creados": 0, "errores": errores}

        with transaction.atomic():
            # ``Vehiculo.save()`` sincroniza el puente legacy activo/estado;
            # no usar bulk_create porque omite deliberadamente ese contrato.
            for item in preparados:
                Vehiculo.objects.create(**{key: value for key, value in item.items() if key != "_fila"})
        return {"exito": True, "creados": len(preparados), "errores": []}

    def _validar_fila(self, fila, data, placas_archivo):
        errores = []
        placa = str(data.get("PLACA") or "").strip().upper()
        marca = str(data.get("MARCA") or "").strip()
        tipo = str(data.get("TIPO") or "").strip().upper()
        estado = str(data.get("ESTADO") or "").strip().upper().replace(" ", "_")
        if not placa:
            errores.append({"fila": fila, "mensaje": "La placa es obligatoria."})
        elif placa in placas_archivo:
            errores.append({"fila": fila, "mensaje": f'La placa "{placa}" está duplicada en el archivo.'})
        else:
            placas_archivo.add(placa)
        if not marca:
            errores.append({"fila": fila, "mensaje": "La marca es obligatoria."})
        if tipo not in Vehiculo.TipoVehiculo.values:
            errores.append({"fila": fila, "mensaje": f'El tipo "{tipo or "(vacío)"}" no es válido.'})
        if estado not in Vehiculo.Estado.values:
            errores.append({"fila": fila, "mensaje": f'El estado "{estado or "(vacío)"}" no es válido.'})
        try:
            ano = int(data["ANO"]) if data.get("ANO") not in (None, "") else None
            if ano is not None and not 1900 <= ano <= 2100:
                raise ValueError
        except (TypeError, ValueError):
            errores.append({"fila": fila, "mensaje": "El año debe estar entre 1900 y 2100."})
            ano = None
        try:
            capacidad = int(data["CAPACIDAD_PERSONAS"]) if data.get("CAPACIDAD_PERSONAS") not in (None, "") else 5
            if capacidad < 1:
                raise ValueError
        except (TypeError, ValueError):
            errores.append({"fila": fila, "mensaje": "La capacidad debe ser de al menos una persona."})
            capacidad = 5
        try:
            costo = Decimal(str(data["COSTO_DIA"])) if data.get("COSTO_DIA") not in (None, "") else Decimal("0")
            if costo < 0:
                raise InvalidOperation
        except (InvalidOperation, ValueError):
            errores.append({"fila": fila, "mensaje": "El costo por día debe ser un número no negativo."})
            costo = Decimal("0")
        if errores:
            return errores, None
        return [], {"_fila": fila, "placa": placa, "marca": marca, "tipo": tipo, "estado": estado,
                    "descripcion": str(data.get("DESCRIPCION") or "").strip(), "modelo": str(data.get("MODELO") or "").strip(),
                    "ano": ano, "capacidad_personas": capacidad, "costo_dia": costo,
                    "observaciones": str(data.get("OBSERVACIONES") or "").strip()}
