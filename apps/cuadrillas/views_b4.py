"""
Views B4 — Carga masiva de Cuadrillas con formato Aviso SAP.

Issue: Indunnova16/Instelec#105

Provee:
- ``CuadrillaUploadView``: formulario que recibe Excel con formato Aviso SAP
  (encabezado de cuadrilla + filas de miembros) y delega en
  ``CuadrillaImporter``.
- ``DescargarPlantillaCuadrillasView``: genera Excel-plantilla con encabezados
  + 3 filas de ejemplo (2 cuadrillas, formato heredado por miembros).

Las URLs se exponen mediante ``urlpatterns`` al final del archivo y son
importadas vía ``try/except`` desde ``apps/cuadrillas/urls.py``.

NOTA convivencia con vista pre-block: ``CuadrillaMasivaUploadView`` (en
``views.py``) usa formato "una cuadrilla por fila + miembros en columnas
H..L". B4 NO la reemplaza — usa la URL ``b4/upload-cuadrillas/`` para no
chocar. Cuando el cliente valide B4, podremos retirar el view legacy.
"""
from io import BytesIO

from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpResponse
from django.shortcuts import render
from django.urls import path
from django.views import View
from django.views.generic import TemplateView

from apps.core.mixins import RoleRequiredMixin

from .importers import (
    CuadrillaImporter,
    ProgramacionS18CuadrillaImporter,
    detectar_formato_cuadrillas,
)


class CuadrillaUploadView(LoginRequiredMixin, RoleRequiredMixin, TemplateView):
    """Carga masiva de cuadrillas desde Excel.

    Auto-detecta el formato del archivo y enruta al importer adecuado:

    - **Programación S18** (#124): filas agrupadas por actividad, encargado
      marcado con ROL=JT/CTA, código de cuadrilla generado automáticamente.
    - **Aviso SAP** (#105): una fila con columna CUADRILLA (código) + miembros.
    """

    template_name = 'cuadrillas/cuadrilla_upload.html'
    # RBAC v2 (#44): roles admin_* pasan vía RoleRequiredMixin automáticamente.
    allowed_roles = ['admin', 'director', 'coordinador', 'ing_residente']

    FORMATO_LABELS = {
        'S18': 'Programación S18 (agrupado por actividad)',
        'AVISO_SAP': 'Carga Masiva Simple (Aviso SAP)',
    }

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Cargar Cuadrillas Masivamente'
        return context

    def post(self, request, *args, **kwargs):
        archivo = request.FILES.get('archivo_cuadrillas')
        if not archivo:
            return render(request, self.template_name, {
                'titulo': 'Cargar Cuadrillas Masivamente',
                'error': 'Por favor selecciona un archivo Excel.',
            })

        # Validar extensión.
        if not archivo.name.lower().endswith(('.xlsx', '.xls')):
            return render(request, self.template_name, {
                'titulo': 'Cargar Cuadrillas Masivamente',
                'error': f'Archivo "{archivo.name}" no es Excel (.xlsx/.xls).',
            })

        # Leer el archivo a memoria una sola vez: lo usamos para detectar el
        # formato y luego para importar (rebobinando el puntero en cada paso).
        from io import BytesIO
        datos = BytesIO(archivo.read())

        datos.seek(0)
        formato = detectar_formato_cuadrillas(datos)

        opciones = {
            'actualizar_existentes': request.POST.get('actualizar_existentes') == 'on',
            'crear_usuarios_faltantes': request.POST.get('crear_usuarios_faltantes') == 'on',
        }

        datos.seek(0)
        if formato == 'S18':
            importer = ProgramacionS18CuadrillaImporter()
        else:
            importer = CuadrillaImporter()
        resultado = importer.importar(datos, opciones)

        return render(request, self.template_name, {
            'titulo': 'Cargar Cuadrillas Masivamente',
            'resultado': resultado,
            'formato_detectado': self.FORMATO_LABELS.get(formato, formato),
            'mensaje_exito': resultado.get('exito', False),
            'error': resultado.get('error') if not resultado.get('exito') else None,
        })


class DescargarPlantillaCuadrillasB4View(LoginRequiredMixin, RoleRequiredMixin, View):
    """Descarga plantilla Excel para CuadrillaUploadView (formato Aviso SAP)."""

    allowed_roles = ['admin', 'director', 'coordinador', 'ing_residente']

    def get(self, request, *args, **kwargs):
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = 'Cuadrillas'

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1F3A93', end_color='1F3A93', fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center')

        headers = [
            '#', 'CUADRILLA', 'LÍNEA', 'SUPERVISOR', 'PERSONAL',
            'CEDULA', 'CARGO', 'CELULAR', 'PLACA', 'ESTADO', 'OBSERVACIONES',
        ]

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        # 3 filas de ejemplo: cuadrilla 1 con 2 miembros, cuadrilla 2 con 1 miembro.
        ejemplo = [
            [1, 'CUA-001', 'L1', 'Juan Pérez', 'Carlos González', '1055688', 'Liniero',
             '3161234567', 'JAK-520', 'Activa', 'Cuadrilla principal sector norte'],
            ['', '', '', '', 'María Rodríguez', '1098765', 'Ayudante',
             '3167654321', '', '', ''],
            [2, 'CUA-002', 'L2', 'Carlos López', 'Luis Martínez', '1033333', 'Supervisor',
             '3197654321', 'HMV-123', 'Activa', ''],
        ]
        for row_idx, row_data in enumerate(ejemplo, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Anchos de columna.
        anchos = [5, 14, 8, 22, 25, 14, 18, 14, 12, 10, 35]
        for idx, ancho in enumerate(anchos, start=1):
            ws.column_dimensions[chr(64 + idx)].width = ancho

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename=plantilla_cuadrillas_aviso_sap.xlsx'
        return response


class DescargarPlantillaProgramacionS18View(LoginRequiredMixin, RoleRequiredMixin, View):
    """Descarga plantilla Excel en formato "Programación S18" (issue #124).

    Una hoja semanal con encabezado en fila 2 y un ejemplo de actividad con
    encargado (ROL=JT/CTA) + 3 miembros, replicando la estructura del archivo
    real del cliente.
    """

    allowed_roles = ['admin', 'director', 'coordinador', 'ing_residente']

    def get(self, request, *args, **kwargs):
        from datetime import date

        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = '18'  # nombre de hoja = número de semana

        # Fila 1: banner.
        ws.cell(row=1, column=1, value='INSTELEC SAS - NIT 890911324')
        ws.cell(row=1, column=13, value='Fecha de envio:')
        ws.cell(row=1, column=16, value=date.today())

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1F3A93', end_color='1F3A93', fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center')

        headers = [
            '#', 'ACTIVIDAD', 'LINEA', 'TRAMO', 'INICIO', 'FIN', 'PERSONAL',
            'CEDULA', 'CELULAR', 'CARGO', 'ROL', 'PLACA', 'AVISOS', 'ORDEN',
            'PT SAP', 'Comentarios',
        ]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        # Ejemplo: 1 actividad, encargado (JT/CTA) + 3 miembros.
        # [#, ACTIVIDAD, LINEA, TRAMO, INICIO, FIN, PERSONAL, CEDULA, CELULAR,
        #  CARGO, ROL, PLACA, AVISOS, ORDEN, PT SAP, Comentarios]
        fi, ff = date(2026, 4, 27), date(2026, 5, 3)
        filas = [
            [1, 'Servidumbre Completa', '817/818', '42 - 75', fi, ff,
             'JUAN PEREZ GOMEZ', '1055688', '3161234567', 'LINIERO I', 'JT/CTA',
             None, '5720794', '35123907', 'T0055348', ''],
            [None, None, None, None, None, None, 'CARLOS GONZALEZ RUIZ', '1004488',
             '3167654321', 'LINIERO II', None, None, None, None, None, ''],
            [None, None, None, None, None, None, 'LUIS MARTINEZ DIAZ', '1098766',
             '3197654321', 'AYUDANTE', None, None, None, None, None, ''],
            [None, None, None, None, None, None, 'PEDRO ROJAS LARA', '72132633',
             '3211234567', 'CONDUCTOR', None, 'GUX-177', None, None, None, ''],
        ]
        for row_idx, row_data in enumerate(filas, start=3):
            for col_idx, value in enumerate(row_data, start=1):
                if value is not None:
                    ws.cell(row=row_idx, column=col_idx, value=value)

        anchos = [4, 22, 12, 12, 11, 11, 28, 13, 13, 12, 9, 10, 12, 12, 12, 18]
        for idx, ancho in enumerate(anchos, start=1):
            ws.column_dimensions[ws.cell(row=2, column=idx).column_letter].width = ancho

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename=plantilla_programacion_s18.xlsx'
        return response


# Re-export para compatibilidad con el spec del issue. Algunos templates legacy
# de Instelec usan el nombre corto ``DescargarPlantillaCuadrillasView``; pero
# ya hay uno con ese nombre en ``views.py`` (formato pre-block). Mantenemos un
# alias explícito B4 para evitar colisión.
DescargarPlantillaCuadrillasView = DescargarPlantillaCuadrillasB4View


# urlpatterns expuesto aquí por contrato con ``apps/cuadrillas/urls.py``
# (que hace ``urlpatterns += views_b4.urlpatterns``). Se mantiene en sync
# con ``urls_b4.py`` que es el archivo "canónico" del bloque.
urlpatterns = [
    path('b4/upload-cuadrillas/', CuadrillaUploadView.as_view(), name='b4_upload_cuadrillas'),
    path('b4/descargar-plantilla/', DescargarPlantillaCuadrillasB4View.as_view(),
         name='b4_descargar_plantilla'),
    path('b4/descargar-plantilla-s18/', DescargarPlantillaProgramacionS18View.as_view(),
         name='b4_descargar_plantilla_s18'),
]
