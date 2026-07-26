"""Tests #178 — Sprint (2026-07-25), sub-item T1: suite de INTEGRACIÓN +
regresión final.

Issue: Indunnova16/Instelec#178

Los 8 sub-items del sprint (M1, F1, B1, A1, D1, F2, A2, C1) ya tienen su
propio archivo de tests aislado con happy + >=2 edge cases cada uno,
incluyendo el mapeo CARGO/ROL no invertido de A2
(``tests_issue_178_a2.py``) y la validación fin<inicio de D1
(``tests_issue_178_d1.py``) — no se duplican acá.

T1 aporta lo que los tests aislados por sub-item NO pueden atrapar: bugs de
INTERACCIÓN entre >=2 sub-items, ejercitando el flujo real (importer real,
endpoints reales vía Django test client) en vez de mocks:

1. A1 -> A2: un bloque importado desde un Excel S18 real (A1 persiste
   ``fecha_fin``) se exporta con ``ProgramacionSemanalHorizontalExporter``
   (A2) — la columna FIN debe salir del ``fecha_fin`` que A1 persistió, no
   quedar vacía ni recalcularse de otro campo. Los tests de A2 (aislados)
   crean la ``Cuadrilla`` directo, sin pasar por el importer real — este es
   el primer test que encadena las DOS piezas.
2. D1 -> C1: un bloque con horario planeado (D1) se reprograma (C1) — se
   afirma EXPLÍCITAMENTE cuál es el comportamiento real observado en el
   código (el origen retiene su horario planeado sin alterarse; el bloque
   nuevo nace SIN horario planeado porque
   ``ProgramacionSemanalBloqueReprogramarView.post`` no lo copia al crear),
   no se asume.
3. F1 + F2: 2 cuadrillas en semanas DISTINTAS, una con ubicación GPS
   automática (origen=auto) y otra con ubicación manual creada vía el
   endpoint real de F2 (``TrackingUbicacionManualCreateView``) — filtrar el
   mapa por semana (F1) deja SOLO la cuadrilla de esa semana, con el origen
   correcto en su marcador.

Al final del archivo: regresión — se corre ``apps/cuadrillas`` completo
(268 tests previos + los de acá) como parte del self-verify de este
sub-item, no dentro de este archivo (ver comando en el docstring de
ejecución).

Ejecutar:
  DJANGO_SETTINGS_MODULE=config.settings.dev_lite \
    venv/bin/python -m pytest apps/cuadrillas/tests_issue_178_demo0725.py -v \
    -o python_files="tests_*.py test_*.py"
"""

import json
from datetime import date, time
from io import BytesIO

from django.contrib.auth import get_user_model
from django.test import TestCase
from django.urls import reverse
from openpyxl import load_workbook

from apps.cuadrillas.importers import ProgramacionS18CuadrillaImporter
from apps.cuadrillas.models import Cuadrilla, TrackingUbicacion
from apps.cuadrillas.tests_s18 import _act, _build_s18_excel, _crear_linea

Usuario = get_user_model()


def _crear_admin(sufijo):
    # Rol admin (NO superuser) para ejercitar el gate real de RoleRequiredMixin
    # en los 3 escenarios (export, reprogramar, mapa/ubicación manual).
    return Usuario.objects.create_user(
        email=f"admin_178t1_{sufijo}@test.com",
        password="testpass123!",
        first_name="Admin",
        last_name="T1",
        rol="admin",
        is_staff=True,
    )


def _crear_usuario(documento, nombre):
    partes = nombre.split(maxsplit=1)
    return Usuario.objects.create(
        email=f"{documento}@test.local",
        documento=documento,
        first_name=partes[0],
        last_name=partes[1] if len(partes) > 1 else "",
        rol="liniero",
        is_active=True,
    )


class TestIntegracionA1ImporterA2Exportador(TestCase):
    """A1 (importer S18 persiste fecha_fin) -> A2 (exportador horizontal lee
    fecha_fin) — flujo real encadenado, sin mocks."""

    def setUp(self):
        self.admin = _crear_admin("a1a2")
        self.client.force_login(self.admin)

    def test_integracion_fin_exportado_coincide_con_fecha_fin_importada(self):
        """happy: importar un bloque S18 con INICIO+FIN reales (A1 persiste
        fecha_fin) -> exportar Excel horizontal (A2) -> la columna FIN del
        Excel exportado coincide EXACTO con la fecha importada -- no vacía,
        no re-derivada de otro campo (ej. fecha de inicio, u observaciones)."""
        _crear_linea("LN178T1")
        _crear_usuario("1143246999", "QA_E2E_ PEDRO SUAREZ")

        # sheet_name="46" -> semana=46 (_numero_semana extrae el dígito del
        # nombre de hoja); fecha_inicio.year=2026 -> anio=2026 (bloque['anio']
        # = fecha_inicio.year). Coincide con los args [2026, 46] del export.
        excel = _build_s18_excel(
            [
                _act(
                    1,
                    "Servidumbre Completa",
                    "178T1",
                    date(2026, 11, 9),
                    date(2026, 11, 13),
                    "QA_E2E_ PEDRO SUAREZ",
                    "1143246999",
                    "LINIERO I",
                    "JT/CTA",
                ),
            ],
            sheet_name="46",
        )

        res = ProgramacionS18CuadrillaImporter().importar(excel)
        self.assertTrue(res["exito"], res.get("errores"))

        cuad = Cuadrilla.objects.get()
        # Pre-condición (ya cubierta por A1 aislado, se re-afirma acá como
        # punto de partida del encadenamiento): A1 persistió fecha_fin real.
        self.assertEqual(cuad.fecha_fin, date(2026, 11, 13))
        self.assertTrue(cuad.codigo.startswith("46-2026-"))

        resp = self.client.get(reverse("cuadrillas:semanal_exportar_horizontal", args=[2026, 46]))
        self.assertEqual(resp.status_code, 200)

        wb = load_workbook(BytesIO(resp.content))
        fila = next(wb.active.iter_rows(min_row=2, values_only=True))
        # Columnas: #(0) ACTIVIDAD(1) LINEA(2) TRAMO(3) INICIO(4) FIN(5) ...
        self.assertEqual(fila[4], "09/11/2026")
        self.assertEqual(fila[5], "13/11/2026")
        self.assertIsNotNone(fila[5])
        self.assertNotEqual(fila[5], "")

        # El export leyó cuadrilla.fecha_fin directo (ProgramacionSemanalHorizontalExporter.
        # generar_excel pasa cuadrilla.fecha_fin a _escribir_bloque) -- confirma
        # que es EL MISMO valor que A1 persistió, no un recálculo paralelo.
        self.assertEqual(fila[5], cuad.fecha_fin.strftime("%d/%m/%Y"))


class TestIntegracionD1HorarioPlaneadoC1Reprogramar(TestCase):
    """D1 (hora_inicio_planeada/hora_fin_planeada) -> C1 (reprogramar) —
    verifica el comportamiento REAL, sin asumirlo."""

    def setUp(self):
        self.admin = _crear_admin("d1c1")
        self.client.force_login(self.admin)

    def test_integracion_origen_conserva_horario_nuevo_nace_sin_horario(self):
        """happy: bloque con horario planeado (D1) reprogramado (C1) ->
        el bloque ORIGEN retiene su horario planeado original sin alterarse
        (ProgramacionSemanalBloqueReprogramarView.post solo actualiza
        `fecha_fin`/`updated_at` del origen vía update_fields, ver
        views_semanal.py) Y el bloque NUEVO nace SIN horario planeado
        (el Cuadrilla.objects.create() del bloque nuevo NO pasa
        hora_inicio_planeada/hora_fin_planeada) -- comportamiento aceptable
        (la nueva actividad/ubicación puede requerir otro horario) pero debe
        quedar afirmado explícitamente, no asumido."""
        origen = Cuadrilla.objects.create(
            codigo="46-2026-0020-MAN",
            nombre="QA_E2E_ Bloque con horario planeado",
            activa=True,
            fecha=date(2026, 11, 9),
            hora_inicio_planeada=time(7, 30),
            hora_fin_planeada=time(16, 0),
        )

        resp = self.client.post(
            reverse("cuadrillas:semanal_bloque_reprogramar", args=[origen.pk]),
            {
                "nombre": "QA_E2E_ Reprogramado tras horario planeado",
                "fecha_desde": "2026-11-11",
                "motivo": "Emergencia sector norte",
            },
        )
        self.assertEqual(resp.status_code, 200)

        origen.refresh_from_db()
        self.assertEqual(origen.fecha_fin, date(2026, 11, 10))  # truncado, D1/C1 ya cubierto por C1
        # Comportamiento afirmado 1: el origen NO pierde su horario planeado.
        self.assertEqual(origen.hora_inicio_planeada, time(7, 30))
        self.assertEqual(origen.hora_fin_planeada, time(16, 0))

        nuevo = Cuadrilla.objects.exclude(pk=origen.pk).get(reprogramado_desde=origen)
        # Comportamiento afirmado 2: el bloque nuevo NO hereda el horario
        # planeado del origen -- nace con ambos campos en None.
        self.assertIsNone(nuevo.hora_inicio_planeada)
        self.assertIsNone(nuevo.hora_fin_planeada)


class TestIntegracionF1FiltroSemanaF2UbicacionManual(TestCase):
    """F1 (filtro por semana en el Mapa) + F2 (entrada manual de
    coordenadas) — 2 cuadrillas en semanas distintas, origen auto vs
    manual."""

    def setUp(self):
        self.admin = _crear_admin("f1f2")
        self.client.force_login(self.admin)

    def _ubicaciones_json(self, **params):
        resp = self.client.get(
            reverse("cuadrillas:mapa_partial"),
            params,
            HTTP_ACCEPT="application/json",
        )
        self.assertEqual(resp.status_code, 200)
        return json.loads(resp.content)["ubicaciones"]

    def test_integracion_filtro_semana_distingue_origen_auto_y_manual(self):
        """happy: cuadrilla A (semana 40) con ping GPS automático (origen=auto,
        NO pasa por F2) y cuadrilla B (semana 41) con ubicación reportada vía
        el endpoint REAL de F2 (origen=manual) -> filtrar el mapa por cada
        semana deja SOLO la cuadrilla correspondiente, con su origen
        correcto."""
        u_auto = _crear_usuario("178T1-100", "QA_E2E_ TRACKER AUTO")

        cuadrilla_auto = Cuadrilla.objects.create(
            codigo="40-2026-0030-MAN",
            nombre="QA_E2E_ Cuadrilla GPS automatico",
            activa=True,
            fecha=date(2026, 9, 28),
        )
        # Ping GPS "de fábrica" -- origen default (AUTO) del modelo, no pasa
        # por el mini-form/endpoint manual de F2.
        TrackingUbicacion.objects.create(
            cuadrilla=cuadrilla_auto,
            usuario=u_auto,
            latitud="4.60000000",
            longitud="-74.08000000",
        )

        cuadrilla_manual = Cuadrilla.objects.create(
            codigo="41-2026-0031-MAN",
            nombre="QA_E2E_ Cuadrilla ubicacion manual",
            activa=True,
            fecha=date(2026, 10, 5),
        )
        resp_manual = self.client.post(
            reverse("cuadrillas:ubicacion_manual_crear", args=[cuadrilla_manual.pk]),
            {"lat": "4.65000000", "lng": "-74.10000000"},
        )
        self.assertEqual(resp_manual.status_code, 200)
        body_manual = json.loads(resp_manual.content)
        self.assertTrue(body_manual["ok"])
        self.assertEqual(body_manual["ubicacion"]["origen"], "manual")

        # Filtro semana 40/2026 -> SOLO la cuadrilla GPS automática, origen 'auto'.
        ubicaciones_40 = {
            u["cuadrilla_codigo"]: u for u in self._ubicaciones_json(anio=2026, semana=40)
        }
        self.assertIn(cuadrilla_auto.codigo, ubicaciones_40)
        self.assertNotIn(cuadrilla_manual.codigo, ubicaciones_40)
        self.assertEqual(ubicaciones_40[cuadrilla_auto.codigo]["origen"], "auto")

        # Filtro semana 41/2026 -> SOLO la cuadrilla con ubicación manual, origen 'manual'.
        ubicaciones_41 = {
            u["cuadrilla_codigo"]: u for u in self._ubicaciones_json(anio=2026, semana=41)
        }
        self.assertIn(cuadrilla_manual.codigo, ubicaciones_41)
        self.assertNotIn(cuadrilla_auto.codigo, ubicaciones_41)
        self.assertEqual(ubicaciones_41[cuadrilla_manual.codigo]["origen"], "manual")
