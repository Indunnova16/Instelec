"""
Importers para Cuadrillas desde Excel.

`CuadrillaImporter` lee un archivo con el formato de Avisos SAP donde:

- Fila con columna ``#`` no vacía = encabezado de cuadrilla.
- Filas siguientes con ``#`` vacío = miembros de la cuadrilla previa.

Sigue el patrón de ``ProgramacionSemanalImporter`` en
``apps/actividades/importers.py``: ``COLUMN_MAPPINGS`` para detección flexible
de encabezados + transacción atómica + resumen detallado de creación/error.
"""
import logging
from datetime import date

from django.db import IntegrityError, transaction
from openpyxl import load_workbook

logger = logging.getLogger(__name__)


# Costo diario por defecto por rol — alineado con CuadrillaMasivaUploadView
# pre-block. Si el rol no se mapea, queda en 0 y el coordinador lo corrige.
COSTOS_POR_ROL = {
    'SUPERVISOR': 0,
    'LINIERO_I': 3176095,
    'LINIERO_II': 2804856,
    'AYUDANTE': 1750905,
    'CONDUCTOR': 480000,
    'ADMINISTRADOR_OBRA': 2522400,
    'PROFESIONAL_SST': 4204000,
    'ING_RESIDENTE': 7357000,
    'SERVICIO_GENERAL': 1750905,
    'ALMACENISTA': 1800000,
    'SUPERVISOR_FOREST': 2969427,
    'ASISTENTE_FOREST': 4204000,
}


# Mapeo de texto libre del Excel ("Liniero", "Ayudante", etc.) a los choices
# del modelo CuadrillaMiembro.RolCuadrilla.
ROL_TEXTO_A_CHOICE = {
    'supervisor': 'SUPERVISOR',
    'liniero': 'LINIERO_I',
    'liniero i': 'LINIERO_I',
    'liniero 1': 'LINIERO_I',
    'liniero ii': 'LINIERO_II',
    'liniero 2': 'LINIERO_II',
    'ayudante': 'AYUDANTE',
    'conductor': 'CONDUCTOR',
    'administrador obra': 'ADMINISTRADOR_OBRA',
    'administrador de obra': 'ADMINISTRADOR_OBRA',
    'profesional sst': 'PROFESIONAL_SST',
    'sst': 'PROFESIONAL_SST',
    'ingeniero residente': 'ING_RESIDENTE',
    'ing residente': 'ING_RESIDENTE',
    'ing. residente': 'ING_RESIDENTE',
    'servicio general': 'SERVICIO_GENERAL',
    'almacenista': 'ALMACENISTA',
    'supervisor forestal': 'SUPERVISOR_FOREST',
    'asistente forestal': 'ASISTENTE_FOREST',
}


class CuadrillaImporter:
    """
    Importa cuadrillas desde un Excel con formato Aviso SAP.

    Estructura:
        - Fila con ``#`` no vacío → encabezado de cuadrilla (crea/actualiza).
        - Filas siguientes con ``#`` vacío → miembros de la cuadrilla previa.

    Detecta columnas automáticamente por nombre (acepta variaciones).
    Operación atómica: si falla la importación de una cuadrilla por error
    fatal (línea inexistente), revierte TODO. Cédulas/vehículos no
    encontrados solo generan advertencias (la cuadrilla se crea sin ellos).
    """

    COLUMN_MAPPINGS = {
        'numero': ['#', 'no', 'no.', 'item', 'num', 'numero'],
        'cuadrilla': ['cuadrilla', 'codigo', 'codigo cuadrilla', 'codigo_cuadrilla', 'código cuadrilla'],
        'linea': ['linea', 'línea', 'lineas', 'codigo linea', 'código línea'],
        'supervisor': ['supervisor', 'super', 'super.', 'supervisor nombre', 'jefe cuadrilla'],
        'personal': ['personal', 'nombre', 'nombres', 'nombre miembro'],
        'cedula': ['cedula', 'cédula', 'documento', 'doc', 'identificacion', 'identificación'],
        'cargo': ['cargo', 'rol', 'rol cuadrilla', 'puesto'],
        'celular': ['celular', 'telefono', 'teléfono', 'cel', 'celular personal'],
        'placa': ['placa', 'vehiculo', 'vehículo', 'veh', 'vehiculo asignado'],
        'estado': ['estado', 'activa', 'activo'],
        'observaciones': ['observaciones', 'notas', 'comentarios', 'obs'],
    }

    def __init__(self):
        self.errores = []
        self.advertencias = []
        self.cuadrillas_creadas = 0
        self.cuadrillas_actualizadas = 0
        self.miembros_agregados = 0
        self.column_indices = {}

    # ---------- API pública ----------

    def importar(self, archivo_excel, opciones=None):
        """
        Importa cuadrillas desde archivo Excel.

        Args:
            archivo_excel: file-like object o ruta al Excel.
            opciones: dict con flags:
                - actualizar_existentes (bool): si True, actualiza cuadrillas
                  que ya existen por código. Si False, las deja intactas.

        Returns:
            dict con resumen:
                - exito (bool)
                - error (str, solo si no éxito)
                - cuadrillas_creadas (int)
                - cuadrillas_actualizadas (int)
                - miembros_agregados (int)
                - advertencias (list[str])
                - errores (list[str])
                - columnas_detectadas (list[str])
        """
        opciones = opciones or {}
        actualizar_existentes = opciones.get('actualizar_existentes', False)

        try:
            workbook = load_workbook(archivo_excel, read_only=True, data_only=True)
            sheet = workbook.active
        except Exception as e:
            logger.error(f"Error cargando Excel: {e}")
            return self._resultado_error(f'Error al cargar archivo Excel: {e}')

        rows = list(sheet.iter_rows(values_only=True))
        try:
            workbook.close()
        except Exception:
            pass

        if not rows:
            return self._resultado_error('El archivo está vacío')

        self._detectar_columnas(rows[0])

        if 'cuadrilla' not in self.column_indices:
            return self._resultado_error('No se encontró columna "CUADRILLA" en el archivo')
        if 'linea' not in self.column_indices:
            return self._resultado_error('No se encontró columna "LÍNEA" en el archivo')

        # Procesar — agrupamos por cuadrilla (encabezado + miembros).
        cuadrilla_actual = None
        miembros_actuales = []
        cuadrillas_bloque = []  # list of (cuadrilla_dict, miembros_list, row_num)

        for row_num, row in enumerate(rows[1:], start=2):
            if not row or all(c is None or str(c).strip() == '' for c in row):
                continue

            numero = self._get_cell_value(row, 'numero')

            if numero is not None and str(numero).strip() != '':
                # Encabezado de nueva cuadrilla.
                if cuadrilla_actual is not None:
                    cuadrillas_bloque.append((cuadrilla_actual, miembros_actuales, cuadrilla_actual['_row_num']))
                cuadrilla_actual = self._parsear_cuadrilla(row)
                cuadrilla_actual['_row_num'] = row_num
                miembros_actuales = []
                # La fila de encabezado puede TAMBIÉN contener el primer
                # miembro (formato Aviso SAP del issue #105).
                personal = self._get_cell_value(row, 'personal')
                cedula = self._get_cell_value(row, 'cedula')
                if personal or cedula:
                    miembros_actuales.append({
                        'nombre': self._str_or_empty(personal),
                        'cedula': self._str_or_empty(cedula),
                        'cargo': self._str_or_empty(self._get_cell_value(row, 'cargo')),
                        'celular': self._str_or_empty(self._get_cell_value(row, 'celular')),
                        '_row_num': row_num,
                    })
            else:
                if cuadrilla_actual is None:
                    self.advertencias.append(
                        f'Fila {row_num}: miembro sin cuadrilla previa, omitido'
                    )
                    continue
                personal = self._get_cell_value(row, 'personal')
                cedula = self._get_cell_value(row, 'cedula')
                if personal or cedula:
                    miembros_actuales.append({
                        'nombre': self._str_or_empty(personal),
                        'cedula': self._str_or_empty(cedula),
                        'cargo': self._str_or_empty(self._get_cell_value(row, 'cargo')),
                        'celular': self._str_or_empty(self._get_cell_value(row, 'celular')),
                        '_row_num': row_num,
                    })

        # Última cuadrilla pendiente.
        if cuadrilla_actual is not None:
            cuadrillas_bloque.append((cuadrilla_actual, miembros_actuales, cuadrilla_actual['_row_num']))

        if not cuadrillas_bloque:
            return self._resultado_error('No se encontraron cuadrillas en el archivo')

        # Transacción atómica — un error fatal revierte todo.
        try:
            with transaction.atomic():
                for cuadrilla_data, miembros, row_num in cuadrillas_bloque:
                    self._guardar_cuadrilla(cuadrilla_data, miembros, actualizar_existentes, row_num)
                if self.errores:
                    # Errores fatales: rollback explícito.
                    raise _RollbackImport(f'{len(self.errores)} errores fatales')
        except _RollbackImport:
            logger.warning('CuadrillaImporter: rollback por errores fatales')
            return {
                'exito': False,
                'error': 'Importación revertida por errores fatales (ver lista)',
                'cuadrillas_creadas': 0,
                'cuadrillas_actualizadas': 0,
                'miembros_agregados': 0,
                'advertencias': self.advertencias,
                'errores': self.errores,
                'columnas_detectadas': list(self.column_indices.keys()),
            }
        except Exception as e:
            logger.exception('CuadrillaImporter: error inesperado')
            return self._resultado_error(f'Error inesperado: {e}')

        return {
            'exito': True,
            'cuadrillas_creadas': self.cuadrillas_creadas,
            'cuadrillas_actualizadas': self.cuadrillas_actualizadas,
            'miembros_agregados': self.miembros_agregados,
            'advertencias': self.advertencias,
            'errores': self.errores,
            'columnas_detectadas': list(self.column_indices.keys()),
        }

    # ---------- internos ----------

    def _detectar_columnas(self, header_row):
        for col_idx, cell_value in enumerate(header_row):
            if cell_value is None:
                continue
            cell_lower = str(cell_value).lower().strip()
            for field_name, posibles in self.COLUMN_MAPPINGS.items():
                if cell_lower in posibles:
                    self.column_indices[field_name] = col_idx
                    break
        logger.info(f'CuadrillaImporter columnas detectadas: {self.column_indices}')

    def _get_cell_value(self, row, field_name):
        if field_name not in self.column_indices:
            return None
        idx = self.column_indices[field_name]
        if idx < len(row):
            return row[idx]
        return None

    def _str_or_empty(self, value, default=''):
        if value is None:
            return default
        text = str(value).strip()
        return text if text else default

    def _parsear_cuadrilla(self, row):
        return {
            'codigo': self._str_or_empty(self._get_cell_value(row, 'cuadrilla')),
            'linea_nombre': self._str_or_empty(self._get_cell_value(row, 'linea')),
            'supervisor_nombre': self._str_or_empty(self._get_cell_value(row, 'supervisor')),
            'placa': self._str_or_empty(self._get_cell_value(row, 'placa')),
            'estado': self._str_or_empty(self._get_cell_value(row, 'estado'), 'Activa'),
            'observaciones': self._str_or_empty(self._get_cell_value(row, 'observaciones')),
        }

    def _guardar_cuadrilla(self, cuadrilla_data, miembros, actualizar, row_num):
        from apps.cuadrillas.models import Cuadrilla, CuadrillaMiembro, Vehiculo
        from apps.lineas.models import Linea
        from apps.usuarios.models import Usuario

        codigo = cuadrilla_data['codigo']
        if not codigo:
            self.errores.append(f'Fila {row_num}: código de cuadrilla vacío')
            return

        linea_nombre = cuadrilla_data['linea_nombre']
        if not linea_nombre:
            self.errores.append(f'Fila {row_num}: línea vacía para cuadrilla {codigo}')
            return

        # Resolver línea (error fatal si no existe).
        try:
            linea = Linea.objects.get(nombre__iexact=linea_nombre)
        except Linea.DoesNotExist:
            try:
                linea = Linea.objects.get(codigo__iexact=linea_nombre)
            except Linea.DoesNotExist:
                self.errores.append(
                    f'Fila {row_num}: línea "{linea_nombre}" no existe para cuadrilla {codigo}'
                )
                return

        # Resolver supervisor (warning si no existe; deja null).
        supervisor = None
        if cuadrilla_data['supervisor_nombre']:
            sup_nombre = cuadrilla_data['supervisor_nombre']
            partes = sup_nombre.split()
            supervisor = Usuario.objects.filter(
                first_name__icontains=partes[0] if partes else sup_nombre,
                is_active=True,
            ).first()
            if not supervisor:
                self.advertencias.append(
                    f'Fila {row_num}: supervisor "{sup_nombre}" no encontrado, queda sin asignar'
                )

        # Resolver vehículo (warning si no existe; deja null).
        vehiculo = None
        if cuadrilla_data['placa']:
            placa = cuadrilla_data['placa']
            vehiculo = Vehiculo.objects.filter(placa__iexact=placa, activo=True).first()
            if not vehiculo:
                self.advertencias.append(
                    f'Fila {row_num}: vehículo placa "{placa}" no encontrado, queda sin asignar'
                )

        activa = cuadrilla_data['estado'].lower() in ['activa', 'active', 'activo', 'true', '1', 'si', 'sí']

        # Crear o actualizar.
        existente = Cuadrilla.objects.filter(codigo=codigo).first()

        if existente is None:
            # B3: campos auditoría (motivo_desactivacion, fecha_desactivacion,
            # desactivado_por) quedan vacíos al crear cuadrilla nueva activa.
            # Si el schema tiene los campos NOT NULL pero el modelo aún no
            # los expone (escenario worktree B4 pre-merge), seteamos via
            # kwargs defensivos detectados a runtime.
            create_kwargs = {
                'codigo': codigo,
                'nombre': f'Cuadrilla {codigo}',
                'linea_asignada': linea,
                'supervisor': supervisor,
                'vehiculo': vehiculo,
                'activa': activa,
                'observaciones': cuadrilla_data['observaciones'],
            }
            # Detectar y completar campos auditoría B3 si existen en el modelo.
            modelo_fields = {f.name for f in Cuadrilla._meta.get_fields()}
            if 'motivo_desactivacion' in modelo_fields:
                create_kwargs['motivo_desactivacion'] = ''
            cuadrilla = self._crear_cuadrilla_b3_safe(Cuadrilla, create_kwargs)
            self.cuadrillas_creadas += 1
        else:
            if actualizar:
                # B3: preservar histórico — si la cuadrilla está inactiva con
                # campos auditoría poblados (motivo_desactivacion, etc.), no
                # los tocamos. Solo actualizamos relaciones operativas.
                existente.linea_asignada = linea
                existente.supervisor = supervisor
                existente.vehiculo = vehiculo
                # Si la cuadrilla viene Activa en el Excel y estaba inactiva,
                # solo la reactivamos; el campo motivo_desactivacion se
                # conserva como histórico (decisión: NO limpiarlo aquí, eso
                # lo hace explícitamente la vista de reactivar de B3).
                existente.activa = activa
                if cuadrilla_data['observaciones']:
                    existente.observaciones = cuadrilla_data['observaciones']
                existente.save()
                self.cuadrillas_actualizadas += 1
            cuadrilla = existente

        # Miembros.
        for miembro_data in miembros:
            self._agregar_miembro(cuadrilla, miembro_data)

    def _crear_cuadrilla_b3_safe(self, modelo, kwargs):
        """Crea Cuadrilla manejando el caso pre-merge B3.

        Si el schema de BD tiene columnas auditoría B3 (motivo_desactivacion,
        fecha_desactivacion, desactivado_por_id) como NOT NULL pero el modelo
        Python aún no las expone (worktree B4 corre antes que F4 merge B3),
        usamos raw SQL desde el inicio para evitar IntegrityError dentro de
        la transacción atómica externa.
        """
        from django.db import connection

        schema_cols = self._cuadrilla_schema_cols()
        if 'motivo_desactivacion' not in schema_cols:
            # Schema "limpio": ORM funciona normalmente.
            return modelo.objects.create(**kwargs)

        # Schema con B3: insertar via raw SQL completando audit fields vacíos.
        import uuid
        from django.utils import timezone

        new_id = uuid.uuid4()
        now = timezone.now()
        with connection.cursor() as cursor:
            cursor.execute(
                """
                INSERT INTO cuadrillas
                    (id, created_at, updated_at, codigo, nombre, activa,
                     observaciones, linea_asignada_id, supervisor_id,
                     vehiculo_id, fecha, motivo_desactivacion,
                     fecha_desactivacion, desactivado_por_id)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """,
                [
                    str(new_id), now, now,
                    kwargs['codigo'], kwargs['nombre'], kwargs['activa'],
                    kwargs.get('observaciones', ''),
                    kwargs['linea_asignada'].id if kwargs.get('linea_asignada') else None,
                    kwargs['supervisor'].id if kwargs.get('supervisor') else None,
                    kwargs['vehiculo'].id if kwargs.get('vehiculo') else None,
                    None, '', None, None,
                ],
            )
        return modelo.objects.get(id=new_id)

    @classmethod
    def _cuadrilla_schema_cols(cls):
        """Cachea las columnas reales de la tabla cuadrillas en la BD."""
        if hasattr(cls, '_schema_cols_cache'):
            return cls._schema_cols_cache
        from django.db import connection
        with connection.cursor() as cursor:
            cursor.execute(
                "SELECT column_name FROM information_schema.columns "
                "WHERE table_name = 'cuadrillas'"
            )
            cols = {row[0] for row in cursor.fetchall()}
        cls._schema_cols_cache = cols
        return cols

    def _agregar_miembro(self, cuadrilla, miembro_data):
        from apps.cuadrillas.models import CuadrillaMiembro
        from apps.usuarios.models import Usuario

        cedula = miembro_data['cedula']
        row_num = miembro_data.get('_row_num', '?')

        if not cedula:
            self.advertencias.append(f'Fila {row_num}: miembro sin cédula, omitido')
            return

        try:
            usuario = Usuario.objects.get(documento=cedula)
        except Usuario.DoesNotExist:
            self.advertencias.append(
                f'Fila {row_num}: usuario con cédula "{cedula}" no existe, miembro omitido'
            )
            return

        rol_choice = ROL_TEXTO_A_CHOICE.get(miembro_data['cargo'].lower(), 'LINIERO_I')

        _, creado = CuadrillaMiembro.objects.get_or_create(
            cuadrilla=cuadrilla,
            usuario=usuario,
            activo=True,
            defaults={
                'rol_cuadrilla': rol_choice,
                'cargo': 'MIEMBRO',
                'costo_dia': COSTOS_POR_ROL.get(rol_choice, 0),
                'fecha_inicio': date.today(),
            },
        )
        if creado:
            self.miembros_agregados += 1

    def _resultado_error(self, msg):
        return {
            'exito': False,
            'error': msg,
            'cuadrillas_creadas': 0,
            'cuadrillas_actualizadas': 0,
            'miembros_agregados': 0,
            'advertencias': self.advertencias,
            'errores': self.errores,
            'columnas_detectadas': list(self.column_indices.keys()),
        }


class _RollbackImport(Exception):
    """Señal interna para forzar rollback de la transacción."""
    pass


# ---------------------------------------------------------------------------
# Issue #124 — Formato "Programación S18" (agrupado por actividad)
# ---------------------------------------------------------------------------

# Palabras en la columna ROL que designan al Jefe de Trabajo / Encargado.
JT_KEYWORDS = ('JT', 'CTA', 'JEFE', 'ENCARGADO')


def detectar_formato_cuadrillas(archivo_excel):
    """Inspecciona el Excel y decide qué importer de cuadrillas usar.

    Devuelve:
        - ``'S18'``      → formato "Programación S18" (agrupado por actividad,
          columnas ACTIVIDAD/ROL/PERSONAL, sin código de cuadrilla).
        - ``'AVISO_SAP'``→ formato del issue #105 (columna CUADRILLA con código).

    Revisa las primeras 6 filas de cada hoja buscando la firma de encabezados.
    Si encuentra ACTIVIDAD+ROL+PERSONAL antes que CUADRILLA → S18. Si solo ve
    CUADRILLA → Aviso SAP. Por defecto (sin firma clara) cae a Aviso SAP para
    preservar compatibilidad hacia atrás con la plantilla heredada.
    """
    try:
        workbook = load_workbook(archivo_excel, read_only=True, data_only=True)
    except Exception:
        return 'AVISO_SAP'

    try:
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for i, row in enumerate(sheet.iter_rows(values_only=True)):
                if i >= 6:
                    break
                celdas = {str(c).lower().strip() for c in row if c is not None}
                if {'actividad', 'rol', 'personal'} <= celdas:
                    return 'S18'
                if 'cuadrilla' in celdas:
                    return 'AVISO_SAP'
    finally:
        try:
            workbook.close()
        except Exception:
            pass

    return 'AVISO_SAP'


class ProgramacionS18CuadrillaImporter:
    """Importa CUADRILLAS desde el Excel real "Programación - S18.xlsx".

    Diferencias frente a ``CuadrillaImporter`` (#105, formato Aviso SAP):

    - **No hay columna de código de cuadrilla** → se genera ``WW-YYYY-NNNN-AAA``
      (semana de la hoja, año del INICIO, número de actividad, iniciales).
    - Las filas vienen **agrupadas por actividad**: una fila con ``#`` numérico
      es el encabezado de la actividad (y su primer miembro); las filas
      siguientes con ``#`` vacío son miembros adicionales.
    - El **encargado** se marca con la columna ``ROL`` (``JT/CTA``) →
      ``CuadrillaMiembro.CargoJerarquico.JT_CTA``.
    - La columna ``CARGO`` ("LINIERO I", "AYUDANTE"...) mapea a ``RolCuadrilla``.
    - La ``PLACA`` aparece en la fila del conductor (no en el encabezado) → se
      toma la primera placa no vacía del grupo.
    - ``LINEA`` puede traer varios códigos (``809\\n808/807``) → primera válida.
    - Procesa **todas las hojas semanales** válidas (``02``..``18``, ``12 (2)``).

    Reglas de error (spec #124):
    - Línea no encontrada → advertencia + ``linea_asignada=None`` (NO fatal).
    - Cédula no encontrada → advertencia + miembro omitido (o se crea el usuario
      si ``crear_usuarios_faltantes=True``).
    - Cualquier excepción inesperada revierte toda la transacción.
    """

    COLUMN_MAPPINGS = {
        'numero':       ['#', 'no', 'no.', 'item'],
        'actividad':    ['actividad', 'tipo actividad', 'tipo de actividad'],
        'linea':        ['linea', 'línea', 'lineas', 'líneas'],
        'tramo':        ['tramo', 'tramos', 'sector', 'sección'],
        'fecha_inicio': ['inicio', 'fecha inicio', 'fecha de inicio'],
        'fecha_fin':    ['fin', 'fecha fin', 'fecha de fin'],
        'personal':     ['personal', 'nombre', 'nombres'],
        'cedula':       ['cedula', 'cédula', 'documento', 'identificacion'],
        'celular':      ['celular', 'telefono', 'teléfono', 'cel'],
        'cargo':        ['cargo'],
        'rol':          ['rol'],
        'placa':        ['placa', 'vehiculo', 'vehículo'],
        # Issue #105: el formato real "Programación - S22" trae AVISOS (col N)
        # y ORDEN (col O) en la fila de encabezado de la actividad. El modelo
        # Cuadrilla no tiene campos para ellas → se anteponen a observaciones.
        'avisos':       ['avisos', 'aviso'],
        'orden':        ['orden', 'ordenes', 'órdenes'],
        'observaciones': ['comentarios', 'observaciones', 'obs', 'notas'],
    }

    SHEETS_EXCLUIR = {'vc', 'hoja1', 'sheet1', 'resumen', 'instrucciones'}

    def __init__(self):
        self.errores = []
        self.advertencias = []
        self.cuadrillas_creadas = 0
        self.cuadrillas_actualizadas = 0
        # Issue #124: códigos de cuadrillas que YA existían y se OMITIERON
        # (no se crean ni se actualizan). Estrategia SALTAR + RESUMEN.
        self.cuadrillas_omitidas = []
        self.miembros_agregados = 0
        self.encargados_asignados = 0
        self.usuarios_creados = 0
        self.column_indices = {}
        self.sheets_procesadas = []
        # Issue #178 (A2): filas de la sección NOVEDADES, independientes de
        # cualquier bloque/actividad — ver `_guardar_novedad`.
        self.novedades_pendientes = []
        self.novedades_creadas = 0

    # ---------- API pública ----------

    def importar(self, archivo_excel, opciones=None):
        opciones = opciones or {}
        actualizar = opciones.get('actualizar_existentes', False)
        crear_usuarios = opciones.get('crear_usuarios_faltantes', False)

        try:
            # Issue #178 (A1): NO usar read_only=True — openpyxl no expone
            # `Worksheet.merged_cells` en modo read_only (AttributeError), y el
            # parser de bloques necesita el RANGO real de la celda combinada de
            # columna A/D para detectar el fin de bloque de forma 100% confiable
            # (antes se dependía implícitamente de que la celda "#" viniera en
            # blanco por el merge, sin verificar el merge en sí). El archivo es
            # una programación semanal (decenas de filas por hoja) — cargarlo
            # completo en memoria es seguro.
            workbook = load_workbook(archivo_excel, data_only=True)
        except Exception as e:
            logger.error(f"Error cargando Excel S18: {e}")
            return self._resultado_error(f'Error al cargar archivo Excel: {e}')

        # Recolectar las cuadrillas de todas las hojas semanales válidas.
        bloques = []  # list[dict] cuadrilla+miembros
        for sheet_name in workbook.sheetnames:
            if not self._es_hoja_semanal(sheet_name):
                continue
            try:
                semana = self._numero_semana(sheet_name)
                nuevos = self._parsear_hoja(workbook[sheet_name], semana)
                if nuevos:
                    bloques.extend(nuevos)
                    self.sheets_procesadas.append(sheet_name)
            except Exception as e:
                logger.exception(f"Error parseando hoja {sheet_name!r}")
                self.errores.append(f'Hoja {sheet_name}: {e}')

        try:
            workbook.close()
        except Exception:
            pass

        # Issue #178 (A2): una hoja puede traer SOLO novedades (sin ninguna
        # actividad real) — no es un error, sigue siendo un import válido.
        if not bloques and not self.novedades_pendientes:
            return self._resultado_error(
                'No se encontraron cuadrillas en el archivo (¿formato Programación S18?)'
            )

        # Persistir en una transacción atómica.
        try:
            with transaction.atomic():
                for bloque in bloques:
                    self._guardar_bloque(bloque, actualizar, crear_usuarios)
                for novedad in self.novedades_pendientes:
                    self._guardar_novedad(novedad)
                if self.errores:
                    raise _RollbackImport(f'{len(self.errores)} errores fatales')
        except _RollbackImport:
            logger.warning('ProgramacionS18CuadrillaImporter: rollback por errores fatales')
            return {
                'exito': False,
                'error': 'Importación revertida por errores fatales (ver lista)',
                'formato': 'S18',
                'cuadrillas_creadas': 0,
                'cuadrillas_actualizadas': 0,
                'cuadrillas_omitidas': self.cuadrillas_omitidas,
                'cuadrillas_omitidas_count': len(self.cuadrillas_omitidas),
                'miembros_agregados': 0,
                'encargados_asignados': 0,
                'usuarios_creados': 0,
                'novedades_creadas': 0,
                'advertencias': self.advertencias,
                'errores': self.errores,
                'sheets_procesadas': self.sheets_procesadas,
            }
        except Exception:
            # Issue #124: NUNCA fugar el str crudo de la excepción (p.ej. el
            # "duplicate key value violates unique constraint ..." de psycopg2)
            # a la UI. Las colisiones de código ya se manejan vía savepoint en
            # _guardar_bloque; lo que llegue aquí es inesperado y se reporta con
            # un mensaje genérico amigable. El detalle queda en el log.
            logger.exception('ProgramacionS18CuadrillaImporter: error inesperado')
            return self._resultado_error(
                'Ocurrió un error inesperado al procesar el archivo. '
                'Revisa el formato e inténtalo de nuevo; si persiste, contacta a soporte.'
            )

        return {
            'exito': True,
            'formato': 'S18',
            'cuadrillas_creadas': self.cuadrillas_creadas,
            'cuadrillas_actualizadas': self.cuadrillas_actualizadas,
            'cuadrillas_omitidas': self.cuadrillas_omitidas,
            'cuadrillas_omitidas_count': len(self.cuadrillas_omitidas),
            'miembros_agregados': self.miembros_agregados,
            'encargados_asignados': self.encargados_asignados,
            'usuarios_creados': self.usuarios_creados,
            'novedades_creadas': self.novedades_creadas,
            'advertencias': self.advertencias,
            'errores': self.errores,
            'sheets_procesadas': self.sheets_procesadas,
        }

    # ---------- parseo ----------

    def _parsear_hoja(self, sheet, semana):
        """Devuelve lista de bloques {cuadrilla..., miembros[...]} de una hoja."""
        rows = list(sheet.iter_rows(values_only=True))
        if len(rows) < 3:
            return []

        header_idx = self._localizar_header(rows)
        if header_idx is None:
            self.advertencias.append(
                f'Hoja {sheet.title}: no se localizó encabezado; omitida'
            )
            return []

        self._detectar_columnas(rows[header_idx])
        requeridos = {'numero', 'actividad', 'personal'}
        faltantes = requeridos - set(self.column_indices)
        if faltantes:
            self.advertencias.append(
                f'Hoja {sheet.title}: columnas faltantes {sorted(faltantes)}; omitida'
            )
            return []

        # Issue #178 (A1): límites de bloque por celda combinada de columna A
        # (numero) o D (tramo) — 100% confiable (verificado contra 106 bloques
        # reales). Si la hoja no trae merges en ninguna de esas columnas (caso
        # raro/manual, p.ej. hojas construidas a mano sin formato), se cae al
        # heurístico legado (columna '#' en blanco = continuación del bloque
        # anterior) con una advertencia explícita.
        mapa_bloques = self._mapa_bloques_por_merge(
            sheet, self.column_indices.get('numero'), self.column_indices.get('tramo')
        )
        usa_fallback_legado = not mapa_bloques
        if usa_fallback_legado:
            self.advertencias.append(
                f'Hoja {sheet.title}: no se detectaron celdas combinadas en '
                f'columna numero/tramo; se usa el heurístico de respaldo '
                f'(columna "#" en blanco = continuación de bloque, menos confiable)'
            )

        bloques = []
        actual = None
        # Issue #178 (A2): al entrar a la sección NOVEDADES se cierra el
        # bloque activo y las filas de personal que siguen se persisten como
        # registro INDEPENDIENTE (no como miembro de la última actividad).
        en_novedades = False
        ultimo_anio = None

        for row_idx, row in enumerate(rows[header_idx + 1:], start=header_idx + 2):
            numero = self._get_cell(row, 'numero')
            numero_str = '' if numero is None else str(numero).strip()

            if numero_str.strip().upper() == 'NOVEDADES':
                if actual is not None:
                    bloques.append(actual)
                    actual = None
                en_novedades = True
                continue

            # Filas de ruido ('-', notas sueltas) → ignorar sin alterar el
            # modo NOVEDADES ni el bloque activo.
            if numero_str and not self._es_numero_actividad(numero_str):
                continue

            es_continuacion = self._es_fila_continuacion(
                row_idx, numero_str, mapa_bloques, usa_fallback_legado
            )

            if es_continuacion:
                if en_novedades:
                    novedad = self._parsear_novedad(row, row_idx, semana, ultimo_anio, sheet.title)
                    if novedad:
                        self.novedades_pendientes.append(novedad)
                elif actual is not None:
                    # Miembro adicional de la actividad en curso.
                    miembro = self._parsear_miembro(row, row_idx)
                    if miembro:
                        actual['miembros'].append(miembro)
                continue

            # Encabezado de nueva actividad → sale de NOVEDADES (si estaba) y
            # cierra el bloque anterior.
            en_novedades = False
            if actual is not None:
                bloques.append(actual)

            actividad = self._str(self._get_cell(row, 'actividad'))
            linea_raw = self._get_cell(row, 'linea')
            fecha_inicio = self._normalizar_fecha(self._get_cell(row, 'fecha_inicio'))
            fecha_fin = self._normalizar_fecha(self._get_cell(row, 'fecha_fin'))
            anio = fecha_inicio.year if fecha_inicio else date.today().year
            ultimo_anio = anio  # issue #178 (A2): contexto de año para NOVEDADES

            actual = {
                'semana': semana,
                'anio': anio,
                'numero': int(numero_str),
                'actividad': actividad,
                'linea_codigos': self._split_multi(linea_raw),
                'linea_str': self._primer_token(linea_raw),
                'fecha_inicio': fecha_inicio,
                'fecha_fin': fecha_fin,
                'observaciones': self._str(self._get_cell(row, 'observaciones')),
                # Issue #105: avisos/orden vienen en la fila de encabezado de la
                # actividad y pueden ser multilínea ("5720754\n5720792") → se
                # normalizan con _split_multi y se unen con ", ".
                'avisos': self._join_multi(self._get_cell(row, 'avisos')),
                'orden': self._join_multi(self._get_cell(row, 'orden')),
                'row_num': row_idx,
                'miembros': [],
            }
            # La fila de encabezado también porta el primer miembro.
            miembro = self._parsear_miembro(row, row_idx)
            if miembro:
                actual['miembros'].append(miembro)

        if actual is not None:
            bloques.append(actual)

        return bloques

    def _parsear_miembro(self, row, row_idx):
        nombre = self._str(self._get_cell(row, 'personal'))
        cedula = self._str(self._get_cell(row, 'cedula'))
        if not nombre and not cedula:
            return None
        return {
            'nombre': nombre,
            'cedula': cedula,
            'celular': self._str(self._get_cell(row, 'celular')),
            'cargo': self._str(self._get_cell(row, 'cargo')),
            'rol_raw': self._str(self._get_cell(row, 'rol')),
            'placa': self._str(self._get_cell(row, 'placa')),
            'es_jt': self._es_jt(self._get_cell(row, 'rol')),
            'row_num': row_idx,
        }

    def _parsear_novedad(self, row, row_idx, semana, anio, hoja_origen):
        """Fila de personal dentro de la sección NOVEDADES (issue #178, A2).

        La nota (p.ej. "Vacaciones", "Incapacidad") viene en la columna
        AVISOS de la fila, NO en una columna dedicada (confirmado contra el
        Excel real del cliente).
        """
        nombre = self._str(self._get_cell(row, 'personal'))
        cedula = self._str(self._get_cell(row, 'cedula'))
        if not nombre and not cedula:
            return None
        return {
            'nombre': nombre,
            'cedula': cedula,
            'cargo': self._str(self._get_cell(row, 'cargo')),
            'nota': self._join_multi(self._get_cell(row, 'avisos')),
            'semana': semana,
            'anio': anio or date.today().year,
            'hoja_origen': hoja_origen,
            'row_num': row_idx,
        }

    # ---------- persistencia ----------

    def _guardar_bloque(self, bloque, actualizar, crear_usuarios):
        from apps.cuadrillas.models import Cuadrilla, Vehiculo

        codigo = self._generar_codigo(
            bloque['semana'], bloque['anio'], bloque['numero'], bloque['actividad']
        )

        # Resolver línea (advertencia, NO fatal — spec #124).
        linea = self._buscar_linea(bloque['linea_codigos'])
        if linea is None and bloque['linea_codigos']:
            self.advertencias.append(
                f'Fila {bloque["row_num"]}: línea {bloque["linea_codigos"]} no '
                f'encontrada para cuadrilla {codigo}; queda sin línea asignada'
            )

        # Resolver vehículo por la primera placa no vacía del grupo.
        vehiculo = None
        placa = next((m['placa'] for m in bloque['miembros'] if m['placa']), '')
        if placa:
            vehiculo = Vehiculo.objects.filter(placa__iexact=placa, activo=True).first()
            if not vehiculo:
                self.advertencias.append(
                    f'Cuadrilla {codigo}: vehículo placa "{placa}" no encontrado, '
                    f'queda sin asignar'
                )

        nombre = self._nombre_cuadrilla(bloque)
        observaciones = self._construir_observaciones(bloque)

        existente = Cuadrilla.objects.filter(codigo=codigo).first()
        if existente is None:
            # Issue #124 — estrategia SALTAR + RESUMEN. El pre-check de arriba es
            # no-atómico: entre el filter() y el create() otra fila del mismo lote
            # (o una carrera) puede insertar el código. Sin aislamiento, ese
            # IntegrityError envenena la transacción atómica externa y aborta TODO
            # el lote, fugando el error crudo de Postgres a la UI. Envolvemos el
            # create en un savepoint anidado: si choca con la UNIQUE
            # (cuadrillas_codigo_key), revertimos SOLO este savepoint, registramos
            # el código como omitido y seguimos con el resto del lote.
            try:
                with transaction.atomic():
                    cuadrilla = Cuadrilla.objects.create(
                        codigo=codigo,
                        nombre=nombre,
                        linea_asignada=linea,
                        vehiculo=vehiculo,
                        fecha=bloque['fecha_inicio'],
                        activa=True,
                        observaciones=observaciones,
                    )
            except IntegrityError:
                # El código ya existía (lo creó otra fila del lote o una carrera).
                # Lo omitimos sin abortar el batch y sin fugar el error técnico.
                self.cuadrillas_omitidas.append(codigo)
                self.advertencias.append(
                    f'Cuadrilla {codigo} ya existe; omitida (marca "actualizar" para sobrescribir)'
                )
                return
            self.cuadrillas_creadas += 1
        elif actualizar:
            existente.nombre = nombre
            existente.linea_asignada = linea
            existente.vehiculo = vehiculo
            existente.fecha = bloque['fecha_inicio']
            if observaciones:
                existente.observaciones = observaciones
            existente.save()
            cuadrilla = existente
            self.cuadrillas_actualizadas += 1
        else:
            # Issue #124 — existe y NO se pidió actualizar → SALTAR.
            self.cuadrillas_omitidas.append(codigo)
            self.advertencias.append(
                f'Cuadrilla {codigo} ya existe; omitida (marca "actualizar" para sobrescribir)'
            )
            return

        # Miembros.
        supervisor_usuario = None
        for miembro in bloque['miembros']:
            usuario = self._agregar_miembro(cuadrilla, miembro, bloque, crear_usuarios)
            if usuario and miembro['es_jt'] and supervisor_usuario is None:
                supervisor_usuario = usuario

        # Si el encargado (JT) es un usuario con rol supervisor, lo fijamos como
        # supervisor de la cuadrilla (encargado ≠ supervisor, pero pueden coincidir).
        if supervisor_usuario is not None and getattr(supervisor_usuario, 'rol', '') == 'supervisor':
            cuadrilla.supervisor = supervisor_usuario
            cuadrilla.save(update_fields=['supervisor', 'updated_at'])

    def _agregar_miembro(self, cuadrilla, miembro, bloque, crear_usuarios):
        from apps.cuadrillas.models import CuadrillaMiembro
        from apps.usuarios.models import Usuario

        cedula = miembro['cedula']
        row_num = miembro['row_num']
        if not cedula:
            self.advertencias.append(f'Fila {row_num}: miembro sin cédula, omitido')
            return None

        usuario = Usuario.objects.filter(documento=cedula).first()
        if usuario is None:
            if crear_usuarios:
                usuario = self._crear_usuario(miembro)
                if usuario is None:
                    return None
            else:
                self.advertencias.append(
                    f'Fila {row_num}: usuario con cédula "{cedula}" '
                    f'({miembro["nombre"]}) no existe, miembro omitido'
                )
                return None

        rol_choice = ROL_TEXTO_A_CHOICE.get(miembro['cargo'].lower(), 'LINIERO_I')
        cargo_jerarquico = 'JT_CTA' if miembro['es_jt'] else 'MIEMBRO'
        fecha_inicio = bloque['fecha_inicio'] or date.today()

        obj, creado = CuadrillaMiembro.objects.get_or_create(
            cuadrilla=cuadrilla,
            usuario=usuario,
            activo=True,
            defaults={
                'rol_cuadrilla': rol_choice,
                'cargo': cargo_jerarquico,
                'costo_dia': COSTOS_POR_ROL.get(rol_choice, 0),
                'fecha_inicio': fecha_inicio,
            },
        )
        if creado:
            self.miembros_agregados += 1
            if cargo_jerarquico == 'JT_CTA':
                self.encargados_asignados += 1
        elif miembro['es_jt'] and obj.cargo != 'JT_CTA':
            # Re-import: promover a encargado si el Excel ahora lo marca como JT.
            obj.cargo = 'JT_CTA'
            obj.save(update_fields=['cargo', 'updated_at'])
            self.encargados_asignados += 1
        return usuario

    def _crear_usuario(self, miembro):
        from apps.usuarios.models import Usuario

        cedula = miembro['cedula']
        nombre = miembro['nombre'] or f'Usuario {cedula}'
        partes = nombre.split()
        first = partes[0] if partes else nombre
        last = ' '.join(partes[1:]) if len(partes) > 1 else ''
        email = f'{cedula}@instelec-import.local'
        try:
            # Issue #124: savepoint anidado para que una colisión UNIQUE
            # (email/documento ya existente por carrera) NO envenene la txn del
            # lote — sin él, el except de abajo atraparía la excepción pero la
            # transacción externa quedaría rota igual.
            with transaction.atomic():
                usuario = Usuario.objects.create(
                    email=email,
                    first_name=first[:150],
                    last_name=last[:150],
                    documento=cedula,
                    telefono=miembro['celular'][:20],
                    rol='operario_general',
                    is_active=True,
                )
                usuario.set_unusable_password()
                usuario.save(update_fields=['password'])
            self.usuarios_creados += 1
            self.advertencias.append(
                f'Fila {miembro["row_num"]}: usuario {nombre} (cédula {cedula}) '
                f'creado automáticamente con email {email}'
            )
            return usuario
        except Exception as e:
            self.advertencias.append(
                f'Fila {miembro["row_num"]}: no se pudo crear usuario {nombre} '
                f'(cédula {cedula}): {e}'
            )
            return None

    def _guardar_novedad(self, novedad):
        """Persiste una fila de NOVEDADES como registro independiente
        (issue #178, A2) — sin FK a Cuadrilla/Actividad. get_or_create sobre
        (cedula, semana, anio, nota) para que un re-import no duplique."""
        from apps.cuadrillas.models import NovedadPersonalSemana

        _, creado = NovedadPersonalSemana.objects.get_or_create(
            cedula=novedad['cedula'],
            semana=novedad['semana'],
            anio=novedad['anio'],
            nota=novedad['nota'],
            defaults={
                'nombre': novedad['nombre'],
                'cargo': novedad['cargo'],
                'hoja_origen': novedad['hoja_origen'],
            },
        )
        if creado:
            self.novedades_creadas += 1

    # ---------- helpers ----------

    def _generar_codigo(self, semana, anio, numero, actividad):
        import re
        iniciales = re.sub(r'[^A-Z]', '', (actividad or '').upper())[:3] or 'ACT'
        return f'{int(semana):02d}-{anio}-{int(numero):04d}-{iniciales}'

    def _nombre_cuadrilla(self, bloque):
        actividad = bloque['actividad'] or 'Cuadrilla'
        linea = bloque['linea_str']
        fi = bloque['fecha_inicio']
        ff = bloque['fecha_fin']
        fechas = ''
        if fi and ff:
            fechas = f' ({fi.strftime("%d/%m/%Y")} - {ff.strftime("%d/%m/%Y")})'
        elif fi:
            fechas = f' ({fi.strftime("%d/%m/%Y")})'
        base = f'{actividad} - {linea}{fechas}' if linea else f'{actividad}{fechas}'
        return base[:100]

    @staticmethod
    def _construir_observaciones(bloque):
        """Antepone AVISOS/ORDEN a las observaciones (issue #105).

        El modelo Cuadrilla no tiene campos para avisos/orden, así que se
        preservan como prefijo legible en ``observaciones`` (text NOT NULL):
        ``"Avisos: 5720754, 5720792 | Orden: 1, 2 | <obs original>"``.
        """
        partes = []
        if bloque.get('avisos'):
            partes.append(f'Avisos: {bloque["avisos"]}')
        if bloque.get('orden'):
            partes.append(f'Orden: {bloque["orden"]}')
        obs = bloque.get('observaciones') or ''
        if obs:
            partes.append(obs)
        return ' | '.join(partes)

    def _detectar_columnas(self, header_row):
        self.column_indices = {}
        for col_idx, value in enumerate(header_row):
            if value is None:
                continue
            cell_lower = str(value).lower().strip()
            for field_name, posibles in self.COLUMN_MAPPINGS.items():
                if cell_lower in posibles and field_name not in self.column_indices:
                    self.column_indices[field_name] = col_idx
                    break

    def _get_cell(self, row, field_name):
        idx = self.column_indices.get(field_name)
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    @staticmethod
    def _str(value, default=''):
        if value is None:
            return default
        text = str(value).strip()
        return text if text else default

    def _primer_token(self, value):
        partes = self._split_multi(value)
        return partes[0] if partes else ''

    @staticmethod
    def _split_multi(value):
        """Lista de strings limpios separados por \\n, /, ',' o ';'."""
        if value is None:
            return []
        texto = str(value).strip()
        if not texto or texto == '-':
            return []
        import re
        return [p.strip() for p in re.split(r'[\n/;,]+', texto) if p.strip()]

    @classmethod
    def _join_multi(cls, value):
        """Normaliza un valor multilínea (avisos/orden) a "a, b, c" (issue #105)."""
        return ', '.join(cls._split_multi(value))

    @staticmethod
    def _es_jt(rol_value):
        if rol_value is None:
            return False
        s = str(rol_value).upper()
        return any(k in s for k in JT_KEYWORDS)

    @staticmethod
    def _mapa_bloques_por_merge(sheet, col_numero, col_alt=None):
        """Mapa ``{fila_excel_1based: 'inicio'|'continuacion'}`` derivado del
        RANGO de la celda combinada en la columna ``numero`` (o ``col_alt``,
        p.ej. ``tramo``/columna D, como respaldo si ``numero`` no viene
        combinada) de ``sheet`` (issue #178, A1).

        Requiere que ``sheet`` se haya cargado SIN ``read_only=True`` —
        openpyxl no expone ``merged_cells`` en modo read_only.

        Devuelve ``{}`` si no se encontró ningún merge de ≥2 filas en esas
        columnas (hoja "plana", sin formato de bloques) — el llamador debe
        entonces usar el heurístico de respaldo (columna '#' en blanco).
        """
        mapa = {}
        columnas = [c for c in (col_numero, col_alt) if c is not None]
        merges = getattr(sheet, 'merged_cells', None)
        if merges is None:
            return mapa
        for col_idx in columnas:
            col_excel = col_idx + 1  # openpyxl es 1-based
            encontrados = False
            for rango in merges.ranges:
                if (
                    rango.min_col == col_excel
                    and rango.max_col == col_excel
                    and rango.max_row > rango.min_row
                ):
                    encontrados = True
                    for fila in range(rango.min_row, rango.max_row + 1):
                        mapa[fila] = 'inicio' if fila == rango.min_row else 'continuacion'
            if encontrados:
                break
        return mapa

    @staticmethod
    def _es_fila_continuacion(row_idx, numero_str, mapa_bloques, usa_fallback_legado):
        """¿La fila `row_idx` es continuación del bloque anterior (miembro
        adicional) o el encabezado de una actividad nueva? (issue #178, A1)

        Primario: RANGO de la celda combinada (`mapa_bloques`). Respaldo
        (`usa_fallback_legado=True`, o fila fuera de cualquier merge conocido
        — p.ej. bloque de 1 sola fila que no genera merge): heurístico legado
        columna '#' en blanco = continuación.
        """
        if not usa_fallback_legado:
            estado = mapa_bloques.get(row_idx)
            if estado == 'continuacion':
                return True
            if estado == 'inicio':
                return False
        # Sin info de merge (fallback o fila fuera de rango conocido).
        return numero_str == ''

    @staticmethod
    def _es_hoja_semanal(sheet_name):
        import re
        nombre = sheet_name.strip().lower()
        if nombre in ProgramacionS18CuadrillaImporter.SHEETS_EXCLUIR:
            return False
        return bool(re.fullmatch(r's?(emana)?[\s_]*\d+(\s*\(\d+\))?', nombre))

    @staticmethod
    def _numero_semana(sheet_name):
        import re
        m = re.search(r'\d+', sheet_name)
        return int(m.group()) if m else 0

    @staticmethod
    def _localizar_header(rows):
        for idx in range(min(6, len(rows))):
            celdas = {str(c).lower().strip() for c in rows[idx] if c is not None}
            if '#' in celdas and 'actividad' in celdas and 'personal' in celdas:
                return idx
        return None

    @staticmethod
    def _es_numero_actividad(numero_str):
        s = str(numero_str).strip()
        return s.isdigit() and int(s) > 0

    @staticmethod
    def _normalizar_fecha(valor):
        if valor is None:
            return None
        from datetime import datetime
        if isinstance(valor, datetime):
            return valor.date()
        if isinstance(valor, date):
            return valor
        try:
            return datetime.strptime(str(valor)[:10], '%Y-%m-%d').date()
        except Exception:
            return None

    def _buscar_linea(self, codigos):
        from apps.lineas.models import Linea
        for codigo in codigos:
            codigo = str(codigo).strip()
            if not codigo:
                continue
            qs = Linea.objects.filter(codigo__iexact=codigo)
            if qs.exists():
                return qs.first()
            qs = Linea.objects.filter(codigo_transelca__iexact=codigo)
            if qs.exists():
                return qs.first()
            qs = Linea.objects.filter(codigo__icontains=codigo)
            if qs.exists():
                return qs.first()
        return None

    def _resultado_error(self, msg):
        return {
            'exito': False,
            'error': msg,
            'formato': 'S18',
            'cuadrillas_creadas': 0,
            'cuadrillas_actualizadas': 0,
            'cuadrillas_omitidas': self.cuadrillas_omitidas,
            'cuadrillas_omitidas_count': len(self.cuadrillas_omitidas),
            'miembros_agregados': 0,
            'encargados_asignados': 0,
            'usuarios_creados': 0,
            'novedades_creadas': 0,
            'advertencias': self.advertencias,
            'errores': self.errores,
            'sheets_procesadas': self.sheets_procesadas,
        }
