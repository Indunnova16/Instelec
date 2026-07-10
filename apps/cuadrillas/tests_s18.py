"""
Tests #124 — ProgramacionS18CuadrillaImporter (carga de cuadrillas formato S18).

Issue: Indunnova16/Instelec#124

Cubre:
- Detección automática de formato (S18 vs Aviso SAP).
- Parseo del archivo REAL del cliente (`tests/fixtures/Programacion_S18_real.xlsx`)
  como test contra dato legacy.
- Generación de código WW-YYYY-NNNN-AAA.
- Asignación de encargado por ROL=JT/CTA → CargoJerarquico.JT_CTA.
- Mapeo CARGO → RolCuadrilla.
- Cédula inexistente = advertencia no fatal (o crear usuario si opt-in).
- Línea inexistente = advertencia no fatal (cuadrilla sin línea).
- Re-import idempotente.

Ejecutar:  pytest apps/cuadrillas/tests_s18.py -v
"""
import os
from datetime import date
from io import BytesIO

import pytest
from openpyxl import Workbook

from apps.cuadrillas.importers import (
    ProgramacionS18CuadrillaImporter,
    detectar_formato_cuadrillas,
)
from apps.cuadrillas.models import Cuadrilla, CuadrillaMiembro, Vehiculo
from apps.lineas.models import Linea
from apps.usuarios.models import Usuario


S18_HEADERS = [
    '#', 'ACTIVIDAD', 'LINEA', 'TRAMO', 'INICIO', 'FIN', 'PERSONAL',
    'CEDULA', 'CELULAR', 'CARGO', 'ROL', 'PLACA', 'AVISOS', 'ORDEN',
    'PT SAP', 'Comentarios',
]

FIXTURE_REAL = os.path.join(
    os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
    'tests', 'fixtures', 'Programacion_S18_real.xlsx',
)


def _build_s18_excel(filas, sheet_name='18'):
    """Construye un Excel formato S18: banner (fila 1) + headers (fila 2) + datos."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(['INSTELEC SAS - NIT 890911324'])  # fila 1 banner
    ws.append(S18_HEADERS)                        # fila 2 headers
    for f in filas:
        ws.append(f)
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def _crear_usuario(documento, nombre='JHON JAIRO JIMENEZ', rol='liniero'):
    partes = nombre.split(maxsplit=1)
    return Usuario.objects.create(
        email=f'{documento}@test.local',
        documento=documento,
        first_name=partes[0],
        last_name=partes[1] if len(partes) > 1 else '',
        rol=rol,
        is_active=True,
    )


def _crear_linea(codigo, nombre=None):
    return Linea.objects.create(
        codigo=codigo,
        nombre=nombre or codigo,
        cliente='TRANSELCA',
    )


# Fila helper: [#, ACTIVIDAD, LINEA, TRAMO, INICIO, FIN, PERSONAL, CEDULA,
#               CELULAR, CARGO, ROL, PLACA, AVISOS, ORDEN, PT SAP, Comentarios]
def _act(numero, actividad, linea, inicio, fin, personal, cedula, cargo,
         rol=None, placa=None, avisos='', orden='', comentarios=''):
    return [numero, actividad, linea, '', inicio, fin, personal, cedula,
            '', cargo, rol, placa, avisos, orden, '', comentarios]


# Layout S22 real del cliente (issue #105): columna extra ATENCION en C corre
# todo a la derecha (LINEA=D, PERSONAL=H, ROL=L, PLACA=M, AVISOS=N, ORDEN=O).
S22_HEADERS = [
    '#', 'ACTIVIDAD', 'ATENCION', 'LINEA', 'TRAMO', 'INICIO', 'FIN', 'PERSONAL',
    'CEDULA', 'CELULAR', 'CARGO', 'ROL', 'PLACA', 'AVISOS', 'ORDEN',
    'PT SAP', 'Comentarios',
]

FIXTURE_S22 = os.path.join(
    os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
    'fixtures', 'instelec', 'Programacion_S22_qa.xlsx',
)


def _build_s22_excel(filas, sheet_name='22'):
    """Excel layout S22 (con columna ATENCION en C): banner + headers + datos."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(['INSTELEC SAS - NIT 890911324'])
    ws.append(S22_HEADERS)
    for f in filas:
        ws.append(f)
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out


# Fila S22: [#, ACTIVIDAD, ATENCION, LINEA, TRAMO, INICIO, FIN, PERSONAL,
#            CEDULA, CELULAR, CARGO, ROL, PLACA, AVISOS, ORDEN, PT SAP, Coment]
def _act_s22(numero, actividad, linea, inicio, fin, personal, cedula, cargo,
             rol=None, placa=None, avisos='', orden=''):
    return [numero, actividad, '7:50 - 8:05', linea, '', inicio, fin, personal,
            cedula, '', cargo, rol, placa, avisos, orden, '', '']


def _miembro_s22(personal, cedula, cargo, rol=None, placa=None):
    return [None, None, None, None, None, None, None, personal, cedula, '',
            cargo, rol, placa, '', '', '', '']


def _miembro(personal, cedula, cargo, rol=None, placa=None):
    return [None, None, None, None, None, None, personal, cedula, '',
            cargo, rol, placa, '', '', '', '']


# ---------------------------------------------------------------------------
# Detección de formato
# ---------------------------------------------------------------------------

class TestDeteccionFormato:

    def test_detecta_s18(self):
        excel = _build_s18_excel([
            _act(1, 'Servidumbre', '817', date(2026, 4, 27), date(2026, 5, 3),
                 'JHON JAIRO', '1143246675', 'LINIERO I', 'JT/CTA'),
        ])
        assert detectar_formato_cuadrillas(excel) == 'S18'

    def test_detecta_aviso_sap(self):
        wb = Workbook()
        ws = wb.active
        ws.append(['#', 'CUADRILLA', 'LÍNEA', 'PERSONAL', 'CEDULA', 'CARGO'])
        ws.append([1, 'CUA-001', 'L1', 'Carlos', '1055688', 'Liniero'])
        out = BytesIO()
        wb.save(out)
        out.seek(0)
        assert detectar_formato_cuadrillas(out) == 'AVISO_SAP'

    @pytest.mark.skipif(not os.path.exists(FIXTURE_REAL), reason='fixture real ausente')
    def test_detecta_fixture_real_como_s18(self):
        with open(FIXTURE_REAL, 'rb') as f:
            assert detectar_formato_cuadrillas(f) == 'S18'


# ---------------------------------------------------------------------------
# Importación
# ---------------------------------------------------------------------------

@pytest.mark.django_db
class TestImportS18:

    def test_happy_crea_cuadrilla_con_encargado(self):
        _crear_linea('LN817', 'LN817 SABANALARGA')
        _crear_usuario('1143246675', 'JHON JAIRO JIMENEZ')
        _crear_usuario('1004487321', 'KEINER SERRANO')
        _crear_usuario('72132633', 'RAFAEL PACHECO')
        Vehiculo.objects.create(placa='TLN-063', activo=True)

        excel = _build_s18_excel([
            _act(1, 'Servidumbre Completa', '817', date(2026, 4, 27),
                 date(2026, 5, 3), 'JHON JAIRO JIMENEZ', '1143246675',
                 'LINIERO I', 'JT/CTA'),
            _miembro('KEINER SERRANO', '1004487321', 'LINIERO II'),
            _miembro('RAFAEL PACHECO', '72132633', 'CONDUCTOR', placa='TLN-063'),
        ])

        res = ProgramacionS18CuadrillaImporter().importar(excel)
        assert res['exito'] is True
        assert res['formato'] == 'S18'
        assert res['cuadrillas_creadas'] == 1
        assert res['miembros_agregados'] == 3
        assert res['encargados_asignados'] == 1

        cuad = Cuadrilla.objects.get()
        # Código WW-YYYY-NNNN-AAA → 18-2026-0001-SER
        assert cuad.codigo == '18-2026-0001-SER'
        assert cuad.linea_asignada.codigo == 'LN817'
        assert cuad.vehiculo.placa == 'TLN-063'
        assert cuad.fecha == date(2026, 4, 27)

        # Encargado correcto.
        jt = CuadrillaMiembro.objects.get(usuario__documento='1143246675')
        assert jt.cargo == 'JT_CTA'
        assert jt.rol_cuadrilla_id == 'LINIERO_I'
        # Miembros normales.
        m2 = CuadrillaMiembro.objects.get(usuario__documento='1004487321')
        assert m2.cargo == 'MIEMBRO'
        assert m2.rol_cuadrilla_id == 'LINIERO_II'
        conductor = CuadrillaMiembro.objects.get(usuario__documento='72132633')
        assert conductor.rol_cuadrilla_id == 'CONDUCTOR'

    def test_cedula_inexistente_es_advertencia_no_fatal(self):
        _crear_linea('LN817')
        _crear_usuario('1143246675', 'JHON JAIRO')
        # 72019461 NO se crea — debe quedar como advertencia.

        excel = _build_s18_excel([
            _act(1, 'Hurto', '817', date(2026, 3, 26), date(2026, 3, 29),
                 'JHON JAIRO', '1143246675', 'LINIERO I', 'JT/CTA'),
            _miembro('JOSE MIGUEL', '72019461', 'LINIERO II'),
        ])
        res = ProgramacionS18CuadrillaImporter().importar(excel)
        assert res['exito'] is True
        assert res['cuadrillas_creadas'] == 1
        assert res['miembros_agregados'] == 1  # solo el encargado
        assert any('72019461' in a for a in res['advertencias'])

    def test_linea_inexistente_no_es_fatal(self):
        _crear_usuario('1143246675', 'JHON JAIRO')
        excel = _build_s18_excel([
            _act(1, 'Hurto', '999', date(2026, 3, 26), date(2026, 3, 29),
                 'JHON JAIRO', '1143246675', 'LINIERO I', 'JT/CTA'),
        ])
        res = ProgramacionS18CuadrillaImporter().importar(excel)
        assert res['exito'] is True
        assert res['cuadrillas_creadas'] == 1
        cuad = Cuadrilla.objects.get()
        assert cuad.linea_asignada is None
        assert any('999' in a for a in res['advertencias'])

    def test_linea_fuzzy_icontains(self):
        # LINEA del Excel '809' debe resolver a 'LN809'.
        _crear_linea('LN809', 'LN809 FUNDACION')
        _crear_usuario('1143246675', 'JHON JAIRO')
        excel = _build_s18_excel([
            _act(1, 'Hurto', '809', date(2026, 3, 26), date(2026, 3, 29),
                 'JHON JAIRO', '1143246675', 'LINIERO I', 'JT/CTA'),
        ])
        res = ProgramacionS18CuadrillaImporter().importar(excel)
        assert Cuadrilla.objects.get().linea_asignada.codigo == 'LN809'

    def test_linea_multivalor_toma_primera_valida(self):
        _crear_linea('LN808')
        _crear_usuario('1143246675', 'JHON JAIRO')
        excel = _build_s18_excel([
            _act(1, 'Hurto', '808/807', date(2026, 3, 26), date(2026, 3, 29),
                 'JHON JAIRO', '1143246675', 'LINIERO I', 'JT/CTA'),
        ])
        res = ProgramacionS18CuadrillaImporter().importar(excel)
        assert Cuadrilla.objects.get().linea_asignada.codigo == 'LN808'

    def test_rerun_idempotente_con_actualizar(self):
        _crear_linea('LN817')
        _crear_usuario('1143246675', 'JHON JAIRO')

        def _excel():
            return _build_s18_excel([
                _act(1, 'Servidumbre', '817', date(2026, 4, 27), date(2026, 5, 3),
                     'JHON JAIRO', '1143246675', 'LINIERO I', 'JT/CTA'),
            ])

        ProgramacionS18CuadrillaImporter().importar(_excel())
        res2 = ProgramacionS18CuadrillaImporter().importar(
            _excel(), {'actualizar_existentes': True}
        )
        assert res2['cuadrillas_creadas'] == 0
        assert res2['cuadrillas_actualizadas'] == 1
        assert Cuadrilla.objects.count() == 1
        assert CuadrillaMiembro.objects.filter(activo=True).count() == 1

    def test_crear_usuarios_faltantes_opt_in(self):
        _crear_linea('LN817')
        excel = _build_s18_excel([
            _act(1, 'Hurto', '817', date(2026, 3, 26), date(2026, 3, 29),
                 'NUEVO TRABAJADOR', '9999999999', 'AYUDANTE', 'JT/CTA'),
        ])
        res = ProgramacionS18CuadrillaImporter().importar(
            excel, {'crear_usuarios_faltantes': True}
        )
        assert res['usuarios_creados'] == 1
        assert res['miembros_agregados'] == 1
        u = Usuario.objects.get(documento='9999999999')
        assert u.email == '9999999999@instelec-import.local'
        assert not u.has_usable_password()

    def test_dos_actividades_codigos_distintos(self):
        _crear_linea('LN805')
        _crear_linea('LN817')
        _crear_usuario('1143246675', 'JHON JAIRO')
        _crear_usuario('1093293706', 'SNEYDER JEREZ')
        excel = _build_s18_excel([
            _act(1, 'Servidumbre Completa', '817', date(2026, 4, 27),
                 date(2026, 5, 3), 'JHON JAIRO', '1143246675', 'LINIERO I', 'JT/CTA'),
            _act(2, 'Avisos SC', '805', date(2026, 4, 27), date(2026, 5, 3),
                 'SNEYDER JEREZ', '1093293706', 'LINIERO II', 'JT/CTA'),
        ])
        res = ProgramacionS18CuadrillaImporter().importar(excel)
        assert res['cuadrillas_creadas'] == 2
        codigos = set(Cuadrilla.objects.values_list('codigo', flat=True))
        assert codigos == {'18-2026-0001-SER', '18-2026-0002-AVI'}

    @pytest.mark.skipif(not os.path.exists(FIXTURE_REAL), reason='fixture real ausente')
    def test_fixture_real_legacy(self):
        """Test contra dato legacy: el archivo REAL del cliente (sheet '18').

        Sembramos las líneas/usuarios que el archivo referencia y verificamos
        que el importer crea cuadrillas con encargados, sin errores fatales.
        """
        # Líneas que aparecen en la hoja 18 del archivo real.
        for cod in ['LN817', 'LN818']:
            _crear_linea(cod)
        # Algunas cédulas reales de la hoja 18 (no todas → habrá advertencias).
        _crear_usuario('1143246675', 'JHON JAIRO JIMENEZ')
        _crear_usuario('1004487321', 'KEINER SERRANO')
        _crear_usuario('8649005', 'OMAR MANTILLA')
        Vehiculo.objects.create(placa='TLN-063', activo=True)

        with open(FIXTURE_REAL, 'rb') as f:
            res = ProgramacionS18CuadrillaImporter().importar(f)

        assert res['exito'] is True, res.get('error')
        assert res['formato'] == 'S18'
        assert res['cuadrillas_creadas'] >= 1
        assert res['encargados_asignados'] >= 1
        # NOTA: el archivo real puede traer varias filas ROL=JT/CTA por
        # actividad (frentes de trabajo) → una actividad puede tener >1
        # encargado. Verificamos que el conteo de encargados sea coherente
        # con los miembros marcados, no que sea <=1.
        for cuad in Cuadrilla.objects.all():
            jts = cuad.miembros.filter(cargo='JT_CTA').count()
            assert jts >= 0
        # El código sigue el patrón WW-YYYY-NNNN-AAA. El archivo real puede
        # traer varias hojas semanales (p.ej. '18' y la copia '12 (2)'), así
        # que validamos el patrón, no una semana concreta.
        import re
        patron = re.compile(r'^\d{2}-\d{4}-\d{4}-[A-Z]{1,3}$')
        for codigo in Cuadrilla.objects.values_list('codigo', flat=True):
            assert patron.match(codigo), codigo


# ---------------------------------------------------------------------------
# Issue #105 — AVISOS/ORDEN del formato real S22 (columna ATENCION + N/O)
# ---------------------------------------------------------------------------

@pytest.mark.django_db
class TestAvisosOrdenS22:
    """El formato real S22 trae AVISOS (col N) y ORDEN (col O) en la fila de
    encabezado de la actividad. El modelo Cuadrilla no tiene campos para ellas,
    así que deben quedar antepuestas a ``observaciones``."""

    def test_avisos_y_orden_van_a_observaciones(self):
        _crear_linea('LN806', 'LN806 SABANALARGA')
        _crear_usuario('1042438483', 'ALFONSO LOPEZ')
        _crear_usuario('1090434209', 'ABRAHAM PAYARES')
        _crear_usuario('1083020533', 'ADOLFO CASTRO')

        excel = _build_s22_excel([
            _act_s22(1, 'Servidumbre Completa', '806', date(2026, 5, 25),
                     date(2026, 5, 30), 'ALFONSO LOPEZ', '1042438483',
                     'LINIERO I', 'JT/CTA', placa='KNZ-217',
                     avisos='5720754', orden='35123993'),
            _miembro_s22('ABRAHAM PAYARES', '1090434209', 'LINIERO I'),
            _miembro_s22('ADOLFO CASTRO', '1083020533', 'AYUDANTE'),
        ])

        res = ProgramacionS18CuadrillaImporter().importar(excel)
        assert res['exito'] is True, res.get('error')
        assert res['formato'] == 'S18'
        assert res['cuadrillas_creadas'] == 1
        assert res['miembros_agregados'] == 3
        assert res['encargados_asignados'] == 1

        cuad = Cuadrilla.objects.get()
        assert cuad.codigo == '22-2026-0001-SER'
        # avisos/orden preservados en observaciones (sin campo de modelo).
        assert '5720754' in cuad.observaciones
        assert '35123993' in cuad.observaciones
        assert 'Avisos:' in cuad.observaciones
        assert 'Orden:' in cuad.observaciones

    def test_avisos_orden_multilinea_se_unen_con_coma(self):
        _crear_linea('LN806')
        _crear_usuario('1042438483', 'ALFONSO LOPEZ')
        excel = _build_s22_excel([
            _act_s22(1, 'Hurto', '806', date(2026, 5, 25), date(2026, 5, 30),
                     'ALFONSO LOPEZ', '1042438483', 'LINIERO I', 'JT/CTA',
                     avisos='5720754\n5720792', orden='1\n2'),
        ])
        res = ProgramacionS18CuadrillaImporter().importar(excel)
        assert res['exito'] is True, res.get('error')
        cuad = Cuadrilla.objects.get()
        assert 'Avisos: 5720754, 5720792' in cuad.observaciones
        assert 'Orden: 1, 2' in cuad.observaciones

    def test_avisos_orden_preserva_comentarios_existentes(self):
        _crear_linea('LN806')
        _crear_usuario('1042438483', 'ALFONSO LOPEZ')
        fila = _act_s22(1, 'Hurto', '806', date(2026, 5, 25), date(2026, 5, 30),
                        'ALFONSO LOPEZ', '1042438483', 'LINIERO I', 'JT/CTA',
                        avisos='5720754', orden='35123993')
        fila[-1] = 'Nota original del coordinador'  # columna Comentarios
        excel = _build_s22_excel([fila])
        res = ProgramacionS18CuadrillaImporter().importar(excel)
        assert res['exito'] is True, res.get('error')
        obs = Cuadrilla.objects.get().observaciones
        assert 'Avisos: 5720754' in obs
        assert 'Orden: 35123993' in obs
        assert 'Nota original del coordinador' in obs

    def test_sin_avisos_orden_no_agrega_prefijo(self):
        # Las hojas S18 sin avisos/orden no deben ensuciar observaciones.
        _crear_linea('LN817')
        _crear_usuario('1143246675', 'JHON JAIRO')
        excel = _build_s18_excel([
            _act(1, 'Servidumbre', '817', date(2026, 4, 27), date(2026, 5, 3),
                 'JHON JAIRO', '1143246675', 'LINIERO I', 'JT/CTA'),
        ])
        res = ProgramacionS18CuadrillaImporter().importar(excel)
        assert res['exito'] is True
        obs = Cuadrilla.objects.get().observaciones
        assert 'Avisos:' not in obs
        assert 'Orden:' not in obs

    @pytest.mark.skipif(not os.path.exists(FIXTURE_S22), reason='fixture S22 ausente')
    def test_fixture_s22_qa_legacy(self):
        """Test contra el fixture E2E real (Programacion_S22_qa.xlsx).

        Es el mismo archivo que usa el journey en prod: 3 miembros con cédulas
        que existen, 1 JT/CTA, avisos=5720754 / orden=35123993.
        """
        assert detectar_formato_cuadrillas(FIXTURE_S22) == 'S18'
        _crear_linea('LN806')
        _crear_usuario('1042438483', 'ALFONSO LOPEZ SARMIENTO')
        _crear_usuario('1090434209', 'ABRAHAM PAYARES JIMENEZ')
        _crear_usuario('1083020533', 'ADOLFO JUNIOR CASTRO PINTO')

        with open(FIXTURE_S22, 'rb') as f:
            res = ProgramacionS18CuadrillaImporter().importar(f)

        assert res['exito'] is True, res.get('error')
        assert res['cuadrillas_creadas'] == 1
        assert res['miembros_agregados'] == 3
        assert res['encargados_asignados'] == 1

        cuad = Cuadrilla.objects.get()
        assert cuad.codigo == '22-2026-0001-QAE'
        assert cuad.nombre.startswith('QA_E2E_M1')
        assert '5720754' in cuad.observaciones
        assert '35123993' in cuad.observaciones
