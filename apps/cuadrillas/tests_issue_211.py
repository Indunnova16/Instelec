"""Tests #211 — Export consolidado de nómina por rango de fechas.

Issue: Indunnova16/Instelec#211

``ProgramacionSemanalExportarRangoView`` (GET
/cuadrillas/semanal/exportar-rango/?fecha_inicio=...&fecha_fin=...) consolida
en UN solo Excel todas las cuadrillas (bloques) activas cuya ``fecha`` cae en
el rango pedido, sin importar la semana ISO de cada una — a diferencia del
export existente (``semanal_exportar_horizontal``), que solo cubre una
semana ISO por vez.

Ejecutar:
  DJANGO_SETTINGS_MODULE=config.settings.dev_lite \
    venv/bin/python -m pytest apps/cuadrillas/tests_issue_211.py -v \
    -o python_files="tests_*.py test_*.py"
"""

from datetime import date
from io import BytesIO

import openpyxl
from django.contrib.auth import get_user_model
from django.test import TestCase
from django.urls import reverse

from apps.actividades.exporters import ProgramacionSemanalHorizontalExporter
from apps.cuadrillas.models import Cuadrilla, CuadrillaMiembro

Usuario = get_user_model()


def _crear_admin():
    return Usuario.objects.create_user(
        email="admin_211@test.com",
        password="testpass123!",
        first_name="Admin",
        last_name="211",
        rol="admin",
        is_staff=True,
    )


def _crear_usuario(documento, nombre):
    partes = nombre.split(maxsplit=1)
    return Usuario.objects.create(
        email=f"{documento}@test.local",
        documento=documento,
        first_name=partes[0],
        last_name=partes[1] if len(partes) > 1 else "",
        rol="liniero",
        is_active=True,
    )


def _crear_bloque(codigo, fecha, miembros, nombre="MANTENIMIENTO - 211"):
    c = Cuadrilla.objects.create(
        codigo=codigo,
        nombre=nombre,
        activa=True,
        observaciones="",
        fecha=fecha,
    )
    for usuario, cargo, rol_cuadrilla in miembros:
        CuadrillaMiembro.objects.create(
            cuadrilla=c,
            usuario=usuario,
            rol_cuadrilla_id=rol_cuadrilla,
            cargo=cargo,
            costo_dia=0,
            fecha_inicio=fecha,
            activo=True,
        )
    return c


def _leer_columna_personal(xlsx_bytes):
    """Devuelve el set de nombres (columna PERSONAL, índice 7) presentes en
    el .xlsx generado — para verificar contenido real fila-por-fila (no solo
    tamaño en bytes, a diferencia del journey E2E que no lee celdas)."""
    wb = openpyxl.load_workbook(BytesIO(xlsx_bytes))
    ws = wb.active
    nombres = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[6]:
            nombres.add(row[6])
    return nombres


class TestExporterGenerarExcelRango(TestCase):
    """Unit: ProgramacionSemanalHorizontalExporter.generar_excel_rango()."""

    def setUp(self):
        self.jt_a = _crear_usuario("111211", "PEDRO PEREZ")
        self.jt_b = _crear_usuario("222211", "JUAN GOMEZ")

    def test_consolida_dos_cuadrillas_de_semanas_iso_distintas(self):
        """Dos bloques en semanas ISO distintas (2 y 5 de 2026) pero dentro
        del mismo rango calendario (enero) -> ambos aparecen en UN Excel."""
        _crear_bloque(
            "02-2026-0001-QAA",
            date(2026, 1, 6),
            [(self.jt_a, "JT_CTA", "LINIERO_I")],
            nombre="RANGO_A",
        )
        _crear_bloque(
            "05-2026-0001-QAB",
            date(2026, 1, 27),
            [(self.jt_b, "JT_CTA", "LINIERO_I")],
            nombre="RANGO_B",
        )
        # Bloque FUERA del rango pedido -- no debe aparecer.
        _crear_bloque(
            "10-2026-0001-QAC",
            date(2026, 3, 1),
            [(self.jt_a, "JT_CTA", "LINIERO_I")],
            nombre="FUERA_DE_RANGO",
        )

        output = ProgramacionSemanalHorizontalExporter().generar_excel_rango(
            date(2026, 1, 1), date(2026, 1, 31)
        )
        nombres = _leer_columna_personal(output.read())
        self.assertIn("PEDRO PEREZ", nombres)
        self.assertIn("JUAN GOMEZ", nombres)

    def test_excluye_cuadrilla_inactiva(self):
        c = _crear_bloque(
            "02-2026-0002-QAD",
            date(2026, 1, 6),
            [(self.jt_a, "JT_CTA", "LINIERO_I")],
            nombre="INACTIVA",
        )
        c.activa = False
        c.save(update_fields=["activa"])

        output = ProgramacionSemanalHorizontalExporter().generar_excel_rango(
            date(2026, 1, 1), date(2026, 1, 31)
        )
        nombres = _leer_columna_personal(output.read())
        self.assertNotIn("PEDRO PEREZ", nombres)

    def test_generar_excel_semanal_existente_sin_cambio_de_comportamiento(self):
        """El refactor a _escribir_hoja NO debe alterar generar_excel()."""
        _crear_bloque(
            "02-2026-0001-QAA",
            date(2026, 1, 6),
            [(self.jt_a, "JT_CTA", "LINIERO_I")],
            nombre="RANGO_A",
        )
        output = ProgramacionSemanalHorizontalExporter().generar_excel(2026, 2)
        wb = openpyxl.load_workbook(BytesIO(output.read()))
        ws = wb.active
        self.assertEqual(ws.title, "Sem 02-2026")
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        self.assertEqual(headers[:2], ["#", "ACTIVIDAD"])


class TestExportarRangoView(TestCase):
    """Integration: GET /cuadrillas/semanal/exportar-rango/."""

    def setUp(self):
        self.admin = _crear_admin()
        self.client.force_login(self.admin)
        self.jt_a = _crear_usuario("333211", "ANA TORRES")
        self.jt_b = _crear_usuario("444211", "LUIS DIAZ")

    def test_export_rango_cruzando_semanas_iso_trae_ambos_bloques(self):
        """Escenario exacto del issue: exportar un rango que cruza 2+ semanas
        ISO distintas y confirmar que el Excel trae ambas."""
        _crear_bloque(
            "02-2026-0010-QAE",
            date(2026, 1, 6),
            [(self.jt_a, "JT_CTA", "LINIERO_I")],
            nombre="VIEW_RANGO_A",
        )
        _crear_bloque(
            "05-2026-0010-QAF",
            date(2026, 1, 27),
            [(self.jt_b, "JT_CTA", "LINIERO_I")],
            nombre="VIEW_RANGO_B",
        )

        resp = self.client.get(
            reverse("cuadrillas:semanal_exportar_rango"),
            {"fecha_inicio": "2026-01-01", "fecha_fin": "2026-01-31"},
        )
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(
            resp["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertGreater(len(resp.content), 2000)
        nombres = _leer_columna_personal(resp.content)
        self.assertIn("ANA TORRES", nombres)
        self.assertIn("LUIS DIAZ", nombres)

    def test_export_rango_sin_fechas_400(self):
        resp = self.client.get(reverse("cuadrillas:semanal_exportar_rango"))
        self.assertEqual(resp.status_code, 400)

    def test_export_rango_fecha_fin_antes_de_inicio_400(self):
        resp = self.client.get(
            reverse("cuadrillas:semanal_exportar_rango"),
            {"fecha_inicio": "2026-01-31", "fecha_fin": "2026-01-01"},
        )
        self.assertEqual(resp.status_code, 400)

    def test_export_rango_fecha_invalida_400(self):
        resp = self.client.get(
            reverse("cuadrillas:semanal_exportar_rango"),
            {"fecha_inicio": "no-es-fecha", "fecha_fin": "2026-01-31"},
        )
        self.assertEqual(resp.status_code, 400)

    def test_export_rango_requiere_login(self):
        self.client.logout()
        resp = self.client.get(
            reverse("cuadrillas:semanal_exportar_rango"),
            {"fecha_inicio": "2026-01-01", "fecha_fin": "2026-01-31"},
        )
        self.assertIn(resp.status_code, (302, 301))

    def test_export_rango_vacio_no_500(self):
        """Rango sin ninguna cuadrilla -> 200 con Excel vacío (solo headers),
        no 500 -- misma robustez que el export semanal existente."""
        resp = self.client.get(
            reverse("cuadrillas:semanal_exportar_rango"),
            {"fecha_inicio": "2026-06-01", "fecha_fin": "2026-06-30"},
        )
        self.assertEqual(resp.status_code, 200)
