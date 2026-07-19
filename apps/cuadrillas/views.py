"""
Views for crew management.
"""
from django.db import models
from django.shortcuts import redirect
from django.contrib import messages
from django.views.generic import ListView, DetailView, TemplateView
from django.views import View
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpResponse, JsonResponse
from apps.core.mixins import HTMXMixin, RoleRequiredMixin
from .models import Asistencia, Cargo, Cuadrilla, CuadrillaMiembro, PersonalCuadrilla, Vehiculo, TrackingUbicacion
from .forms_personal import PersonalCuadrillaForm
from .forms_cargo import CargoForm


class CuadrillaListView(LoginRequiredMixin, RoleRequiredMixin, HTMXMixin, ListView):
    """List all crews, organized by week."""
    model = Cuadrilla
    template_name = 'cuadrillas/lista.html'
    partial_template_name = 'cuadrillas/partials/lista_cuadrillas.html'
    context_object_name = 'cuadrillas'
    paginate_by = None
    allowed_roles = ['admin', 'director', 'coordinador', 'ing_residente', 'supervisor']

    @staticmethod
    def _parse_semana(codigo):
        """Extract (week, year) from code format WW-YYYY-XXX."""
        try:
            parts = codigo.split('-')
            if len(parts) >= 2:
                semana = int(parts[0])
                ano = int(parts[1])
                if 1 <= semana <= 53 and 2000 <= ano <= 2100:
                    return semana, ano
        except (ValueError, IndexError):
            pass
        return None, None

    def get_queryset(self):
        qs = Cuadrilla.objects.filter(activa=True).select_related(
            'supervisor', 'vehiculo', 'linea_asignada'
        ).prefetch_related('miembros__usuario')

        # Filter by week if parameter provided
        semana_param = self.request.GET.get('semana', '').strip()
        if semana_param:
            # Format: WW-YYYY
            try:
                parts = semana_param.split('-')
                sem = parts[0].zfill(2)
                ano = parts[1]
                qs = qs.filter(codigo__startswith=f'{sem}-{ano}-')
            except (IndexError, ValueError):
                pass

        return qs

    def get_context_data(self, **kwargs):
        import json
        from collections import OrderedDict
        context = super().get_context_data(**kwargs)

        # Build list of available weeks from all active cuadrillas
        todas = Cuadrilla.objects.filter(activa=True).values_list('codigo', flat=True)
        semanas_set = set()
        for codigo in todas:
            sem, ano = self._parse_semana(codigo)
            if sem is not None:
                semanas_set.add((ano, sem))

        # Sort descending: most recent first
        semanas_disponibles = sorted(semanas_set, reverse=True)
        context['semanas_disponibles'] = [
            {'value': f'{s[1]}-{s[0]}', 'label': f'Semana {s[1]} - {s[0]}'}
            for s in semanas_disponibles
        ]

        # Current filter
        semana_param = self.request.GET.get('semana', '').strip()
        context['semana_actual'] = semana_param

        # Group cuadrillas by week for display
        cuadrillas_por_semana = OrderedDict()
        sin_semana = []
        for cuadrilla in context['cuadrillas']:
            sem, ano = self._parse_semana(cuadrilla.codigo)
            if sem is not None:
                key = f'Semana {sem} - {ano}'
                cuadrillas_por_semana.setdefault(key, []).append(cuadrilla)
            else:
                sin_semana.append(cuadrilla)

        if sin_semana:
            cuadrillas_por_semana['Otras'] = sin_semana

        context['cuadrillas_por_semana'] = cuadrillas_por_semana

        # Stats
        all_active = Cuadrilla.objects.filter(activa=True)
        context['total_cuadrillas'] = all_active.count()
        context['cuadrillas_activas'] = all_active.count()

        # Mapa de cuadrillas por UBICACIÓN DEL PROYECTO asignado (#155 sub-5).
        # El cliente pide ubicar cada cuadrilla por su proyecto asignado (no por
        # GPS). La relación es Cuadrilla → ProgramacionSemanalCuadrilla.proyecto →
        # ProyectoConstruccion(latitud, longitud). Tomamos la programación más
        # reciente con proyecto geolocalizado. Robusto a 0 puntos: una cuadrilla
        # sin proyecto/coords simplemente no aporta marcador (mapa inicializa
        # vacío sin pageerror — el JS de lista.html ya tolera lista vacía).
        ubicaciones = self._build_ubicaciones_proyecto(context['cuadrillas'])
        context['cuadrillas_ubicaciones_json'] = json.dumps(ubicaciones)

        # Issue #188 (A9): fusión con /cuadrillas/semanal/ -- la pantalla
        # fusionada agrega el grid EDITABLE (bloques cards) de una semana
        # puntual, reusando _contexto_semana/_choices_form_bloque de
        # views_semanal.py como fuente única de verdad (en vez de duplicar el
        # modelo de acceso a datos). Reusa el MISMO `semana_param` que ya
        # filtra el listado plano de abajo (mismo formato WW-YYYY); si no
        # viene, usa la semana con datos más reciente. Preserva TODO lo de
        # arriba (filtros activas/inactivas de B3, mapa por proyecto, listado
        # agrupado) sin tocarlo -- el grid es una sección nueva, no un
        # reemplazo.
        from .views_semanal import (
            _choices_form_bloque,
            _contexto_semana,
            _semana_mas_reciente_con_datos,
        )

        grid_anio = grid_semana = None
        if semana_param:
            try:
                partes = semana_param.split('-')
                grid_semana = int(partes[0])
                grid_anio = int(partes[1])
            except (IndexError, ValueError):
                grid_anio = grid_semana = None
        if grid_anio is None or grid_semana is None:
            grid_anio, grid_semana = _semana_mas_reciente_con_datos()

        context.update(_contexto_semana(grid_anio, grid_semana))
        context.update(_choices_form_bloque())
        context['confirmar_duplicado'] = self.request.GET.get('confirmar_duplicado') == '1'
        return context

    @staticmethod
    def _build_ubicaciones_proyecto(cuadrillas):
        """Construye los marcadores del mapa desde la ubicación del proyecto.

        Por cada cuadrilla, busca su programación más reciente cuyo proyecto
        tenga latitud y longitud cargadas, y emite un punto. Sin proyecto o sin
        coordenadas → la cuadrilla se omite (no rompe el mapa).
        """
        from .models_pc import ProgramacionSemanalCuadrilla

        ubicaciones = []
        for cuadrilla in cuadrillas:
            prog = (
                ProgramacionSemanalCuadrilla.objects
                .filter(
                    cuadrilla=cuadrilla,
                    proyecto__isnull=False,
                    proyecto__latitud__isnull=False,
                    proyecto__longitud__isnull=False,
                )
                .select_related('proyecto')
                .order_by('-anio', '-semana')
                .first()
            )
            if not prog or not prog.proyecto:
                continue
            proyecto = prog.proyecto
            ubicaciones.append({
                'cuadrilla_id': str(cuadrilla.id),
                'cuadrilla_codigo': cuadrilla.codigo,
                'proyecto_nombre': proyecto.nombre,
                'lat': float(proyecto.latitud),
                'lng': float(proyecto.longitud),
            })
        return ubicaciones


class CuadrillaDetailView(LoginRequiredMixin, RoleRequiredMixin, HTMXMixin, DetailView):
    """Detail view for a crew."""
    model = Cuadrilla
    template_name = 'cuadrillas/detalle.html'
    context_object_name = 'cuadrilla'
    allowed_roles = ['admin', 'director', 'coordinador', 'ing_residente', 'supervisor']

    def get_context_data(self, **kwargs):
        from datetime import date, timedelta
        from decimal import Decimal
        from apps.usuarios.models import Usuario

        context = super().get_context_data(**kwargs)
        miembros = list(self.object.miembros.filter(activo=True).select_related('usuario'))

        # If the cuadrilla has a supervisor and they're not already a member,
        # create a virtual member entry so they appear in the attendance list
        supervisor = self.object.supervisor
        supervisor_is_member = False
        if supervisor:
            supervisor_is_member = any(m.usuario_id == supervisor.id for m in miembros)
            if not supervisor_is_member:
                # Try to reactivate an existing inactive record first
                existing = CuadrillaMiembro.objects.filter(
                    cuadrilla=self.object, usuario=supervisor, activo=False
                ).first()
                if existing:
                    existing.activo = True
                    existing.rol_cuadrilla_id = 'SUPERVISOR'
                    existing.cargo = 'JT_CTA'
                    existing.save(update_fields=['activo', 'rol_cuadrilla', 'cargo', 'updated_at'])
                    miembros.insert(0, existing)
                else:
                    try:
                        miembro = CuadrillaMiembro.objects.create(
                            cuadrilla=self.object,
                            usuario=supervisor,
                            rol_cuadrilla_id='SUPERVISOR',
                            cargo='JT_CTA',
                            costo_dia=0,
                            fecha_inicio=date.today(),
                            activo=True,
                        )
                        miembros.insert(0, miembro)
                    except Exception:
                        pass

        context['miembros'] = miembros

        # Total daily cost
        context['costo_total_dia'] = sum(
            (m.costo_dia for m in miembros), Decimal('0')
        )

        # Available users for add member form (legacy — se mantiene por si
        # algun template/reporte externo aun lo consume).
        miembros_ids = [m.usuario_id for m in miembros]
        context['usuarios_disponibles'] = Usuario.objects.filter(
            is_active=True
        ).exclude(id__in=miembros_ids).order_by('first_name', 'last_name')

        # Picklist de colaboradores disponibles (issue #176, A4): origen de
        # datos = PersonalCuadrilla.activo=True, excluyendo los que ya son
        # miembros activos de esta cuadrilla (match por documento, no por
        # Usuario -- un PersonalCuadrilla puede no tener Usuario vinculado
        # aun la primera vez que se asigna).
        miembros_documentos = {
            m.usuario.documento for m in miembros if m.usuario and m.usuario.documento
        }
        context['personal_disponible'] = PersonalCuadrilla.objects.filter(
            activo=True
        ).exclude(documento__in=miembros_documentos).order_by('nombre')

        # Choices for form selects
        # Issue #176 (A4): RolCuadrilla (TextChoices) eliminado, catalogo
        # ahora es Cargo. Solo activos -- coincide con el picklist real.
        context['roles_cuadrilla'] = list(
            Cargo.objects.filter(activo=True).values_list('codigo', 'nombre')
        )
        context['cargos_jerarquicos'] = CuadrillaMiembro.CargoJerarquico.choices

        # Last known location
        ultima_ubicacion = TrackingUbicacion.objects.filter(
            cuadrilla=self.object
        ).order_by('-created_at').first()
        context['ultima_ubicacion'] = ultima_ubicacion

        # Weekly attendance calendar
        semana, ano = self._get_semana_from_codigo(self.object.codigo)
        if semana and ano:
            # ISO week: Monday=0 to Sunday=6
            lunes = date.fromisocalendar(ano, semana, 1)
        else:
            # Fallback: current week
            hoy = date.today()
            lunes = hoy - timedelta(days=hoy.weekday())

        dias_semana = [lunes + timedelta(days=i) for i in range(7)]
        context['dias_semana'] = dias_semana
        context['semana_lunes'] = lunes

        # Load existing attendance for this week
        asistencias = Asistencia.objects.filter(
            cuadrilla=self.object,
            fecha__in=dias_semana,
        ).select_related('usuario')

        # Build dict: {usuario_id: {fecha_iso: {tipo_novedad, viaticos, horas_extra, observacion, viatico_aplica}}}
        asistencia_por_usuario = {}
        for a in asistencias:
            uid = str(a.usuario_id)
            if uid not in asistencia_por_usuario:
                asistencia_por_usuario[uid] = {}
            asistencia_por_usuario[uid][a.fecha.isoformat()] = {
                'tipo_novedad': a.tipo_novedad,
                'viaticos': float(a.viaticos),
                'horas_extra': float(a.horas_extra),
                'observacion': a.observacion,
                'viatico_aplica': a.viatico_aplica,
                'he_diurna': float(a.he_diurna),
                'he_nocturna': float(a.he_nocturna),
                'he_dominical_diurna': float(a.he_dominical_diurna),
                'he_dominical_nocturna': float(a.he_dominical_nocturna),
            }

        # Build template-friendly structure
        filas_asistencia = []
        for miembro in miembros:
            uid = str(miembro.usuario_id)
            usuario_asistencia = asistencia_por_usuario.get(uid, {})
            dias = []
            total_viaticos = Decimal('0')
            total_horas_extra = Decimal('0')
            total_horas_ordinarias = Decimal('0')
            for dia in dias_semana:
                fecha_iso = dia.isoformat()
                info = usuario_asistencia.get(fecha_iso, {})
                viaticos_val = info.get('viaticos', 0)
                horas_extra_val = info.get('horas_extra', 0)
                if info.get('viatico_aplica', False):
                    total_viaticos += Decimal(str(viaticos_val))
                total_horas_extra += Decimal(str(horas_extra_val))
                # Ordinary hours: jornada regular if PRESENTE
                jornada = Asistencia.JORNADA_POR_DIA.get(dia.weekday(), 0)
                if info.get('tipo_novedad') == 'PRESENTE':
                    total_horas_ordinarias += Decimal(str(jornada))
                dias.append({
                    'fecha': fecha_iso,
                    'fecha_display': dia.strftime('%d/%m'),
                    'tipo_novedad': info.get('tipo_novedad', ''),
                    'viaticos': viaticos_val,
                    'horas_extra': horas_extra_val,
                    'observacion': info.get('observacion', ''),
                    'viatico_aplica': info.get('viatico_aplica', False),
                    'he_diurna': info.get('he_diurna', 0),
                    'he_nocturna': info.get('he_nocturna', 0),
                    'he_dominical_diurna': info.get('he_dominical_diurna', 0),
                    'he_dominical_nocturna': info.get('he_dominical_nocturna', 0),
                    'dia_semana': dia.weekday(),
                })
            filas_asistencia.append({
                'miembro': miembro,
                'dias': dias,
                'total_viaticos': total_viaticos,
                'total_horas_extra': total_horas_extra,
                'total_horas_ordinarias': total_horas_ordinarias,
                'total_horas_total': total_horas_ordinarias + total_horas_extra,
            })

        context['filas_asistencia'] = filas_asistencia
        context['tipos_novedad'] = Asistencia.TipoNovedad.choices

        return context

    @staticmethod
    def _get_semana_from_codigo(codigo):
        """Extract (week, year) from code format WW-YYYY-XXX."""
        try:
            parts = codigo.split('-')
            if len(parts) >= 2:
                semana = int(parts[0])
                ano = int(parts[1])
                if 1 <= semana <= 53 and 2000 <= ano <= 2100:
                    return semana, ano
        except (ValueError, IndexError):
            pass
        return None, None


class CuadrillaEditView(LoginRequiredMixin, RoleRequiredMixin, HTMXMixin, DetailView):
    """View for editing a crew."""
    model = Cuadrilla
    template_name = 'cuadrillas/editar.html'
    context_object_name = 'cuadrilla'
    allowed_roles = ['admin', 'director', 'coordinador']

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from apps.usuarios.models import Usuario
        from apps.lineas.models import Linea
        context['supervisores'] = Usuario.objects.filter(rol='supervisor', is_active=True)
        context['lineas'] = Linea.objects.filter(activa=True)
        context['vehiculos'] = Vehiculo.objects.filter(activo=True)
        return context

    def post(self, request, *args, **kwargs):
        """Handle form submission to update a crew."""
        from apps.usuarios.models import Usuario
        from apps.lineas.models import Linea

        cuadrilla = self.get_object()

        codigo = request.POST.get('codigo', '').strip()
        nombre = request.POST.get('nombre', '').strip()

        if not codigo or not nombre:
            messages.error(request, 'Código y nombre son obligatorios.')
            return self.get(request, *args, **kwargs)

        if Cuadrilla.objects.filter(codigo=codigo).exclude(pk=cuadrilla.pk).exists():
            messages.error(request, f'Ya existe otra cuadrilla con el código {codigo}.')
            return self.get(request, *args, **kwargs)

        try:
            cuadrilla.codigo = codigo
            cuadrilla.nombre = nombre

            supervisor_id = request.POST.get('supervisor') or None
            cuadrilla.supervisor = Usuario.objects.get(pk=supervisor_id) if supervisor_id else None

            vehiculo_id = request.POST.get('vehiculo') or None
            cuadrilla.vehiculo = Vehiculo.objects.get(pk=vehiculo_id) if vehiculo_id else None

            linea_id = request.POST.get('linea_asignada') or None
            cuadrilla.linea_asignada = Linea.objects.get(pk=linea_id) if linea_id else None

            fecha_str = request.POST.get('fecha', '').strip()
            if fecha_str:
                from datetime import date as date_cls
                cuadrilla.fecha = date_cls.fromisoformat(fecha_str)
            else:
                cuadrilla.fecha = None

            cuadrilla.activa = request.POST.get('activa') == 'on'
            cuadrilla.observaciones = request.POST.get('observaciones', '').strip()
            cuadrilla.save()
            messages.success(request, f'Cuadrilla {cuadrilla.codigo} actualizada exitosamente.')
            return redirect('cuadrillas:detalle', pk=cuadrilla.pk)
        except (Usuario.DoesNotExist, Linea.DoesNotExist, Vehiculo.DoesNotExist) as e:
            messages.error(request, f'Referencia inválida: {str(e)}')
            return self.get(request, *args, **kwargs)
        except Exception as e:
            messages.error(request, f'Error al actualizar la cuadrilla: {str(e)}')
            return self.get(request, *args, **kwargs)


class MapaCuadrillasView(LoginRequiredMixin, RoleRequiredMixin, TemplateView):
    """Real-time map of all crews."""
    template_name = 'cuadrillas/mapa.html'
    allowed_roles = ['admin', 'director', 'coordinador', 'ing_residente', 'supervisor', 'liniero', 'auxiliar']


class MapaCuadrillasPartialView(LoginRequiredMixin, RoleRequiredMixin, TemplateView):
    """Partial view for HTMX polling of crew locations."""
    template_name = 'cuadrillas/partials/mapa_cuadrillas.html'
    allowed_roles = ['admin', 'director', 'coordinador', 'ing_residente', 'supervisor', 'liniero', 'auxiliar']

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Get latest location for each active crew
        cuadrillas = Cuadrilla.objects.filter(activa=True)
        ubicaciones = []

        for cuadrilla in cuadrillas:
            ultima = TrackingUbicacion.objects.filter(
                cuadrilla=cuadrilla
            ).order_by('-created_at').first()

            if ultima:
                ubicaciones.append({
                    'cuadrilla_id': str(cuadrilla.id),
                    'cuadrilla_codigo': cuadrilla.codigo,
                    'cuadrilla_nombre': cuadrilla.nombre,
                    'lat': float(ultima.latitud),
                    'lng': float(ultima.longitud),
                    'precision': float(ultima.precision_metros) if ultima.precision_metros else None,
                    'timestamp': ultima.created_at.isoformat(),
                })

        context['ubicaciones'] = ubicaciones
        return context

    def render_to_response(self, context, **response_kwargs):
        if self.request.headers.get('Accept') == 'application/json':
            return JsonResponse({'ubicaciones': context['ubicaciones']})
        return super().render_to_response(context, **response_kwargs)


class CuadrillaCreateView(LoginRequiredMixin, RoleRequiredMixin, HTMXMixin, TemplateView):
    """View for creating a new crew."""
    template_name = 'cuadrillas/crear.html'
    partial_template_name = 'cuadrillas/partials/form_cuadrilla.html'
    allowed_roles = ['admin', 'director', 'coordinador']

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from apps.usuarios.models import Usuario
        from apps.lineas.models import Linea
        context['supervisores'] = Usuario.objects.filter(rol='supervisor', is_active=True)
        context['lineas'] = Linea.objects.filter(activa=True)
        return context

    def post(self, request, *args, **kwargs):
        """Handle form submission to create a new crew."""
        from django.shortcuts import redirect
        from django.contrib import messages
        from apps.usuarios.models import Usuario
        from apps.lineas.models import Linea

        codigo = request.POST.get('codigo', '').strip()
        nombre = request.POST.get('nombre', '').strip()
        supervisor_id = request.POST.get('supervisor') or None
        linea_id = request.POST.get('linea_asignada') or None

        # Validation
        if not codigo or not nombre:
            messages.error(request, 'Código y nombre son obligatorios.')
            return self.get(request, *args, **kwargs)

        if Cuadrilla.objects.filter(codigo=codigo).exists():
            messages.error(request, f'Ya existe una cuadrilla con el código {codigo}.')
            return self.get(request, *args, **kwargs)

        # Get related objects
        supervisor = None
        if supervisor_id:
            try:
                supervisor = Usuario.objects.get(pk=supervisor_id)
            except Usuario.DoesNotExist:
                pass

        linea_asignada = None
        if linea_id:
            try:
                linea_asignada = Linea.objects.get(pk=linea_id)
            except Linea.DoesNotExist:
                pass

        # Create the crew
        try:
            cuadrilla = Cuadrilla.objects.create(
                codigo=codigo,
                nombre=nombre,
                supervisor=supervisor,
                linea_asignada=linea_asignada,
            )
            messages.success(request, f'Cuadrilla {cuadrilla.codigo} creada exitosamente.')
            return redirect('cuadrillas:detalle', pk=cuadrilla.pk)
        except Exception as e:
            messages.error(request, f'Error al crear la cuadrilla: {str(e)}')
            return self.get(request, *args, **kwargs)


class CuadrillaMiembroAddView(LoginRequiredMixin, RoleRequiredMixin, DetailView):
    """Add a member to a cuadrilla.

    Refactor issue #176 (A4): el origen de datos del picker pasa de
    `Usuario` a `PersonalCuadrilla` (maestro de Colaboradores). El usuario
    ingresa el `documento` de un colaborador ACTIVO (autocompletado AJAX
    vía `PersonalCuadrillaAPIView`); el `rol_cuadrilla` (Cargo) NO es
    editable en el form -- se toma siempre del maestro `PersonalCuadrilla`,
    ignorando cualquier valor distinto que venga en el POST. Al hacer
    submit, se resuelve/crea el `Usuario` correspondiente por
    documento+nombre (mismo patrón de `apps.cuadrillas.importers._crear_usuario`:
    email sintético, password no utilizable). `CuadrillaMiembro.usuario`
    se mantiene como FK a Usuario -- no se toca el modelo de datos (otros
    flujos como TrackingUbicacion/fotos/reportes dependen de él).
    """
    model = Cuadrilla
    allowed_roles = ['admin', 'director', 'coordinador', 'ing_residente', 'supervisor']

    @staticmethod
    def _resolver_o_crear_usuario(personal):
        """Issue #188 (A5): extraído a ``apps.cuadrillas.services`` para
        reusarlo también desde ``ProgramacionSemanalMiembroAgregarView``
        (grid editable). Este wrapper se conserva tal cual (mismo nombre,
        mismo comportamiento) para no romper llamadas existentes."""
        from .services import resolver_o_crear_usuario

        return resolver_o_crear_usuario(personal)

    def post(self, request, *args, **kwargs):
        from datetime import date

        cuadrilla = self.get_object()
        documento = (request.POST.get('documento') or '').strip()
        costo_dia = request.POST.get('costo_dia', '0') or '0'

        if not documento:
            messages.error(request, 'Debe ingresar el documento de un colaborador.')
            return redirect('cuadrillas:detalle', pk=cuadrilla.pk)

        personal = PersonalCuadrilla.objects.filter(documento=documento, activo=True).first()
        if not personal:
            messages.error(
                request,
                f'No se encontró un colaborador activo con documento "{documento}". '
                'Verifique el documento o dé de alta el colaborador en el maestro.',
            )
            return redirect('cuadrillas:detalle', pk=cuadrilla.pk)

        # Cargo (rol_cuadrilla) NO es editable en el form de asignación: se
        # toma siempre del maestro PersonalCuadrilla, ignorando cualquier
        # valor distinto que venga en el POST (issue #176, A4). rol_id es
        # el string del codigo (Cargo.codigo), NO el objeto Cargo.
        rol_id = personal.rol_cuadrilla_id
        cargo = request.POST.get('cargo', 'MIEMBRO')

        usuario = self._resolver_o_crear_usuario(personal)

        if CuadrillaMiembro.objects.filter(
            cuadrilla=cuadrilla, usuario=usuario, activo=True
        ).exists():
            messages.error(request, f'{usuario.get_full_name()} ya es miembro activo.')
            return redirect('cuadrillas:detalle', pk=cuadrilla.pk)

        try:
            CuadrillaMiembro.objects.create(
                cuadrilla=cuadrilla,
                usuario=usuario,
                rol_cuadrilla_id=rol_id,
                cargo=cargo if cargo in dict(CuadrillaMiembro.CargoJerarquico.choices) else 'MIEMBRO',
                costo_dia=float(costo_dia),
                fecha_inicio=date.today(),
            )
            messages.success(request, f'{usuario.get_full_name()} agregado a la cuadrilla.')
        except Exception as e:
            messages.error(request, f'Error al agregar miembro: {str(e)}')

        return redirect('cuadrillas:detalle', pk=cuadrilla.pk)


class AsistenciaUpdateView(LoginRequiredMixin, RoleRequiredMixin, DetailView):
    """Update attendance for a crew member on a specific day (HTMX)."""
    model = Cuadrilla
    allowed_roles = ['admin', 'director', 'coordinador', 'ing_residente', 'supervisor']

    def post(self, request, *args, **kwargs):
        from datetime import date as date_cls
        from decimal import Decimal, InvalidOperation
        from django.http import HttpResponse

        cuadrilla = self.get_object()
        usuario_id = request.POST.get('usuario_id', '').strip()
        fecha_str = request.POST.get('fecha', '').strip()
        tipo_novedad = request.POST.get('tipo_novedad', '').strip()
        viaticos_str = request.POST.get('viaticos', '').strip()
        horas_extra_str = request.POST.get('horas_extra', '').strip()
        observacion = request.POST.get('observacion', '').strip()
        viatico_aplica = request.POST.get('viatico_aplica') == 'on'

        if not usuario_id or not fecha_str:
            return HttpResponse('Datos incompletos', status=400)

        try:
            fecha = date_cls.fromisoformat(fecha_str)
        except ValueError:
            return HttpResponse('Fecha invalida', status=400)

        # Verify user is member of the cuadrilla
        if not CuadrillaMiembro.objects.filter(
            cuadrilla=cuadrilla, usuario_id=usuario_id, activo=True
        ).exists():
            return HttpResponse('Usuario no es miembro', status=400)

        # Parse viaticos
        try:
            viaticos = Decimal(viaticos_str) if viaticos_str else Decimal('0')
        except InvalidOperation:
            viaticos = Decimal('0')

        # Parse horas extra detalladas
        def _parse_dec(s):
            try:
                return Decimal(s) if s else Decimal('0')
            except InvalidOperation:
                return Decimal('0')

        he_diurna = _parse_dec(request.POST.get('he_diurna', '').strip())
        he_nocturna = _parse_dec(request.POST.get('he_nocturna', '').strip())
        he_dominical_diurna = _parse_dec(request.POST.get('he_dominical_diurna', '').strip())
        he_dominical_nocturna = _parse_dec(request.POST.get('he_dominical_nocturna', '').strip())
        horas_extra = he_diurna + he_nocturna + he_dominical_diurna + he_dominical_nocturna

        # If viatico_aplica, calculate viaticos from CostoRecurso
        if viatico_aplica:
            from apps.financiero.models import CostoRecurso
            costo_viatico = CostoRecurso.objects.filter(
                tipo='VIATICO', activo=True
            ).order_by('-vigencia_desde').first()
            if costo_viatico:
                viaticos = costo_viatico.costo_unitario
            else:
                viaticos = Decimal('136941')
        else:
            viaticos = Decimal('0')

        tipos_validos = dict(Asistencia.TipoNovedad.choices)

        if not tipo_novedad and not viatico_aplica:
            # Empty selection and no viatico: remove attendance record
            Asistencia.objects.filter(
                usuario_id=usuario_id, cuadrilla=cuadrilla, fecha=fecha
            ).delete()
            viaticos = Decimal('0')
        elif not tipo_novedad and viatico_aplica:
            # No novedad but viatico checked: save with PRESENTE default
            Asistencia.objects.update_or_create(
                usuario_id=usuario_id,
                cuadrilla=cuadrilla,
                fecha=fecha,
                defaults={
                    'tipo_novedad': Asistencia.TipoNovedad.PRESENTE,
                    'viaticos': viaticos,
                    'horas_extra': horas_extra,
                    'he_diurna': he_diurna,
                    'he_nocturna': he_nocturna,
                    'he_dominical_diurna': he_dominical_diurna,
                    'he_dominical_nocturna': he_dominical_nocturna,
                    'observacion': observacion,
                    'viatico_aplica': viatico_aplica,
                    'registrado_por': request.user,
                }
            )
            tipo_novedad = 'PRESENTE'
        elif tipo_novedad in tipos_validos:
            Asistencia.objects.update_or_create(
                usuario_id=usuario_id,
                cuadrilla=cuadrilla,
                fecha=fecha,
                defaults={
                    'tipo_novedad': tipo_novedad,
                    'viaticos': viaticos,
                    'horas_extra': horas_extra,
                    'he_diurna': he_diurna,
                    'he_nocturna': he_nocturna,
                    'he_dominical_diurna': he_dominical_diurna,
                    'he_dominical_nocturna': he_dominical_nocturna,
                    'observacion': observacion,
                    'viatico_aplica': viatico_aplica,
                    'registrado_por': request.user,
                }
            )
        else:
            return HttpResponse('Tipo de novedad invalido', status=400)

        # Return the updated cell content
        color_map = {
            'PRESENTE': 'text-green-600 bg-green-50 border-green-300',
            'AUSENTE': 'text-red-600 bg-red-50 border-red-300',
            'VACACIONES': 'text-blue-600 bg-blue-50 border-blue-300',
            'INCAPACIDAD': 'text-orange-600 bg-orange-50 border-orange-300',
            'PERMISO': 'text-purple-600 bg-purple-50 border-purple-300',
            'LICENCIA': 'text-yellow-700 bg-yellow-50 border-yellow-300',
            'CAPACITACION': 'text-teal-600 bg-teal-50 border-teal-300',
            'COMPENSATORIO': 'text-cyan-600 bg-cyan-50 border-cyan-300',
            'DESCANSO': 'text-slate-600 bg-slate-50 border-slate-300',
        }
        css = color_map.get(tipo_novedad, 'text-gray-400 bg-white border-gray-200')

        options_html = '<option value="">---</option>'
        for val, lbl in Asistencia.TipoNovedad.choices:
            sel = ' selected' if val == tipo_novedad else ''
            options_html += f'<option value="{val}"{sel}>{lbl}</option>'

        viatico_checked = ' checked' if viatico_aplica else ''
        obs_escaped = observacion.replace('"', '&quot;')

        # Calculate weekly totals for OOB swap
        from datetime import timedelta
        semana, ano = CuadrillaDetailView._get_semana_from_codigo(cuadrilla.codigo)
        if semana and ano:
            lunes = date_cls.fromisocalendar(ano, semana, 1)
        else:
            hoy = date_cls.today()
            lunes = hoy - timedelta(days=hoy.weekday())
        dias_semana = [lunes + timedelta(days=i) for i in range(7)]

        week_asistencias = Asistencia.objects.filter(
            usuario_id=usuario_id,
            cuadrilla=cuadrilla,
            fecha__in=dias_semana,
        )
        totals = week_asistencias.aggregate(
            total_viaticos=models.Sum('viaticos', filter=models.Q(viatico_aplica=True)),
            total_he=models.Sum('horas_extra'),
        )
        total_viaticos_fmt = int(totals['total_viaticos'] or 0)
        total_he_fmt = float(totals['total_he'] or 0)

        # Calculate ordinary hours (jornada regular for PRESENTE days)
        JORNADA_DIA = {0: 8.0, 1: 7.5, 2: 7.5, 3: 7.5, 4: 7.5, 5: 6.0, 6: 0.0}
        total_hord = 0.0
        for a in week_asistencias.filter(tipo_novedad='PRESENTE'):
            total_hord += JORNADA_DIA.get(a.fecha.weekday(), 0)
        total_htotal = total_hord + total_he_fmt

        # Build observation field (visible when not PRESENTE)
        if tipo_novedad and tipo_novedad != 'PRESENTE':
            obs_field = (
                f'<input type="text" name="observacion" value="{obs_escaped}" '
                f'placeholder="Motivo de ausencia..." '
                f'hx-post="{request.path}" '
                f'hx-target="closest .asistencia-cell" '
                f'hx-swap="innerHTML" '
                f'hx-include="closest .asistencia-cell" '
                f'hx-trigger="change" '
                f'class="mt-1 text-xs rounded border border-yellow-300 bg-yellow-50 px-1 py-0.5 w-full '
                f'text-gray-700 dark:bg-gray-700 dark:border-gray-600 dark:text-gray-200">'
            )
        else:
            obs_field = f'<input type="hidden" name="observacion" value="{obs_escaped}">'

        # Build Alpine.js overtime section
        JORNADA = {0: 8.0, 1: 7.5, 2: 7.5, 3: 7.5, 4: 7.5, 5: 6.0, 6: 0.0}
        jornada = JORNADA.get(fecha.weekday(), 0)

        he_vals = {
            'he_diurna': float(he_diurna),
            'he_nocturna': float(he_nocturna),
            'he_dominical_diurna': float(he_dominical_diurna),
            'he_dominical_nocturna': float(he_dominical_nocturna),
        }
        tiene_he = any(v > 0 for v in he_vals.values())

        # Build initial rows for Alpine
        initial_rows = []
        for field_name, val in he_vals.items():
            if val > 0:
                initial_rows.append(f"{{tipo:'{field_name}',horas:{val}}}")
        rows_js = ','.join(initial_rows) if initial_rows else ""

        he_section = (
            f'<div x-data="{{'
            f'showHE:{str(tiene_he).lower()},'
            f'rows:[{rows_js}],'
            f'calcField(tipo){{return this.rows.filter(r=>r.tipo===tipo).reduce((s,r)=>s+(parseFloat(r.horas)||0),0)}},'
            f'totalHE(){{return this.rows.reduce((s,r)=>s+(parseFloat(r.horas)||0),0)}},'
            f'syncAndSubmit(){{this.$nextTick(()=>{{let c=this.$el.closest(\'.asistencia-cell\');let s=c.querySelector(\'[name=tipo_novedad]\');htmx.trigger(s,\'change\')}})}}'
            f'}}" class="mt-1">'
            f'<div class="flex items-center gap-1">'
            f'<input type="checkbox" x-model="showHE" '
            f'@change="if(showHE&&rows.length===0)rows.push({{tipo:\'he_diurna\',horas:0}});if(!showHE){{rows=[];syncAndSubmit()}}" '
            f'class="rounded border-gray-300 text-orange-600 cursor-pointer">'
            f'<span class="text-xs text-gray-500">HE</span>'
            f'<span class="text-[10px] text-orange-400">({jornada}h)</span>'
            f'</div>'
            f'<div x-show="showHE" x-collapse class="mt-1 space-y-1">'
            f'<template x-for="(row,idx) in rows" :key="idx">'
            f'<div class="flex items-center gap-0.5">'
            f'<select x-model="row.tipo" '
            f'class="text-[10px] rounded border border-orange-200 bg-orange-50 px-0.5 py-0.5 flex-1 '
            f'dark:bg-gray-700 dark:border-gray-600 dark:text-orange-300">'
            f'<option value="he_diurna">Diurna</option>'
            f'<option value="he_nocturna">Nocturna</option>'
            f'<option value="he_dominical_diurna">Dom.D</option>'
            f'<option value="he_dominical_nocturna">Dom.N</option>'
            f'</select>'
            f'<input type="number" x-model.number="row.horas" step="0.5" min="0" max="16" '
            f'@change="syncAndSubmit()" '
            f'class="text-[10px] rounded border border-orange-200 bg-orange-50 px-0.5 py-0.5 w-[35px] '
            f'text-center text-orange-700 dark:bg-gray-700 dark:border-gray-600 dark:text-orange-300">'
            f'<button type="button" '
            f'@click="rows.splice(idx,1);if(rows.length===0)showHE=false;syncAndSubmit()" '
            f'class="text-red-400 hover:text-red-600 text-xs leading-none">&times;</button>'
            f'</div>'
            f'</template>'
            f'<button type="button" '
            f'@click="if(rows.length<4)rows.push({{tipo:\'he_diurna\',horas:0}})" '
            f'x-show="rows.length<4" '
            f'class="text-[10px] text-orange-500 hover:text-orange-700 cursor-pointer">+ Agregar</button>'
            f'</div>'
            f'<input type="hidden" name="he_diurna" :value="calcField(\'he_diurna\')">'
            f'<input type="hidden" name="he_nocturna" :value="calcField(\'he_nocturna\')">'
            f'<input type="hidden" name="he_dominical_diurna" :value="calcField(\'he_dominical_diurna\')">'
            f'<input type="hidden" name="he_dominical_nocturna" :value="calcField(\'he_dominical_nocturna\')">'
            f'<input type="hidden" name="horas_extra" :value="totalHE()">'
            f'</div>'
        )

        html = (
            f'<select name="tipo_novedad" '
            f'hx-post="{request.path}" '
            f'hx-target="closest .asistencia-cell" '
            f'hx-swap="innerHTML" '
            f'hx-include="closest .asistencia-cell" '
            f'class="text-xs rounded border px-1 py-1 w-full cursor-pointer {css} dark:bg-gray-700 dark:border-gray-600 dark:text-gray-200">'
            f'{options_html}</select>'
            f'{obs_field}'
            f'<div class="mt-1 flex items-center gap-1">'
            f'<input type="checkbox" name="viatico_aplica" {viatico_checked} '
            f'hx-post="{request.path}" '
            f'hx-target="closest .asistencia-cell" '
            f'hx-swap="innerHTML" '
            f'hx-include="closest .asistencia-cell" '
            f'class="rounded border-gray-300 text-green-600 cursor-pointer">'
            f'<span class="text-xs text-gray-500">V</span>'
            f'</div>'
            f'{he_section}'
            f'<input type="hidden" name="usuario_id" value="{usuario_id}">'
            f'<input type="hidden" name="fecha" value="{fecha_str}">'
            f'<input type="hidden" name="viaticos" value="{float(viaticos)}">'
        )
        import json
        response = HttpResponse(html)
        response['HX-Trigger'] = json.dumps({
            'updateTotalViaticos': {
                'usuario_id': str(usuario_id),
                'total': total_viaticos_fmt,
            },
            'updateTotalHorasExtra': {
                'usuario_id': str(usuario_id),
                'total': total_he_fmt,
            },
            'updateTotalHorasOrd': {
                'usuario_id': str(usuario_id),
                'total_hord': total_hord,
                'total_htotal': total_htotal,
            },
        })
        return response


class CuadrillaMiembroRemoveView(LoginRequiredMixin, RoleRequiredMixin, DetailView):
    """Remove a member from a cuadrilla (soft delete)."""
    model = Cuadrilla
    allowed_roles = ['admin', 'director', 'coordinador']

    def post(self, request, *args, **kwargs):
        from datetime import date

        cuadrilla = self.get_object()
        miembro_pk = self.kwargs['miembro_pk']

        try:
            miembro = CuadrillaMiembro.objects.get(pk=miembro_pk, cuadrilla=cuadrilla)
            nombre = miembro.usuario.get_full_name()
            miembro.activo = False
            miembro.fecha_fin = date.today()
            miembro.save(update_fields=['activo', 'fecha_fin', 'updated_at'])
            messages.success(request, f'{nombre} removido de la cuadrilla.')
        except CuadrillaMiembro.DoesNotExist:
            messages.error(request, 'Miembro no encontrado.')

        return redirect('cuadrillas:detalle', pk=cuadrilla.pk)


class CuadrillaMiembroUploadView(LoginRequiredMixin, RoleRequiredMixin, DetailView):
    """Upload Excel to batch-update member roles in a cuadrilla."""
    model = Cuadrilla
    allowed_roles = ['admin', 'director', 'coordinador', 'ing_residente']

    def post(self, request, *args, **kwargs):
        import openpyxl
        from io import BytesIO

        cuadrilla = self.get_object()
        archivo = request.FILES.get('archivo')
        if not archivo:
            messages.error(request, 'Debe seleccionar un archivo.')
            return redirect('cuadrillas:detalle', pk=cuadrilla.pk)

        # Issue #176 (A4): RolCuadrilla (TextChoices) eliminado, catalogo
        # ahora es Cargo (dinamico, editable via /cuadrillas/cargos/).
        roles_validos = dict(Cargo.objects.filter(activo=True).values_list('codigo', 'nombre'))
        roles_por_nombre = {v.upper(): k for k, v in roles_validos.items()}

        # Costos fijos por rol
        costos = {
            'SUPERVISOR': 0, 'LINIERO_I': 3176095, 'LINIERO_II': 2804856,
            'AYUDANTE': 1750905, 'CONDUCTOR': 480000, 'ADMINISTRADOR_OBRA': 2522400,
            'PROFESIONAL_SST': 4204000, 'ING_RESIDENTE': 7357000,
            'SERVICIO_GENERAL': 1750905, 'ALMACENISTA': 1800000,
            'SUPERVISOR_FOREST': 2969427, 'ASISTENTE_FOREST': 4204000,
        }

        try:
            wb = openpyxl.load_workbook(BytesIO(archivo.read()), read_only=True)
            ws = wb.active

            # Build lookup of current members by normalized name
            miembros = CuadrillaMiembro.objects.filter(
                cuadrilla=cuadrilla, activo=True
            ).select_related('usuario')
            miembros_por_nombre = {}
            for m in miembros:
                nombre_norm = m.usuario.get_full_name().strip().upper()
                miembros_por_nombre[nombre_norm] = m

            actualizados = 0
            no_encontrados = []

            for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row or not row[0]:
                    continue

                nombre = str(row[0]).strip()
                cargo_raw = str(row[1]).strip().upper() if len(row) > 1 and row[1] else ''

                if not cargo_raw:
                    continue

                # Resolve role code
                rol_code = cargo_raw if cargo_raw in roles_validos else roles_por_nombre.get(cargo_raw, '')
                if not rol_code:
                    continue

                # Match by name
                nombre_norm = nombre.upper()
                miembro = miembros_por_nombre.get(nombre_norm)

                if miembro:
                    miembro.rol_cuadrilla_id = rol_code
                    miembro.costo_dia = costos.get(rol_code, 0)
                    miembro.save(update_fields=['rol_cuadrilla', 'costo_dia', 'updated_at'])
                    actualizados += 1
                else:
                    no_encontrados.append(nombre)

                # Also update PersonalCuadrilla catalog
                documento = str(row[2]).strip() if len(row) > 2 and row[2] else ''
                if documento:
                    PersonalCuadrilla.objects.update_or_create(
                        documento=documento,
                        defaults={'nombre': nombre, 'rol_cuadrilla_id': rol_code, 'activo': True},
                    )

            msg = f'{actualizados} miembro(s) actualizado(s).'
            if no_encontrados:
                msg += f' No encontrados: {", ".join(no_encontrados[:5])}'
                if len(no_encontrados) > 5:
                    msg += f' y {len(no_encontrados) - 5} más.'
            messages.success(request, msg)

        except Exception as e:
            messages.error(request, f'Error al procesar archivo: {str(e)}')

        return redirect('cuadrillas:detalle', pk=cuadrilla.pk)


class ExportarAsistenciaView(LoginRequiredMixin, RoleRequiredMixin, View):
    """Export weekly attendance to Excel."""
    allowed_roles = ['admin', 'director', 'coordinador', 'ing_residente', 'supervisor']

    def get(self, request, pk, *args, **kwargs):
        from datetime import date, timedelta
        from decimal import Decimal
        from io import BytesIO

        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        try:
            cuadrilla = Cuadrilla.objects.get(pk=pk)
        except Cuadrilla.DoesNotExist:
            return HttpResponse('Cuadrilla no encontrada', status=404)

        # Determine week from cuadrilla code
        semana, ano = self._get_semana_from_codigo(cuadrilla.codigo)
        if semana and ano:
            lunes = date.fromisocalendar(ano, semana, 1)
        else:
            hoy = date.today()
            lunes = hoy - timedelta(days=hoy.weekday())

        dias_semana = [lunes + timedelta(days=i) for i in range(7)]
        dias_nombres = ['Lun', 'Mar', 'Mie', 'Jue', 'Vie', 'Sab', 'Dom']

        miembros = cuadrilla.miembros.filter(activo=True).select_related('usuario')
        asistencias = Asistencia.objects.filter(
            cuadrilla=cuadrilla,
            fecha__in=dias_semana,
        ).select_related('usuario')

        # Build dict
        asist_dict = {}
        for a in asistencias:
            uid = str(a.usuario_id)
            if uid not in asist_dict:
                asist_dict[uid] = {}
            asist_dict[uid][a.fecha.isoformat()] = a

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Asistencia'

        # Styles
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        center = Alignment(horizontal='center', vertical='center')

        # Title row
        ws.merge_cells('A1:R1')
        ws['A1'] = f'Asistencia Semanal - {cuadrilla.codigo} - {cuadrilla.nombre}'
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A2:R2')
        ws['A2'] = f'Semana: {dias_semana[0].strftime("%d/%m/%Y")} - {dias_semana[6].strftime("%d/%m/%Y")}'
        ws['A2'].font = Font(size=11)

        # Headers (row 4)
        headers = ['Nombre', 'Documento', 'Cargo', 'Rol']
        for dia, nombre in zip(dias_semana, dias_nombres):
            headers.append(f'{nombre} {dia.strftime("%d/%m")}')
        headers.extend([
            'Total Viaticos', 'H. Extra Total',
            'HE Diurna', 'HE Nocturna', 'HE Dom.D', 'HE Dom.N',
            'Observaciones',
        ])

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = thin_border

        # Data rows
        novedad_labels = dict(Asistencia.TipoNovedad.choices)
        row = 5
        for miembro in miembros:
            uid = str(miembro.usuario_id)
            user_asist = asist_dict.get(uid, {})

            ws.cell(row=row, column=1, value=miembro.usuario.get_full_name()).border = thin_border
            ws.cell(row=row, column=2, value=getattr(miembro.usuario, 'documento', '')).border = thin_border
            ws.cell(row=row, column=3, value=miembro.get_rol_cuadrilla_display()).border = thin_border
            ws.cell(row=row, column=4, value=miembro.get_cargo_display()).border = thin_border

            total_viaticos = Decimal('0')
            total_horas_extra = Decimal('0')
            total_he_diurna = Decimal('0')
            total_he_nocturna = Decimal('0')
            total_he_dom_diurna = Decimal('0')
            total_he_dom_nocturna = Decimal('0')
            observaciones_semana = []

            for i, dia in enumerate(dias_semana):
                asist = user_asist.get(dia.isoformat())
                col = 5 + i
                if asist:
                    cell = ws.cell(row=row, column=col, value=novedad_labels.get(asist.tipo_novedad, ''))
                    cell.alignment = center
                    cell.border = thin_border
                    total_viaticos += asist.viaticos
                    total_horas_extra += asist.horas_extra
                    total_he_diurna += asist.he_diurna
                    total_he_nocturna += asist.he_nocturna
                    total_he_dom_diurna += asist.he_dominical_diurna
                    total_he_dom_nocturna += asist.he_dominical_nocturna
                    if asist.observacion:
                        observaciones_semana.append(f'{dias_nombres[i]}: {asist.observacion}')

                    # Color coding
                    color_map = {
                        'PRESENTE': '92D050',
                        'AUSENTE': 'FF6B6B',
                        'VACACIONES': '6BB5FF',
                        'INCAPACIDAD': 'FFB366',
                        'PERMISO': 'C39BD3',
                        'LICENCIA': 'F7DC6F',
                        'CAPACITACION': '76D7C4',
                        'COMPENSATORIO': '67E8F9',
                        'DESCANSO': 'CBD5E1',
                    }
                    fill_color = color_map.get(asist.tipo_novedad)
                    if fill_color:
                        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                else:
                    cell = ws.cell(row=row, column=col, value='---')
                    cell.alignment = center
                    cell.border = thin_border

            # Totals
            cell_v = ws.cell(row=row, column=12, value=float(total_viaticos))
            cell_v.number_format = '$#,##0'
            cell_v.alignment = center
            cell_v.border = thin_border

            cell_h = ws.cell(row=row, column=13, value=float(total_horas_extra))
            cell_h.number_format = '0.0'
            cell_h.alignment = center
            cell_h.border = thin_border

            # Detailed overtime columns
            for col_offset, val in enumerate([total_he_diurna, total_he_nocturna, total_he_dom_diurna, total_he_dom_nocturna]):
                c = ws.cell(row=row, column=14 + col_offset, value=float(val))
                c.number_format = '0.0'
                c.alignment = center
                c.border = thin_border

            ws.cell(row=row, column=18, value='; '.join(observaciones_semana)).border = thin_border

            row += 1

        # Auto-fit column widths (MergedCell no tiene column_letter).
        from openpyxl.utils import get_column_letter
        for col in ws.columns:
            max_len = 0
            col_idx = col[0].column
            if col_idx is None:
                continue
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 35)

        # Write to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f'asistencia_{cuadrilla.codigo}_{lunes.strftime("%Y%m%d")}.xlsx'
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    @staticmethod
    def _get_semana_from_codigo(codigo):
        try:
            parts = codigo.split('-')
            if len(parts) >= 2:
                semana = int(parts[0])
                ano = int(parts[1])
                if 1 <= semana <= 53 and 2000 <= ano <= 2100:
                    return semana, ano
        except (ValueError, IndexError):
            pass
        return None, None


class PersonalCuadrillaUploadView(LoginRequiredMixin, RoleRequiredMixin, View):
    """Upload crew personnel from Excel/CSV.

    Columnas soportadas (issue #176, A5): Nombre | Documento | Cargo |
    Salario Base | Fecha Ingreso | Fecha Salida. Las 3 columnas nuevas son
    opcionales (retrocompatible con el formato original de 3 columnas) --
    si faltan, quedan en su default del modelo (salario_base=0,
    fecha_ingreso/fecha_salida=None -> activo=True).
    """
    allowed_roles = ['admin', 'director', 'coordinador']

    @staticmethod
    def _parse_fecha(valor):
        """Acepta date/datetime nativos de openpyxl o texto 'YYYY-MM-DD'."""
        from datetime import date, datetime

        if valor in (None, ''):
            return None
        if isinstance(valor, datetime):
            return valor.date()
        if isinstance(valor, date):
            return valor
        texto = str(valor).strip()
        if not texto:
            return None
        for fmt in ('%Y-%m-%d', '%d/%m/%Y'):
            try:
                return datetime.strptime(texto, fmt).date()
            except ValueError:
                continue
        return None

    @staticmethod
    def _parse_decimal(valor):
        from decimal import Decimal, InvalidOperation

        if valor in (None, ''):
            return Decimal('0')
        try:
            return Decimal(str(valor).replace(',', '').strip() or '0')
        except InvalidOperation:
            return Decimal('0')

    def post(self, request, *args, **kwargs):
        import openpyxl
        from io import BytesIO

        archivo = request.FILES.get('archivo')
        if not archivo:
            messages.error(request, 'Debe seleccionar un archivo.')
            return redirect('cuadrillas:colaboradores_lista')

        # Issue #176 (A4): RolCuadrilla (TextChoices) eliminado, catalogo
        # ahora es Cargo (dinamico, editable via /cuadrillas/cargos/).
        roles_validos = dict(Cargo.objects.filter(activo=True).values_list('codigo', 'nombre'))
        # Also build reverse map: display name -> key
        roles_por_nombre = {v.upper(): k for k, v in roles_validos.items()}

        try:
            wb = openpyxl.load_workbook(BytesIO(archivo.read()))
            ws = wb.active
            creados = 0
            actualizados = 0
            errores = []

            for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row or not row[0]:
                    continue

                try:
                    nombre = str(row[0]).strip()
                    documento = str(row[1]).strip() if len(row) > 1 and row[1] else ''
                    rol_raw = str(row[2]).strip().upper() if len(row) > 2 and row[2] else ''
                    salario_raw = row[3] if len(row) > 3 else None
                    fecha_ingreso_raw = row[4] if len(row) > 4 else None
                    fecha_salida_raw = row[5] if len(row) > 5 else None

                    if not nombre or not documento:
                        errores.append(f'Fila {idx}: nombre o documento vacío')
                        continue

                    # Resolve role
                    rol = rol_raw if rol_raw in roles_validos else roles_por_nombre.get(rol_raw, 'LINIERO_I')
                    salario_base = self._parse_decimal(salario_raw)
                    fecha_ingreso = self._parse_fecha(fecha_ingreso_raw)
                    fecha_salida = self._parse_fecha(fecha_salida_raw)

                    defaults = {
                        'nombre': nombre,
                        'rol_cuadrilla_id': rol,
                        'salario_base': salario_base,
                        'fecha_ingreso': fecha_ingreso,
                        'fecha_salida': fecha_salida,
                        # activo se resuelve en PersonalCuadrilla.save(): True salvo
                        # que fecha_salida venga poblada (issue #176, A2).
                        'activo': True,
                    }

                    obj, created = PersonalCuadrilla.objects.update_or_create(
                        documento=documento,
                        defaults=defaults,
                    )
                    if created:
                        creados += 1
                    else:
                        actualizados += 1
                except Exception as row_exc:
                    # Documento duplicado intra-archivo u otro error de fila:
                    # se reporta y se sigue con el resto (no se aborta el lote).
                    errores.append(f'Fila {idx}: {row_exc}')
                    continue

            msg = f'Personal cargado: {creados} nuevos, {actualizados} actualizados.'
            if errores:
                msg += f' Errores: {len(errores)} filas.'
            messages.success(request, msg)

        except Exception as e:
            messages.error(request, f'Error al procesar archivo: {str(e)}')

        # Issue #188 (A11): el trigger de esta carga vive ahora en
        # /cuadrillas/colaboradores/ (antes en /cuadrillas/) — redirige ahí
        # para que el usuario vea el maestro actualizado en el mismo lugar
        # donde subió el archivo.
        return redirect('cuadrillas:colaboradores_lista')


class PersonalCuadrillaAPIView(LoginRequiredMixin, View):
    """API endpoint to get crew personnel info by user selection."""

    def get(self, request, *args, **kwargs):
        documento = request.GET.get('documento', '').strip()
        personal_id = request.GET.get('id', '').strip()

        personal = None
        if personal_id:
            personal = PersonalCuadrilla.objects.filter(id=personal_id, activo=True).first()
        elif documento:
            personal = PersonalCuadrilla.objects.filter(documento=documento, activo=True).first()

        if personal:
            return JsonResponse({
                'nombre': personal.nombre,
                'documento': personal.documento,
                'rol_cuadrilla': personal.rol_cuadrilla_id,
                # Issue #188 (A5): celular nuevo en el maestro Colaboradores,
                # expuesto para el autocompletado del grid editable.
                'celular': personal.celular,
            })
        return JsonResponse({}, status=404)


class PersonalCuadrillaListAPIView(LoginRequiredMixin, View):
    """API endpoint to list all active crew personnel."""

    def get(self, request, *args, **kwargs):
        personal = PersonalCuadrilla.objects.filter(activo=True).order_by('nombre')
        data = [
            {
                'id': str(p.id),
                'nombre': p.nombre,
                'documento': p.documento,
                'rol_cuadrilla': p.rol_cuadrilla_id,
            }
            for p in personal
        ]
        return JsonResponse(data, safe=False)


class CostoRolAPIView(LoginRequiredMixin, RoleRequiredMixin, View):
    """API endpoint: costo/salario base de un Cargo.

    Retrofit issue #176 (bounce 3, A2): antes tenia un dict hardcoded de
    costos por rol (huerfano -- grep 2026-07-19 confirmo CERO consumidores
    en templates/JS antes de este retrofit). Se retrofitea para leer
    Cargo.salario_base y se convierte en el backend real del autocompletado
    de colaboradores_form.html (A3). Mantiene el contrato JSON original
    (costo_dia/es_conductor/conductor_interno) para minimizar diff -- NO se
    renombra `costo_dia` pese a que hoy contiene un valor MENSUAL (mismo
    nombre confuso que ya traia el dict original, fuera de scope
    corregirlo).
    """
    allowed_roles = ['admin', 'director', 'coordinador', 'ing_residente', 'supervisor']

    def get(self, request, *args, **kwargs):
        rol = request.GET.get('rol', '').strip()
        if not rol:
            return JsonResponse({'costo_dia': 0})

        cargo = Cargo.objects.filter(codigo=rol, activo=True).first()
        # float(): Decimal NO es JSON-serializable por default -- JsonResponse
        # truena con TypeError si se pasa el Decimal crudo (issue #176 A2).
        costo = float(cargo.salario_base) if cargo else 0

        # Para conductor, diferenciar interno/externo
        es_conductor = rol == 'CONDUCTOR'
        conductor_interno = request.GET.get('conductor_interno', 'true') == 'true'

        return JsonResponse({
            'costo_dia': costo,
            'es_conductor': es_conductor,
            'conductor_interno': conductor_interno if es_conductor else None,
        })


class CuadrillaMasivaUploadView(LoginRequiredMixin, RoleRequiredMixin, TemplateView):
    """Bulk upload of cuadrillas and their members from Excel."""
    template_name = 'cuadrillas/masiva_upload.html'
    allowed_roles = ['admin', 'director', 'coordinador']

    COSTOS = {
        'SUPERVISOR': 0, 'LINIERO_I': 3176095, 'LINIERO_II': 2804856,
        'AYUDANTE': 1750905, 'CONDUCTOR': 480000, 'ADMINISTRADOR_OBRA': 2522400,
        'PROFESIONAL_SST': 4204000, 'ING_RESIDENTE': 7357000,
        'SERVICIO_GENERAL': 1750905, 'ALMACENISTA': 1800000,
        'SUPERVISOR_FOREST': 2969427, 'ASISTENTE_FOREST': 4204000,
    }

    def post(self, request, *args, **kwargs):
        import openpyxl
        from io import BytesIO
        from datetime import date
        from apps.usuarios.models import Usuario
        from apps.lineas.models import Linea

        archivo = request.FILES.get('archivo')
        if not archivo:
            messages.error(request, 'Debe seleccionar un archivo Excel.')
            return self.get(request, *args, **kwargs)

        try:
            wb = openpyxl.load_workbook(BytesIO(archivo.read()), read_only=True)
            ws = wb.active

            cuadrillas_creadas = 0
            miembros_agregados = 0
            errores = []

            # Expected columns:
            # A: Cuadrilla (numero), B: Año, C: Actividad, D: Fecha,
            # E: Supervisor, F: Linea asignada, G: Vehiculo,
            # H: Miembro1, I: Miembro2, J: Miembro3, K: Miembro4, L: Miembro5

            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row or not row[0]:
                    continue

                try:
                    cuadrilla_num = str(row[0]).strip()
                    ano = str(row[1]).strip() if len(row) > 1 and row[1] else str(date.today().year)
                    actividad = str(row[2]).strip() if len(row) > 2 and row[2] else ''
                    fecha_str = str(row[3]).strip() if len(row) > 3 and row[3] else ''
                    supervisor_name = str(row[4]).strip() if len(row) > 4 and row[4] else ''
                    linea_name = str(row[5]).strip() if len(row) > 5 and row[5] else ''
                    vehiculo_placa = str(row[6]).strip() if len(row) > 6 and row[6] else ''

                    # Generate week number from date or current week
                    from datetime import datetime
                    fecha = None
                    semana = date.today().isocalendar()[1]
                    if fecha_str:
                        try:
                            if isinstance(row[3], datetime):
                                fecha = row[3].date()
                            else:
                                fecha = date.fromisoformat(fecha_str)
                            semana = fecha.isocalendar()[1]
                        except (ValueError, TypeError):
                            pass

                    # Build cuadrilla code: WW-YYYY-NNN
                    codigo = f'{semana:02d}-{ano}-{cuadrilla_num.zfill(3)}'
                    nombre = f'Cuadrilla {cuadrilla_num}'
                    if actividad:
                        nombre = f'{nombre} - {actividad}'

                    # Find supervisor
                    supervisor = None
                    if supervisor_name:
                        supervisor = Usuario.objects.filter(
                            is_active=True,
                            rol='supervisor',
                        ).filter(
                            models.Q(first_name__icontains=supervisor_name.split()[0]) if supervisor_name.split() else models.Q()
                        ).first()

                    # Find linea
                    linea = None
                    if linea_name:
                        linea = Linea.objects.filter(
                            activa=True,
                        ).filter(
                            models.Q(codigo__icontains=linea_name) | models.Q(nombre__icontains=linea_name)
                        ).first()

                    # Find vehiculo
                    vehiculo = None
                    if vehiculo_placa:
                        vehiculo = Vehiculo.objects.filter(placa__icontains=vehiculo_placa, activo=True).first()

                    # Create or update cuadrilla
                    cuadrilla, created = Cuadrilla.objects.update_or_create(
                        codigo=codigo,
                        defaults={
                            'nombre': nombre,
                            'supervisor': supervisor,
                            'linea_asignada': linea,
                            'vehiculo': vehiculo,
                            'fecha': fecha,
                            'activa': True,
                        }
                    )
                    if created:
                        cuadrillas_creadas += 1

                    # Add members (columns H onwards)
                    for col_idx in range(7, min(len(row), 17)):
                        miembro_ref = str(row[col_idx]).strip() if row[col_idx] else ''
                        if not miembro_ref:
                            continue

                        # Try to find by documento first, then by name
                        usuario = None
                        personal = PersonalCuadrilla.objects.filter(
                            models.Q(documento=miembro_ref) | models.Q(nombre__icontains=miembro_ref),
                            activo=True
                        ).first()

                        if personal:
                            # Find corresponding Usuario
                            usuario = Usuario.objects.filter(documento=personal.documento).first()
                            # Issue #176 (A4): normalizar a rol_cuadrilla_id
                            # (string) -- antes mezclaba objeto Cargo (esta
                            # rama) con string (rama else), lo que hacia que
                            # self.COSTOS.get(rol, 0) matcheara silenciosamente
                            # a 0 cuando rol era un objeto.
                            rol = personal.rol_cuadrilla_id
                        else:
                            # Try direct user lookup
                            usuario = Usuario.objects.filter(
                                models.Q(documento=miembro_ref) |
                                models.Q(first_name__icontains=miembro_ref.split()[0] if miembro_ref.split() else ''),
                                is_active=True
                            ).first()
                            rol = 'LINIERO_I'

                        if usuario:
                            _, member_created = CuadrillaMiembro.objects.get_or_create(
                                cuadrilla=cuadrilla,
                                usuario=usuario,
                                activo=True,
                                defaults={
                                    'rol_cuadrilla_id': rol,
                                    'cargo': 'MIEMBRO',
                                    'costo_dia': self.COSTOS.get(rol, 0),
                                    'fecha_inicio': fecha or date.today(),
                                }
                            )
                            if member_created:
                                miembros_agregados += 1
                        else:
                            errores.append(f'Fila {row_num}: Miembro "{miembro_ref}" no encontrado.')

                except Exception as e:
                    errores.append(f'Fila {row_num}: {str(e)}')

            wb.close()

            msg = f'Carga completada: {cuadrillas_creadas} cuadrillas creadas, {miembros_agregados} miembros agregados.'
            if errores:
                msg += f' {len(errores)} advertencias.'
            messages.success(request, msg)

            for err in errores[:10]:
                messages.warning(request, err)

        except Exception as e:
            messages.error(request, f'Error al procesar el archivo: {str(e)}')

        return self.get(request, *args, **kwargs)


class DescargarPlantillaCuadrillasView(LoginRequiredMixin, RoleRequiredMixin, View):
    """
    Descarga plantilla Excel para carga masiva de cuadrillas.
    Agregado: 1 abril 2026
    """
    allowed_roles = ['admin', 'director', 'coordinador', 'ing_residente']

    def get(self, request):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        from datetime import date

        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Cuadrillas"

        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Headers
        headers = [
            'CuadrillaNum', 'Año', 'Actividad', 'Fecha',
            'Supervisor', 'Línea', 'Vehículo',
            'Miembro1', 'Miembro2', 'Miembro3', 'Miembro4', 'Miembro5',
            'Miembro6', 'Miembro7', 'Miembro8', 'Miembro9', 'Miembro10'
        ]

        # Escribir headers con estilo
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Fila de ejemplo
        ejemplo = [
            1,                                      # CuadrillaNum
            date.today().year,                      # Año
            'Poda de vegetación',                   # Actividad
            date.today().strftime('%Y-%m-%d'),      # Fecha
            'Juan Pérez',                           # Supervisor
            'LT-811',                               # Línea
            'ABC-123',                              # Vehículo
            'Carlos López',                         # Miembro1
            'María García',                         # Miembro2
            '',                                     # Miembro3
            '',                                     # Miembro4
            '',                                     # Miembro5
            '',                                     # Miembro6
            '',                                     # Miembro7
            '',                                     # Miembro8
            '',                                     # Miembro9
            '',                                     # Miembro10
        ]
        ws.append(ejemplo)

        # Ajustar anchos de columna
        column_widths = {
            'A': 14,  # CuadrillaNum
            'B': 8,   # Año
            'C': 25,  # Actividad
            'D': 12,  # Fecha
            'E': 20,  # Supervisor
            'F': 15,  # Línea
            'G': 12,  # Vehículo
        }
        for i in range(8, 18):  # Miembros 1-10
            column_widths[get_column_letter(i)] = 18

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Agregar instrucciones en hoja separada
        ws_instrucciones = wb.create_sheet("Instrucciones")
        instrucciones = [
            "INSTRUCCIONES PARA CARGA MASIVA DE CUADRILLAS",
            "",
            "1. Complete la hoja 'Cuadrillas' con los datos de sus cuadrillas",
            "",
            "2. Formato de columnas:",
            "   - CuadrillaNum: Número secuencial de cuadrilla (ej: 1, 2, 3...)",
            "   - Año: Año de la cuadrilla (ej: 2026)",
            "   - Actividad: Nombre de la actividad principal",
            "   - Fecha: Fecha en formato YYYY-MM-DD (ej: 2026-04-15)",
            "   - Supervisor: Nombre completo del supervisor",
            "   - Línea: Código de la línea (ej: LT-811)",
            "   - Vehículo: Placa del vehículo (ej: ABC-123)",
            "   - Miembro1 a Miembro10: Nombres completos de los miembros",
            "",
            "3. El supervisor debe estar previamente registrado en el sistema",
            "",
            "4. Los miembros se buscarán por nombre. Si no existen, se crearán automáticamente",
            "",
            "5. Puede dejar en blanco las columnas de miembros que no necesite",
            "",
            "6. Guarde el archivo y súbalo en la opción 'Carga Masiva' del sistema",
            "",
            "IMPORTANTE:",
            "- No modifique los nombres de las columnas",
            "- Use el formato de fecha especificado",
            "- Asegúrese de que la línea existe en el sistema",
        ]

        for row_num, instruccion in enumerate(instrucciones, 1):
            cell = ws_instrucciones.cell(row=row_num, column=1)
            cell.value = instruccion
            if row_num == 1:
                cell.font = Font(bold=True, size=14)
            elif instruccion.startswith("IMPORTANTE:"):
                cell.font = Font(bold=True, color="FF0000")

        ws_instrucciones.column_dimensions['A'].width = 80

        # Preparar respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'plantilla_cuadrillas_{date.today().strftime("%Y%m%d")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename={filename}'

        wb.save(response)
        return response


# ---------------------------------------------------------------------------
# Colaboradores — CRUD sobre PersonalCuadrilla (issue #176, A3)
# ---------------------------------------------------------------------------
class ColaboradorListView(LoginRequiredMixin, RoleRequiredMixin, ListView):
    """Listado del maestro Colaboradores, con activos e inactivos."""
    model = PersonalCuadrilla
    template_name = 'cuadrillas/colaboradores_lista.html'
    context_object_name = 'colaboradores'
    allowed_roles = ['admin', 'director', 'coordinador', 'ing_residente']

    def get_queryset(self):
        qs = PersonalCuadrilla.objects.all()
        estado = self.request.GET.get('estado', '').strip()
        if estado == 'activos':
            qs = qs.filter(activo=True)
        elif estado == 'inactivos':
            qs = qs.filter(activo=False)
        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['estado_filtro'] = self.request.GET.get('estado', '')
        context['total_activos'] = PersonalCuadrilla.objects.filter(activo=True).count()
        context['total_inactivos'] = PersonalCuadrilla.objects.filter(activo=False).count()
        return context


class ColaboradorCreateView(LoginRequiredMixin, RoleRequiredMixin, HTMXMixin, TemplateView):
    """Crear un nuevo Colaborador (PersonalCuadrilla)."""
    template_name = 'cuadrillas/colaboradores_form.html'
    allowed_roles = ['admin', 'director', 'coordinador']

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.setdefault('form', PersonalCuadrillaForm())
        context['modo'] = 'crear'
        return context

    def post(self, request, *args, **kwargs):
        form = PersonalCuadrillaForm(request.POST)
        if form.is_valid():
            colaborador = form.save()
            messages.success(request, f'Colaborador "{colaborador.nombre}" creado exitosamente.')
            return redirect('cuadrillas:colaboradores_lista')
        messages.error(request, 'Revise los errores del formulario.')
        return self.render_to_response(self.get_context_data(form=form))


class ColaboradorEditView(LoginRequiredMixin, RoleRequiredMixin, HTMXMixin, DetailView):
    """Editar un Colaborador (PersonalCuadrilla) existente."""
    model = PersonalCuadrilla
    template_name = 'cuadrillas/colaboradores_form.html'
    context_object_name = 'colaborador'
    allowed_roles = ['admin', 'director', 'coordinador']

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.setdefault('form', PersonalCuadrillaForm(instance=self.object))
        context['modo'] = 'editar'
        return context

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        form = PersonalCuadrillaForm(request.POST, instance=self.object)
        if form.is_valid():
            colaborador = form.save()
            messages.success(request, f'Colaborador "{colaborador.nombre}" actualizado exitosamente.')
            return redirect('cuadrillas:colaboradores_lista')
        messages.error(request, 'Revise los errores del formulario.')
        return self.render_to_response(self.get_context_data(form=form))


class ColaboradorInactivarView(LoginRequiredMixin, RoleRequiredMixin, DetailView):
    """Alterna activo/inactivo de un Colaborador. Nunca borra el registro.

    Al reactivar (de inactivo a activo), limpia fecha_salida (si no,
    save() la volveria a marcar inactivo inmediatamente — ver A2).
    """
    model = PersonalCuadrilla
    allowed_roles = ['admin', 'director', 'coordinador']

    def post(self, request, *args, **kwargs):
        colaborador = self.get_object()
        if colaborador.activo:
            from datetime import date
            colaborador.fecha_salida = date.today()
            colaborador.save(update_fields=['fecha_salida', 'activo', 'updated_at'])
            messages.success(
                request,
                f'Colaborador "{colaborador.nombre}" inactivado. '
                'No aparecerá en el picklist de asignación a cuadrillas.',
            )
        else:
            colaborador.fecha_salida = None
            colaborador.activo = True
            colaborador.save(update_fields=['fecha_salida', 'activo', 'updated_at'])
            messages.success(request, f'Colaborador "{colaborador.nombre}" reactivado.')
        return redirect('cuadrillas:colaboradores_lista')


# ---------------------------------------------------------------------------
# Cargos — CRUD sobre Cargo, Maestro 3 (issue #176, bounce 2)
# ---------------------------------------------------------------------------
class CargoListView(LoginRequiredMixin, RoleRequiredMixin, ListView):
    """Listado del maestro Cargos, con activos e inactivos."""
    model = Cargo
    template_name = 'cuadrillas/cargos_lista.html'
    context_object_name = 'cargos'
    allowed_roles = ['admin', 'director', 'coordinador', 'ing_residente']

    def get_queryset(self):
        qs = Cargo.objects.all()
        estado = self.request.GET.get('estado', '').strip()
        if estado == 'activos':
            qs = qs.filter(activo=True)
        elif estado == 'inactivos':
            qs = qs.filter(activo=False)
        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['estado_filtro'] = self.request.GET.get('estado', '')
        context['total_activos'] = Cargo.objects.filter(activo=True).count()
        context['total_inactivos'] = Cargo.objects.filter(activo=False).count()
        return context


class CargoCreateView(LoginRequiredMixin, RoleRequiredMixin, HTMXMixin, TemplateView):
    """Crear un nuevo Cargo."""
    template_name = 'cuadrillas/cargos_form.html'
    allowed_roles = ['admin', 'director', 'coordinador']

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.setdefault('form', CargoForm())
        context['modo'] = 'crear'
        return context

    def post(self, request, *args, **kwargs):
        form = CargoForm(request.POST)
        if form.is_valid():
            cargo = form.save()
            messages.success(request, f'Cargo "{cargo.nombre}" creado exitosamente.')
            return redirect('cuadrillas:cargos_lista')
        messages.error(request, 'Revise los errores del formulario.')
        return self.render_to_response(self.get_context_data(form=form))


class CargoEditView(LoginRequiredMixin, RoleRequiredMixin, HTMXMixin, DetailView):
    """Editar un Cargo existente. `codigo` queda de solo lectura (ver CargoForm)."""
    model = Cargo
    template_name = 'cuadrillas/cargos_form.html'
    context_object_name = 'cargo'
    allowed_roles = ['admin', 'director', 'coordinador']

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.setdefault('form', CargoForm(instance=self.object))
        context['modo'] = 'editar'
        return context

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        form = CargoForm(request.POST, instance=self.object)
        if form.is_valid():
            cargo = form.save()
            messages.success(request, f'Cargo "{cargo.nombre}" actualizado exitosamente.')
            return redirect('cuadrillas:cargos_lista')
        messages.error(request, 'Revise los errores del formulario.')
        return self.render_to_response(self.get_context_data(form=form))


class CargoInactivarView(LoginRequiredMixin, RoleRequiredMixin, DetailView):
    """Alterna activo/inactivo de un Cargo. Nunca borra el registro
    (no hay delete duro — Cargo puede estar referenciado por
    PersonalCuadrilla/CuadrillaMiembro via FK PROTECT, ver A3)."""
    model = Cargo
    allowed_roles = ['admin', 'director', 'coordinador']

    def post(self, request, *args, **kwargs):
        cargo = self.get_object()
        cargo.activo = not cargo.activo
        cargo.save(update_fields=['activo', 'updated_at'])
        if cargo.activo:
            messages.success(request, f'Cargo "{cargo.nombre}" reactivado.')
        else:
            messages.success(
                request,
                f'Cargo "{cargo.nombre}" inactivado. '
                'No aparecerá en el picklist de asignación de Colaboradores/Miembros.',
            )
        return redirect('cuadrillas:cargos_lista')


class CargoUploadView(LoginRequiredMixin, RoleRequiredMixin, View):
    """Carga masiva de Cargos (issue #176, A5). Upsert por código: crea el
    código si no existe, actualiza nombre+salario_base si ya existe (NUNCA
    reasigna el propio código -- es el campo de lookup)."""
    allowed_roles = ['admin', 'director', 'coordinador']

    def post(self, request, *args, **kwargs):
        import openpyxl
        from decimal import Decimal, InvalidOperation
        from io import BytesIO

        archivo = request.FILES.get('archivo')
        if not archivo:
            messages.error(request, 'Debe seleccionar un archivo.')
            return redirect('cuadrillas:cargos_lista')

        creados = actualizados = 0
        errores = []
        try:
            wb = openpyxl.load_workbook(BytesIO(archivo.read()))
            ws = wb.active
            for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row or not row[0]:
                    continue
                try:
                    codigo = str(row[0]).strip().upper()
                    if not codigo:
                        errores.append(f'Fila {idx}: código vacío')
                        continue
                    nombre = str(row[1]).strip() if len(row) > 1 and row[1] else codigo
                    try:
                        salario_raw = row[2] if len(row) > 2 else None
                        salario_base = (
                            Decimal(str(salario_raw).replace(',', '').strip() or '0')
                            if salario_raw not in (None, '')
                            else Decimal('0')
                        )
                    except InvalidOperation:
                        salario_base = Decimal('0')

                    _, created = Cargo.objects.update_or_create(
                        codigo=codigo,
                        defaults={'nombre': nombre, 'salario_base': salario_base, 'activo': True},
                    )
                    if created:
                        creados += 1
                    else:
                        actualizados += 1
                except Exception as row_exc:
                    errores.append(f'Fila {idx}: {row_exc}')

            msg = f'Cargos cargados: {creados} nuevos, {actualizados} actualizados.'
            if errores:
                msg += f' Errores: {len(errores)} filas.'
            messages.success(request, msg)
        except Exception as e:
            messages.error(request, f'Error al procesar archivo: {str(e)}')

        return redirect('cuadrillas:cargos_lista')


class CargoExportView(LoginRequiredMixin, RoleRequiredMixin, View):
    """Exporta el catálogo Cargo a xlsx (issue #176, A5)."""
    allowed_roles = ['admin', 'director', 'coordinador', 'ing_residente']

    def get(self, request, *args, **kwargs):
        import openpyxl
        from io import BytesIO

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Cargos'
        ws.append(['Código', 'Nombre', 'Salario Base'])
        for cargo in Cargo.objects.all().order_by('nombre'):
            ws.append([cargo.codigo, cargo.nombre, float(cargo.salario_base)])

        buffer = BytesIO()
        wb.save(buffer)
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="cargos.xlsx"'
        return response


class ColaboradorExportView(LoginRequiredMixin, RoleRequiredMixin, View):
    """Exporta el maestro de Colaboradores a xlsx (issue #176, A6)."""
    allowed_roles = ['admin', 'director', 'coordinador', 'ing_residente']

    def get(self, request, *args, **kwargs):
        import openpyxl
        from io import BytesIO

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Colaboradores'
        ws.append(['Documento', 'Nombre', 'Cargo', 'Salario Base', 'Fecha Ingreso', 'Fecha Salida'])
        for p in PersonalCuadrilla.objects.select_related('rol_cuadrilla').all().order_by('nombre'):
            ws.append([
                p.documento,
                p.nombre,
                p.rol_cuadrilla.nombre if p.rol_cuadrilla_id else '',
                float(p.salario_base),
                p.fecha_ingreso.strftime('%Y-%m-%d') if p.fecha_ingreso else '',
                p.fecha_salida.strftime('%Y-%m-%d') if p.fecha_salida else '',
            ])

        buffer = BytesIO()
        wb.save(buffer)
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="colaboradores.xlsx"'
        return response
