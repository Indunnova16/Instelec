"""
Tests issue #176 — Maestros editables: Tipos de Actividad, Colaboradores y
Cargos.

Cubre (dentro de apps/cuadrillas):
- A2 (bounce 1): migración PersonalCuadrilla (salario_base, fecha_ingreso,
  fecha_salida) + save()/signal que fija activo=False cuando se registra
  fecha_salida.
- A3 (bounce 1): CRUD Colaboradores en /cuadrillas/colaboradores/.
- A5 (bounce 1): Importer de colaboradores (extiende PersonalCuadrillaUploadView).
- A4 (bounce 1): Refactor de asignación a cuadrilla (picklist activo=True,
  cargo bloqueado, autocompletado AJAX, resolución/creación de Usuario).
- TestMaestro3A1CargoModeloYSeed (bounce 2, Maestro 3): modelo Cargo +
  migración de seed 0019_seed_cargos (14 códigos de la unión de ambos
  RolCuadrilla). Las clases `TestMaestro3A*` cubren los sub-items A1-A6 del
  plan SPRINTS/PLAN_2026-07-10_maestro_cargos.md — usan el prefijo
  "Maestro3" para no colisionar con los nombres TestA2/A3/A4/A5 del
  bounce 1 de arriba (letras reusadas en el plan nuevo).

Issue: Indunnova16/Instelec#176

Ejecutar con:
    pytest apps/cuadrillas/tests_issue_176.py -v
"""

import importlib
from datetime import date, timedelta
from decimal import Decimal
from io import BytesIO

from django.apps import apps as django_apps
from django.contrib.auth import get_user_model
from django.test import Client, TestCase
from django.urls import reverse

from apps.cuadrillas.importers import ProgramacionS18CuadrillaImporter
from apps.cuadrillas.models import Cargo, Cuadrilla, CuadrillaMiembro, PersonalCuadrilla
from apps.cuadrillas.tests_s18 import _act, _build_s18_excel, _crear_linea, _crear_usuario, _miembro

Usuario = get_user_model()


def _crear_admin():
    return Usuario.objects.create_user(
        email="admin_176@test.com",
        password="testpass123!",
        first_name="Admin",
        last_name="Test176",
        rol="admin",
        is_staff=True,
        is_superuser=True,
    )


# ---------------------------------------------------------------------------
# A2 — Migración PersonalCuadrilla: campos nuevos + save()/signal fecha_salida
# ---------------------------------------------------------------------------
class TestA2PersonalCuadrillaCamposNuevos(TestCase):
    """A2: salario_base/fecha_ingreso/fecha_salida + activo=False automático."""

    def test_migracion_aplica_sobre_datos_legacy_defaults_sensatos(self):
        """Un PersonalCuadrilla creado SIN los campos nuevos usa defaults sensatos
        (salario_base=0, fecha_ingreso/fecha_salida=None) — simula un registro
        legacy pre-existente antes de esta migración."""
        legacy = PersonalCuadrilla.objects.create(
            nombre="Colaborador Legacy",
            documento="LEG-0001",
            rol_cuadrilla_id="LINIERO_I",
        )
        legacy.refresh_from_db()
        self.assertEqual(legacy.salario_base, Decimal("0"))
        self.assertIsNone(legacy.fecha_ingreso)
        self.assertIsNone(legacy.fecha_salida)
        self.assertTrue(legacy.activo)

    def test_registrar_fecha_salida_inactiva_automaticamente(self):
        """Al registrar fecha_salida, activo pasa a False vía save()."""
        persona = PersonalCuadrilla.objects.create(
            nombre="Juan Perez",
            documento="176-0001",
            rol_cuadrilla_id="AYUDANTE",
            salario_base=Decimal("1750905"),
            fecha_ingreso=date(2025, 1, 15),
            activo=True,
        )
        self.assertTrue(persona.activo)

        persona.fecha_salida = date.today()
        persona.save()
        persona.refresh_from_db()

        self.assertFalse(persona.activo)
        self.assertIsNotNone(persona.fecha_salida)

    def test_sin_fecha_salida_permanece_activo(self):
        """Sin fecha_salida, el colaborador permanece activo=True (no se toca)."""
        persona = PersonalCuadrilla.objects.create(
            nombre="Maria Gomez",
            documento="176-0002",
            rol_cuadrilla_id="LINIERO_II",
            salario_base=Decimal("2804856"),
            fecha_ingreso=date(2025, 3, 1),
        )
        persona.refresh_from_db()
        self.assertTrue(persona.activo)
        self.assertIsNone(persona.fecha_salida)

        # Editar otro campo sin tocar fecha_salida no debe desactivar.
        persona.salario_base = Decimal("2900000")
        persona.save()
        persona.refresh_from_db()
        self.assertTrue(persona.activo)


# ---------------------------------------------------------------------------
# A3 — CRUD Colaboradores en /cuadrillas/colaboradores/
# ---------------------------------------------------------------------------
class TestA3CRUDColaboradores(TestCase):
    """A3: crear/editar/inactivar sobre PersonalCuadrilla extendido."""

    def setUp(self):
        self.admin = _crear_admin()
        self.client = Client()
        self.client.force_login(self.admin)

    def test_crear_colaborador_persiste(self):
        url = reverse("cuadrillas:colaboradores_crear")
        resp = self.client.post(
            url,
            {
                "nombre": "Carlos Ruiz",
                "documento": "176-1001",
                "rol_cuadrilla": "CONDUCTOR",
                "salario_base": "1800000",
                "fecha_ingreso": "2026-01-10",
            },
        )
        self.assertIn(resp.status_code, (200, 302))
        self.assertTrue(PersonalCuadrilla.objects.filter(documento="176-1001").exists())
        creado = PersonalCuadrilla.objects.get(documento="176-1001")
        self.assertEqual(creado.nombre, "Carlos Ruiz")
        self.assertEqual(creado.salario_base, Decimal("1800000"))
        self.assertTrue(creado.activo)

    def test_editar_colaborador_existente_actualiza_campos(self):
        persona = PersonalCuadrilla.objects.create(
            nombre="Ana Torres",
            documento="176-1002",
            rol_cuadrilla_id="AYUDANTE",
            salario_base=Decimal("1750905"),
        )
        url = reverse("cuadrillas:colaboradores_editar", args=[persona.pk])
        resp = self.client.post(
            url,
            {
                "nombre": "Ana Torres Actualizada",
                "documento": "176-1002",
                "rol_cuadrilla": "LINIERO_I",
                "salario_base": "3176095",
                "fecha_ingreso": "2025-05-01",
            },
        )
        self.assertIn(resp.status_code, (200, 302))
        persona.refresh_from_db()
        self.assertEqual(persona.nombre, "Ana Torres Actualizada")
        self.assertEqual(persona.rol_cuadrilla_id, "LINIERO_I")
        self.assertEqual(persona.salario_base, Decimal("3176095"))

    def test_documento_duplicado_rechazado_con_mensaje_dominio(self):
        PersonalCuadrilla.objects.create(
            nombre="Pedro Existing",
            documento="176-1003",
            rol_cuadrilla_id="LINIERO_I",
        )
        url = reverse("cuadrillas:colaboradores_crear")
        resp = self.client.post(
            url,
            {
                "nombre": "Otro Pedro",
                "documento": "176-1003",
                "rol_cuadrilla": "AYUDANTE",
                "salario_base": "1000000",
            },
        )
        # No debe lanzar IntegrityError (500); debe re-renderizar el form con error.
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(PersonalCuadrilla.objects.filter(documento="176-1003").count(), 1)
        form = resp.context["form"]
        self.assertFalse(form.is_valid())
        self.assertIn("documento", form.errors)

    def test_inactivar_via_fecha_salida_reusa_signal_a2(self):
        persona = PersonalCuadrilla.objects.create(
            nombre="Luis Mora",
            documento="176-1004",
            rol_cuadrilla_id="LINIERO_I",
            activo=True,
        )
        url = reverse("cuadrillas:colaboradores_editar", args=[persona.pk])
        resp = self.client.post(
            url,
            {
                "nombre": persona.nombre,
                "documento": persona.documento,
                "rol_cuadrilla": persona.rol_cuadrilla_id,
                "salario_base": "0",
                "fecha_salida": date.today().isoformat(),
            },
        )
        self.assertIn(resp.status_code, (200, 302))
        persona.refresh_from_db()
        self.assertFalse(persona.activo)

    def test_colaborador_inactivo_no_aparece_en_lista_activa(self):
        PersonalCuadrilla.objects.create(
            nombre="Activo Uno",
            documento="176-1005",
            rol_cuadrilla_id="LINIERO_I",
            activo=True,
        )
        PersonalCuadrilla.objects.create(
            nombre="Inactivo Uno",
            documento="176-1006",
            rol_cuadrilla_id="LINIERO_I",
            fecha_salida=date.today(),
        )
        # El listado por defecto muestra todos (activos + inactivos) para que
        # el admin pueda reactivar; el filtro estado=activos es el que
        # excluye inactivos — es el usado por el picklist real (A4).
        url = reverse("cuadrillas:colaboradores_lista")
        resp = self.client.get(url, {"estado": "activos"})
        self.assertEqual(resp.status_code, 200)
        nombres = [p.nombre for p in resp.context["colaboradores"]]
        self.assertIn("Activo Uno", nombres)
        self.assertNotIn("Inactivo Uno", nombres)


# ---------------------------------------------------------------------------
# A5 — Importer de colaboradores (extiende PersonalCuadrillaUploadView)
# ---------------------------------------------------------------------------
class TestA5ImporterColaboradores(TestCase):
    """A5: PersonalCuadrillaUploadView soporta columnas nuevas."""

    def setUp(self):
        self.admin = _crear_admin()
        self.client = Client()
        self.client.force_login(self.admin)

    @staticmethod
    def _build_workbook(rows):
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Nombre", "Documento", "Cargo", "Salario Base", "Fecha Ingreso", "Fecha Salida"])
        for row in rows:
            ws.append(row)
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        buf.name = "colaboradores.xlsx"
        return buf

    def test_importa_fila_con_campos_nuevos(self):
        archivo = self._build_workbook(
            [
                ["Jose Herrera", "176-2001", "AYUDANTE", 1750905, "2025-02-01", ""],
            ]
        )
        url = reverse("cuadrillas:personal_upload")
        resp = self.client.post(url, {"archivo": archivo})
        self.assertIn(resp.status_code, (200, 302))
        persona = PersonalCuadrilla.objects.get(documento="176-2001")
        self.assertEqual(persona.nombre, "Jose Herrera")
        self.assertEqual(persona.salario_base, Decimal("1750905"))
        self.assertEqual(persona.fecha_ingreso, date(2025, 2, 1))
        self.assertIsNone(persona.fecha_salida)
        self.assertTrue(persona.activo)

    def test_fila_sin_fecha_salida_queda_activo(self):
        archivo = self._build_workbook(
            [
                ["Sin Salida", "176-2002", "LINIERO_I", 3176095, "2026-01-01", ""],
            ]
        )
        url = reverse("cuadrillas:personal_upload")
        self.client.post(url, {"archivo": archivo})
        persona = PersonalCuadrilla.objects.get(documento="176-2002")
        self.assertTrue(persona.activo)

    def test_fila_con_fecha_salida_queda_inactivo(self):
        archivo = self._build_workbook(
            [
                ["Con Salida", "176-2003", "LINIERO_I", 3176095, "2024-01-01", "2026-01-01"],
            ]
        )
        url = reverse("cuadrillas:personal_upload")
        self.client.post(url, {"archivo": archivo})
        persona = PersonalCuadrilla.objects.get(documento="176-2003")
        self.assertFalse(persona.activo)

    def test_documento_duplicado_en_archivo_se_reporta_sin_abortar(self):
        # Pre-existente en BD para forzar duplicado ya conocido + duplicado intra-archivo
        PersonalCuadrilla.objects.create(
            nombre="Ya Existe",
            documento="176-2004",
            rol_cuadrilla_id="LINIERO_I",
        )
        archivo = self._build_workbook(
            [
                ["Ya Existe Actualizado", "176-2004", "AYUDANTE", 1800000, "2025-06-01", ""],
                ["Fila Valida", "176-2005", "CONDUCTOR", 480000, "2025-06-01", ""],
            ]
        )
        url = reverse("cuadrillas:personal_upload")
        resp = self.client.post(url, {"archivo": archivo})
        self.assertIn(resp.status_code, (200, 302))
        # El "duplicado" en este importer es update_or_create (no error) — pero
        # la fila válida siguiente SIEMPRE debe procesarse (no abortar el resto).
        self.assertTrue(PersonalCuadrilla.objects.filter(documento="176-2005").exists())
        actualizado = PersonalCuadrilla.objects.get(documento="176-2004")
        self.assertEqual(actualizado.nombre, "Ya Existe Actualizado")


# ---------------------------------------------------------------------------
# Issue #237 — A2: encabezados canónicos de Nombre y Documento
# ---------------------------------------------------------------------------
class TestIssue237A2ImportadorPorEncabezado(TestCase):
    def setUp(self):
        self.admin = _crear_admin()
        self.client = Client()
        self.client.force_login(self.admin)

    @staticmethod
    def _workbook(headers, rows):
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(headers)
        for row in rows:
            ws.append(row)
        archivo = BytesIO()
        wb.save(archivo)
        archivo.seek(0)
        archivo.name = 'colaboradores_237.xlsx'
        return archivo

    def test_adjunto_exportado_documento_nombre_se_importa_por_encabezado(self):
        """El adjunto real invierte el orden legacy y debe seguir siendo importable."""
        PersonalCuadrilla.objects.create(
            nombre='Registro Legacy', documento='237-LEGACY-1', rol_cuadrilla_id='LINIERO_I'
        )

        exportado = self.client.get(reverse('cuadrillas:colaboradores_export'))
        self.assertEqual(exportado.status_code, 200)
        archivo = BytesIO(exportado.content)
        archivo.name = 'colaboradores.xlsx'
        respuesta = self.client.post(reverse('cuadrillas:personal_upload'), {'archivo': archivo})

        self.assertIn(respuesta.status_code, (200, 302))
        persona = PersonalCuadrilla.objects.get(documento='237-LEGACY-1')
        self.assertEqual(persona.nombre, 'Registro Legacy')
        self.assertEqual(PersonalCuadrilla.objects.filter(documento='237-LEGACY-1').count(), 1)

    def test_plantilla_legacy_nombre_documento_conserva_posiciones_historicas(self):
        archivo = self._workbook(
            ['Nombre', 'Documento', 'Cargo'],
            [['Formato Legacy', '237-LEGACY-2', 'AYUDANTE']],
        )

        self.client.post(reverse('cuadrillas:personal_upload'), {'archivo': archivo})

        persona = PersonalCuadrilla.objects.get(documento='237-LEGACY-2')
        self.assertEqual(persona.nombre, 'Formato Legacy')

    def test_columnas_reordenadas_usan_nombre_y_documento_canonicos(self):
        archivo = self._workbook(
            ['Cargo', 'Documento', 'Fecha Salida', 'Nombre', 'Salario Base'],
            [['AYUDANTE', '237-ORDEN-1', '2026-08-01', 'Columnas Reordenadas', 1500000]],
        )

        self.client.post(reverse('cuadrillas:personal_upload'), {'archivo': archivo})

        persona = PersonalCuadrilla.objects.get(documento='237-ORDEN-1')
        self.assertEqual(persona.nombre, 'Columnas Reordenadas')
        self.assertEqual(persona.salario_base, Decimal('1500000'))
        self.assertFalse(persona.activo)

    def test_encabezado_canonico_incompleto_no_caen_en_posiciones_legacy(self):
        archivo = self._workbook(
            ['Documento', 'Cargo'],
            [['237-INVALIDO-1', 'AYUDANTE']],
        )

        self.client.post(reverse('cuadrillas:personal_upload'), {'archivo': archivo})

        self.assertFalse(PersonalCuadrilla.objects.filter(documento='237-INVALIDO-1').exists())


# ---------------------------------------------------------------------------
# Issue #237 — A3: Cargo explícito en el importador de colaboradores
# ---------------------------------------------------------------------------
class TestIssue237A3CargoImportacion(TestCase):
    def setUp(self):
        self.admin = _crear_admin()
        self.client = Client()
        self.client.force_login(self.admin)
        self.cargo = Cargo.objects.create(
            codigo='A3-CODIGO', nombre='Cargo A3 Válido', activo=True,
        )
        self.inactivo = Cargo.objects.create(
            codigo='A3-INACTIVO', nombre='Cargo A3 Inactivo', activo=False,
        )

    @staticmethod
    def _workbook(rows):
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['Nombre', 'Documento', 'Cargo'])
        for row in rows:
            ws.append(row)
        archivo = BytesIO()
        wb.save(archivo)
        archivo.seek(0)
        archivo.name = 'cargos_237.xlsx'
        return archivo

    def _importar(self, rows):
        return self.client.post(
            reverse('cuadrillas:personal_upload'),
            {'archivo': self._workbook(rows)},
            follow=True,
        )

    def test_importa_cargo_activo_por_codigo(self):
        self._importar([['Por Código', '237-A3-01', self.cargo.codigo]])

        persona = PersonalCuadrilla.objects.get(documento='237-A3-01')
        self.assertEqual(persona.rol_cuadrilla_id, self.cargo.codigo)

    def test_importa_cargo_activo_por_nombre_normalizado(self):
        self._importar([['Por Nombre', '237-A3-02', ' cargo a3 valido ']])

        persona = PersonalCuadrilla.objects.get(documento='237-A3-02')
        self.assertEqual(persona.rol_cuadrilla_id, self.cargo.codigo)

    def test_cargo_desconocido_no_persiste_y_reporta_fila(self):
        respuesta = self._importar([['Cargo Inválido', '237-A3-03', 'NO EXISTE']])

        self.assertFalse(PersonalCuadrilla.objects.filter(documento='237-A3-03').exists())
        self.assertContains(respuesta, 'Fila 2: cargo inválido (NO EXISTE)')

    def test_cargo_inactivo_se_rechaza_y_no_impide_fila_valida_siguiente(self):
        respuesta = self._importar([
            ['Cargo Inactivo', '237-A3-04', self.inactivo.codigo],
            ['Cargo Válido', '237-A3-05', self.cargo.codigo],
        ])

        self.assertFalse(PersonalCuadrilla.objects.filter(documento='237-A3-04').exists())
        self.assertTrue(PersonalCuadrilla.objects.filter(documento='237-A3-05').exists())
        self.assertContains(respuesta, 'Fila 2: cargo inválido (A3-INACTIVO)')

    def test_modal_indica_rechazo_y_enlace_a_cargos(self):
        respuesta = self.client.get(reverse('cuadrillas:colaboradores_lista'))

        self.assertContains(respuesta, reverse('cuadrillas:cargos_lista'))
        self.assertContains(respuesta, 'la fila se rechaza')


# ---------------------------------------------------------------------------
# Issue #237 — A4: upsert por documento sobre colaboradores legacy
# ---------------------------------------------------------------------------
class TestIssue237A4UpsertPorDocumento(TestCase):
    """La reimportación reemplaza todos los datos del mismo documento."""

    def setUp(self):
        self.admin = _crear_admin()
        self.client = Client()
        self.client.force_login(self.admin)

    @staticmethod
    def _workbook(rows):
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append([
            'Nombre', 'Documento', 'Área', 'Cargo', 'Salario Base',
            'Fecha Ingreso', 'Fecha Salida',
        ])
        for row in rows:
            ws.append(row)
        archivo = BytesIO()
        wb.save(archivo)
        archivo.seek(0)
        archivo.name = 'upsert_colaboradores_237.xlsx'
        return archivo

    def _importar(self, rows):
        return self.client.post(
            reverse('cuadrillas:personal_upload'),
            {'archivo': self._workbook(rows)},
            follow=True,
        )

    def test_reimportar_documento_legacy_actualiza_todos_los_campos_sin_duplicar(self):
        legacy = PersonalCuadrilla.objects.create(
            nombre='Nombre Legacy', documento='237-A4-LEGACY',
            rol_cuadrilla_id='LINIERO_I', area='', salario_base=Decimal('1'),
            fecha_ingreso=date(2020, 1, 1), fecha_salida=date(2024, 1, 1),
        )
        self._importar([
            ['Segundo Documento', '237-A4-OTRO', 'Mantenimiento', 'AYUDANTE', 900000, '2025-01-01', ''],
            ['Nombre Primera Carga', legacy.documento, 'Construcción', 'AYUDANTE', 1200000, '2025-02-01', '2025-12-31'],
        ])
        self._importar([
            ['Nombre Actualizado', legacy.documento, 'Mantenimiento', 'CONDUCTOR', 1850000, '2026-03-01', ''],
        ])

        actualizado = PersonalCuadrilla.objects.get(documento=legacy.documento)
        self.assertEqual(PersonalCuadrilla.objects.filter(documento=legacy.documento).count(), 1)
        self.assertEqual(PersonalCuadrilla.objects.filter(documento='237-A4-OTRO').count(), 1)
        self.assertEqual(actualizado.nombre, 'Nombre Actualizado')
        self.assertEqual(actualizado.area, 'MANTENIMIENTO')
        self.assertEqual(actualizado.rol_cuadrilla_id, 'CONDUCTOR')
        self.assertEqual(actualizado.salario_base, Decimal('1850000'))
        self.assertEqual(actualizado.fecha_ingreso, date(2026, 3, 1))
        self.assertIsNone(actualizado.fecha_salida)
        self.assertTrue(actualizado.activo)

    def test_cargo_invalido_no_muta_el_registro_legacy_existente(self):
        legacy = PersonalCuadrilla.objects.create(
            nombre='No Modificar', documento='237-A4-INVALIDO',
            rol_cuadrilla_id='AYUDANTE', area='CONSTRUCCION',
            salario_base=Decimal('700000'), fecha_ingreso=date(2024, 1, 1),
        )

        respuesta = self._importar([
            ['Intento Inválido', legacy.documento, 'Mantenimiento', 'CARGO AUSENTE', 999999, '2026-01-01', ''],
        ])

        legacy.refresh_from_db()
        self.assertContains(respuesta, 'Fila 2: cargo inválido (CARGO AUSENTE)')
        self.assertEqual(legacy.nombre, 'No Modificar')
        self.assertEqual(legacy.area, 'CONSTRUCCION')
        self.assertEqual(legacy.salario_base, Decimal('700000'))
        self.assertEqual(legacy.fecha_ingreso, date(2024, 1, 1))

    def test_fecha_salida_reimportada_deja_inactivo_al_mismo_documento(self):
        PersonalCuadrilla.objects.create(
            nombre='Activo Antes', documento='237-A4-SALIDA',
            rol_cuadrilla_id='LINIERO_I', activo=True,
        )

        self._importar([
            ['Con Salida', '237-A4-SALIDA', 'Construcción', 'AYUDANTE', 800000, '2025-01-01', '2026-08-01'],
        ])

        actualizado = PersonalCuadrilla.objects.get(documento='237-A4-SALIDA')
        self.assertEqual(PersonalCuadrilla.objects.filter(documento='237-A4-SALIDA').count(), 1)
        self.assertEqual(actualizado.fecha_salida, date(2026, 8, 1))
        self.assertFalse(actualizado.activo)


# ---------------------------------------------------------------------------
# Issue #237 — A5: documento con separador de miles en lista y export
# ---------------------------------------------------------------------------
class TestIssue237A5DocumentoConMiles(TestCase):
    def setUp(self):
        self.admin = _crear_admin()
        self.client = Client()
        self.client.force_login(self.admin)
        Cargo.objects.get_or_create(
            codigo='LINIERO_I', defaults={'nombre': 'Liniero I', 'activo': True},
        )

    def test_lista_muestra_dos_documentos_numericos_con_miles(self):
        for documento in ('2370001', '2370002'):
            PersonalCuadrilla.objects.create(
                nombre=f'Colaborador {documento}', documento=documento,
                rol_cuadrilla_id='LINIERO_I',
            )

        respuesta = self.client.get(reverse('cuadrillas:colaboradores_lista'))

        self.assertContains(respuesta, '2.370.001')
        self.assertContains(respuesta, '2.370.002')
        self.assertEqual(
            list(PersonalCuadrilla.objects.filter(documento__in=('2370001', '2370002'))
                 .order_by('documento').values_list('documento', flat=True)),
            ['2370001', '2370002'],
        )

    def test_export_numero_documento_usa_formato_de_miles_y_es_reimportable(self):
        legacy = PersonalCuadrilla.objects.create(
            nombre='Registro Legacy', documento='2370001',
            rol_cuadrilla_id='LINIERO_I', salario_base=Decimal('15000'),
        )

        exportado = self.client.get(reverse('cuadrillas:colaboradores_export'))
        self.assertEqual(exportado.status_code, 200)

        import openpyxl

        wb = openpyxl.load_workbook(BytesIO(exportado.content))
        ws = wb.active
        fila = next(row for row in ws.iter_rows(min_row=2) if row[0].value == 2370001)
        self.assertEqual(fila[0].number_format, '#,##0')
        self.assertEqual(fila[0].value, 2370001)

        archivo = BytesIO(exportado.content)
        archivo.name = 'colaboradores_exportados.xlsx'
        respuesta = self.client.post(
            reverse('cuadrillas:personal_upload'), {'archivo': archivo}, follow=True,
        )

        self.assertEqual(respuesta.status_code, 200)
        self.assertEqual(PersonalCuadrilla.objects.filter(documento=legacy.documento).count(), 1)
        self.assertEqual(PersonalCuadrilla.objects.get(documento=legacy.documento).nombre, legacy.nombre)

    def test_export_conserva_como_texto_documentos_no_numericos_o_con_cero_inicial(self):
        PersonalCuadrilla.objects.create(
            nombre='Con Cero Inicial', documento='001234567', rol_cuadrilla_id='LINIERO_I',
        )
        PersonalCuadrilla.objects.create(
            nombre='Documento Legacy', documento='LEG-237-1', rol_cuadrilla_id='LINIERO_I',
        )

        exportado = self.client.get(reverse('cuadrillas:colaboradores_export'))
        import openpyxl

        ws = openpyxl.load_workbook(BytesIO(exportado.content)).active
        filas = {row[0].value: row[0] for row in ws.iter_rows(min_row=2)}
        self.assertEqual(filas['001234567'].data_type, 's')
        self.assertEqual(filas['LEG-237-1'].data_type, 's')


# ---------------------------------------------------------------------------
# A4 — Refactor asignación a cuadrilla
# ---------------------------------------------------------------------------
class TestA4RefactorAsignacionCuadrilla(TestCase):
    """A4: picklist activo=True, cargo bloqueado, AJAX, resolución de Usuario."""

    def setUp(self):
        self.admin = _crear_admin()
        self.client = Client()
        self.client.force_login(self.admin)
        self.cuadrilla = Cuadrilla.objects.create(
            codigo="01-2026-TST",
            nombre="Cuadrilla Test 176",
            activa=True,
        )
        self.personal_activo = PersonalCuadrilla.objects.create(
            nombre="Roberto Activo",
            documento="176-3001",
            rol_cuadrilla_id="LINIERO_I",
            salario_base=Decimal("3176095"),
            activo=True,
        )
        self.personal_inactivo = PersonalCuadrilla.objects.create(
            nombre="Sergio Inactivo",
            documento="176-3002",
            rol_cuadrilla_id="AYUDANTE",
            fecha_salida=date.today(),
        )

    def test_colaborador_inactivo_no_aparece_en_picklist(self):
        url = reverse("cuadrillas:detalle", args=[self.cuadrilla.pk])
        resp = self.client.get(url)
        self.assertEqual(resp.status_code, 200)
        disponibles = resp.context["personal_disponible"]
        ids = [str(p.id) for p in disponibles]
        self.assertIn(str(self.personal_activo.id), ids)
        self.assertNotIn(str(self.personal_inactivo.id), ids)

    def test_autocompletado_ajax_responde_nombre_y_cargo_por_documento(self):
        url = reverse("cuadrillas:personal_detalle_api")
        resp = self.client.get(url, {"documento": "176-3001"})
        self.assertEqual(resp.status_code, 200)
        data = resp.json()
        self.assertEqual(data["nombre"], "Roberto Activo")
        self.assertEqual(data["rol_cuadrilla"], "LINIERO_I")

    def test_cargo_bloqueado_post_con_rol_distinto_usa_el_del_maestro(self):
        url = reverse("cuadrillas:miembro_agregar", args=[self.cuadrilla.pk])
        resp = self.client.post(
            url,
            {
                "documento": "176-3001",
                "rol_cuadrilla": "SUPERVISOR",  # intento de forzar un rol distinto
                "cargo": "MIEMBRO",
            },
        )
        self.assertIn(resp.status_code, (200, 302))
        miembro = CuadrillaMiembro.objects.filter(
            cuadrilla=self.cuadrilla,
            usuario__documento="176-3001",
        ).first()
        self.assertIsNotNone(miembro)
        # Se ignora el rol del POST; se usa el del maestro PersonalCuadrilla.
        self.assertEqual(miembro.rol_cuadrilla_id, "LINIERO_I")

    def test_usuario_se_crea_o_reusa_al_asignar_primera_vez(self):
        self.assertFalse(Usuario.objects.filter(documento="176-3001").exists())
        url = reverse("cuadrillas:miembro_agregar", args=[self.cuadrilla.pk])
        self.client.post(
            url,
            {
                "documento": "176-3001",
                "rol_cuadrilla": "LINIERO_I",
                "cargo": "MIEMBRO",
            },
        )
        usuario = Usuario.objects.get(documento="176-3001")
        self.assertEqual(usuario.get_full_name(), "Roberto Activo")
        self.assertFalse(usuario.is_active)
        self.assertFalse(usuario.has_usable_password())

        miembro = CuadrillaMiembro.objects.get(cuadrilla=self.cuadrilla, usuario=usuario)
        self.assertEqual(miembro.usuario_id, usuario.id)

    def test_colaborador_inactivo_no_puede_ser_asignado(self):
        url = reverse("cuadrillas:miembro_agregar", args=[self.cuadrilla.pk])
        resp = self.client.post(
            url,
            {
                "documento": "176-3002",
                "rol_cuadrilla": "AYUDANTE",
                "cargo": "MIEMBRO",
            },
        )
        self.assertIn(resp.status_code, (200, 302))
        self.assertFalse(
            CuadrillaMiembro.objects.filter(
                cuadrilla=self.cuadrilla, usuario__documento="176-3002"
            ).exists()
        )

    def test_cuadrillamiembro_legacy_sigue_renderizando_sin_error(self):
        """Miembro legado (Usuario sin PersonalCuadrilla correspondiente)
        sigue apareciendo en la tabla de miembros sin romper la vista."""
        legacy_usuario = Usuario.objects.create_user(
            email="legacy176@test.com",
            password="testpass123!",
            first_name="Legacy",
            last_name="Miembro",
            rol="liniero",
        )
        CuadrillaMiembro.objects.create(
            cuadrilla=self.cuadrilla,
            usuario=legacy_usuario,
            rol_cuadrilla_id="LINIERO_I",
            fecha_inicio=date.today() - timedelta(days=30),
            activo=True,
        )
        url = reverse("cuadrillas:detalle", args=[self.cuadrilla.pk])
        resp = self.client.get(url)
        self.assertEqual(resp.status_code, 200)
        nombres = [m.usuario.get_full_name() for m in resp.context["miembros"]]
        self.assertIn("Legacy Miembro", nombres)


# ---------------------------------------------------------------------------
# Reproceso bounce=1 (FIX_INCOMPLETO) — navbar Parametrizacion sin los 2
# maestros nuevos, formato de moneda sin intcomma y boton "Subir Excel de
# Cargos" mezclado con el card "Agregar Miembro".
# ---------------------------------------------------------------------------
class TestReprocesoNavbarFormatoBoton(TestCase):
    """QA #176 bounce=1: reporto que los maestros de A1/A3 no eran
    descubribles por navegacion normal (solo por URL directa), que los
    montos monetarios de Colaboradores no tenian separador de miles, y que
    el boton de carga masiva de Excel quedo dentro del card de alta
    individual de miembro."""

    def setUp(self):
        self.admin = _crear_admin()
        self.client = Client()
        self.client.force_login(self.admin)

    def test_sidebar_incluye_links_a_tipos_actividad_y_colaboradores(self):
        url = reverse("cuadrillas:colaboradores_lista")
        resp = self.client.get(url)
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, reverse("actividades:tipos_lista"))
        self.assertContains(resp, reverse("cuadrillas:colaboradores_lista"))

    def test_salario_base_se_muestra_con_separador_de_miles(self):
        PersonalCuadrilla.objects.create(
            nombre="Colaborador Moneda",
            documento="176-4001",
            rol_cuadrilla_id="LINIERO_I",
            salario_base=Decimal("3176095"),
            activo=True,
        )
        url = reverse("cuadrillas:colaboradores_lista")
        resp = self.client.get(url)
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "3.176.095")

    def test_costo_dia_miembro_se_muestra_con_separador_de_miles(self):
        cuadrilla = Cuadrilla.objects.create(
            codigo="02-2026-TST",
            nombre="Cuadrilla Moneda 176",
            activa=True,
        )
        usuario = Usuario.objects.create_user(
            email="miembro176moneda@test.com",
            password="testpass123!",
            first_name="Roberto",
            last_name="Moneda",
            rol="liniero",
            documento="176-4002",
        )
        CuadrillaMiembro.objects.create(
            cuadrilla=cuadrilla,
            usuario=usuario,
            rol_cuadrilla_id="LINIERO_I",
            costo_dia=Decimal("3176095"),
            fecha_inicio=date.today(),
            activo=True,
        )
        url = reverse("cuadrillas:detalle", args=[cuadrilla.pk])
        resp = self.client.get(url)
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "3.176.095")

    def test_boton_subir_excel_cargos_queda_en_card_propio(self):
        """El boton + form colapsable de carga masiva debe quedar en su
        propio card, cerrado ANTES de que abra el card 'Agregar Miembro'
        (antes de este fix compartian el mismo div)."""
        cuadrilla = Cuadrilla.objects.create(
            codigo="03-2026-TST",
            nombre="Cuadrilla Reubicacion 176",
            activa=True,
        )
        url = reverse("cuadrillas:detalle", args=[cuadrilla.pk])
        resp = self.client.get(url)
        self.assertEqual(resp.status_code, 200)
        content = resp.content.decode()
        idx_subir = content.index("Subir Excel de Cargos")
        idx_agregar = content.index("Agregar Miembro")
        idx_cierre_form_upload = content.index("</form>", idx_subir)
        self.assertLess(idx_subir, idx_agregar)
        # El form de upload (dentro de su propio card) debe cerrarse antes
        # de que empiece el texto "Agregar Miembro" del card siguiente.
        self.assertLess(idx_cierre_form_upload, idx_agregar)


# ---------------------------------------------------------------------------
# Maestro 3: Cargos (issue #176, bounce 2) — A1: modelo Cargo + seed
# ---------------------------------------------------------------------------
# Los 14 códigos esperados: union de PersonalCuadrilla.RolCuadrilla y
# CuadrillaMiembro.RolCuadrilla (idénticos, confirmado en F2). Se corre la
# función de la migración 0019_seed_cargos directamente (importlib, el
# nombre del módulo empieza con dígitos y no es importable con `import`
# normal) contra el registro de apps real — el schema de Cargo no ha
# cambiado de forma desde 0018, así que apps.get_model('cuadrillas','Cargo')
# resuelve igual con el apps registry real que con el histórico.
_SEED_MODULE = importlib.import_module("apps.cuadrillas.migrations.0019_seed_cargos")


class TestMaestro3A1CargoModeloYSeed(TestCase):
    """A1: modelo Cargo (codigo/nombre/activo) + data migration de seed."""

    def test_modelo_cargo_tiene_los_campos_esperados(self):
        cargo = Cargo.objects.create(codigo="SOLDADOR", nombre="Soldador")
        self.assertTrue(cargo.activo)
        self.assertEqual(str(cargo), "SOLDADOR - Soldador")

    def test_codigo_es_unico(self):
        from django.db import transaction
        from django.db.utils import IntegrityError

        # Código fuera del catálogo de 14 (el conftest autouse ya los
        # siembra a todos) para no colisionar con la siembra global.
        Cargo.objects.create(codigo="TEST_CODIGO_UNICO", nombre="Test Único")
        with self.assertRaises(IntegrityError), transaction.atomic():
            Cargo.objects.create(codigo="TEST_CODIGO_UNICO", nombre="Otro Nombre")

    def test_seed_crea_los_14_codigos_de_la_union_de_ambos_enums(self):
        """Los 14 códigos vienen de la unión de PersonalCuadrilla.RolCuadrilla
        / CuadrillaMiembro.RolCuadrilla (idénticos, confirmado en F2) — ese
        TextChoices se ELIMINÓ en A3 (la FK a Cargo lo reemplaza), así que
        acá se verifica contra la copia congelada de esos 14 códigos (ya
        no hay un enum vivo en el código contra el cual comparar)."""
        _SEED_MODULE.seed_cargos(django_apps, None)
        self.assertEqual(Cargo.objects.count(), 14)
        codigos_esperados = {
            "SUPERVISOR",
            "LINIERO_I",
            "LINIERO_II",
            "AYUDANTE",
            "CONDUCTOR",
            "ADMINISTRADOR_OBRA",
            "PROFESIONAL_SST",
            "ING_RESIDENTE",
            "SERVICIO_GENERAL",
            "ALMACENISTA",
            "SUPERVISOR_FOREST",
            "ASISTENTE_FOREST",
            "MALACATERO",
            "COORDINADOR_HSQ",
        }
        codigos_sembrados = set(Cargo.objects.values_list("codigo", flat=True))
        self.assertEqual(codigos_sembrados, codigos_esperados)

    def test_seed_labels_coinciden_con_cargos_data_de_la_migracion(self):
        _SEED_MODULE.seed_cargos(django_apps, None)
        labels_esperados = dict(_SEED_MODULE.CARGOS)
        for cargo in Cargo.objects.all():
            self.assertEqual(cargo.nombre, labels_esperados[cargo.codigo])

    def test_seed_es_idempotente_correrlo_dos_veces_no_duplica(self):
        _SEED_MODULE.seed_cargos(django_apps, None)
        _SEED_MODULE.seed_cargos(django_apps, None)
        self.assertEqual(Cargo.objects.count(), 14)

    def test_seed_no_pisa_activo_false_si_ya_fue_inactivado_por_el_usuario(self):
        """get_or_create solo aplica defaults en creación — si el
        coordinador ya inactivó un cargo sembrado, re-correr la migración
        (idempotencia) no debe reactivarlo."""
        _SEED_MODULE.seed_cargos(django_apps, None)
        cargo = Cargo.objects.get(codigo="ASISTENTE_FOREST")
        cargo.activo = False
        cargo.save(update_fields=["activo"])
        _SEED_MODULE.seed_cargos(django_apps, None)
        cargo.refresh_from_db()
        self.assertFalse(cargo.activo)
        self.assertEqual(Cargo.objects.count(), 14)


# ---------------------------------------------------------------------------
# Maestro 3: Cargos (issue #176, bounce 2) — A2: CRUD Cargo
# ---------------------------------------------------------------------------
class TestMaestro3A2CRUDCargo(TestCase):
    """A2: crear/editar/inactivar sobre Cargo + codigo read-only en edicion."""

    def setUp(self):
        self.admin = _crear_admin()
        self.client = Client()
        self.client.force_login(self.admin)

    def test_crear_cargo_persiste(self):
        url = reverse("cuadrillas:cargos_crear")
        resp = self.client.post(url, {"codigo": "SOLDADOR", "nombre": "Soldador", "activo": "on"})
        self.assertIn(resp.status_code, (200, 302))
        self.assertTrue(Cargo.objects.filter(codigo="SOLDADOR").exists())
        creado = Cargo.objects.get(codigo="SOLDADOR")
        self.assertEqual(creado.nombre, "Soldador")
        self.assertTrue(creado.activo)

    def test_codigo_duplicado_rechazado_con_mensaje_dominio(self):
        Cargo.objects.create(codigo="TECNICO", nombre="Técnico")
        url = reverse("cuadrillas:cargos_crear")
        resp = self.client.post(
            url, {"codigo": "TECNICO", "nombre": "Otro Técnico", "activo": "on"}
        )
        # No debe lanzar IntegrityError (500); debe re-renderizar el form con error.
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(Cargo.objects.filter(codigo="TECNICO").count(), 1)
        form = resp.context["form"]
        self.assertFalse(form.is_valid())
        self.assertIn("codigo", form.errors)

    def test_editar_cargo_actualiza_nombre_pero_codigo_es_readonly(self):
        cargo = Cargo.objects.create(codigo="AYUDANTE_ESP", nombre="Ayudante Especial")
        url = reverse("cuadrillas:cargos_editar", args=[cargo.pk])
        # GET: el widget de codigo debe estar deshabilitado en el form renderizado.
        resp_get = self.client.get(url)
        self.assertEqual(resp_get.status_code, 200)
        self.assertTrue(resp_get.context["form"].fields["codigo"].disabled)
        # POST intentando cambiar el codigo: al estar disabled, Django ignora
        # el valor del POST y conserva el original.
        resp = self.client.post(
            url,
            {"codigo": "OTRO_CODIGO", "nombre": "Ayudante Especial Actualizado", "activo": "on"},
        )
        self.assertIn(resp.status_code, (200, 302))
        cargo.refresh_from_db()
        self.assertEqual(cargo.codigo, "AYUDANTE_ESP")
        self.assertEqual(cargo.nombre, "Ayudante Especial Actualizado")

    def test_inactivar_cargo_no_borra_el_registro(self):
        cargo = Cargo.objects.create(codigo="AUXILIAR_TEMP", nombre="Auxiliar Temporal")
        url = reverse("cuadrillas:cargos_inactivar", args=[cargo.pk])
        resp = self.client.post(url)
        self.assertIn(resp.status_code, (200, 302))
        cargo.refresh_from_db()
        self.assertFalse(cargo.activo)
        self.assertTrue(Cargo.objects.filter(pk=cargo.pk).exists())

    def test_reactivar_cargo_previamente_inactivo(self):
        cargo = Cargo.objects.create(
            codigo="AUXILIAR_TEMP2", nombre="Auxiliar Temporal 2", activo=False
        )
        url = reverse("cuadrillas:cargos_inactivar", args=[cargo.pk])
        resp = self.client.post(url)
        self.assertIn(resp.status_code, (200, 302))
        cargo.refresh_from_db()
        self.assertTrue(cargo.activo)

    def test_cargo_inactivo_no_aparece_en_filtro_activos(self):
        Cargo.objects.create(codigo="ACTIVO_X", nombre="Cargo Activo X", activo=True)
        Cargo.objects.create(codigo="INACTIVO_X", nombre="Cargo Inactivo X", activo=False)
        url = reverse("cuadrillas:cargos_lista")
        resp = self.client.get(url, {"estado": "activos"})
        self.assertEqual(resp.status_code, 200)
        codigos = {c.codigo for c in resp.context["cargos"]}
        self.assertIn("ACTIVO_X", codigos)
        self.assertNotIn("INACTIVO_X", codigos)


# ---------------------------------------------------------------------------
# Maestro 3: Cargos (issue #176, bounce 2) — A3: FK rol_cuadrilla -> Cargo
# ---------------------------------------------------------------------------
class TestMaestro3A3FKRolCuadrilla(TestCase):
    """A3: PersonalCuadrilla/CuadrillaMiembro.rol_cuadrilla convertido a
    FK(Cargo, to_field='codigo'). Cubre el contrato documentado en el plan:
    `instance.rol_cuadrilla_id` sigue siendo el string del código,
    `instance.rol_cuadrilla` es el objeto Cargo, get_rol_cuadrilla_display()
    es un shim manual, y filtros __in/values/order_by contra strings siguen
    funcionando sin joins explícitos (riesgo #3 del plan, línea
    financiero/views.py:1692-1694)."""

    def setUp(self):
        # Sembrar el catalogo completo (14 codigos) — la FK real exige que
        # el codigo referenciado exista en `cargos` antes de poder crear
        # PersonalCuadrilla/CuadrillaMiembro con ese rol_cuadrilla_id.
        _SEED_MODULE.seed_cargos(django_apps, None)

    def test_rol_cuadrilla_id_sigue_siendo_el_string_del_codigo(self):
        persona = PersonalCuadrilla.objects.create(
            nombre="FK Test",
            documento="176-fk-001",
            rol_cuadrilla_id="SUPERVISOR",
        )
        persona.refresh_from_db()
        self.assertEqual(persona.rol_cuadrilla_id, "SUPERVISOR")
        self.assertIsInstance(persona.rol_cuadrilla_id, str)

    def test_rol_cuadrilla_sin_id_es_el_objeto_cargo_completo(self):
        persona = PersonalCuadrilla.objects.create(
            nombre="FK Test 2",
            documento="176-fk-002",
            rol_cuadrilla_id="LINIERO_I",
        )
        persona.refresh_from_db()
        self.assertIsInstance(persona.rol_cuadrilla, Cargo)
        self.assertEqual(persona.rol_cuadrilla.codigo, "LINIERO_I")

    def test_get_rol_cuadrilla_display_shim_personal_cuadrilla(self):
        persona = PersonalCuadrilla.objects.create(
            nombre="FK Test 3",
            documento="176-fk-003",
            rol_cuadrilla_id="AYUDANTE",
        )
        self.assertEqual(persona.get_rol_cuadrilla_display(), "Ayudante")

    def test_get_rol_cuadrilla_display_shim_cuadrilla_miembro(self):
        cuadrilla = Cuadrilla.objects.create(codigo="FK-TEST-001", nombre="Cuadrilla FK Test")
        usuario = Usuario.objects.create_user(
            email="fkmiembro176@test.com",
            password="testpass123!",
            first_name="FK",
            last_name="Miembro",
            rol="liniero",
            documento="176-fk-004",
        )
        miembro = CuadrillaMiembro.objects.create(
            cuadrilla=cuadrilla,
            usuario=usuario,
            rol_cuadrilla_id="CONDUCTOR",
            fecha_inicio=date.today(),
        )
        self.assertEqual(miembro.get_rol_cuadrilla_display(), "Conductor")

    def test_default_sigue_siendo_liniero_i(self):
        persona = PersonalCuadrilla.objects.create(nombre="FK Default", documento="176-fk-005")
        self.assertEqual(persona.rol_cuadrilla_id, "LINIERO_I")

    def test_filter_in_por_string_sigue_funcionando_sin_joins_explicitos(self):
        """Riesgo #3 del plan: financiero/views.py usa
        rol_cuadrilla__in=[...] / order_by('rol_cuadrilla') / values(...,
        'rol_cuadrilla') contra strings — Django debe resolver esto
        directo contra el string con to_field='codigo', sin requerir
        .rol_cuadrilla__codigo__in=[...]."""
        PersonalCuadrilla.objects.create(
            nombre="A", documento="176-fk-006", rol_cuadrilla_id="SUPERVISOR"
        )
        PersonalCuadrilla.objects.create(
            nombre="B", documento="176-fk-007", rol_cuadrilla_id="LINIERO_I"
        )
        PersonalCuadrilla.objects.create(
            nombre="C", documento="176-fk-008", rol_cuadrilla_id="AYUDANTE"
        )
        qs = PersonalCuadrilla.objects.filter(rol_cuadrilla__in=["SUPERVISOR", "LINIERO_I"])
        self.assertEqual(qs.count(), 2)
        valores = list(
            PersonalCuadrilla.objects.filter(documento="176-fk-006").values("rol_cuadrilla")
        )
        self.assertEqual(valores[0]["rol_cuadrilla"], "SUPERVISOR")

    def test_on_delete_protect_bloquea_borrar_cargo_referenciado(self):
        from django.db import transaction
        from django.db.models import ProtectedError

        persona = PersonalCuadrilla.objects.create(
            nombre="FK Protect", documento="176-fk-009", rol_cuadrilla_id="ALMACENISTA"
        )
        cargo_referenciado = Cargo.objects.get(codigo="ALMACENISTA")
        with self.assertRaises(ProtectedError), transaction.atomic():
            cargo_referenciado.delete()
        # Confirma que sigue existiendo (PROTECT bloqueo el delete).
        self.assertTrue(PersonalCuadrilla.objects.filter(pk=persona.pk).exists())
        self.assertTrue(Cargo.objects.filter(codigo="ALMACENISTA").exists())

    def test_codigo_jerarquico_cargo_no_se_confunde_con_rol_cuadrilla(self):
        """CuadrillaMiembro.cargo (CargoJerarquico: JT_CTA/MIEMBRO) es un
        campo DISTINTO de rol_cuadrilla — no debe verse afectado por A3."""
        cuadrilla = Cuadrilla.objects.create(codigo="FK-TEST-002", nombre="Cuadrilla FK Test 2")
        usuario = Usuario.objects.create_user(
            email="fkjerarquico176@test.com",
            password="testpass123!",
            first_name="FK",
            last_name="Jerarquico",
            rol="liniero",
            documento="176-fk-010",
        )
        miembro = CuadrillaMiembro.objects.create(
            cuadrilla=cuadrilla,
            usuario=usuario,
            rol_cuadrilla_id="SUPERVISOR",
            cargo=CuadrillaMiembro.CargoJerarquico.JT_CTA,
            fecha_inicio=date.today(),
        )
        self.assertEqual(miembro.cargo, "JT_CTA")
        self.assertEqual(miembro.rol_cuadrilla_id, "SUPERVISOR")


# ---------------------------------------------------------------------------
# Maestro 3: Cargos (issue #176, bounce 2) — A4: retrofit de call sites
# ---------------------------------------------------------------------------
class TestMaestro3A4RetrofitCallSites(TestCase):
    """A4: cubre el Retrofit Ledger del plan -- JSON APIs no serializan el
    objeto Cargo (TypeError), carga masiva Excel (2 vistas) resuelve roles
    via Cargo dinamico, autocompletado AJAX, nomina/detalle template badges
    (comparacion == 'STRING' contra .codigo), y seed_data sin IntegrityError."""

    def setUp(self):
        _SEED_MODULE.seed_cargos(django_apps, None)
        self.admin = _crear_admin()
        self.client = Client()
        self.client.force_login(self.admin)

    def test_personal_detalle_api_serializa_rol_como_string(self):
        PersonalCuadrilla.objects.create(
            nombre="API Test", documento="176-a4-001", rol_cuadrilla_id="SUPERVISOR"
        )
        url = reverse("cuadrillas:personal_detalle_api")
        resp = self.client.get(url, {"documento": "176-a4-001"})
        self.assertEqual(resp.status_code, 200)
        data = resp.json()
        self.assertEqual(data["rol_cuadrilla"], "SUPERVISOR")
        self.assertIsInstance(data["rol_cuadrilla"], str)

    def test_personal_list_api_serializa_rol_como_string(self):
        PersonalCuadrilla.objects.create(
            nombre="API List Test", documento="176-a4-002", rol_cuadrilla_id="AYUDANTE"
        )
        url = reverse("cuadrillas:personal_list_api")
        resp = self.client.get(url)
        self.assertEqual(resp.status_code, 200)
        data = resp.json()
        encontrado = next(p for p in data if p["documento"] == "176-a4-002")
        self.assertEqual(encontrado["rol_cuadrilla"], "AYUDANTE")

    def test_ninja_api_cuadrilla_detalle_no_revienta_serializando_cargo(self):
        cuadrilla = Cuadrilla.objects.create(codigo="A4-API-001", nombre="Cuadrilla API")
        usuario = Usuario.objects.create_user(
            email="a4api176@test.com",
            password="testpass123!",
            first_name="A4",
            last_name="Api",
            rol="liniero",
            documento="176-a4-003",
        )
        CuadrillaMiembro.objects.create(
            cuadrilla=cuadrilla,
            usuario=usuario,
            rol_cuadrilla_id="LINIERO_I",
            fecha_inicio=date.today(),
        )
        from rest_framework_simplejwt.tokens import RefreshToken

        access_token = str(RefreshToken.for_user(self.admin).access_token)
        resp = self.client.get(
            f"/api/cuadrillas/cuadrillas/{cuadrilla.pk}",
            HTTP_AUTHORIZATION=f"Bearer {access_token}",
        )
        self.assertEqual(resp.status_code, 200)
        data = resp.json()
        self.assertEqual(data["miembros"][0]["rol_cuadrilla"], "LINIERO_I")

    def test_bulk_upload_personal_excel_resuelve_rol_desde_catalogo_dinamico(self):
        """PersonalCuadrillaUploadView (masivo, /cuadrillas/personal/subir/)."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Nombre", "Documento", "Cargo"])
        ws.append(["Excel Bulk Persona", "176-a4-004", "SUPERVISOR"])
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        buf.name = "personal.xlsx"

        url = reverse("cuadrillas:personal_upload")
        resp = self.client.post(url, {"archivo": buf})
        self.assertIn(resp.status_code, (200, 302))
        creado = PersonalCuadrilla.objects.get(documento="176-a4-004")
        self.assertEqual(creado.rol_cuadrilla_id, "SUPERVISOR")

    def test_bulk_upload_miembros_cuadrilla_excel_resuelve_rol_desde_catalogo(self):
        """CuadrillaMiembroUploadView (/cuadrillas/<uuid>/miembros/subir/)."""
        import openpyxl

        cuadrilla = Cuadrilla.objects.create(codigo="A4-BULK-001", nombre="Cuadrilla Bulk")
        usuario = Usuario.objects.create_user(
            email="a4bulk176@test.com",
            password="testpass123!",
            first_name="Bulk",
            last_name="Miembro",
            rol="liniero",
            documento="176-a4-005",
        )
        miembro = CuadrillaMiembro.objects.create(
            cuadrilla=cuadrilla,
            usuario=usuario,
            rol_cuadrilla_id="LINIERO_I",
            fecha_inicio=date.today(),
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Nombre", "Cargo", "Documento"])
        ws.append(["Bulk Miembro", "SUPERVISOR", "176-a4-005"])
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        buf.name = "miembros.xlsx"

        url = reverse("cuadrillas:miembros_upload", args=[cuadrilla.pk])
        resp = self.client.post(url, {"archivo": buf})
        self.assertIn(resp.status_code, (200, 302))
        miembro.refresh_from_db()
        self.assertEqual(miembro.rol_cuadrilla_id, "SUPERVISOR")

    def test_detalle_cuadrilla_badge_supervisor_correcto_no_cae_siempre_gris(self):
        """Riesgo #2 del plan: comparacion == 'STRING' contra objeto Cargo
        siempre caia al else (badge gris). Verifica que con .codigo el
        badge morado de Supervisor SI aparece."""
        cuadrilla = Cuadrilla.objects.create(codigo="A4-BADGE-001", nombre="Cuadrilla Badge")
        usuario = Usuario.objects.create_user(
            email="a4badge176@test.com",
            password="testpass123!",
            first_name="Badge",
            last_name="Supervisor",
            rol="liniero",
            documento="176-a4-006",
        )
        CuadrillaMiembro.objects.create(
            cuadrilla=cuadrilla,
            usuario=usuario,
            rol_cuadrilla_id="SUPERVISOR",
            fecha_inicio=date.today(),
        )
        url = reverse("cuadrillas:detalle", args=[cuadrilla.pk])
        resp = self.client.get(url)
        self.assertEqual(resp.status_code, 200)
        content = resp.content.decode()
        # El badge morado de Supervisor debe aparecer (no siempre gris).
        self.assertIn("bg-purple-100 text-purple-800", content)
        self.assertIn("Supervisor", content)

    def test_roles_cuadrilla_context_es_iterable_value_label(self):
        """context['roles_cuadrilla'] pasa de RolCuadrilla.choices a una
        lista de tuplas (codigo, nombre) desde Cargo -- el template hace
        `{% for value, label in roles_cuadrilla %}` (JS ROL_CUADRILLA_LABELS)."""
        cuadrilla = Cuadrilla.objects.create(codigo="A4-CTX-001", nombre="Cuadrilla Ctx")
        url = reverse("cuadrillas:detalle", args=[cuadrilla.pk])
        resp = self.client.get(url)
        self.assertEqual(resp.status_code, 200)
        roles = resp.context["roles_cuadrilla"]
        self.assertGreater(len(roles), 0)
        primer_codigo, primer_label = roles[0]
        self.assertIsInstance(primer_codigo, str)
        self.assertIsInstance(primer_label, str)

    def test_seed_data_create_cuadrillas_sin_integrity_error(self):
        """Issue #176 (A4): seed_data.py tenia strings invalidos
        preexistentes ('supervisor'/'liniero' minuscula) que CharField+
        choices toleraba silenciosamente -- con la FK real, un codigo
        invalido revienta con IntegrityError. Se corrigio a los codigos
        reales (SUPERVISOR/LINIERO_I); este test corre exactamente los
        3 metodos del comando relacionados (_create_users/_vehiculos/
        _cuadrillas) y confirma que no truena."""
        from apps.core.management.commands.seed_data import Command

        cmd = Command()
        cmd.stdout = type("_N", (), {"write": lambda self, *a, **k: None})()
        cmd.style = type("_S", (), {"SUCCESS": lambda self, x: x})()
        cmd._create_users()
        cmd._create_vehiculos()
        cmd._create_cuadrillas()  # no debe lanzar IntegrityError

        miembros_supervisor = CuadrillaMiembro.objects.filter(rol_cuadrilla_id="SUPERVISOR")
        miembros_liniero = CuadrillaMiembro.objects.filter(rol_cuadrilla_id="LINIERO_I")
        self.assertTrue(miembros_supervisor.exists() or miembros_liniero.exists())


# ---------------------------------------------------------------------------
# Maestro 3: Cargos (issue #176, bounce 2) — A5: retrofit importer S18
# ---------------------------------------------------------------------------
class TestMaestro3A5ImporterS18CargoDinamico(TestCase):
    """A5: ProgramacionS18CuadrillaImporter reconoce un cargo NUEVO del
    catálogo Cargo (no presente en ROL_TEXTO_A_CHOICE) vía el fallback de
    matching dinámico, en vez de caer silenciosamente al default
    LINIERO_I. Ver apps/cuadrillas/importers.py::_resolver_rol_desde_cargo_raw."""

    def setUp(self):
        _SEED_MODULE.seed_cargos(django_apps, None)

    def test_cargo_nuevo_del_catalogo_se_reconoce_no_cae_a_fallback(self):
        """Cargo 'TECNICO_EMPALMES' dado de alta en el catálogo (vía CRUD
        /cuadrillas/cargos/, mismo efecto que un INSERT directo aquí) debe
        reconocerse en la carga S18 con solo escribir 'Tecnico Empalmes'
        en el Excel -- sin tocar ROL_TEXTO_A_CHOICE."""
        Cargo.objects.create(codigo="TECNICO_EMPALMES", nombre="Técnico Empalmes", activo=True)
        _crear_linea("LN817")
        _crear_usuario("1143246675", "JHON JAIRO")
        _crear_usuario("11122299", "NUEVO EMPALMADOR")

        excel = _build_s18_excel(
            [
                _act(
                    1,
                    "Servidumbre",
                    "817",
                    date(2026, 4, 27),
                    date(2026, 5, 3),
                    "JHON JAIRO",
                    "1143246675",
                    "LINIERO I",
                    "JT/CTA",
                ),
                _miembro("NUEVO EMPALMADOR", "11122299", "Tecnico Empalmes"),
            ]
        )
        res = ProgramacionS18CuadrillaImporter().importar(excel)
        self.assertTrue(res["exito"], res.get("error"))
        miembro = CuadrillaMiembro.objects.get(usuario__documento="11122299")
        self.assertEqual(miembro.rol_cuadrilla_id, "TECNICO_EMPALMES")
        # Anti falso-verde: confirma que NO cayó en el fallback LINIERO_I.
        self.assertNotEqual(miembro.rol_cuadrilla_id, "LINIERO_I")

    def test_cargo_verdaderamente_desconocido_sigue_con_fallback_y_mensaje_nuevo(self):
        """Un CARGO que no está ni en ROL_TEXTO_A_CHOICE ni en el catálogo
        Cargo sigue cayendo a LINIERO_I (comportamiento histórico #124),
        pero el mensaje de advertencia ahora apunta a la UI del maestro
        (/cuadrillas/cargos/), no al dict hardcoded."""
        _crear_linea("LN817")
        _crear_usuario("1143246675", "JHON JAIRO")
        _crear_usuario("11122300", "CARGO INVENTADO")

        excel = _build_s18_excel(
            [
                _act(
                    1,
                    "Servidumbre",
                    "817",
                    date(2026, 4, 27),
                    date(2026, 5, 3),
                    "JHON JAIRO",
                    "1143246675",
                    "LINIERO I",
                    "JT/CTA",
                ),
                _miembro("CARGO INVENTADO", "11122300", "Puesto Que No Existe"),
            ]
        )
        res = ProgramacionS18CuadrillaImporter().importar(excel)
        self.assertTrue(res["exito"], res.get("error"))
        miembro = CuadrillaMiembro.objects.get(usuario__documento="11122300")
        self.assertEqual(miembro.rol_cuadrilla_id, "LINIERO_I")
        self.assertTrue(
            any("/cuadrillas/cargos/" in a for a in res["advertencias"]),
            res["advertencias"],
        )

    def test_cargo_nuevo_normaliza_espacios_y_mayusculas(self):
        """'Tecnico Empalmes' (texto libre con espacios/minusculas) debe
        normalizar a 'TECNICO_EMPALMES' para matchear el catálogo -- no
        hace falta que el Excel traiga el código exacto."""
        Cargo.objects.create(codigo="SOLDADOR_ESPECIAL", nombre="Soldador Especial", activo=True)
        _crear_linea("LN818")
        _crear_usuario("1143246676", "PEDRO PABLO")
        _crear_usuario("11122301", "NUEVO SOLDADOR")

        excel = _build_s18_excel(
            [
                _act(
                    1,
                    "Mtto",
                    "818",
                    date(2026, 4, 27),
                    date(2026, 5, 3),
                    "PEDRO PABLO",
                    "1143246676",
                    "LINIERO I",
                    "JT/CTA",
                ),
                _miembro("NUEVO SOLDADOR", "11122301", "soldador especial"),
            ]
        )
        res = ProgramacionS18CuadrillaImporter().importar(excel)
        self.assertTrue(res["exito"], res.get("error"))
        miembro = CuadrillaMiembro.objects.get(usuario__documento="11122301")
        self.assertEqual(miembro.rol_cuadrilla_id, "SOLDADOR_ESPECIAL")


class TestAreaAsignacionIssue176(TestCase):
    """El área del usuario limita asignación, API e importación sin romper legacy."""

    def setUp(self):
        self.usuario = _crear_admin()
        self.usuario.area = "MANTENIMIENTO"
        self.usuario.save(update_fields=["area"])
        self.client = Client()
        self.client.force_login(self.usuario)
        self.cuadrilla = Cuadrilla.objects.create(
            codigo="02-2026-AREA176", nombre="Cuadrilla área 176", activa=True
        )

    def test_picklist_y_api_excluyen_otra_area(self):
        visible = PersonalCuadrilla.objects.create(
            nombre="Visible Mantenimiento", documento="176-AREA-1",
            rol_cuadrilla_id="LINIERO_I", area="MANTENIMIENTO",
        )
        oculta = PersonalCuadrilla.objects.create(
            nombre="Oculta Construcción", documento="176-AREA-2",
            rol_cuadrilla_id="LINIERO_I", area="CONSTRUCCION",
        )
        resp = self.client.get(reverse("cuadrillas:detalle", args=[self.cuadrilla.pk]))
        ids = {str(persona.id) for persona in resp.context["personal_disponible"]}
        self.assertIn(str(visible.id), ids)
        self.assertNotIn(str(oculta.id), ids)
        api = self.client.get(reverse("cuadrillas:personal_detalle_api"), {"documento": oculta.documento})
        self.assertEqual(api.status_code, 404)

    def test_importa_area_y_conserva_formato_legacy(self):
        archivo = self._build_workbook_area([
            ["Importado Área", "176-AREA-3", "CONSTRUCCION", "AYUDANTE", 0, "", ""],
        ])
        self.client.post(reverse("cuadrillas:personal_upload"), {"archivo": archivo})
        persona = PersonalCuadrilla.objects.get(documento="176-AREA-3")
        self.assertEqual(persona.area, "CONSTRUCCION")

        legacy = self._build_workbook_area([
            ["Importado Legacy", "176-AREA-4", "AYUDANTE", 0, "", ""],
        ], headers=["Nombre", "Documento", "Cargo", "Salario Base", "Fecha Ingreso", "Fecha Salida"])
        self.client.post(reverse("cuadrillas:personal_upload"), {"archivo": legacy})
        self.assertEqual(PersonalCuadrilla.objects.get(documento="176-AREA-4").area, "")

    def test_importa_area_de_construccion_con_valor_acento(self):
        archivo = self._build_workbook_area(
            [["Construcción", "237-AREA-1", "Construcción", "AYUDANTE", 0, "", ""]],
            headers=["Nombre", "Documento", "Área de Construcción", "Cargo", "Salario Base", "Fecha Ingreso", "Fecha Salida"],
        )

        self.client.post(reverse("cuadrillas:personal_upload"), {"archivo": archivo})

        self.assertEqual(PersonalCuadrilla.objects.get(documento="237-AREA-1").area, "CONSTRUCCION")

    def test_importa_encabezado_area_en_mayusculas(self):
        archivo = self._build_workbook_area(
            [["Mayúscula", "237-AREA-2", "Construcción", "AYUDANTE", 0, "", ""]],
            headers=["Nombre", "Documento", "ÁREA", "Cargo", "Salario Base", "Fecha Ingreso", "Fecha Salida"],
        )

        self.client.post(reverse("cuadrillas:personal_upload"), {"archivo": archivo})

        self.assertEqual(PersonalCuadrilla.objects.get(documento="237-AREA-2").area, "CONSTRUCCION")

    def test_importa_encabezado_area_en_minusculas_sin_tilde(self):
        archivo = self._build_workbook_area(
            [["Minúscula", "237-AREA-3", "Construccion", "AYUDANTE", 0, "", ""]],
            headers=["Nombre", "Documento", "area", "Cargo", "Salario Base", "Fecha Ingreso", "Fecha Salida"],
        )

        self.client.post(reverse("cuadrillas:personal_upload"), {"archivo": archivo})

        self.assertEqual(PersonalCuadrilla.objects.get(documento="237-AREA-3").area, "CONSTRUCCION")

    @staticmethod
    def _build_workbook_area(rows, headers=None):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(headers or ["Nombre", "Documento", "Área", "Cargo", "Salario Base", "Fecha Ingreso", "Fecha Salida"])
        for row in rows:
            ws.append(row)
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        buf.name = "colaboradores_area.xlsx"
        return buf
