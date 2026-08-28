"""Tests #239 — Exportar nómina/asistencia por rango de fechas.

Issue: Indunnova16/Instelec#239

``ExportarAsistenciaRangoView`` (GET
/cuadrillas/exportar-asistencia-rango/?fecha_inicio=...&fecha_fin=...)
generaliza ``ExportarAsistenciaView`` (limitada a 1 sola cuadrilla + 1
semana derivada del código) a un rango de fechas arbitrario que consolida
en UN solo Excel TODOS los colaboradores de TODAS las cuadrillas activas
cuya ``fecha`` cae dentro del rango pedido -- mismo patrón que #211 usó
para el export de programación por rango.

Ejecutar:
  DJANGO_SETTINGS_MODULE=config.settings.dev_lite \
    venv/bin/python -m pytest apps/cuadrillas/tests_issue_239.py -v \
    -o python_files="tests_*.py test_*.py"
"""

from datetime import date
from io import BytesIO

import openpyxl
from django.contrib.auth import get_user_model
from django.test import TestCase
from django.urls import reverse

from apps.cuadrillas.models import Asistencia, Cuadrilla, CuadrillaMiembro

Usuario = get_user_model()


def _crear_admin():
    return Usuario.objects.create_user(
        email="admin_239@test.com",
        password="testpass123!",
        first_name="Admin",
        last_name="239",
        rol="admin",
        is_staff=True,
    )


def _crear_usuario(documento, nombre):
    partes = nombre.split(maxsplit=1)
    return Usuario.objects.create(
        email=f"{documento}@test.local",
        documento=documento,
        first_name=partes[0],
        last_name=partes[1] if len(partes) > 1 else "",
        rol="liniero",
        is_active=True,
    )


def _crear_cuadrilla(codigo, fecha, nombre="MANTENIMIENTO - 239"):
    return Cuadrilla.objects.create(
        codigo=codigo,
        nombre=nombre,
        activa=True,
        observaciones="",
        fecha=fecha,
    )


def _agregar_miembro(cuadrilla, usuario, fecha_inicio, cargo="MIEMBRO", rol_cuadrilla="LINIERO_I"):
    return CuadrillaMiembro.objects.create(
        cuadrilla=cuadrilla,
        usuario=usuario,
        rol_cuadrilla_id=rol_cuadrilla,
        cargo=cargo,
        costo_dia=0,
        fecha_inicio=fecha_inicio,
        activo=True,
    )


def _registrar_asistencia(cuadrilla, usuario, fecha, tipo_novedad="PRESENTE", **extra):
    return Asistencia.objects.create(
        usuario=usuario,
        cuadrilla=cuadrilla,
        fecha=fecha,
        tipo_novedad=tipo_novedad,
        registrado_por=None,
        **extra,
    )


def _leer_columna_nombre(xlsx_bytes):
    """Devuelve el set de valores de la columna Nombre (índice 1, tras la
    columna Cuadrilla en índice 0) presentes en el .xlsx generado."""
    wb = openpyxl.load_workbook(BytesIO(xlsx_bytes))
    ws = wb.active
    nombres = set()
    for row in ws.iter_rows(min_row=5, values_only=True):
        if row[1]:
            nombres.add(row[1])
    return nombres


class TestExportarAsistenciaRangoView(TestCase):
    """Integration: GET /cuadrillas/exportar-asistencia-rango/."""

    def setUp(self):
        self.admin = _crear_admin()
        self.client.force_login(self.admin)
        self.usuario_a = _crear_usuario("111239", "PEDRO PEREZ")
        self.usuario_b = _crear_usuario("222239", "JUAN GOMEZ")

    def test_rango_con_datos_de_2_cuadrillas_distintas_trae_ambas(self):
        """Escenario exacto del issue: rango que cruza 2 cuadrillas
        distintas -> el Excel trae los colaboradores de AMBAS."""
        c1 = _crear_cuadrilla("02-2026-0001-QA9", date(2026, 1, 6), nombre="CUADRILLA_A")
        _agregar_miembro(c1, self.usuario_a, date(2026, 1, 6))
        _registrar_asistencia(c1, self.usuario_a, date(2026, 1, 6))

        c2 = _crear_cuadrilla("05-2026-0001-QA9", date(2026, 1, 27), nombre="CUADRILLA_B")
        _agregar_miembro(c2, self.usuario_b, date(2026, 1, 27))
        _registrar_asistencia(c2, self.usuario_b, date(2026, 1, 27), tipo_novedad="AUSENTE")

        resp = self.client.get(
            reverse("cuadrillas:exportar_asistencia_rango"),
            {"fecha_inicio": "2026-01-01", "fecha_fin": "2026-01-31"},
        )
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(
            resp["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        nombres = _leer_columna_nombre(resp.content)
        self.assertIn("PEDRO PEREZ", nombres)
        self.assertIn("JUAN GOMEZ", nombres)

    def test_rango_sin_datos_no_crashea_excel_con_headers(self):
        resp = self.client.get(
            reverse("cuadrillas:exportar_asistencia_rango"),
            {"fecha_inicio": "2026-06-01", "fecha_fin": "2026-06-30"},
        )
        self.assertEqual(resp.status_code, 200)
        wb = openpyxl.load_workbook(BytesIO(resp.content))
        ws = wb.active
        headers = [c.value for c in next(ws.iter_rows(min_row=4, max_row=4))]
        self.assertEqual(headers[:5], ["Cuadrilla", "Nombre", "Documento", "Cargo", "Rol"])
        # No debe haber filas de datos (solo título + headers).
        self.assertIsNone(ws.cell(row=5, column=1).value)

    def test_fecha_fin_antes_de_inicio_error_claro_no_500(self):
        resp = self.client.get(
            reverse("cuadrillas:exportar_asistencia_rango"),
            {"fecha_inicio": "2026-01-31", "fecha_fin": "2026-01-01"},
        )
        self.assertEqual(resp.status_code, 400)
        self.assertIn("no puede ser anterior", resp.content.decode())

    def test_sin_fechas_400(self):
        resp = self.client.get(reverse("cuadrillas:exportar_asistencia_rango"))
        self.assertEqual(resp.status_code, 400)

    def test_fecha_invalida_400(self):
        resp = self.client.get(
            reverse("cuadrillas:exportar_asistencia_rango"),
            {"fecha_inicio": "no-es-fecha", "fecha_fin": "2026-01-31"},
        )
        self.assertEqual(resp.status_code, 400)

    def test_requiere_login(self):
        self.client.logout()
        resp = self.client.get(
            reverse("cuadrillas:exportar_asistencia_rango"),
            {"fecha_inicio": "2026-01-01", "fecha_fin": "2026-01-31"},
        )
        self.assertIn(resp.status_code, (302, 301))

    def test_cuadrilla_inactiva_excluida(self):
        c = _crear_cuadrilla("02-2026-0002-QA9", date(2026, 1, 6), nombre="INACTIVA")
        c.activa = False
        c.save(update_fields=["activa"])
        _agregar_miembro(c, self.usuario_a, date(2026, 1, 6))
        _registrar_asistencia(c, self.usuario_a, date(2026, 1, 6))

        resp = self.client.get(
            reverse("cuadrillas:exportar_asistencia_rango"),
            {"fecha_inicio": "2026-01-01", "fecha_fin": "2026-01-31"},
        )
        self.assertEqual(resp.status_code, 200)
        nombres = _leer_columna_nombre(resp.content)
        self.assertNotIn("PEDRO PEREZ", nombres)

    def test_columnas_de_dia_cubren_todo_el_rango_no_solo_7_dias(self):
        """Rango de 10 días -> 10 columnas de día (no fijo a 7 como el
        export semanal de 1 cuadrilla)."""
        c1 = _crear_cuadrilla("02-2026-0003-QA9", date(2026, 1, 6), nombre="RANGO_LARGO")
        _agregar_miembro(c1, self.usuario_a, date(2026, 1, 6))
        for i in range(10):
            _registrar_asistencia(c1, self.usuario_a, date(2026, 1, 6 + i))

        resp = self.client.get(
            reverse("cuadrillas:exportar_asistencia_rango"),
            {"fecha_inicio": "2026-01-06", "fecha_fin": "2026-01-15"},
        )
        self.assertEqual(resp.status_code, 200)
        wb = openpyxl.load_workbook(BytesIO(resp.content))
        ws = wb.active
        headers = [c.value for c in next(ws.iter_rows(min_row=4, max_row=4))]
        # 5 columnas fijas (Cuadrilla..Rol) + 10 dias + 6 totales + Observaciones = 22
        self.assertEqual(len(headers), 5 + 10 + 6 + 1)
