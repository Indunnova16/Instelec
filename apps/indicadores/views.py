"""
Views for KPIs and SLA dashboard.
"""
import json
from datetime import timedelta

from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.cache import cache
from django.db.models import Count
from django.db.models.functions import TruncMonth
from django.http import HttpResponse
from django.utils import timezone
from django.views import View
from django.views.generic import DetailView, ListView, TemplateView

from apps.core.mixins import HTMXMixin, RoleRequiredMixin
from apps.core.utils import get_unidad_negocio

from .models import ActaSeguimiento, Indicador, MedicionIndicador


class DashboardView(LoginRequiredMixin, HTMXMixin, TemplateView):
    """KPI Dashboard."""
    template_name = 'indicadores/dashboard.html'
    partial_template_name = 'indicadores/partials/dashboard_content.html'

    def get_template_names(self):
        # Check if this is an HTMX request for a specific chart
        chart = self.request.GET.get('chart')
        if self.request.headers.get('HX-Request') and chart:
            return [f'indicadores/partials/chart_{chart}.html']
        return super().get_template_names()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        import json

        from django.db.models import Avg
        from django.utils import timezone

        from apps.actividades.models import Actividad
        from apps.cuadrillas.models import Cuadrilla
        from apps.lineas.models import Linea

        hoy = timezone.now()
        try:
            mes = int(self.request.GET.get('mes', hoy.month))
        except (ValueError, TypeError):
            mes = hoy.month
        try:
            anio = int(self.request.GET.get('anio', hoy.year))
        except (ValueError, TypeError):
            anio = hoy.year

        # Get all active indicators
        indicadores = Indicador.objects.filter(activo=True)
        context['indicadores'] = indicadores

        # Get measurements for current period
        mediciones = MedicionIndicador.objects.filter(
            anio=anio,
            mes=mes
        ).select_related('indicador', 'linea')

        context['mediciones'] = mediciones

        # Calculate summary
        context['promedio_cumplimiento'] = mediciones.aggregate(
            promedio=Avg('valor_calculado')
        )['promedio'] or 0

        context['en_alerta'] = mediciones.filter(en_alerta=True).count()
        context['cumplen_meta'] = mediciones.filter(cumple_meta=True).count()

        context['mes'] = mes
        context['anio'] = anio

        # Get activities for KPIs
        actividades = Actividad.objects.filter(
            fecha_programada__year=anio,
            fecha_programada__month=mes
        )
        total_actividades = actividades.count()
        completadas = actividades.filter(estado='COMPLETADA').count()

        # KPIs
        cumplimiento = (completadas / total_actividades * 100) if total_actividades > 0 else 0
        context['kpis'] = {
            'cumplimiento': cumplimiento,
            'actividades_completadas': completadas,
            'actividades_programadas': total_actividades,
            'dias_sin_accidentes': 45,  # Placeholder - should come from safety model
            'record_dias_sin_accidentes': 120,
            'informes_tiempo': 92.5,  # Placeholder
        }

        # Period filters
        context['periodos'] = [
            {'value': 'mes', 'label': 'Este mes'},
            {'value': 'semana', 'label': 'Esta semana'},
            {'value': 'trimestre', 'label': 'Este trimestre'},
        ]
        context['periodo_actual'] = self.request.GET.get('periodo', 'mes')

        # Lines for filter
        context['lineas'] = Linea.objects.filter(activa=True)

        # Cumplimiento por cuadrilla
        cuadrillas = Cuadrilla.objects.filter(activa=True)
        cuadrillas_labels = []
        cuadrillas_data = []
        for cuadrilla in cuadrillas[:10]:
            cuadrillas_labels.append(cuadrilla.codigo)
            acts = actividades.filter(cuadrilla=cuadrilla)
            total = acts.count()
            comp = acts.filter(estado='COMPLETADA').count()
            pct = (comp / total * 100) if total > 0 else 0
            cuadrillas_data.append(round(pct, 1))

        context['cuadrillas_labels'] = json.dumps(cuadrillas_labels)
        context['cuadrillas_data'] = json.dumps(cuadrillas_data)

        # Tendencia mensual (últimos 6 meses)
        meses_labels = []
        planeado_data = []
        ejecutado_data = []
        meses_nombres = ['Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun', 'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic']
        for i in range(5, -1, -1):
            m = mes - i
            a = anio
            if m <= 0:
                m += 12
                a -= 1
            meses_labels.append(f"{meses_nombres[m-1]}")
            acts_mes = Actividad.objects.filter(fecha_programada__year=a, fecha_programada__month=m)
            planeado_data.append(acts_mes.count())
            ejecutado_data.append(acts_mes.filter(estado='COMPLETADA').count())

        context['meses_labels'] = json.dumps(meses_labels)
        context['planeado_data'] = json.dumps(planeado_data)
        context['ejecutado_data'] = json.dumps(ejecutado_data)

        # Por tipo de actividad
        from apps.actividades.models import TipoActividad
        tipos = TipoActividad.objects.filter(activo=True)
        tipo_data = []
        for tipo in tipos[:8]:
            count = actividades.filter(tipo_actividad=tipo).count()
            if count > 0:
                tipo_data.append({'value': count, 'name': tipo.nombre})
        context['tipo_data'] = json.dumps(tipo_data)

        # Por prioridad
        context['prioridad_data'] = {
            'urgente': actividades.filter(prioridad='URGENTE').count(),
            'alta': actividades.filter(prioridad='ALTA').count(),
            'normal': actividades.filter(prioridad='NORMAL').count(),
            'baja': actividades.filter(prioridad='BAJA').count(),
        }

        # Actividades recientes
        context['actividades_recientes'] = Actividad.objects.select_related(
            'torre'
        ).order_by('-updated_at')[:5]

        # Data for charts
        context['indicadores_data'] = [
            {
                'nombre': m.indicador.nombre,
                'valor': float(m.valor_calculado),
                'meta': float(m.indicador.meta),
            }
            for m in mediciones
        ]

        # #122 (rebote): los 6 KPIs técnico-financieros + ANS deben verse en ESTE
        # dashboard (el que usa el cliente), no solo en /financiero/. Reutilizamos
        # el cálculo del módulo financiero (misma fuente de verdad) filtrando por
        # la línea seleccionada (o todas).
        #
        # mes=0 fuerza la agregación ANUAL (los 12 meses), igual que /financiero/
        # (que usa mes=0 por default). Pasar mes=hoy.month ventaneaba los KPIs al
        # mes actual: si el presupuesto REAL del año vive en otro mes (p.ej. marzo
        # 2026 = $57M, junio = $0), el mes actual sumaba $0 y los 6 KPIs daban
        # 0.00%. La paridad con /financiero/ (lo que el cliente compara) exige
        # agregar el año completo.
        try:
            from apps.financiero.indicadores_finv2 import contexto_indicadores_finv2
            linea_sel = None
            linea_id = self.request.GET.get('linea')
            if linea_id:
                linea_sel = Linea.objects.filter(id=linea_id).first()
            context.update(contexto_indicadores_finv2(
                anio=anio, mes=0, contrato=None, linea=linea_sel))
        except Exception:
            # Nunca romper el dashboard de indicadores si el cálculo financiero falla.
            context.setdefault('indicadores_tecnico_financieros', [])
            context.setdefault('indicadores_ans', [])
            context.setdefault('resumen_ans', None)

        return context


class IndicadorDetailView(LoginRequiredMixin, DetailView):
    """Indicator detail with history."""
    model = Indicador
    template_name = 'indicadores/detalle.html'
    context_object_name = 'indicador'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Get historical measurements
        context['historial'] = MedicionIndicador.objects.filter(
            indicador=self.object
        ).order_by('-anio', '-mes')[:12]

        return context


class ActaListView(LoginRequiredMixin, RoleRequiredMixin, ListView):
    """List follow-up meeting minutes."""
    model = ActaSeguimiento
    template_name = 'indicadores/actas.html'
    context_object_name = 'actas'
    allowed_roles = ['admin', 'director', 'coordinador', 'ing_residente']

    def get_queryset(self):
        return super().get_queryset().select_related('linea')


class ActaDetailView(LoginRequiredMixin, RoleRequiredMixin, DetailView):
    """Meeting minutes detail."""
    model = ActaSeguimiento
    template_name = 'indicadores/acta_detalle.html'
    context_object_name = 'acta'
    allowed_roles = ['admin', 'director', 'coordinador', 'ing_residente']


def _resumen_mantenimiento(unidad: str) -> dict:
    """Computa métricas del dashboard de mantenimiento. Cacheado 5 min por unidad."""
    from apps.actividades.models import HistorialIntervencion
    from apps.campo.models import RegistroCampo
    from apps.lineas.models import Linea

    cache_key = f'dashboard_mtto:{unidad}'
    cached = cache.get(cache_key)
    if cached:
        return cached

    hoy = timezone.now().date()
    hace_30d = hoy - timedelta(days=30)
    hace_6m = hoy - timedelta(days=180)

    lineas = Linea.objects.filter(activa=True)
    registros = RegistroCampo.objects.filter(sincronizado=True)
    intervenciones = HistorialIntervencion.objects.all()
    if unidad in ('MANTENIMIENTO', 'CONSTRUCCION'):
        lineas = lineas.filter(contrato__unidad_negocio=unidad)
        registros = registros.filter(actividad__linea__contrato__unidad_negocio=unidad)
        intervenciones = intervenciones.filter(linea__contrato__unidad_negocio=unidad)

    total_lineas = lineas.count()
    vencidas = lineas.filter(inspection_status='VENCIDA').count()
    criticas = lineas.filter(inspection_status='CRITICA').count()
    proximas = lineas.filter(inspection_status='PROXIMA').count()

    # Distribución de severidad de los últimos 30 días.
    severidad_qs = registros.filter(
        fecha_inicio__date__gte=hace_30d,
    ).exclude(severidad='').values('severidad').annotate(n=Count('id'))
    distribucion_severidad = {row['severidad']: row['n'] for row in severidad_qs}

    # Inspecciones por tipo (últimos 30 días).
    tipo_qs = registros.filter(
        fecha_inicio__date__gte=hace_30d,
    ).values('actividad__tipo_actividad__nombre').annotate(n=Count('id')).order_by('-n')[:10]
    por_tipo = [(row['actividad__tipo_actividad__nombre'] or 'N/A', row['n']) for row in tipo_qs]

    # Tendencia mensual últimos 6 meses (registros).
    tendencia_qs = registros.filter(
        fecha_inicio__date__gte=hace_6m,
    ).annotate(mes=TruncMonth('fecha_inicio')).values('mes').annotate(n=Count('id')).order_by('mes')
    tendencia = [(row['mes'].strftime('%Y-%m'), row['n']) for row in tendencia_qs if row['mes']]

    data = {
        'total_lineas': total_lineas,
        'vencidas': vencidas,
        'criticas': criticas,
        'proximas': proximas,
        'al_dia': total_lineas - vencidas - criticas - proximas,
        'distribucion_severidad': distribucion_severidad,
        'por_tipo': por_tipo,
        'tendencia': tendencia,
        'total_registros_30d': registros.filter(fecha_inicio__date__gte=hace_30d).count(),
        'total_intervenciones_30d': intervenciones.filter(fecha_intervencion__date__gte=hace_30d).count(),
    }
    cache.set(cache_key, data, 300)
    return data


class DashboardMantenimientoView(LoginRequiredMixin, RoleRequiredMixin, TemplateView):
    """Dashboard específico de mantenimiento (#43).

    Métricas: líneas vencidas/críticas, severidad, inspecciones por tipo, tendencia.
    Filtra por sesión (`get_unidad_negocio`).
    """
    template_name = 'indicadores/dashboard_mantenimiento.html'
    allowed_roles = ['admin', 'director', 'coordinador', 'ing_residente', 'ing_ambiental', 'supervisor']

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        unidad = get_unidad_negocio(self.request)
        data = _resumen_mantenimiento(unidad)
        ctx['data'] = data
        ctx['unidad_actual'] = unidad
        # JSON para ECharts.
        ctx['severidad_json'] = json.dumps([
            {'name': k, 'value': v} for k, v in data['distribucion_severidad'].items()
        ])
        ctx['por_tipo_json'] = json.dumps({
            'labels': [t[0] for t in data['por_tipo']],
            'values': [t[1] for t in data['por_tipo']],
        })
        ctx['tendencia_json'] = json.dumps({
            'labels': [t[0] for t in data['tendencia']],
            'values': [t[1] for t in data['tendencia']],
        })
        return ctx


class ExportarDashboardMantenimientoExcelView(LoginRequiredMixin, RoleRequiredMixin, View):
    """Exporta el dashboard de mantenimiento a XLSX."""
    allowed_roles = ['admin', 'director', 'coordinador']

    def get(self, request, *args, **kwargs):
        from openpyxl import Workbook
        from openpyxl.styles import Font

        unidad = get_unidad_negocio(request)
        data = _resumen_mantenimiento(unidad)

        wb = Workbook()

        # Resumen.
        ws = wb.active
        ws.title = 'Resumen'
        bold = Font(bold=True)
        ws['A1'] = 'Indicador'
        ws['B1'] = 'Valor'
        ws['A1'].font = bold
        ws['B1'].font = bold
        filas = [
            ('Unidad de negocio', unidad),
            ('Total líneas activas', data['total_lineas']),
            ('Vencidas', data['vencidas']),
            ('Críticas', data['criticas']),
            ('Próximas a vencer', data['proximas']),
            ('Al día', data['al_dia']),
            ('Registros últimos 30d', data['total_registros_30d']),
            ('Intervenciones últimos 30d', data['total_intervenciones_30d']),
        ]
        for i, (k, v) in enumerate(filas, start=2):
            ws.cell(row=i, column=1, value=k)
            ws.cell(row=i, column=2, value=v)

        # Severidad.
        ws_sev = wb.create_sheet('Severidad')
        ws_sev.append(['Severidad', 'Registros'])
        for k, v in data['distribucion_severidad'].items():
            ws_sev.append([k, v])

        # Por tipo.
        ws_tipo = wb.create_sheet('Por tipo')
        ws_tipo.append(['Tipo de actividad', 'Inspecciones'])
        for label, n in data['por_tipo']:
            ws_tipo.append([label, n])

        # Tendencia.
        ws_tend = wb.create_sheet('Tendencia mensual')
        ws_tend.append(['Mes', 'Registros'])
        for label, n in data['tendencia']:
            ws_tend.append([label, n])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="dashboard_mantenimiento_{unidad}.xlsx"'
        wb.save(response)
        return response
