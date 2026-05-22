"""Importer Excel para indicadores KPI mensuales (#45).

Acepta un Excel con estructura simple:
- Hoja con encabezado en fila 1: codigo_indicador, año, mes, valor_real
- Filas 2+ con datos. Cada fila crea/actualiza una MedicionIndicador.

Más adelante, cuando Janeth envíe el template final, este importer se
extiende con las columnas que use ese formato (rentabilidad, variacion
presupuesto, ejecución presupuestal, avance, cumplimiento, revisiones,
fotos, etc.) — cada uno mapeado a un `Indicador.codigo` ya existente
en el modelo `Indicador` actual.
"""
import logging
from datetime import date

from django.db import transaction
from openpyxl import load_workbook

logger = logging.getLogger(__name__)


def importar_kpi_mensual(archivo_excel, usuario=None):
    """Carga MedicionIndicador desde xlsx con columnas:
    codigo_indicador, anio, mes, valor_real, [valor_meta opcional]."""
    from .models import Indicador, MedicionIndicador

    resultado = {
        'creadas': 0, 'actualizadas': 0, 'omitidas': 0,
        'errores': [], 'advertencias': [],
    }

    try:
        wb = load_workbook(archivo_excel, read_only=True, data_only=True)
        ws = wb.active
    except Exception as e:
        return {**resultado, 'error': f'Error al cargar Excel: {e}'}

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return {**resultado, 'error': 'Excel vacío'}

    header = [str(c or '').lower().strip() for c in rows[0]]
    idx = {}
    for col in ('codigo_indicador', 'anio', 'mes', 'valor_real', 'valor_meta'):
        for i, h in enumerate(header):
            if h == col or h.replace('_', ' ') == col.replace('_', ' '):
                idx[col] = i
                break

    requeridos = {'codigo_indicador', 'anio', 'mes', 'valor_real'}
    faltantes = requeridos - set(idx.keys())
    if faltantes:
        return {**resultado, 'error': f'columnas faltantes: {sorted(faltantes)}'}

    with transaction.atomic():
        for n, row in enumerate(rows[1:], start=2):
            try:
                codigo = str(row[idx['codigo_indicador']] or '').strip()
                anio = int(row[idx['anio']])
                mes = int(row[idx['mes']])
                valor = row[idx['valor_real']]
                if valor is None or codigo == '':
                    resultado['omitidas'] += 1
                    continue
                indicador = Indicador.objects.filter(codigo__iexact=codigo).first()
                if not indicador:
                    resultado['advertencias'].append({
                        'fila': n,
                        'mensaje': f'indicador {codigo} no existe — fila omitida',
                    })
                    resultado['omitidas'] += 1
                    continue
                from decimal import Decimal
                _, created = MedicionIndicador.objects.update_or_create(
                    indicador=indicador,
                    anio=anio,
                    mes=mes,
                    linea=None,
                    defaults={
                        'valor_calculado': Decimal(str(valor)),
                        'cumple_meta': Decimal(str(valor)) >= indicador.meta,
                        'en_alerta': Decimal(str(valor)) < indicador.umbral_alerta,
                    },
                )
                if created:
                    resultado['creadas'] += 1
                else:
                    resultado['actualizadas'] += 1
            except Exception as e:
                resultado['errores'].append({'fila': n, 'error': str(e)})

    return resultado
