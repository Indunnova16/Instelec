"""Archivo Excel para Programación Semanal de Construcción (#225, B4)."""
from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime, time
from io import BytesIO

from django.core.exceptions import ValidationError
from django.db import transaction
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from apps.cuadrillas.models import PersonalCuadrilla, Vehiculo
from apps.usuarios.models import Usuario

from .models import (
    ProgramacionSemanalConstruccion,
    ProgramacionSemanalConstruccionPersonal,
    ProgramacionSemanalConstruccionVehiculo,
    ProyectoConstruccion,
)
from .services_psc_disponibilidad import personal_elegible, validar_personal_elegible


HEADERS = (
    'Proyecto', 'Tipo Actividad', 'Sub-Actividad', 'Supervisor', 'Personal',
    'Vehículos', 'Fecha Inicio', 'Fecha Fin', 'Hora Inicio', 'Hora Fin', 'Observaciones',
)
MAX_IMPORT_ROWS = 500


@dataclass
class ImportResult:
    created: int = 0
    errors: list[dict] = field(default_factory=list)

    @property
    def ok(self):
        return not self.errors


def _workbook_response(rows=()):
    book = Workbook()
    sheet = book.active
    sheet.title = 'Programación semanal'
    sheet.append(HEADERS)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='1D4ED8')
    sheet.freeze_panes = 'A2'
    sheet.auto_filter.ref = f'A1:K{max(2, len(rows) + 1)}'
    for row in rows:
        sheet.append(row)
    for column, width in {'A': 28, 'B': 24, 'C': 30, 'D': 28, 'E': 35, 'F': 25,
                          'G': 14, 'H': 14, 'I': 12, 'J': 12, 'K': 40}.items():
        sheet.column_dimensions[column].width = width
    output = BytesIO()
    book.save(output)
    output.seek(0)
    return output


def plantilla_programacion_semanal():
    """Devuelve la plantilla XLSX con las once columnas contratadas."""
    return _workbook_response()


def exportar_programacion_semanal():
    """Exporta todas las programaciones, incluyendo personal y vehículos."""
    programaciones = ProgramacionSemanalConstruccion.objects.select_related(
        'proyecto', 'supervisor',
    ).prefetch_related('asignaciones_personal__personal', 'asignaciones_vehiculo__vehiculo')
    rows = []
    for item in programaciones:
        personal = ', '.join(
            asignacion.personal.documento or asignacion.personal.nombre
            for asignacion in item.asignaciones_personal.all()
        )
        vehiculos = ', '.join(asignacion.vehiculo.placa for asignacion in item.asignaciones_vehiculo.all())
        rows.append((
            item.proyecto.nombre, item.get_tipo_actividad_display(), item.subactividad,
            item.supervisor.email if item.supervisor else '', personal, vehiculos,
            item.fecha_inicio, item.fecha_fin, item.hora_inicio, item.hora_fin,
            item.observaciones,
        ))
    return _workbook_response(rows)


def _as_date(value, field_name):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        try:
            return date.fromisoformat(value.strip())
        except ValueError:
            pass
    raise ValidationError(f'{field_name} debe ser una fecha válida (AAAA-MM-DD).')


def _as_time(value, field_name):
    if value in (None, ''):
        return None
    if isinstance(value, datetime):
        return value.time().replace(second=0, microsecond=0)
    if isinstance(value, time):
        return value.replace(second=0, microsecond=0)
    if isinstance(value, str):
        try:
            return time.fromisoformat(value.strip())
        except ValueError:
            pass
    raise ValidationError(f'{field_name} debe ser una hora válida (HH:MM).')


def _tokens(value):
    return [part.strip() for part in str(value or '').split(',') if part.strip()]


def _one(queryset, value, description):
    matches = list(queryset)
    if not matches:
        raise ValidationError(f'{description} "{value}" no existe.')
    if len(matches) > 1:
        raise ValidationError(f'{description} "{value}" es ambiguo; use un identificador único.')
    return matches[0]


def _tipo(value):
    normalized = str(value or '').strip().casefold()
    for code, label in ProgramacionSemanalConstruccion.TipoActividad.choices:
        if normalized in (code.casefold(), label.casefold()):
            return code
    raise ValidationError('Tipo Actividad no es válido.')


def _parse_row(row):
    proyecto_nombre, tipo, subactividad, supervisor, personal, vehiculos, inicio, fin, hora_inicio, hora_fin, observaciones = row
    proyecto = _one(ProyectoConstruccion.objects.filter(nombre__iexact=str(proyecto_nombre or '').strip()), proyecto_nombre, 'Proyecto')
    fecha_inicio, fecha_fin = _as_date(inicio, 'Fecha Inicio'), _as_date(fin, 'Fecha Fin')
    if fecha_fin < fecha_inicio:
        raise ValidationError('Fecha Fin no puede ser anterior a Fecha Inicio.')
    inicio_hora, fin_hora = _as_time(hora_inicio, 'Hora Inicio'), _as_time(hora_fin, 'Hora Fin')
    if fecha_inicio == fecha_fin and inicio_hora and fin_hora and fin_hora <= inicio_hora:
        raise ValidationError('Hora Fin debe ser posterior a Hora Inicio.')
    tipo_actividad = _tipo(tipo)
    subactividad = str(subactividad or '').strip()
    if tipo_actividad == ProgramacionSemanalConstruccion.TipoActividad.COMPLEMENTARIAS:
        if not subactividad:
            raise ValidationError('Sub-Actividad es obligatoria para actividades complementarias.')
        actividad_complementaria = subactividad
    elif not subactividad:
        raise ValidationError('Sub-Actividad es obligatoria.')
    else:
        actividad_complementaria = ''
    supervisor = str(supervisor or '').strip()
    supervisor_obj = _one(Usuario.objects.filter(email__iexact=supervisor), supervisor, 'Supervisor') if supervisor else None
    people = []
    for token in _tokens(personal):
        people.append(_one(PersonalCuadrilla.objects.filter(documento__iexact=token), token, 'Personal'))
    if len({str(person.pk) for person in people}) != len(people):
        raise ValidationError('Personal contiene una persona repetida.')
    vehicle_objects = [_one(Vehiculo.objects.filter(placa__iexact=token), token, 'Vehículo') for token in _tokens(vehiculos)]
    if len({str(vehicle.pk) for vehicle in vehicle_objects}) != len(vehicle_objects):
        raise ValidationError('Vehículos contiene una placa repetida.')
    return {
        'proyecto': proyecto, 'tipo_actividad': tipo_actividad, 'subactividad': subactividad,
        'actividad_complementaria': actividad_complementaria, 'supervisor': supervisor_obj,
        'personal': people, 'vehiculos': vehicle_objects, 'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin, 'hora_inicio': inicio_hora, 'hora_fin': fin_hora,
        'observaciones': str(observaciones or '').strip(),
    }


def importar_programacion_semanal(uploaded_file):
    """Valida todo el XLSX y persiste sus filas en una única transacción."""
    result = ImportResult()
    try:
        workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
    except Exception as exc:
        result.errors.append({'row': 0, 'error': f'No se pudo leer el archivo XLSX: {exc}'})
        return result
    if not rows or tuple(rows[0]) != HEADERS:
        result.errors.append({'row': 1, 'error': 'Las columnas deben coincidir exactamente con la plantilla descargada.'})
        return result
    data_rows = [(number, row) for number, row in enumerate(rows[1:], start=2) if any(value not in (None, '') for value in row)]
    if not data_rows:
        result.errors.append({'row': 0, 'error': 'El archivo no contiene programaciones para importar.'})
        return result
    if len(data_rows) > MAX_IMPORT_ROWS:
        result.errors.append({'row': 0, 'error': f'El archivo supera el máximo de {MAX_IMPORT_ROWS} filas.'})
        return result
    parsed = []
    for number, row in data_rows:
        if len(row) != len(HEADERS):
            result.errors.append({'row': number, 'error': 'La fila no contiene las 11 columnas requeridas.'})
            continue
        try:
            parsed.append((number, _parse_row(row)))
        except ValidationError as exc:
            result.errors.append({'row': number, 'error': '; '.join(exc.messages)})
    occupied = {}
    for number, item in parsed:
        for person in item['personal']:
            for other_start, other_end, other_row in occupied.get(person.pk, []):
                if item['fecha_inicio'] <= other_end and item['fecha_fin'] >= other_start:
                    result.errors.append({'row': number, 'error': f'Personal {person.documento} se cruza con la fila {other_row}.'})
            occupied.setdefault(person.pk, []).append((item['fecha_inicio'], item['fecha_fin'], number))
        eligible_ids = set(personal_elegible(item['proyecto'].pk, item['fecha_inicio'], item['fecha_fin']).values_list('pk', flat=True))
        ineligible = [person.documento for person in item['personal'] if person.pk not in eligible_ids]
        if ineligible:
            result.errors.append({'row': number, 'error': 'Personal no elegible o ya ocupado: ' + ', '.join(ineligible)})
    if result.errors:
        return result
    with transaction.atomic():
        for _, item in parsed:
            people, vehicles = item.pop('personal'), item.pop('vehiculos')
            programacion = ProgramacionSemanalConstruccion.objects.create(**item)
            validar_personal_elegible(programacion, [person.pk for person in people])
            ProgramacionSemanalConstruccionPersonal.objects.bulk_create([
                ProgramacionSemanalConstruccionPersonal(programacion=programacion, personal=person)
                for person in people
            ])
            ProgramacionSemanalConstruccionVehiculo.objects.bulk_create([
                ProgramacionSemanalConstruccionVehiculo(programacion=programacion, vehiculo=vehicle)
                for vehicle in vehicles
            ])
    result.created = len(parsed)
    return result
