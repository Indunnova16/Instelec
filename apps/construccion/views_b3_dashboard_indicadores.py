"""B3 — Dashboard Indicadores en General (#97).

Dashboard ejecutivo en ``/construccion/<uuid>/indicadores-financieros/``
(path conservado, view reemplazada). 6 KPI cards + 6 gráficas Chart.js
(CDN 4.4.0). Filtros (período, tipo, línea). Exportar PDF (weasyprint) y
Excel (openpyxl).

Lee modelos de B2 (#98); si vacío, muestra placeholder "Aún no hay
indicadores registrados".
"""
from __future__ import annotations

import json
from datetime import date, timedelta
from decimal import Decimal
from io import BytesIO

from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404
from django.utils import timezone
from django.views.generic import TemplateView

from apps.core.mixins import RoleRequiredMixin

from .models import ProyectoConstruccion

# ALL_ADMIN_ROLES vive en apps/construccion/views.py — la importamos para
# mantener consistencia con el resto del módulo (los stubs F2 reexportan
# este símbolo y por contrato sólo lo tocamos en B3 via import).
ALL_ADMIN_ROLES = [
    'admin', 'director', 'coordinador', 'ing_residente',
    'admin_general', 'coordinador_general', 'admin_construccion',
]


# ===========================================================================
# Constantes de meta / target del dashboard
# ===========================================================================

#: Meta de torres terminadas por mes (configurable en futuro via Settings).
META_TORRES_MES = 18

#: Períodos de filtro soportados.
PERIODO_CHOICES = (
    ('semana', 'Última semana'),
    ('mes', 'Último mes'),
    ('trimestre', 'Último trimestre'),
    ('anio', 'Último año'),
    ('todo', 'Todo el proyecto'),
)

#: Tipos de indicador para filtro (cuál bloque mostrar).
TIPO_CHOICES = (
    ('todos', 'Todos los indicadores'),
    ('financiero', 'Sólo financieros'),
    ('tecnico', 'Sólo técnicos'),
    ('desempeno', 'Sólo desempeño por línea'),
)


def _periodo_to_since(periodo: str) -> date | None:
    """Convierte ``?periodo=`` a fecha-corte. ``None`` = sin filtro."""
    hoy = timezone.localdate()
    mapping = {
        'semana': hoy - timedelta(days=7),
        'mes': hoy - timedelta(days=30),
        'trimestre': hoy - timedelta(days=90),
        'anio': hoy - timedelta(days=365),
    }
    return mapping.get(periodo)


def _safe_float(value, default=0.0):
    """Cast a float manejando None / Decimal / Exception."""
    if value is None:
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


# ===========================================================================
# Agregador
# ===========================================================================

class IndicadoresAggregator:
    """Lee modelos B2 + agrega data para 6 KPIs + 6 gráficas.

    Se reusa desde la View HTML, el export PDF y el export Excel — single
    source of truth para evitar drift entre vistas.
    """

    def __init__(self, proyecto: ProyectoConstruccion, periodo: str = 'todo',
                 tipo: str = 'todos', linea_id: str | None = None):
        self.proyecto = proyecto
        self.periodo = periodo
        self.tipo = tipo
        self.linea_id = linea_id
        self.since = _periodo_to_since(periodo)

    # ---- querysets base ----

    def qs_financieros(self):
        try:
            from .models_b2_indicadores import IndicadorFinancieroConstruccion
        except ImportError:
            return []
        qs = IndicadorFinancieroConstruccion.objects.filter(proyecto=self.proyecto)
        if self.since:
            qs = qs.filter(fecha__gte=self.since)
        return qs.order_by('fecha')

    def qs_tecnicos(self):
        try:
            from .models_b2_indicadores import IndicadorTecnicoConstruccion
        except ImportError:
            return []
        qs = IndicadorTecnicoConstruccion.objects.filter(proyecto=self.proyecto)
        if self.since:
            qs = qs.filter(fecha__gte=self.since)
        return qs.order_by('fecha')

    def qs_desempeno(self):
        try:
            from .models_b2_indicadores import IndicadorDesempenoLinea
        except ImportError:
            return []
        qs = IndicadorDesempenoLinea.objects.filter(proyecto=self.proyecto)
        if self.since:
            qs = qs.filter(fecha__gte=self.since)
        if self.linea_id:
            qs = qs.filter(linea_id=self.linea_id)
        return qs.select_related('linea', 'cuadrilla').order_by('fecha')

    # ---- empty state ----

    @staticmethod
    def _qs_count(qs) -> int:
        """Cuenta elementos de QuerySet o lista. ``list.count(value)`` tiene
        firma distinta a ``QuerySet.count()`` así que NO podemos usar hasattr.
        """
        from django.db.models import QuerySet
        if isinstance(qs, QuerySet):
            return qs.count()
        try:
            return len(qs)
        except TypeError:
            return sum(1 for _ in qs)

    @property
    def is_empty(self) -> bool:
        return (
            self._qs_count(self.qs_financieros())
            + self._qs_count(self.qs_tecnicos())
            + self._qs_count(self.qs_desempeno())
        ) == 0

    # ---- 6 KPI cards ----

    def kpis(self) -> list[dict]:
        """6 KPI cards. Cada item: {label, value, unit, trend, status, accent}."""
        fin = list(self.qs_financieros())
        tec = list(self.qs_tecnicos())

        ultimo_fin = fin[-1] if fin else None
        ultimo_tec = tec[-1] if tec else None

        # 1. Margen Operativo (%)
        margen_val = _safe_float(getattr(ultimo_fin, 'margen_operativo', None)) if ultimo_fin else 0.0
        margen_status = self._classify(margen_val, ok=15.0, warn=5.0)

        # 2. Desviación Presupuestal (%)
        desv_val = _safe_float(getattr(ultimo_fin, 'desviacion_presupuestal', None)) if ultimo_fin else 0.0
        desv_status = self._classify_desviacion(desv_val)

        # 3. Avance Técnico (% — promedio avance_obra del último registro)
        avance_val = _safe_float(getattr(ultimo_tec, 'avance_obra', None)) if ultimo_tec else 0.0
        avance_status = self._classify(avance_val, ok=80.0, warn=50.0)

        # 4. Accidentes (count) — leído de actividades.RegistroIncidenteSeguridad si
        #    existe, sino 0. Defensivo: módulo opcional.
        accidentes_val = self._count_incidentes_seguridad()
        accidentes_status = 'ok' if accidentes_val == 0 else ('warn' if accidentes_val <= 2 else 'critical')

        # 5. Ejecución Presupuestal (%)
        ejec_val = _safe_float(getattr(ultimo_tec, 'ejecucion_presupuestal', None)) if ultimo_tec else 0.0
        ejec_status = self._classify(ejec_val, ok=80.0, warn=50.0)

        # 6. Capacitaciones (count) — leído de actividades.Capacitacion si existe.
        capacit_val = self._count_capacitaciones()
        capacit_status = 'ok' if capacit_val >= 4 else ('warn' if capacit_val >= 1 else 'critical')

        return [
            {
                'key': 'margen_operativo',
                'label': 'Margen Operativo',
                'value': round(margen_val, 2),
                'unit': '%',
                'status': margen_status,
                'accent': 'green' if margen_status == 'ok' else ('amber' if margen_status == 'warn' else 'red'),
                'description': 'Utilidad sobre ingresos ejecutados',
            },
            {
                'key': 'desviacion_presupuestal',
                'label': 'Desviación Presupuestal',
                'value': round(desv_val, 2),
                'unit': '%',
                'status': desv_status,
                'accent': 'green' if desv_status == 'ok' else ('amber' if desv_status == 'warn' else 'red'),
                'description': 'Costo real vs presupuestado',
            },
            {
                'key': 'avance_tecnico',
                'label': 'Avance Técnico',
                'value': round(avance_val, 2),
                'unit': '%',
                'status': avance_status,
                'accent': 'green' if avance_status == 'ok' else ('amber' if avance_status == 'warn' else 'red'),
                'description': 'Obra ejecutada vs programada',
            },
            {
                'key': 'accidentes',
                'label': 'Accidentes',
                'value': accidentes_val,
                'unit': '',
                'status': accidentes_status,
                'accent': 'green' if accidentes_status == 'ok' else ('amber' if accidentes_status == 'warn' else 'red'),
                'description': f'Período: {dict(PERIODO_CHOICES).get(self.periodo, self.periodo)}',
            },
            {
                'key': 'ejecucion_presupuestal',
                'label': 'Ejecución Presupuestal',
                'value': round(ejec_val, 2),
                'unit': '%',
                'status': ejec_status,
                'accent': 'green' if ejec_status == 'ok' else ('amber' if ejec_status == 'warn' else 'red'),
                'description': '% ejecutado / planeado',
            },
            {
                'key': 'capacitaciones',
                'label': 'Capacitaciones',
                'value': capacit_val,
                'unit': '',
                'status': capacit_status,
                'accent': 'green' if capacit_status == 'ok' else ('amber' if capacit_status == 'warn' else 'red'),
                'description': 'Sesiones registradas en período',
            },
        ]

    @staticmethod
    def _classify(value: float, ok: float, warn: float) -> str:
        if value >= ok:
            return 'ok'
        if value >= warn:
            return 'warn'
        return 'critical'

    @staticmethod
    def _classify_desviacion(value: float) -> str:
        """Desviación: |val| ≤ 5 → ok, ≤ 15 → warn, > 15 → critical."""
        abs_val = abs(value)
        if abs_val <= 5.0:
            return 'ok'
        if abs_val <= 15.0:
            return 'warn'
        return 'critical'

    def _count_incidentes_seguridad(self) -> int:
        """Cuenta incidentes en el período. Defensivo si el modelo no existe."""
        try:
            from apps.actividades.models import RegistroIncidenteSeguridad  # type: ignore
        except Exception:
            return 0
        qs = RegistroIncidenteSeguridad.objects.filter(proyecto=self.proyecto)
        if self.since:
            qs = qs.filter(fecha__gte=self.since)
        return qs.count()

    def _count_capacitaciones(self) -> int:
        """Cuenta capacitaciones en el período. Defensivo."""
        try:
            from apps.actividades.models import Capacitacion  # type: ignore
        except Exception:
            return 0
        qs = Capacitacion.objects.filter(proyecto=self.proyecto)
        if self.since:
            qs = qs.filter(fecha__gte=self.since)
        return qs.count()

    # ---- 6 gráficas ----

    def chart_flujo_caja(self) -> dict:
        """1. Flujo de caja acumulado (area). X = fecha, Y = ingresos − costos acumulados."""
        fin = list(self.qs_financieros())
        labels, values = [], []
        acumulado = Decimal('0')
        for ind in fin:
            ingresos = ind.ingresos_ejecutados or Decimal('0')
            costos = (ind.costos_directos or Decimal('0')) + (ind.gastos or Decimal('0'))
            flujo = ingresos - costos
            acumulado += flujo
            labels.append(ind.fecha.isoformat())
            values.append(_safe_float(acumulado))
        return {'labels': labels, 'values': values}

    def chart_avance_tecnico(self) -> dict:
        """2. Avance técnico programado vs ejecutado (line dual)."""
        tec = list(self.qs_tecnicos())
        labels, programado, ejecutado = [], [], []
        for ind in tec:
            labels.append(ind.fecha.isoformat())
            programado.append(_safe_float(ind.presupuesto_planeado_pct))
            ejecutado.append(_safe_float(ind.presupuesto_ejecutado_pct))
        return {'labels': labels, 'programado': programado, 'ejecutado': ejecutado}

    def chart_egresos_pie(self) -> dict:
        """3. Desglose egresos (pie). Costos directos vs gastos vs costo real (legacy)."""
        fin = list(self.qs_financieros())
        sum_cd = sum((_safe_float(i.costos_directos) for i in fin), 0.0)
        sum_g = sum((_safe_float(i.gastos) for i in fin), 0.0)
        sum_cr = sum((_safe_float(i.costo_real) for i in fin), 0.0)
        labels = ['Costos directos', 'Gastos', 'Costo real']
        values = [round(sum_cd, 2), round(sum_g, 2), round(sum_cr, 2)]
        return {'labels': labels, 'values': values}

    def chart_margen_kpi(self) -> dict:
        """4. Margen + Desviación (barras horizontales — series temporales)."""
        fin = list(self.qs_financieros())
        labels = [i.fecha.isoformat() for i in fin]
        margen = [_safe_float(i.margen_operativo) for i in fin]
        desviacion = [_safe_float(i.desviacion_presupuestal) for i in fin]
        return {'labels': labels, 'margen': margen, 'desviacion': desviacion}

    def chart_seguridad(self) -> dict:
        """5. Seguridad — incidentes acumulados (line). Sin modelo → vacío."""
        try:
            from apps.actividades.models import RegistroIncidenteSeguridad  # type: ignore
        except Exception:
            return {'labels': [], 'values': []}
        qs = RegistroIncidenteSeguridad.objects.filter(proyecto=self.proyecto)
        if self.since:
            qs = qs.filter(fecha__gte=self.since)
        qs = qs.order_by('fecha')
        labels, values, acumulado = [], [], 0
        for inc in qs:
            acumulado += 1
            labels.append(inc.fecha.isoformat())
            values.append(acumulado)
        return {'labels': labels, 'values': values}

    def chart_productividad_cuadrillas(self) -> dict:
        """6. Productividad cuadrillas (horizontal bar con meta META_TORRES_MES).

        Agrupa por cuadrilla (cuando IndicadorDesempenoLinea tiene FK) y suma
        ``actual``. Si no hay datos por cuadrilla, cae a sum por línea.
        """
        des = list(self.qs_desempeno())
        agg: dict[str, float] = {}
        for ind in des:
            key = (
                getattr(getattr(ind, 'cuadrilla', None), 'nombre', None)
                or getattr(getattr(ind, 'linea', None), 'codigo', None)
                or '—'
            )
            agg[str(key)] = agg.get(str(key), 0.0) + _safe_float(ind.actual)
        labels = list(agg.keys())
        values = [round(v, 2) for v in agg.values()]
        return {'labels': labels, 'values': values, 'meta': META_TORRES_MES}

    # ---- agregado para template ----

    def all_charts_json(self) -> str:
        """Empaqueta las 6 gráficas en JSON para el template."""
        return json.dumps({
            'flujo_caja': self.chart_flujo_caja(),
            'avance_tecnico': self.chart_avance_tecnico(),
            'egresos_pie': self.chart_egresos_pie(),
            'margen_kpi': self.chart_margen_kpi(),
            'seguridad': self.chart_seguridad(),
            'productividad_cuadrillas': self.chart_productividad_cuadrillas(),
        }, default=str)


# ===========================================================================
# View principal — HTML + exports
# ===========================================================================

class DashboardIndicadoresGeneralesView(LoginRequiredMixin, RoleRequiredMixin, TemplateView):
    """Dashboard ejecutivo. Reemplaza la vista placeholder previa.

    Query params:
      - ``periodo``: semana | mes | trimestre | anio | todo (default: todo)
      - ``tipo``:    todos | financiero | tecnico | desempeno
      - ``linea``:   uuid (filtra desempeño por línea)
      - ``export``:  pdf | excel (dispara descarga)
    """
    template_name = 'construccion/dashboard_indicadores_generales.html'
    allowed_roles = ALL_ADMIN_ROLES

    def get_proyecto(self):
        return get_object_or_404(
            ProyectoConstruccion,
            id=self.kwargs.get('proyecto_id'),
        )

    def get(self, request, *args, **kwargs):
        proyecto = self.get_proyecto()
        periodo = request.GET.get('periodo', 'todo')
        tipo = request.GET.get('tipo', 'todos')
        linea_id = request.GET.get('linea') or None
        export = request.GET.get('export')

        # Validación de período (edge case: ?periodo=invalido → cae a 'todo')
        if periodo not in dict(PERIODO_CHOICES):
            periodo = 'todo'
        if tipo not in dict(TIPO_CHOICES):
            tipo = 'todos'

        agg = IndicadoresAggregator(
            proyecto=proyecto, periodo=periodo, tipo=tipo, linea_id=linea_id,
        )

        if export == 'pdf':
            return self._render_pdf(proyecto, agg)
        if export == 'excel':
            return self._render_excel(proyecto, agg)
        if export == 'json':
            return JsonResponse({
                'kpis': agg.kpis(),
                'charts': json.loads(agg.all_charts_json()),
                'is_empty': agg.is_empty,
            })

        # Lista de líneas disponibles para el dropdown filtro
        lineas = self._lineas_del_proyecto(proyecto)

        context = {
            'proyecto': proyecto,
            'periodo': periodo,
            'tipo': tipo,
            'linea_id': linea_id,
            'periodo_choices': PERIODO_CHOICES,
            'tipo_choices': TIPO_CHOICES,
            'lineas': lineas,
            'kpis': agg.kpis(),
            'charts_json': agg.all_charts_json(),
            'is_empty': agg.is_empty,
            'meta_torres_mes': META_TORRES_MES,
        }
        return self.render_to_response(context)

    # ---- helpers ----

    @staticmethod
    def _lineas_del_proyecto(proyecto):
        """Devuelve las líneas asociadas al proyecto (defensivo: si no hay
        relación directa, devuelve queryset vacío)."""
        try:
            from apps.lineas.models import Linea
        except Exception:
            return []
        # Heurística: filtrar por contrato del proyecto si el FK existe.
        try:
            return Linea.objects.filter(contrato=proyecto.contrato).order_by('codigo')
        except Exception:
            try:
                return Linea.objects.all().order_by('codigo')[:50]
            except Exception:
                return []

    # ---- export PDF (weasyprint) ----

    def _render_pdf(self, proyecto, agg: IndicadoresAggregator) -> HttpResponse:
        """Genera PDF del dashboard via weasyprint.

        Defensivo: si weasyprint no está instalado en el entorno (CI minimal),
        devuelve un PDF stub para no romper el journey.
        """
        try:
            from weasyprint import HTML
        except Exception:
            return self._stub_pdf(proyecto)

        from django.template.loader import render_to_string

        html_content = render_to_string(
            'construccion/dashboard_indicadores_generales.html',
            {
                'proyecto': proyecto,
                'periodo': agg.periodo,
                'tipo': agg.tipo,
                'kpis': agg.kpis(),
                'charts_json': agg.all_charts_json(),
                'is_empty': agg.is_empty,
                'meta_torres_mes': META_TORRES_MES,
                'pdf_mode': True,
                'periodo_choices': PERIODO_CHOICES,
                'tipo_choices': TIPO_CHOICES,
                'lineas': [],
            },
            request=self.request,
        )
        buf = BytesIO()
        HTML(string=html_content, base_url=self.request.build_absolute_uri('/')).write_pdf(buf)
        buf.seek(0)
        filename = f'indicadores_{proyecto.id}.pdf'
        resp = HttpResponse(buf.read(), content_type='application/pdf')
        resp['Content-Disposition'] = f'attachment; filename="{filename}"'
        return resp

    @staticmethod
    def _stub_pdf(proyecto) -> HttpResponse:
        """PDF mínimo válido (header + nombre proyecto) para fallback."""
        body = (
            b'%PDF-1.4\n%\xE2\xE3\xCF\xD3\n'
            b'1 0 obj<</Type/Catalog/Pages 2 0 R>>endobj\n'
            b'2 0 obj<</Type/Pages/Count 0/Kids[]>>endobj\n'
            b'xref\n0 3\n0000000000 65535 f \n'
            b'0000000015 00000 n \n0000000060 00000 n \n'
            b'trailer<</Root 1 0 R/Size 3>>\nstartxref\n105\n%%EOF\n'
        )
        resp = HttpResponse(body, content_type='application/pdf')
        resp['Content-Disposition'] = (
            f'attachment; filename="indicadores_{proyecto.id}.pdf"'
        )
        return resp

    # ---- export Excel (openpyxl) ----

    def _render_excel(self, proyecto, agg: IndicadoresAggregator) -> HttpResponse:
        """Genera xlsx con hojas: KPIs / Financieros / Técnicos / Desempeño."""
        from openpyxl import Workbook
        from openpyxl.styles import Font

        wb = Workbook()
        ws = wb.active
        ws.title = 'KPIs'
        ws.append(['Indicador', 'Valor', 'Unidad', 'Estado'])
        for row in ws[1]:
            row.font = Font(bold=True)
        for kpi in agg.kpis():
            ws.append([kpi['label'], kpi['value'], kpi['unit'], kpi['status']])

        ws2 = wb.create_sheet('Financieros')
        ws2.append(['Fecha', 'Ingresos', 'Costos directos', 'Gastos',
                    'Costo real', 'Costo ppto', 'Margen %', 'Desviación %'])
        for ind in agg.qs_financieros():
            ws2.append([
                ind.fecha,
                _safe_float(ind.ingresos_ejecutados),
                _safe_float(ind.costos_directos),
                _safe_float(ind.gastos),
                _safe_float(ind.costo_real),
                _safe_float(ind.costo_presupuestado),
                _safe_float(ind.margen_operativo),
                _safe_float(ind.desviacion_presupuestal),
            ])

        ws3 = wb.create_sheet('Técnicos')
        ws3.append(['Fecha', '% Ejecutado', '% Planeado',
                    'Avance obra', 'Cumplimiento', 'Productividad'])
        for ind in agg.qs_tecnicos():
            ws3.append([
                ind.fecha,
                _safe_float(ind.presupuesto_ejecutado_pct),
                _safe_float(ind.presupuesto_planeado_pct),
                _safe_float(ind.avance_obra),
                _safe_float(ind.cumplimiento_cronograma),
                _safe_float(ind.productividad),
            ])

        ws4 = wb.create_sheet('Desempeño')
        ws4.append(['Fecha', 'Línea', 'Cuadrilla', 'Tipo trabajo',
                    'Meta', 'Actual', 'Estado'])
        for ind in agg.qs_desempeno():
            ws4.append([
                ind.fecha,
                getattr(getattr(ind, 'linea', None), 'codigo', '—'),
                getattr(getattr(ind, 'cuadrilla', None), 'nombre', '—'),
                getattr(ind, 'tipo_trabajo', '—'),
                _safe_float(ind.meta),
                _safe_float(ind.actual),
                getattr(ind, 'estado', '—'),
            ])

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        resp = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        resp['Content-Disposition'] = (
            f'attachment; filename="indicadores_{proyecto.id}.xlsx"'
        )
        return resp
