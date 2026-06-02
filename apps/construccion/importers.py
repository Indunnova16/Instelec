"""B5 (#123 Fase 4) — Importadores de Excel del Módulo Financiero de Construcción.

Tres importadores + un detector de formato, que leen archivos .xlsx y escriben
en los modelos B3 (``models_fin``):

1. ``PresupuestoConstruccionExcelImporter``  → ``PresupuestoDetalladoConstruccion.datos``
   (estructura por secciones ingreso/variables/fijos con valores mensuales).
2. ``ContableConstruccionExcelImporter``      → ``PresupuestoDetalladoConstruccion.datos``
   bajo la llave ``finv2_bd`` (agrupa la BD contable por cuenta equivalente,
   reusando ``apps.financiero.importers_finv2.ContableCompleteImporter``).
3. ``CostosConstruccionExcelImporter``        → registros ``CostosConstruccion``
   (una fila de Excel = un costo ejecutado).
4. ``detect_excel_format_construccion(archivo)`` → 'presupuesto' | 'contable' |
   'costos' | None (heurística por nombres de hoja/encabezados).

Reuso (issue #123 Fase 4 "Adaptación de importadores existentes"):
- ``apps.financiero.importers_finv2``: ``_norm`` (normalización de encabezados),
  ``_to_number`` (celda → float seguro), ``MAX_UPLOAD_BYTES``,
  ``ContableCompleteImporter`` (toda la lógica de agrupación contable).

Contrato de retorno (espejo de ``ContableCompleteImporter.procesar_*``): un dict
con ``exito``, ``error`` (❌), ``advertencia`` (⚠️), ``mensaje`` (✅) y datos
auxiliares. Los importadores **no** hacen ``save()`` ellos mismos cuando el
destino es un JSONField (devuelven ``datos`` para que la vista lo persista); el
importador de costos sí persiste filas (devuelve ``creados``/``omitidos``).

Validación de archivo (.xlsx + ≤ 20 MB) centralizada en ``_validar_archivo``.
"""
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

from openpyxl import load_workbook

from apps.financiero.importers_finv2 import (
    MAX_UPLOAD_BYTES,
    ContableCompleteImporter,
    _norm,
    _to_number,
)

from .models_fin import (
    CostosConstruccion,
    PresupuestoDetalladoConstruccion,
)


# Meses reconocidos en encabezados de presupuesto (normalizados, sin acento).
_MESES = (
    'enero', 'febrero', 'marzo', 'abril', 'mayo', 'junio', 'julio',
    'agosto', 'septiembre', 'setiembre', 'octubre', 'noviembre', 'diciembre',
)

# Palabras clave que marcan la sección de una fila de presupuesto.
_KW_INGRESO = ('ingreso', 'ingresos', 'facturacion', 'venta', 'ventas')
_KW_VARIABLES = ('variable', 'variables', 'directo', 'directos', 'costo directo')
_KW_FIJOS = ('fijo', 'fijos', 'gasto', 'gastos', 'indirecto', 'indirectos')

# Encabezados candidatos para el importador de costos.
_COL_CONCEPTO = ('concepto', 'descripcion', 'detalle', 'item')
_COL_TIPO = ('tipo recurso', 'tipo de recurso', 'tipo', 'recurso')
_COL_CANTIDAD = ('cantidad', 'cant', 'qty')
_COL_UNITARIO = ('costo unitario', 'valor unitario', 'precio unitario', 'unitario', 'vr unitario')
_COL_TOTAL = ('costo total', 'valor total', 'total')
_COL_FECHA = ('fecha', 'fecha costo', 'fecha registro')

# Mapa de texto libre → choice de TipoRecurso.
_TIPO_RECURSO_ALIASES = {
    'material': CostosConstruccion.TipoRecurso.MATERIAL,
    'materiales': CostosConstruccion.TipoRecurso.MATERIAL,
    'mano de obra': CostosConstruccion.TipoRecurso.MANO_OBRA,
    'mano obra': CostosConstruccion.TipoRecurso.MANO_OBRA,
    'manodeobra': CostosConstruccion.TipoRecurso.MANO_OBRA,
    'mo': CostosConstruccion.TipoRecurso.MANO_OBRA,
    'equipo': CostosConstruccion.TipoRecurso.EQUIPOS,
    'equipos': CostosConstruccion.TipoRecurso.EQUIPOS,
    'maquinaria': CostosConstruccion.TipoRecurso.EQUIPOS,
    'subcontrata': CostosConstruccion.TipoRecurso.SUBCONTRATA,
    'subcontratos': CostosConstruccion.TipoRecurso.SUBCONTRATA,
    'subcontrato': CostosConstruccion.TipoRecurso.SUBCONTRATA,
    'otro': CostosConstruccion.TipoRecurso.OTROS,
    'otros': CostosConstruccion.TipoRecurso.OTROS,
}


# ===========================================================================
# Helpers compartidos
# ===========================================================================
def _validar_archivo(archivo):
    """Valida extensión .xlsx + tamaño ≤ 20 MB.

    Devuelve ``None`` si es válido, o un dict de error (contrato estándar) si no.
    """
    nombre = (getattr(archivo, 'name', '') or '')
    if not nombre.lower().endswith('.xlsx'):
        return _resultado_error(
            'Archivo inválido. Verifique que sea .xlsx con estructura correcta.'
        )
    tamano = getattr(archivo, 'size', None)
    if tamano is not None and tamano > MAX_UPLOAD_BYTES:
        return _resultado_error(
            'Archivo inválido. El archivo excede el tamaño máximo (20 MB).'
        )
    return None


def _resultado_error(msg):
    return {
        'exito': False, 'error': msg, 'advertencia': None, 'mensaje': None,
        'datos': None, 'filas': 0,
    }


def _resultado_advertencia(msg):
    return {
        'exito': False, 'error': None, 'advertencia': msg, 'mensaje': None,
        'datos': None, 'filas': 0,
    }


def _to_decimal(valor) -> Decimal:
    """Celda → Decimal seguro (nunca lanza)."""
    if isinstance(valor, Decimal):
        return valor
    if valor is None:
        return Decimal('0')
    try:
        return Decimal(str(valor))
    except (InvalidOperation, ValueError, TypeError):
        return Decimal('0')


def _locate_columns(sheet, mapa_candidatos):
    """Detecta índices 1-based de columnas por encabezado normalizado (fila 1).

    ``mapa_candidatos`` es ``{logical_name: (candidato1, candidato2, ...)}``.
    Devuelve ``{logical_name: col_index | None}``.
    """
    headers = {}
    for col in range(1, 40):
        val = sheet.cell(row=1, column=col).value
        if val:
            headers[_norm(val)] = col

    resultado = {}
    for logical, candidatos in mapa_candidatos.items():
        idx = None
        for cand in candidatos:
            if _norm(cand) in headers:
                idx = headers[_norm(cand)]
                break
        resultado[logical] = idx
    return resultado


def _abrir_workbook(archivo):
    """Abre el workbook read-only/data-only o devuelve (None, error_dict)."""
    try:
        wb = load_workbook(archivo, read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        return None, _resultado_error(
            f'Archivo inválido. Verifique que sea .xlsx con estructura '
            f'correcta ({exc}).'
        )
    return wb, None


# ===========================================================================
# 1. PRESUPUESTO (ingreso / variables / fijos por mes) → datos JSON
# ===========================================================================
class PresupuestoConstruccionExcelImporter:
    """Lee un Excel de presupuesto y construye la estructura ``datos``.

    Layout esperado (flexible): la columna A/primera contiene el rubro/concepto
    de la fila; las columnas de meses (enero..diciembre, detectadas por
    encabezado) contienen los valores. Cada fila se clasifica en una sección
    (``ingreso`` / ``variables`` / ``fijos``) según palabras clave del concepto.

    Estructura de salida (consumida por ``ProyectoFinMixin._resumen_presupuesto``
    en views_fin.py y por el template ``_financiero_presupuesto_tabla.html``)::

        {"ingreso":   {"<concepto>": {"<mes>": valor, ...}, ...},
         "variables": {...},
         "fijos":     {...}}
    """

    def __init__(self):
        self.warnings = []

    def _seccion_para(self, concepto_norm):
        for kw in _KW_INGRESO:
            if kw in concepto_norm:
                return 'ingreso'
        for kw in _KW_VARIABLES:
            if kw in concepto_norm:
                return 'variables'
        for kw in _KW_FIJOS:
            if kw in concepto_norm:
                return 'fijos'
        # Por defecto, un costo sin clasificar se trata como variable (directo).
        return 'variables'

    def procesar(self, archivo):
        err = _validar_archivo(archivo)
        if err:
            return err

        wb, err = _abrir_workbook(archivo)
        if err:
            return err

        sheet = wb.active
        # Detectar columnas de meses por encabezado.
        meses_cols = {}  # nombre_mes_norm -> col_index
        concepto_col = 1
        for col in range(1, 40):
            val = sheet.cell(row=1, column=col).value
            if not val:
                continue
            nval = _norm(val)
            if nval in _MESES:
                meses_cols[nval] = col
            elif nval in ('concepto', 'rubro', 'descripcion', 'item', 'detalle'):
                concepto_col = col

        # Edge case 1: ninguna columna de mes detectada → archivo no es de presupuesto.
        if not meses_cols:
            wb.close()
            return _resultado_advertencia(
                'Archivo sin columnas de meses reconocibles (enero..diciembre). '
                'Verifique la estructura del presupuesto.'
            )

        datos = {'ingreso': {}, 'variables': {}, 'fijos': {}}
        filas_validas = 0
        max_col = max(max(meses_cols.values()), concepto_col)

        for row in sheet.iter_rows(min_row=2, max_col=max_col, values_only=True):
            concepto_val = row[concepto_col - 1] if len(row) >= concepto_col else None
            if concepto_val is None:
                continue
            concepto = str(concepto_val).strip()
            if not concepto:
                continue

            seccion = self._seccion_para(_norm(concepto))
            fila_meses = {}
            tiene_valor = False
            for mes_norm, col in meses_cols.items():
                celda = row[col - 1] if len(row) >= col else None
                valor = _to_number(celda)
                if valor:
                    tiene_valor = True
                fila_meses[mes_norm] = round(valor, 2)

            # Edge case 2: fila de concepto sin ningún valor mensual → se omite.
            if not tiene_valor:
                continue

            datos[seccion][concepto] = fila_meses
            filas_validas += 1

        wb.close()

        # Edge case 3: archivo con encabezados de mes pero sin filas con valores.
        if filas_validas == 0:
            return _resultado_advertencia(
                'Archivo sin filas de presupuesto con valores. Revise que los '
                'montos mensuales estén diligenciados.'
            )

        secciones_no_vacias = [s for s, v in datos.items() if v]
        mensaje = (
            f'Presupuesto importado. {filas_validas} conceptos en '
            f'{len(secciones_no_vacias)} secciones ({", ".join(secciones_no_vacias)}).'
        )
        return {
            'exito': True, 'error': None, 'advertencia': None,
            'mensaje': mensaje, 'datos': datos, 'filas': filas_validas,
            'warnings': self.warnings,
        }


# ===========================================================================
# 2. CONTABLE (BD contable agrupada por cuenta) → datos['finv2_bd']
# ===========================================================================
class ContableConstruccionExcelImporter:
    """Adapta ``ContableCompleteImporter`` (#120) al financiero de construcción.

    Delega 100 % de la lógica de lectura/agrupación a
    ``apps.financiero.importers_finv2.ContableCompleteImporter`` (lee hoja 'BD',
    agrupa por columna O sumando Neto de la columna C, mapea a rubros). Solo
    re-empaqueta el resultado al contrato estándar de construcción.

    La salida ``datos`` (``{'finv2_bd': {...}}``) se guarda tal cual en
    ``PresupuestoDetalladoConstruccion.datos`` (no colisiona con las secciones
    ingreso/variables/fijos del importador de presupuesto: viven en llaves
    distintas del mismo JSONField).
    """

    def __init__(self):
        self._inner = ContableCompleteImporter()
        self.warnings = self._inner.warnings

    def procesar(self, archivo, mapeo=None):
        err = _validar_archivo(archivo)
        if err:
            return err
        res = self._inner.procesar_bd_completa(archivo, mapeo=mapeo)
        # ContableCompleteImporter ya retorna exito/error/advertencia/mensaje/datos;
        # normalizamos la llave 'filas' (usa 'cuentas').
        res.setdefault('filas', res.get('cuentas', 0))
        return res


# ===========================================================================
# 3. COSTOS (una fila = un CostosConstruccion) → persiste registros
# ===========================================================================
class CostosConstruccionExcelImporter:
    """Lee un Excel de costos ejecutados y crea registros ``CostosConstruccion``.

    Cada fila con concepto + (cantidad × unitario | total) crea un costo del
    proyecto. ``costo_total`` lo recalcula el ``save()`` del modelo, así que solo
    necesitamos cantidad/unitario; si el Excel trae 'total' pero no cantidad, se
    usa cantidad=1 y unitario=total para preservar el monto.

    A diferencia de los otros dos importadores (que devuelven ``datos`` para que
    la vista persista el JSON), este SÍ persiste filas (es un alta masiva de un
    modelo relacional). Por eso requiere ``proyecto``.
    """

    def __init__(self, proyecto):
        self.proyecto = proyecto
        self.warnings = []

    def _resolver_tipo(self, valor):
        nval = _norm(valor)
        return _TIPO_RECURSO_ALIASES.get(
            nval, CostosConstruccion.TipoRecurso.MATERIAL
        )

    def _resolver_fecha(self, valor):
        if isinstance(valor, datetime):
            return valor.date()
        if isinstance(valor, date):
            return valor
        # Texto YYYY-MM-DD o DD/MM/YYYY → intento simple, fallback hoy.
        if valor:
            txt = str(valor).strip()
            for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%Y/%m/%d'):
                try:
                    return datetime.strptime(txt, fmt).date()
                except ValueError:
                    continue
        return date.today()

    def procesar(self, archivo):
        err = _validar_archivo(archivo)
        if err:
            return err

        wb, err = _abrir_workbook(archivo)
        if err:
            return err

        sheet = wb.active
        cols = _locate_columns(sheet, {
            'concepto': _COL_CONCEPTO,
            'tipo': _COL_TIPO,
            'cantidad': _COL_CANTIDAD,
            'unitario': _COL_UNITARIO,
            'total': _COL_TOTAL,
            'fecha': _COL_FECHA,
        })

        # Edge case 1: sin columna de concepto NI de total → no es archivo de costos.
        if cols['concepto'] is None and cols['total'] is None:
            wb.close()
            return _resultado_advertencia(
                'Archivo sin columnas reconocibles de costos (concepto/total). '
                'Verifique la estructura.'
            )

        concepto_col = cols['concepto'] or 1
        max_col = max(c for c in cols.values() if c) or concepto_col

        nuevos = []
        omitidos = 0
        for row in sheet.iter_rows(min_row=2, max_col=max_col, values_only=True):
            def _cell(key):
                idx = cols[key]
                if idx and len(row) >= idx:
                    return row[idx - 1]
                return None

            concepto_val = row[concepto_col - 1] if len(row) >= concepto_col else None
            concepto = str(concepto_val).strip() if concepto_val is not None else ''

            cantidad = _to_decimal(_cell('cantidad'))
            unitario = _to_decimal(_cell('unitario'))
            total = _to_decimal(_cell('total'))

            # Edge case 2: fila sin concepto y sin ningún monto → se omite.
            if not concepto and cantidad == 0 and unitario == 0 and total == 0:
                continue
            if not concepto:
                concepto = 'Sin concepto'

            # Si solo viene el total, lo expresamos como cantidad 1 × unitario total.
            if (cantidad == 0 or unitario == 0) and total > 0:
                cantidad = Decimal('1')
                unitario = total

            # Edge case 3: fila completamente en cero (ni cantidad ni total) → omitir.
            if cantidad == 0 and unitario == 0:
                omitidos += 1
                continue

            nuevos.append(CostosConstruccion(
                proyecto=self.proyecto,
                concepto=concepto[:300],
                tipo_recurso=self._resolver_tipo(_cell('tipo')),
                cantidad=cantidad,
                costo_unitario=unitario,
                fecha=self._resolver_fecha(_cell('fecha')),
            ))

        wb.close()

        if not nuevos:
            return _resultado_advertencia(
                'Archivo sin filas de costo válidas. Revise que haya conceptos '
                'con cantidad/unitario o un total mayor a cero.'
            )

        # save() de cada instancia recalcula costo_total (no usar bulk_create:
        # bulk_create se salta save() y costo_total quedaría en 0).
        creados = 0
        for obj in nuevos:
            obj.save()
            creados += 1

        mensaje = f'Costos importados. {creados} registros creados.'
        if omitidos:
            self.warnings.append(f'{omitidos} filas en cero se omitieron.')
        return {
            'exito': True, 'error': None, 'advertencia': None,
            'mensaje': mensaje, 'datos': None, 'filas': creados,
            'creados': creados, 'omitidos': omitidos, 'warnings': self.warnings,
        }


# ===========================================================================
# 4. DETECTOR DE FORMATO
# ===========================================================================
def detect_excel_format_construccion(archivo):
    """Heurística: inspecciona hojas/encabezados y devuelve el formato probable.

    Returns: 'contable' | 'presupuesto' | 'costos' | None.

    - 'contable': existe una hoja 'BD'/'base de datos' o un encabezado
      'Cta equivalente'.
    - 'presupuesto': hay ≥ 2 columnas con nombres de mes (enero..diciembre).
    - 'costos': hay encabezados de costo (concepto + tipo/cantidad/total).
    - None: no se reconoce (archivo inválido o ambiguo).
    """
    if _validar_archivo(archivo):
        return None
    wb, err = _abrir_workbook(archivo)
    if err:
        return None

    try:
        # 1. Contable: hoja BD o encabezado 'cta equivalente'.
        nombres_hoja = {_norm(n) for n in wb.sheetnames}
        if nombres_hoja & {'bd', 'base de datos', 'base datos'}:
            return 'contable'

        sheet = wb.active
        headers = set()
        meses_detectados = 0
        for col in range(1, 40):
            val = sheet.cell(row=1, column=col).value
            if not val:
                continue
            nval = _norm(val)
            headers.add(nval)
            if nval in _MESES:
                meses_detectados += 1

        if 'cta equivalente' in headers or any('cta' in h for h in headers):
            return 'contable'
        # 2. Presupuesto: varias columnas de mes.
        if meses_detectados >= 2:
            return 'presupuesto'
        # 3. Costos: concepto + alguna de tipo/cantidad/unitario/total.
        tiene_concepto = bool(headers & {_norm(c) for c in _COL_CONCEPTO})
        tiene_monto = bool(
            headers & {_norm(c) for grupo in (_COL_TIPO, _COL_CANTIDAD,
                                              _COL_UNITARIO, _COL_TOTAL)
                       for c in grupo}
        )
        if tiene_concepto and tiene_monto:
            return 'costos'
        return None
    finally:
        wb.close()
