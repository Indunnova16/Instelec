"""Importador del Excel PDEO (#103).

Lee el workbook PDEO y pueblja:
  - PeriodoFinanciero  (uno por (proyecto, año, mes) presente)
  - MovimientoFinanciero  (uno por (periodo, categoría, tipo PRESUPUESTO|REAL))
  - TransaccionContable  (uno por fila de la hoja BD)

Idempotente: identifica transacciones por (numero_factura, nit_proveedor,
fecha, valor) y omite duplicados — re-correr el mismo Excel no duplica datos.

Las hojas Res EP y pyg son derivadas (cálculo) → no se cargan.
"""
from __future__ import annotations

from decimal import Decimal, InvalidOperation
from typing import Any

import openpyxl

from .models import (
    CategoriaFinanciera,
    MovimientoFinanciero,
    PeriodoFinanciero,
    TransaccionContable,
)


# Mapeo de Auxiliar (código contable) → código de CategoriaFinanciera del seed PDEO.
# El Excel real tiene Auxiliares como 41250501, 42100501, 42950501… y la
# CategoriaFinanciera tiene códigos como 'INGRESOS', 'PERSONAL', 'ADMIN', etc.
# Mapeamos por prefijo de cuenta contable (PUC colombiano):
#   41xx  → INGRESOS
#   42xx  → INGRESOS (otros ingresos)
#   51xx  → ADMIN (gastos administrativos)
#   52xx  → PERSONAL (gastos de personal)
#   53xx  → FINANCIEROS
#   54xx  → IMPUESTOS
#   72xx  → CIF (costos indirectos)
#   73xx  → MATERIALES
PUC_PREFIX_MAP = {
    '41': 'INGRESOS',
    '42': 'INGRESOS',
    '51': 'ADMIN',
    '52': 'PERSONAL',
    '53': 'FINANCIEROS',
    '54': 'IMPUESTOS',
    '72': 'CIF',
    '73': 'MATERIALES',
}


def _to_decimal(v: Any) -> Decimal:
    if v is None or v == '':
        return Decimal('0')
    try:
        return Decimal(str(v).replace(',', '').strip())
    except (InvalidOperation, ValueError):
        return Decimal('0')


def _categoria_para_auxiliar(auxiliar: str | None) -> CategoriaFinanciera | None:
    """Resuelve la CategoriaFinanciera a partir del código auxiliar PUC.

    Si no matchea ningún prefijo, retorna la categoría 'OTROS' (la creamos si no existe).
    """
    if not auxiliar:
        codigo = 'OTROS'
    else:
        prefix = str(auxiliar)[:2]
        codigo = PUC_PREFIX_MAP.get(prefix, 'OTROS')
    cat, _ = CategoriaFinanciera.objects.get_or_create(
        codigo=codigo,
        defaults={
            'nombre': codigo.title(),
            'tipo': 'INGRESO' if codigo == 'INGRESOS' else 'GASTO',
            'orden': 999,
        },
    )
    return cat


def _get_or_create_periodo(proyecto, anio: int, mes: int) -> PeriodoFinanciero:
    p, _ = PeriodoFinanciero.objects.get_or_create(
        proyecto=proyecto, anio=anio, mes=mes)
    return p


def _get_or_create_movimiento(periodo, categoria, tipo: str) -> MovimientoFinanciero:
    """tipo ∈ {'PRESUPUESTO','REAL'}."""
    m, _ = MovimientoFinanciero.objects.get_or_create(
        periodo=periodo, categoria=categoria, tipo=tipo,
        defaults={'valor': Decimal('0')})
    return m


def import_pdeo_workbook(file_obj, proyecto, usuario=None) -> dict:
    """Lee el workbook PDEO y pueblja modelos. Retorna stats."""
    wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
    stats = {
        'transacciones_creadas': 0,
        'transacciones_omitidas': 0,
        'movimientos_actualizados': 0,
        'hojas_procesadas': [],
    }
    if 'BD' in wb.sheetnames:
        _import_hoja_bd(wb['BD'], proyecto, usuario, stats)
    return stats


def _import_hoja_bd(ws, proyecto, usuario, stats):
    """Hoja BD: 23K transacciones con NIT, factura, neto, fecha, periodo."""
    stats['hojas_procesadas'].append('BD')
    # Header en row 1: Auxiliar | Desc. auxiliar | Neto | Fecha | Docto. |
    # Periodo | Tercero movto. | Razón social tercero movto. | Desc. C.O. movto.
    # | Usuario creación | C.O. movto. | Notas | C.Costo | Desc. C.Costo |
    # Cta equivalente | Cargo | Subcontratista | SUBCONTRATA | Q
    rows = ws.iter_rows(min_row=2, values_only=True)

    # Cache de movimientos por (periodo_id, categoria_id, tipo) para evitar
    # queries repetidos en 23K filas.
    mov_cache: dict[tuple, MovimientoFinanciero] = {}
    # Cache de existencia (numero_factura, nit, fecha, valor) en BD para idempotencia.
    existentes = set(
        TransaccionContable.objects.filter(
            movimiento__periodo__proyecto=proyecto
        ).values_list('numero_factura', 'nit_proveedor', 'fecha', 'valor')
    )

    a_crear: list[TransaccionContable] = []
    movimientos_a_actualizar: dict[int, Decimal] = {}

    for row in rows:
        if not row or row[0] is None:
            continue
        auxiliar = row[0]
        # desc_auxiliar = row[1]
        neto = _to_decimal(row[2])
        fecha = row[3]
        docto = (row[4] or '').strip() if isinstance(row[4], str) else (row[4] or '')
        periodo_str = str(row[5]) if row[5] is not None else ''
        tercero_id = (row[6] or '').strip() if isinstance(row[6], str) else (str(row[6]) if row[6] is not None else '')
        razon = (row[7] or '').strip() if isinstance(row[7], str) else (row[7] or '')
        desc_co = (row[8] or '').strip() if isinstance(row[8], str) else (row[8] or '')
        ccosto = (row[12] or '').strip() if isinstance(row[12], str) else (row[12] or '')

        if not fecha or not neto:
            continue
        # fecha puede venir como datetime o date
        from datetime import datetime
        if isinstance(fecha, datetime):
            fecha_d = fecha.date()
        else:
            fecha_d = fecha
        if not (hasattr(fecha_d, 'year') and hasattr(fecha_d, 'month')):
            continue

        # Idempotencia
        key = (str(docto), str(tercero_id), fecha_d, neto)
        if key in existentes:
            stats['transacciones_omitidas'] += 1
            continue
        existentes.add(key)

        # Período: preferir el campo Periodo de la hoja (YYYYMM); si no, derivar de fecha
        if periodo_str and len(periodo_str) == 6 and periodo_str.isdigit():
            anio = int(periodo_str[:4])
            mes = int(periodo_str[4:6])
        else:
            anio = fecha_d.year
            mes = fecha_d.month

        categoria = _categoria_para_auxiliar(auxiliar)
        periodo = _get_or_create_periodo(proyecto, anio, mes)
        cache_key = (periodo.id, categoria.id, 'REAL')
        mov = mov_cache.get(cache_key)
        if mov is None:
            mov = _get_or_create_movimiento(periodo, categoria, 'REAL')
            mov_cache[cache_key] = mov
        movimientos_a_actualizar[mov.id] = (
            movimientos_a_actualizar.get(mov.id, Decimal('0')) + neto)

        a_crear.append(TransaccionContable(
            movimiento=mov,
            fecha=fecha_d,
            descripcion=(str(desc_co) or str(razon) or 'Sin descripción')[:400],
            nit_proveedor=str(tercero_id)[:30],
            nombre_proveedor=str(razon)[:200],
            numero_factura=str(docto)[:50],
            valor=neto,
            iva=Decimal('0'),
            centro_costo=str(ccosto)[:50],
            usuario=usuario,
            notas='',
        ))
        # Bulk create cada 1000 para no aguantar 23K en memoria
        if len(a_crear) >= 1000:
            TransaccionContable.objects.bulk_create(a_crear)
            stats['transacciones_creadas'] += len(a_crear)
            a_crear = []

    if a_crear:
        TransaccionContable.objects.bulk_create(a_crear)
        stats['transacciones_creadas'] += len(a_crear)

    # Aplicar acumulado a MovimientoFinanciero.valor
    for mov_id, suma in movimientos_a_actualizar.items():
        mov = MovimientoFinanciero.objects.get(id=mov_id)
        mov.valor = (mov.valor or Decimal('0')) + suma
        mov.save(update_fields=['valor'])
        stats['movimientos_actualizados'] += 1
