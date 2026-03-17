"""
Views for environmental management.
"""
from django.views.generic import ListView, DetailView, TemplateView, View
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import JsonResponse, HttpResponse
from apps.core.mixins import HTMXMixin, RoleRequiredMixin
from .models import InformeAmbiental, PermisoServidumbre


class InformeListView(LoginRequiredMixin, RoleRequiredMixin, ListView):
    """List environmental reports."""
    model = InformeAmbiental
    template_name = 'ambiental/lista.html'
    context_object_name = 'informes'
    allowed_roles = ['admin', 'director', 'coordinador', 'ing_ambiental']

    def get_queryset(self):
        return super().get_queryset().select_related('linea', 'elaborado_por')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        informes = self.get_queryset()
        context['borradores'] = informes.filter(estado='BORRADOR').count()
        context['aprobados'] = informes.filter(estado='APROBADO').count()
        context['enviados'] = informes.filter(estado='ENVIADO').count()
        return context


class InformeDetailView(LoginRequiredMixin, RoleRequiredMixin, DetailView):
    """Detail view for an environmental report."""
    model = InformeAmbiental
    template_name = 'ambiental/detalle.html'
    context_object_name = 'informe'
    allowed_roles = ['admin', 'director', 'coordinador', 'ing_ambiental']

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Get activities included in this report
        from apps.actividades.models import Actividad
        context['actividades'] = Actividad.objects.filter(
            linea=self.object.linea,
            fecha_programada__year=self.object.periodo_anio,
            fecha_programada__month=self.object.periodo_mes,
            estado='COMPLETADA'
        ).select_related('torre', 'tipo_actividad')
        return context


class GenerarInformeView(LoginRequiredMixin, RoleRequiredMixin, View):
    """Generate PDF/Excel report."""
    allowed_roles = ['admin', 'director', 'coordinador', 'ing_ambiental']

    def post(self, request, pk):
        from .tasks import generar_informe_ambiental

        informe = InformeAmbiental.objects.get(pk=pk)

        # Trigger async generation
        generar_informe_ambiental.delay(str(informe.id))

        return JsonResponse({
            'status': 'processing',
            'message': 'El informe se esta generando. Recibira una notificacion cuando este listo.'
        })


class PermisoListView(LoginRequiredMixin, RoleRequiredMixin, ListView):
    """List easement permissions."""
    model = PermisoServidumbre
    template_name = 'ambiental/permisos.html'
    context_object_name = 'permisos'
    allowed_roles = ['admin', 'director', 'coordinador', 'ing_ambiental']

    def get_queryset(self):
        return super().get_queryset().select_related('torre__linea')


class ConsolidadoView(LoginRequiredMixin, RoleRequiredMixin, TemplateView):
    """Consolidated view of field data for report generation."""
    template_name = 'ambiental/consolidado.html'
    allowed_roles = ['admin', 'director', 'coordinador', 'ing_ambiental']

    def _get_registros_filtrados(self):
        """Get filtered registros based on query params."""
        from apps.campo.models import RegistroCampo

        registros = RegistroCampo.objects.filter(sincronizado=True)

        mes_param = self.request.GET.get('mes')
        anio_param = self.request.GET.get('anio')
        linea = self.request.GET.get('linea')

        if mes_param and anio_param:
            try:
                mes = int(mes_param)
                anio = int(anio_param)
                registros = registros.filter(
                    fecha_inicio__year=anio,
                    fecha_inicio__month=mes
                )
            except (ValueError, TypeError):
                pass

        if linea:
            from uuid import UUID
            try:
                UUID(linea)
                registros = registros.filter(actividad__linea_id=linea)
            except ValueError:
                pass

        return registros

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from django.db.models import Count, Sum
        from apps.lineas.models import Linea

        registros = self._get_registros_filtrados()

        registros_qs = registros.select_related(
            'actividad__linea',
            'actividad__torre',
            'actividad__tipo_actividad',
            'usuario'
        ).prefetch_related('evidencias')

        context['registros'] = registros_qs

        context['stats'] = registros.aggregate(
            total=Count('id'),
        )

        # KPI calculations
        from apps.campo.models import Evidencia
        context['total_evidencias'] = Evidencia.objects.filter(
            registro__in=registros
        ).count()

        context['torres_intervenidas'] = registros.filter(
            actividad__torre__isnull=False
        ).values('actividad__torre').distinct().count()

        context['lineas_intervenidas'] = registros.filter(
            actividad__linea__isnull=False
        ).values('actividad__linea').distinct().count()

        # Especies de vegetacion (unique species from JSON datos_formulario)
        especies = set()
        for reg in registros_qs:
            datos = reg.datos_formulario or {}
            for veg in datos.get('reporte_vegetacion', []):
                especie = veg.get('especie', '').strip()
                if especie:
                    especies.add(especie.lower())
        context['especies_vegetacion'] = len(especies)

        # Porcentaje de avance (actividades completadas vs total para las lineas/periodo filtrados)
        from apps.actividades.models import Actividad
        actividad_filter = {}
        mes_param = self.request.GET.get('mes')
        anio_param = self.request.GET.get('anio')
        linea_param = self.request.GET.get('linea')
        if mes_param and anio_param:
            try:
                actividad_filter['fecha_programada__month'] = int(mes_param)
                actividad_filter['fecha_programada__year'] = int(anio_param)
            except (ValueError, TypeError):
                pass
        if linea_param:
            from uuid import UUID
            try:
                UUID(linea_param)
                actividad_filter['linea_id'] = linea_param
            except ValueError:
                pass

        if actividad_filter:
            total_act = Actividad.objects.filter(**actividad_filter).count()
            completadas_act = Actividad.objects.filter(
                estado='COMPLETADA', **actividad_filter
            ).count()
            context['porcentaje_avance'] = (
                round(completadas_act / total_act * 100)
                if total_act > 0 else 0
            )
        else:
            context['porcentaje_avance'] = 0

        context['lineas'] = Linea.objects.filter(activa=True)
        context['mes_actual'] = self.request.GET.get('mes', '')
        context['anio_actual'] = self.request.GET.get('anio', '')
        context['linea_actual'] = self.request.GET.get('linea', '')

        return context


class ExportarConsolidadoView(LoginRequiredMixin, RoleRequiredMixin, View):
    """Export consolidated environmental report as Excel."""
    allowed_roles = ['admin', 'director', 'coordinador', 'ing_ambiental']

    def get(self, request, *args, **kwargs):
        import io
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from apps.campo.models import RegistroCampo

        # Apply same filters as ConsolidadoView
        registros = RegistroCampo.objects.filter(sincronizado=True)

        mes_param = request.GET.get('mes')
        anio_param = request.GET.get('anio')
        linea_param = request.GET.get('linea')

        if mes_param and anio_param:
            try:
                registros = registros.filter(
                    fecha_inicio__year=int(anio_param),
                    fecha_inicio__month=int(mes_param)
                )
            except (ValueError, TypeError):
                pass

        if linea_param:
            from uuid import UUID
            try:
                UUID(linea_param)
                registros = registros.filter(actividad__linea_id=linea_param)
            except ValueError:
                pass

        registros = registros.select_related(
            'actividad__linea',
            'actividad__torre',
            'actividad__tipo_actividad',
            'usuario'
        ).prefetch_related('evidencias')

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Consolidado Ambiental"

        # Styles
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        center = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Title
        periodo = ''
        if mes_param and anio_param:
            meses = ['', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
                     'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
            try:
                periodo = f'{meses[int(mes_param)]} {anio_param}'
            except (ValueError, IndexError):
                periodo = f'{mes_param}/{anio_param}'

        ws.merge_cells('A1:K1')
        ws['A1'] = f'Informe Ambiental Consolidado{" - " + periodo if periodo else ""}'
        ws['A1'].font = Font(bold=True, size=14)

        ws.merge_cells('A2:K2')
        ws['A2'] = f'Total registros: {registros.count()}'
        ws['A2'].font = Font(size=11)

        # Headers
        headers = [
            'Fecha', 'Linea', 'Vano Desde', 'Vano Hasta',
            'Tipo Actividad', 'Diligenciado Por', 'Trabajo Ejecutado',
            'Propietario', 'Vereda/Municipio',
            'Observaciones', 'Total Evidencias'
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = thin_border

        # Data rows
        row_num = 5
        for registro in registros:
            datos = registro.datos_formulario or {}

            ws.cell(row=row_num, column=1,
                    value=registro.fecha_inicio.strftime('%Y-%m-%d') if registro.fecha_inicio else '').border = thin_border

            linea_nombre = ''
            if registro.actividad and registro.actividad.linea:
                linea_nombre = registro.actividad.linea.codigo
            ws.cell(row=row_num, column=2, value=linea_nombre).border = thin_border

            ws.cell(row=row_num, column=3,
                    value=datos.get('vano_torre_desde', '')).border = thin_border
            ws.cell(row=row_num, column=4,
                    value=datos.get('vano_torre_hasta', '')).border = thin_border

            tipo = ''
            if registro.actividad and registro.actividad.tipo_actividad:
                tipo = registro.actividad.tipo_actividad.nombre
            ws.cell(row=row_num, column=5, value=tipo).border = thin_border

            ws.cell(row=row_num, column=6,
                    value=datos.get('diligenciado_por', '')).border = thin_border
            ws.cell(row=row_num, column=7,
                    value=datos.get('trabajo_ejecutado', '')).border = thin_border

            contacto = datos.get('contacto_permiso', {})
            ws.cell(row=row_num, column=8,
                    value=contacto.get('propietario', '')).border = thin_border

            vereda_municipio = ''
            if contacto.get('vereda') or contacto.get('municipio'):
                vereda_municipio = f"{contacto.get('vereda', '')} / {contacto.get('municipio', '')}"
            ws.cell(row=row_num, column=9, value=vereda_municipio).border = thin_border

            ws.cell(row=row_num, column=10,
                    value=registro.observaciones or '').border = thin_border
            ws.cell(row=row_num, column=11,
                    value=registro.evidencias.count()).border = thin_border

            row_num += 1

        # Vegetation detail sheet
        ws_veg = wb.create_sheet("Detalle Vegetacion")
        veg_headers = ['Fecha', 'Linea', 'Vano', 'Especie', 'Cantidad',
                       'DAP (cm)', 'Altura (m)', 'Tipo Manejo']

        for col, header in enumerate(veg_headers, 1):
            cell = ws_veg.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = thin_border

        veg_row = 2
        for registro in registros:
            datos = registro.datos_formulario or {}
            reporte_veg = datos.get('reporte_vegetacion', [])
            linea_cod = registro.actividad.linea.codigo if registro.actividad and registro.actividad.linea else ''
            vano = f"{datos.get('vano_torre_desde', '')}-{datos.get('vano_torre_hasta', '')}"
            fecha = registro.fecha_inicio.strftime('%Y-%m-%d') if registro.fecha_inicio else ''

            for veg in reporte_veg:
                ws_veg.cell(row=veg_row, column=1, value=fecha).border = thin_border
                ws_veg.cell(row=veg_row, column=2, value=linea_cod).border = thin_border
                ws_veg.cell(row=veg_row, column=3, value=vano).border = thin_border
                ws_veg.cell(row=veg_row, column=4, value=veg.get('especie', '')).border = thin_border
                ws_veg.cell(row=veg_row, column=5, value=veg.get('cantidad', '')).border = thin_border
                ws_veg.cell(row=veg_row, column=6, value=veg.get('dap', '')).border = thin_border
                ws_veg.cell(row=veg_row, column=7, value=veg.get('altura', '')).border = thin_border
                ws_veg.cell(row=veg_row, column=8, value=veg.get('tipo_manejo', '')).border = thin_border
                veg_row += 1

        # Auto-fit column widths
        for sheet in [ws, ws_veg]:
            for col in sheet.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                sheet.column_dimensions[col_letter].width = min(max_len + 2, 40)

        # Write to response
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f'consolidado_ambiental_{periodo.replace(" ", "_") or "todos"}.xlsx'
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
