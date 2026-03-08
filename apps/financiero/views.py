"""
Views for financial management.
"""
from datetime import date
from decimal import Decimal

from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Sum
from django.http import JsonResponse
from django.urls import reverse_lazy
from django.views.generic import CreateView, DetailView, ListView, TemplateView, UpdateView

from apps.core.mixins import HTMXMixin, RoleRequiredMixin

from .models import (
    ArchivoChecklist,
    ArchivoPeriodoFacturacion,
    ChecklistFacturacion,
    CicloFacturacion,
    EjecucionCosto,
    Presupuesto,
    PresupuestoDetallado,
)


def _extract_presupuesto_summary(datos, mes_indices):
    """Extract summary from PresupuestoDetallado datos for given month indices.

    Args:
        datos: the JSON datos from PresupuestoDetallado
        mes_indices: list of 0-based month indices (e.g. [0] for Jan, range(12) for all)

    Returns dict with ingreso, category-level breakdown, totals.
    """
    meses_list = [
        'enero', 'febrero', 'marzo', 'abril', 'mayo', 'junio',
        'julio', 'agosto', 'septiembre', 'octubre', 'noviembre', 'diciembre',
    ]
    selected_months = [meses_list[i] for i in mes_indices]

    # Income
    ing = datos.get('ingreso_proyectado', {})
    ingreso = sum(ing.get(m, 0) for m in selected_months)

    # Cost breakdown by category
    desglose = []
    total_variables = 0
    total_fijos = 0

    for seccion_key in ('costos_variables', 'costos_fijos'):
        seccion_data = datos.get(seccion_key, {})
        estructura = ESTRUCTURA_COSTOS.get(seccion_key, {})
        is_variable = seccion_key == 'costos_variables'

        for cat in estructura.get('categorias', []):
            cat_data = seccion_data.get(cat['codigo'], {})
            cat_total = 0
            for item_name in cat['items']:
                item_data = cat_data.get(item_name, {})
                cat_total += sum(item_data.get(m, 0) for m in selected_months)

            desglose.append({
                'categoria': cat['nombre'],
                'codigo': cat['codigo'],
                'seccion': 'Variable' if is_variable else 'Fijo',
                'valor': cat_total,
            })

            if is_variable:
                total_variables += cat_total
            else:
                total_fijos += cat_total

    total_gastos = total_variables + total_fijos
    resultado = ingreso - total_gastos
    utilidad_pct = ((ingreso - total_gastos) / ingreso * 100) if ingreso else 0

    return {
        'ingreso': ingreso,
        'total_variables': total_variables,
        'total_fijos': total_fijos,
        'total_gastos': total_gastos,
        'resultado': resultado,
        'utilidad_pct': utilidad_pct,
        'desglose': desglose,
    }


class DashboardFinancieroView(LoginRequiredMixin, RoleRequiredMixin, TemplateView):
    """Financial dashboard with year/month filters and Planeado vs Real comparison."""
    template_name = 'financiero/dashboard.html'
    allowed_roles = ['admin', 'director', 'coordinador']

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        import json

        from django.utils import timezone

        from apps.contratos.models import Contrato

        hoy = timezone.now()
        anio = int(self.request.GET.get('anio', hoy.year))
        mes = int(self.request.GET.get('mes', 0))  # 0 = all months
        unidad_filter = self.request.GET.get('unidad', '')

        # Contract filter
        contrato = None
        contrato_id = self.request.GET.get('contrato')
        if contrato_id:
            try:
                contrato = Contrato.objects.get(pk=contrato_id)
            except Contrato.DoesNotExist:
                pass

        contratos_qs = Contrato.objects.all()
        if unidad_filter:
            contratos_qs = contratos_qs.filter(unidad_negocio=unidad_filter)
        context['unidad_filter'] = unidad_filter
        context['contrato_seleccionado'] = contrato
        context['contratos_disponibles'] = contratos_qs
        context['unidades_negocio'] = Contrato.UnidadNegocio.choices

        # Available years and months for filters
        anios_db = list(
            PresupuestoDetallado.objects.values_list('anio', flat=True).distinct()
        )
        anios_disponibles = sorted(set(anios_db) | set(range(anio - 2, anio + 3)))

        meses_nombres = [
            'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
            'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre',
        ]
        meses_cortos = ['Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun',
                        'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic']

        context['anio'] = anio
        context['mes'] = mes
        context['anios_disponibles'] = anios_disponibles
        context['meses_opciones'] = [{'value': 0, 'label': 'Todo el año'}] + [
            {'value': i + 1, 'label': meses_nombres[i]} for i in range(12)
        ]

        # Determine which months to aggregate
        if mes == 0:
            mes_indices = list(range(12))
            periodo_label = f'Año {anio}'
        else:
            mes_indices = [mes - 1]
            periodo_label = f'{meses_nombres[mes - 1]} {anio}'
        context['periodo_label'] = periodo_label

        # Fetch PresupuestoDetallado for both types
        planeado_obj = PresupuestoDetallado.objects.filter(
            anio=anio, tipo='PLANEADO', contrato=contrato,
        ).first()
        real_obj = PresupuestoDetallado.objects.filter(
            anio=anio, tipo='REAL', contrato=contrato,
        ).first()

        datos_planeado = planeado_obj.datos if planeado_obj else _build_empty_datos()
        datos_real = real_obj.datos if real_obj else _build_empty_datos()

        resumen_planeado = _extract_presupuesto_summary(datos_planeado, mes_indices)
        resumen_real = _extract_presupuesto_summary(datos_real, mes_indices)

        context['resumen_planeado'] = resumen_planeado
        context['resumen_real'] = resumen_real

        # Compute differences
        context['diff_ingreso'] = resumen_real['ingreso'] - resumen_planeado['ingreso']
        context['diff_gastos'] = resumen_real['total_gastos'] - resumen_planeado['total_gastos']
        context['diff_resultado'] = resumen_real['resultado'] - resumen_planeado['resultado']

        # Build comparison detail table
        detalle_comparacion = []
        desglose_p = {d['codigo'] + d['seccion']: d for d in resumen_planeado['desglose']}
        for item_r in resumen_real['desglose']:
            key = item_r['codigo'] + item_r['seccion']
            item_p = desglose_p.get(key, {'valor': 0})
            planeado_val = item_p['valor']
            real_val = item_r['valor']
            diff = real_val - planeado_val
            pct_desviacion = ((diff / planeado_val) * 100) if planeado_val else 0
            detalle_comparacion.append({
                'categoria': item_r['categoria'],
                'seccion': item_r['seccion'],
                'planeado': planeado_val,
                'real': real_val,
                'diferencia': diff,
                'pct_desviacion': pct_desviacion,
            })

        # Add income row at the top
        diff_ing = resumen_real['ingreso'] - resumen_planeado['ingreso']
        pct_ing = ((diff_ing / resumen_planeado['ingreso']) * 100) if resumen_planeado['ingreso'] else 0
        detalle_comparacion.insert(0, {
            'categoria': 'Ingreso Proyectado',
            'seccion': 'Ingreso',
            'planeado': resumen_planeado['ingreso'],
            'real': resumen_real['ingreso'],
            'diferencia': diff_ing,
            'pct_desviacion': pct_ing,
        })

        context['detalle_comparacion'] = detalle_comparacion

        # Chart data - monthly trend for the selected year (Planeado vs Real gastos)
        planeado_mensual = []
        real_mensual = []
        for i in range(12):
            p_sum = _extract_presupuesto_summary(datos_planeado, [i])
            r_sum = _extract_presupuesto_summary(datos_real, [i])
            planeado_mensual.append(p_sum['total_gastos'])
            real_mensual.append(r_sum['total_gastos'])

        context['meses_labels'] = json.dumps(meses_cortos)
        context['planeado_mensual'] = json.dumps(planeado_mensual)
        context['real_mensual'] = json.dumps(real_mensual)

        # Chart data - income monthly trend
        ingreso_planeado_mensual = []
        ingreso_real_mensual = []
        for i in range(12):
            ing_p = datos_planeado.get('ingreso_proyectado', {}).get(MESES[i], 0)
            ing_r = datos_real.get('ingreso_proyectado', {}).get(MESES[i], 0)
            ingreso_planeado_mensual.append(ing_p)
            ingreso_real_mensual.append(ing_r)

        context['ingreso_planeado_mensual'] = json.dumps(ingreso_planeado_mensual)
        context['ingreso_real_mensual'] = json.dumps(ingreso_real_mensual)

        # Chart data - category comparison (bar chart)
        cat_labels = []
        cat_planeado = []
        cat_real = []
        for item in detalle_comparacion:
            if item['seccion'] != 'Ingreso' and (item['planeado'] > 0 or item['real'] > 0):
                cat_labels.append(item['categoria'])
                cat_planeado.append(item['planeado'])
                cat_real.append(item['real'])

        context['cat_labels'] = json.dumps(cat_labels)
        context['cat_planeado'] = json.dumps(cat_planeado)
        context['cat_real'] = json.dumps(cat_real)

        # Alerts
        alertas = []
        if resumen_real['total_gastos'] > resumen_planeado['total_gastos'] and resumen_planeado['total_gastos'] > 0:
            pct_exceso = ((resumen_real['total_gastos'] - resumen_planeado['total_gastos'])
                          / resumen_planeado['total_gastos'] * 100)
            if pct_exceso > 10:
                alertas.append({
                    'tipo': 'danger',
                    'titulo': 'Gastos reales superan lo planeado',
                    'mensaje': f'Los gastos reales superan el presupuesto planeado en {pct_exceso:.1f}%',
                    'color': 'red',
                })
            elif pct_exceso > 0:
                alertas.append({
                    'tipo': 'warning',
                    'titulo': 'Gastos reales por encima de lo planeado',
                    'mensaje': f'Los gastos reales exceden lo planeado en {pct_exceso:.1f}%',
                    'color': 'yellow',
                })

        if resumen_real['ingreso'] < resumen_planeado['ingreso'] and resumen_planeado['ingreso'] > 0:
            pct_deficit = ((resumen_planeado['ingreso'] - resumen_real['ingreso'])
                           / resumen_planeado['ingreso'] * 100)
            if pct_deficit > 20:
                alertas.append({
                    'tipo': 'warning',
                    'titulo': 'Ingreso real por debajo de lo planeado',
                    'mensaje': f'El ingreso real es {pct_deficit:.1f}% menor que lo planeado',
                    'color': 'yellow',
                })

        if resumen_real['resultado'] < 0 and resumen_planeado['resultado'] >= 0:
            alertas.append({
                'tipo': 'danger',
                'titulo': 'Resultado negativo',
                'mensaje': 'El resultado real es negativo mientras el planeado era positivo',
                'color': 'red',
            })

        context['alertas'] = alertas

        return context


class ExportarDashboardExcelView(LoginRequiredMixin, RoleRequiredMixin, TemplateView):
    """Export the financial dashboard detail as an Excel file."""
    template_name = 'financiero/dashboard.html'  # fallback, never rendered
    allowed_roles = ['admin', 'director', 'coordinador']

    def get(self, request, *args, **kwargs):
        import io
        from django.http import HttpResponse
        from django.utils import timezone
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter

        hoy = timezone.now()
        periodo = request.GET.get('periodo', 'mes')

        # Reuse dashboard logic for period
        if periodo == 'trimestre':
            q = (hoy.month - 1) // 3
            mes_inicio = q * 3 + 1
            mes_fin = mes_inicio + 2
        elif periodo == 'anio':
            mes_inicio, mes_fin = 1, 12
        else:
            mes_inicio = mes_fin = hoy.month

        anio = hoy.year
        presupuestos = Presupuesto.objects.filter(
            anio=anio, mes__gte=mes_inicio, mes__lte=mes_fin,
        )

        pillar_aggs = presupuestos.aggregate(
            personal=Sum('costo_dias_hombre'),
            vehiculos=Sum('costo_vehiculos'),
            viaticos=Sum('viaticos_planeados'),
            herramientas=Sum('costo_herramientas'),
            ambientales=Sum('costo_ambientales'),
            subcontratistas=Sum('costo_subcontratistas'),
            transporte=Sum('costo_transporte'),
            materiales=Sum('costo_materiales'),
            garantia=Sum('costo_garantia'),
            otros=Sum('otros_costos'),
        )

        rows = [
            ('Mano de Obra', 'Días hombre', pillar_aggs['personal']),
            ('Vehículos', 'Equipos y vehículos', pillar_aggs['vehiculos']),
            ('Viáticos', 'Viáticos del personal', pillar_aggs['viaticos']),
            ('Herramientas', 'Herramientas y equipos menores', pillar_aggs['herramientas']),
            ('Ambientales', 'Gestión ambiental y permisos', pillar_aggs['ambientales']),
            ('Subcontratistas', 'Subcontratistas y terceros', pillar_aggs['subcontratistas']),
            ('Transporte', 'Transporte adicional', pillar_aggs['transporte']),
            ('Materiales', 'Materiales e insumos', pillar_aggs['materiales']),
            ('Garantía', 'Garantías y pólizas', pillar_aggs['garantia']),
            ('Otros', 'Otros costos', pillar_aggs['otros']),
        ]
        rows = [(cat, conc, float(val or 0)) for cat, conc, val in rows if val]

        total_pres = float(presupuestos.aggregate(t=Sum('total_presupuestado'))['t'] or 0)
        total_ejec = float(presupuestos.aggregate(t=Sum('total_ejecutado'))['t'] or 0)

        wb = Workbook()
        ws = wb.active
        ws.title = 'Detalle Costos'

        # Styles
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        total_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
        total_font = Font(bold=True, size=11)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )
        money_fmt = '#,##0'
        pct_fmt = '0.0"%"'

        # Title
        meses_nombres = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
                         'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
        if mes_inicio == mes_fin:
            titulo_periodo = f'{meses_nombres[mes_inicio - 1]} {anio}'
        else:
            titulo_periodo = f'{meses_nombres[mes_inicio - 1]} - {meses_nombres[mes_fin - 1]} {anio}'

        ws.merge_cells('A1:F1')
        title_cell = ws['A1']
        title_cell.value = f'Detalle de Ejecucion Presupuestal - {titulo_periodo}'
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center')

        ws.merge_cells('A2:F2')
        ws['A2'].value = f'Generado: {hoy.strftime("%d/%m/%Y %H:%M")}'
        ws['A2'].alignment = Alignment(horizontal='center')
        ws['A2'].font = Font(italic=True, color='666666')

        # Headers (row 4)
        headers = ['Categoría', 'Concepto', 'Presupuesto ($)', 'Ejecutado ($)', '% Ejecución', 'Disponible ($)']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='right' if col > 2 else 'left')

        # Data rows
        for i, (cat, conc, val) in enumerate(rows, 5):
            pres = val
            ejec = val
            pct = 100.0 if pres > 0 else 0.0
            disp = pres - ejec

            ws.cell(row=i, column=1, value=cat).border = thin_border
            ws.cell(row=i, column=2, value=conc).border = thin_border
            c = ws.cell(row=i, column=3, value=pres)
            c.number_format = money_fmt
            c.border = thin_border
            c.alignment = Alignment(horizontal='right')
            c = ws.cell(row=i, column=4, value=ejec)
            c.number_format = money_fmt
            c.border = thin_border
            c.alignment = Alignment(horizontal='right')
            c = ws.cell(row=i, column=5, value=pct)
            c.number_format = pct_fmt
            c.border = thin_border
            c.alignment = Alignment(horizontal='right')
            c = ws.cell(row=i, column=6, value=disp)
            c.number_format = money_fmt
            c.border = thin_border
            c.alignment = Alignment(horizontal='right')

        # Total row
        total_row = 5 + len(rows)
        pct_total = (total_ejec / total_pres * 100) if total_pres > 0 else 0

        for col in range(1, 7):
            cell = ws.cell(row=total_row, column=col)
            cell.fill = total_fill
            cell.font = total_font
            cell.border = thin_border

        ws.cell(row=total_row, column=1, value='TOTAL')
        ws.cell(row=total_row, column=2, value='')
        c = ws.cell(row=total_row, column=3, value=total_pres)
        c.number_format = money_fmt
        c.alignment = Alignment(horizontal='right')
        c = ws.cell(row=total_row, column=4, value=total_ejec)
        c.number_format = money_fmt
        c.alignment = Alignment(horizontal='right')
        c = ws.cell(row=total_row, column=5, value=pct_total)
        c.number_format = pct_fmt
        c.alignment = Alignment(horizontal='right')
        c = ws.cell(row=total_row, column=6, value=total_pres - total_ejec)
        c.number_format = money_fmt
        c.alignment = Alignment(horizontal='right')

        # ---- Sheet 2: Detalle por Linea ----
        ws2 = wb.create_sheet('Por Linea')
        from apps.lineas.models import Linea
        lineas = Linea.objects.filter(activa=True)

        ws2.merge_cells('A1:G1')
        ws2['A1'].value = f'Presupuesto vs Ejecutado por Línea - {titulo_periodo}'
        ws2['A1'].font = Font(bold=True, size=14)
        ws2['A1'].alignment = Alignment(horizontal='center')

        headers2 = ['Línea', 'Código', 'Presupuestado ($)', 'Ejecutado ($)', '% Ejecución', 'Disponible ($)', 'Estado']
        for col, h in enumerate(headers2, 1):
            cell = ws2.cell(row=3, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border

        row_num = 4
        for linea in lineas:
            pres_linea = Presupuesto.objects.filter(
                linea=linea, anio=anio,
                mes__gte=mes_inicio, mes__lte=mes_fin,
            )
            pres_val = float(pres_linea.aggregate(t=Sum('total_presupuestado'))['t'] or 0)
            ejec_val = float(pres_linea.aggregate(t=Sum('total_ejecutado'))['t'] or 0)
            if pres_val == 0 and ejec_val == 0:
                continue
            pct = (ejec_val / pres_val * 100) if pres_val > 0 else 0
            estado = 'OK' if pct <= 90 else ('Alerta' if pct <= 100 else 'Sobrecosto')

            ws2.cell(row=row_num, column=1, value=linea.nombre).border = thin_border
            ws2.cell(row=row_num, column=2, value=linea.codigo).border = thin_border
            c = ws2.cell(row=row_num, column=3, value=pres_val)
            c.number_format = money_fmt
            c.border = thin_border
            c = ws2.cell(row=row_num, column=4, value=ejec_val)
            c.number_format = money_fmt
            c.border = thin_border
            c = ws2.cell(row=row_num, column=5, value=pct)
            c.number_format = pct_fmt
            c.border = thin_border
            c = ws2.cell(row=row_num, column=6, value=pres_val - ejec_val)
            c.number_format = money_fmt
            c.border = thin_border
            ws2.cell(row=row_num, column=7, value=estado).border = thin_border
            row_num += 1

        # Auto-width for both sheets
        for sheet in [ws, ws2]:
            for col_idx in range(1, sheet.max_column + 1):
                max_len = 0
                col_letter = get_column_letter(col_idx)
                for row in sheet.iter_rows(min_col=col_idx, max_col=col_idx, values_only=False):
                    for cell in row:
                        if cell.value:
                            max_len = max(max_len, len(str(cell.value)))
                sheet.column_dimensions[col_letter].width = min(max_len + 4, 40)

        # Write to response
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f'detalle_costos_{anio}_{mes_inicio:02d}.xlsx'
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


class PresupuestoListView(LoginRequiredMixin, RoleRequiredMixin, ListView):
    """List budgets."""
    model = Presupuesto
    template_name = 'financiero/presupuestos.html'
    context_object_name = 'presupuestos'
    allowed_roles = ['admin', 'director', 'coordinador']

    def get_queryset(self):
        return super().get_queryset().select_related('linea')


class PresupuestoDetailView(LoginRequiredMixin, RoleRequiredMixin, DetailView):
    """Budget detail view."""
    model = Presupuesto
    template_name = 'financiero/presupuesto_detalle.html'
    context_object_name = 'presupuesto'
    allowed_roles = ['admin', 'director', 'coordinador']

    def get_queryset(self):
        return super().get_queryset().select_related('linea').prefetch_related(
            'ejecuciones',
            'ejecuciones__actividad',
            'ejecuciones__actividad__torre',
            'ejecuciones__actividad__tipo_actividad'
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # ejecuciones already prefetched via get_queryset
        context['ejecuciones'] = self.object.ejecuciones.all()
        return context


class PresupuestoCreateView(LoginRequiredMixin, RoleRequiredMixin, CreateView):
    """Create a new budget."""
    model = Presupuesto
    template_name = 'financiero/presupuesto_form.html'
    allowed_roles = ['admin', 'director', 'coordinador']
    success_url = reverse_lazy('financiero:dashboard')

    def get_form_class(self):
        from .forms import PresupuestoForm
        return PresupuestoForm


class PresupuestoUpdateView(LoginRequiredMixin, RoleRequiredMixin, UpdateView):
    """Edit an existing budget."""
    model = Presupuesto
    template_name = 'financiero/presupuesto_form.html'
    allowed_roles = ['admin', 'director', 'coordinador']

    def get_form_class(self):
        from .forms import PresupuestoForm
        return PresupuestoForm

    def get_success_url(self):
        return reverse_lazy('financiero:presupuesto_detalle', kwargs={'pk': self.object.pk})


class CuadroCostosView(LoginRequiredMixin, RoleRequiredMixin, TemplateView):
    """Generate billing cost table."""
    template_name = 'financiero/cuadro_costos.html'
    allowed_roles = ['admin', 'director', 'coordinador']

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        mes = self.request.GET.get('mes')
        anio = self.request.GET.get('anio')
        linea = self.request.GET.get('linea')

        if mes and anio and linea:
            ejecuciones = EjecucionCosto.objects.filter(
                presupuesto__mes=mes,
                presupuesto__anio=anio,
                presupuesto__linea_id=linea
            ).select_related('actividad__torre', 'actividad__tipo_actividad')

            context['ejecuciones'] = ejecuciones
            context['total'] = ejecuciones.aggregate(total=Sum('costo_total'))['total'] or 0

        return context


class FacturacionView(LoginRequiredMixin, RoleRequiredMixin, ListView):
    """Billing cycles view."""
    model = CicloFacturacion
    template_name = 'financiero/facturacion.html'
    context_object_name = 'ciclos'
    allowed_roles = ['admin', 'director', 'coordinador']

    def get_queryset(self):
        return super().get_queryset().select_related('presupuesto__linea')


class CostosCuadrillaView(LoginRequiredMixin, RoleRequiredMixin, HTMXMixin, TemplateView):
    """View for crew costs filtered by day or week."""
    template_name = 'financiero/costos_cuadrilla.html'
    partial_template_name = 'financiero/partials/costos_cuadrilla_tabla.html'
    allowed_roles = ['admin', 'director', 'coordinador', 'ing_residente']

    def get_context_data(self, **kwargs):
        import json
        from collections import OrderedDict

        from apps.cuadrillas.models import Cuadrilla, CuadrillaMiembro

        context = super().get_context_data(**kwargs)

        filtro = self.request.GET.get('filtro', 'semana')  # 'dia' or 'semana'
        semana_param = self.request.GET.get('semana', '').strip()
        fecha_param = self.request.GET.get('fecha', '').strip()

        context['filtro'] = filtro
        context['semana_param'] = semana_param
        context['fecha_param'] = fecha_param

        cuadrillas_data = []
        gran_total_personal = Decimal('0')
        gran_total_vehiculo = Decimal('0')
        gran_total = Decimal('0')

        if filtro == 'dia' and fecha_param:
            # Filter cuadrillas by fecha field
            try:
                fecha_filtro = date.fromisoformat(fecha_param)
            except ValueError:
                fecha_filtro = None

            if fecha_filtro:
                cuadrillas = Cuadrilla.objects.filter(
                    activa=True, fecha=fecha_filtro
                ).select_related('supervisor', 'vehiculo').prefetch_related(
                    'miembros__usuario'
                )

                for cuadrilla in cuadrillas:
                    miembros = cuadrilla.miembros.filter(activo=True).select_related('usuario')
                    costo_personal = sum((m.costo_dia for m in miembros), Decimal('0'))
                    costo_vehiculo = cuadrilla.vehiculo.costo_dia if cuadrilla.vehiculo else Decimal('0')
                    total = costo_personal + costo_vehiculo

                    gran_total_personal += costo_personal
                    gran_total_vehiculo += costo_vehiculo
                    gran_total += total

                    miembros_list = [{
                        'nombre': m.usuario.get_full_name(),
                        'rol': m.get_rol_cuadrilla_display(),
                        'cargo': m.get_cargo_display(),
                        'costo_dia': m.costo_dia,
                    } for m in miembros]

                    cuadrillas_data.append({
                        'cuadrilla': cuadrilla,
                        'miembros': miembros_list,
                        'costo_personal': costo_personal,
                        'costo_vehiculo': costo_vehiculo,
                        'total': total,
                    })

        elif filtro == 'semana' and semana_param:
            # Filter by week code prefix (WW-YYYY)
            try:
                parts = semana_param.split('-')
                sem = parts[0].zfill(2)
                ano = parts[1]
                prefix = f'{sem}-{ano}-'
            except (IndexError, ValueError):
                prefix = None

            if prefix:
                cuadrillas = Cuadrilla.objects.filter(
                    activa=True, codigo__startswith=prefix
                ).select_related('supervisor', 'vehiculo').prefetch_related(
                    'miembros__usuario'
                )

                for cuadrilla in cuadrillas:
                    miembros = cuadrilla.miembros.filter(activo=True).select_related('usuario')
                    costo_personal = sum((m.costo_dia for m in miembros), Decimal('0'))
                    costo_vehiculo = cuadrilla.vehiculo.costo_dia if cuadrilla.vehiculo else Decimal('0')
                    total = costo_personal + costo_vehiculo

                    gran_total_personal += costo_personal
                    gran_total_vehiculo += costo_vehiculo
                    gran_total += total

                    miembros_list = [{
                        'nombre': m.usuario.get_full_name(),
                        'rol': m.get_rol_cuadrilla_display(),
                        'cargo': m.get_cargo_display(),
                        'costo_dia': m.costo_dia,
                    } for m in miembros]

                    cuadrillas_data.append({
                        'cuadrilla': cuadrilla,
                        'miembros': miembros_list,
                        'costo_personal': costo_personal,
                        'costo_vehiculo': costo_vehiculo,
                        'total': total,
                    })

        context['cuadrillas_data'] = cuadrillas_data
        context['gran_total_personal'] = gran_total_personal
        context['gran_total_vehiculo'] = gran_total_vehiculo
        context['gran_total'] = gran_total

        # Build list of available weeks for the filter
        todas = Cuadrilla.objects.filter(activa=True).values_list('codigo', flat=True)
        semanas_set = set()
        for codigo in todas:
            try:
                parts = codigo.split('-')
                if len(parts) >= 2:
                    semana = int(parts[0])
                    ano = int(parts[1])
                    if 1 <= semana <= 53 and 2000 <= ano <= 2100:
                        semanas_set.add((ano, semana))
            except (ValueError, IndexError):
                pass

        semanas_disponibles = sorted(semanas_set, reverse=True)
        context['semanas_disponibles'] = [
            {'value': f'{s[1]}-{s[0]}', 'label': f'Semana {s[1]} - {s[0]}'}
            for s in semanas_disponibles
        ]

        return context



class ChecklistFacturacionView(LoginRequiredMixin, RoleRequiredMixin, HTMXMixin, TemplateView):
    """Checklist for tracking billing status of completed activities."""
    template_name = 'financiero/checklist_facturacion.html'
    allowed_roles = ['admin', 'director', 'coordinador']

    def get_context_data(self, **kwargs):
        from django.utils import timezone

        from apps.actividades.models import Actividad
        from apps.lineas.models import Linea

        context = super().get_context_data(**kwargs)

        hoy = timezone.now()
        mes = int(self.request.GET.get('mes', hoy.month))
        anio = int(self.request.GET.get('anio', hoy.year))
        linea_id = self.request.GET.get('linea', '')

        # Get completed activities for the selected month
        qs = Actividad.objects.filter(
            estado='COMPLETADA',
            fecha_programada__year=anio,
            fecha_programada__month=mes,
        ).select_related('linea', 'tipo_actividad', 'cuadrilla')

        if linea_id:
            qs = qs.filter(linea_id=linea_id)

        # Build checklist items (create if not exist)
        items = []
        total_facturado = 0
        total_pendiente = 0
        monto_total = Decimal('0')

        for actividad in qs:
            checklist, _ = ChecklistFacturacion.objects.get_or_create(
                actividad=actividad,
                defaults={'facturado': False}
            )
            monto = getattr(actividad, 'valor_facturacion', Decimal('0')) or Decimal('0')
            monto_total += monto

            if checklist.facturado:
                total_facturado += 1
            else:
                total_pendiente += 1

            items.append({
                'actividad': actividad,
                'checklist': checklist,
                'monto': monto,
            })

        context['items'] = items
        context['total_actividades'] = len(items)
        context['total_facturado'] = total_facturado
        context['total_pendiente'] = total_pendiente
        context['monto_total'] = monto_total

        # Period-level files
        archivos_periodo_qs = ArchivoPeriodoFacturacion.objects.filter(
            anio=anio, mes=mes
        )
        if linea_id:
            archivos_periodo_qs = archivos_periodo_qs.filter(linea_id=linea_id)
        context['archivos_periodo'] = archivos_periodo_qs

        from .forms import ArchivoPeriodoForm
        context['periodo_upload_form'] = ArchivoPeriodoForm()

        # Filters
        context['lineas'] = Linea.objects.filter(activa=True)
        context['filtro_mes'] = mes
        context['filtro_anio'] = anio
        context['filtro_linea'] = linea_id
        context['meses'] = [
            (1, 'Enero'), (2, 'Febrero'), (3, 'Marzo'), (4, 'Abril'),
            (5, 'Mayo'), (6, 'Junio'), (7, 'Julio'), (8, 'Agosto'),
            (9, 'Septiembre'), (10, 'Octubre'), (11, 'Noviembre'), (12, 'Diciembre'),
        ]
        context['anios'] = list(range(hoy.year - 1, hoy.year + 2))

        return context


class ToggleFacturadoView(LoginRequiredMixin, RoleRequiredMixin, DetailView):
    """Toggle billing status of a checklist item via HTMX."""
    model = ChecklistFacturacion
    allowed_roles = ['admin', 'director', 'coordinador']

    def post(self, request, *args, **kwargs):
        from django.http import HttpResponse

        checklist = self.get_object()
        checklist.facturado = not checklist.facturado
        if checklist.facturado:
            checklist.fecha_facturacion = date.today()
        else:
            checklist.fecha_facturacion = None
        checklist.save(update_fields=['facturado', 'fecha_facturacion', 'updated_at'])

        if checklist.facturado:
            html = (
                f'<button hx-post="/financiero/checklist-facturacion/{checklist.pk}/toggle/" '
                f'hx-target="closest .checklist-toggle" hx-swap="innerHTML" '
                f'class="px-3 py-1 bg-green-100 text-green-800 rounded-full text-xs font-medium hover:bg-green-200 transition">'
                f'Facturado</button>'
            )
        else:
            html = (
                f'<button hx-post="/financiero/checklist-facturacion/{checklist.pk}/toggle/" '
                f'hx-target="closest .checklist-toggle" hx-swap="innerHTML" '
                f'class="px-3 py-1 bg-red-100 text-red-800 rounded-full text-xs font-medium hover:bg-red-200 transition">'
                f'Pendiente</button>'
            )
        return HttpResponse(html)


class ChecklistDetallePartialView(LoginRequiredMixin, RoleRequiredMixin, DetailView):
    """HTMX partial: expanded detail for a checklist item."""
    model = ChecklistFacturacion
    template_name = 'financiero/partials/checklist_detalle.html'
    context_object_name = 'checklist'
    allowed_roles = ['admin', 'director', 'coordinador']

    def get_queryset(self):
        return super().get_queryset().select_related(
            'actividad__linea', 'actividad__tipo_actividad',
            'actividad__cuadrilla',
        ).prefetch_related('archivos')

    def get_context_data(self, **kwargs):
        from .forms import ChecklistEditForm

        context = super().get_context_data(**kwargs)
        context['edit_form'] = ChecklistEditForm(instance=self.object)
        context['archivos'] = self.object.archivos.all()
        return context


class ChecklistEditarView(LoginRequiredMixin, RoleRequiredMixin, DetailView):
    """Save numero_factura and observaciones via HTMX POST."""
    model = ChecklistFacturacion
    allowed_roles = ['admin', 'director', 'coordinador']

    def post(self, request, *args, **kwargs):
        import json

        from django.http import HttpResponse
        from django.shortcuts import render

        from .forms import ChecklistEditForm

        checklist = self.get_object()
        form = ChecklistEditForm(request.POST, instance=checklist)
        if form.is_valid():
            form.save()
            response = HttpResponse()
            response['HX-Trigger'] = json.dumps({
                'showToast': {'message': 'Datos guardados', 'type': 'success'},
            })
            # Re-render the edit form with saved data
            return render(request, 'financiero/partials/checklist_edit_form.html', {
                'edit_form': ChecklistEditForm(instance=checklist),
                'checklist': checklist,
            }, headers={'HX-Trigger': json.dumps({
                'showToast': {'message': 'Datos guardados', 'type': 'success'},
            })})
        return render(request, 'financiero/partials/checklist_edit_form.html', {
            'edit_form': form, 'checklist': checklist,
        })


class ChecklistSubirArchivoView(LoginRequiredMixin, RoleRequiredMixin, DetailView):
    """Upload files to a checklist item via HTMX POST."""
    model = ChecklistFacturacion
    allowed_roles = ['admin', 'director', 'coordinador']

    ALLOWED_EXTENSIONS = {'.pdf', '.xlsx', '.xls', '.jpg', '.jpeg', '.png', '.webp'}
    MAX_FILE_SIZE = 20 * 1024 * 1024  # 20 MB

    def post(self, request, *args, **kwargs):
        import json
        import os

        from django.http import HttpResponse
        from django.template.loader import render_to_string

        checklist = self.get_object()
        files = request.FILES.getlist('archivos')

        if not files:
            response = HttpResponse(status=400)
            response['HX-Trigger'] = json.dumps({
                'showToast': {'message': 'No se seleccionaron archivos', 'type': 'error'},
            })
            return response

        uploaded = 0
        for f in files:
            _, ext = os.path.splitext(f.name)
            if ext.lower() not in self.ALLOWED_EXTENSIONS:
                continue
            if f.size > self.MAX_FILE_SIZE:
                continue
            ArchivoChecklist.objects.create(
                checklist=checklist,
                archivo=f,
                nombre_original=f.name,
                tipo_archivo=f.content_type or '',
                tamanio=f.size,
            )
            uploaded += 1

        archivos = checklist.archivos.all()
        html = render_to_string('financiero/partials/checklist_archivos.html', {
            'archivos': archivos, 'checklist': checklist,
        }, request=request)

        response = HttpResponse(html)
        response['HX-Trigger'] = json.dumps({
            'showToast': {'message': f'{uploaded} archivo(s) subido(s)', 'type': 'success'},
        })
        return response


class ChecklistEliminarArchivoView(LoginRequiredMixin, RoleRequiredMixin, DetailView):
    """Delete a file from a checklist item via HTMX DELETE."""
    model = ArchivoChecklist
    allowed_roles = ['admin', 'director', 'coordinador']

    def delete(self, request, *args, **kwargs):
        import json

        from django.http import HttpResponse
        from django.template.loader import render_to_string

        archivo = self.get_object()
        checklist = archivo.checklist
        archivo.archivo.delete(save=False)
        archivo.delete()

        archivos = checklist.archivos.all()
        html = render_to_string('financiero/partials/checklist_archivos.html', {
            'archivos': archivos, 'checklist': checklist,
        }, request=request)

        response = HttpResponse(html)
        response['HX-Trigger'] = json.dumps({
            'showToast': {'message': 'Archivo eliminado', 'type': 'success'},
        })
        return response


class PeriodoSubirArchivoView(LoginRequiredMixin, RoleRequiredMixin, TemplateView):
    """Upload general files for a billing period."""
    template_name = 'financiero/checklist_facturacion.html'
    allowed_roles = ['admin', 'director', 'coordinador']

    ALLOWED_EXTENSIONS = {'.pdf', '.xlsx', '.xls', '.jpg', '.jpeg', '.png', '.webp'}
    MAX_FILE_SIZE = 50 * 1024 * 1024  # 50 MB

    def post(self, request, *args, **kwargs):
        import json
        import os

        from django.http import HttpResponse
        from django.template.loader import render_to_string

        mes = int(request.POST.get('mes', 0))
        anio = int(request.POST.get('anio', 0))
        linea_id = request.POST.get('linea') or None
        descripcion = request.POST.get('descripcion', '')
        files = request.FILES.getlist('archivos')

        if not files or not mes or not anio:
            response = HttpResponse(status=400)
            response['HX-Trigger'] = json.dumps({
                'showToast': {'message': 'Datos incompletos', 'type': 'error'},
            })
            return response

        uploaded = 0
        for f in files:
            _, ext = os.path.splitext(f.name)
            if ext.lower() not in self.ALLOWED_EXTENSIONS:
                continue
            if f.size > self.MAX_FILE_SIZE:
                continue
            ArchivoPeriodoFacturacion.objects.create(
                anio=anio,
                mes=mes,
                linea_id=linea_id if linea_id else None,
                archivo=f,
                nombre_original=f.name,
                descripcion=descripcion,
                tipo_archivo=f.content_type or '',
                tamanio=f.size,
            )
            uploaded += 1

        archivos_periodo = ArchivoPeriodoFacturacion.objects.filter(
            anio=anio, mes=mes
        )
        if linea_id:
            archivos_periodo = archivos_periodo.filter(linea_id=linea_id)

        html = render_to_string('financiero/partials/periodo_archivos.html', {
            'archivos_periodo': archivos_periodo,
            'filtro_mes': mes, 'filtro_anio': anio, 'filtro_linea': linea_id or '',
        }, request=request)

        response = HttpResponse(html)
        response['HX-Trigger'] = json.dumps({
            'showToast': {'message': f'{uploaded} archivo(s) del periodo subido(s)', 'type': 'success'},
        })
        return response


class PeriodoEliminarArchivoView(LoginRequiredMixin, RoleRequiredMixin, DetailView):
    """Delete a period-level file via HTMX DELETE."""
    model = ArchivoPeriodoFacturacion
    allowed_roles = ['admin', 'director', 'coordinador']

    def delete(self, request, *args, **kwargs):
        import json

        from django.http import HttpResponse
        from django.template.loader import render_to_string

        archivo = self.get_object()
        mes = archivo.mes
        anio = archivo.anio
        linea_id = request.GET.get('linea', '')

        archivo.archivo.delete(save=False)
        archivo.delete()

        archivos_periodo = ArchivoPeriodoFacturacion.objects.filter(
            anio=anio, mes=mes
        )
        if linea_id:
            archivos_periodo = archivos_periodo.filter(linea_id=linea_id)

        html = render_to_string('financiero/partials/periodo_archivos.html', {
            'archivos_periodo': archivos_periodo,
            'filtro_mes': mes, 'filtro_anio': anio, 'filtro_linea': linea_id,
        }, request=request)

        response = HttpResponse(html)
        response['HX-Trigger'] = json.dumps({
            'showToast': {'message': 'Archivo del periodo eliminado', 'type': 'success'},
        })
        return response


# ---------------------------------------------------------------------------
# Cost structure constants (mirrors the Excel PRESUPUESTO sheet)
# ---------------------------------------------------------------------------
MESES = [
    'enero', 'febrero', 'marzo', 'abril', 'mayo', 'junio',
    'julio', 'agosto', 'septiembre', 'octubre', 'noviembre', 'diciembre',
]

ESTRUCTURA_COSTOS = {
    'costos_variables': {
        'titulo': 'Costos Variables',
        'categorias': [
            {
                'codigo': 'MO',
                'nombre': 'Mano de Obra',
                'items': [
                    'Nómina operación',
                    'Tiempo extra operación',
                    'Tiempo festivo operación',
                    'Beneficios operación',
                    'Aportes y parafiscales operación',
                    'Prestaciones sociales operación',
                    'FIC operación',
                    'Seguro de vida operación',
                    'Viáticos reembolsables operación',
                    'Viáticos NO reembolsables operación',
                    'Viáticos descanso operación',
                    'Auxilio alimentación/alojamiento operación',
                    'Hidratación operación',
                    'Campamento operación',
                    'Alimentación operación',
                    'Celaduría campo',
                ],
            },
            {
                'codigo': 'SST',
                'nombre': 'SST y Ambiental',
                'items': [
                    'EPP consumibles',
                    'EPP alturas',
                    'EPP servidumbre',
                    'Seguridad grupal',
                    'Bioseguridad',
                    'Dotación operación',
                    'Examen médico operación',
                    'Certificación operación',
                    'Capacitaciones',
                    'Ambiental',
                ],
            },
            {
                'codigo': 'TA',
                'nombre': 'Transporte y Administración',
                'items': [
                    'Transporte operación',
                    'Transporte de materiales y herramientas',
                    'Transporte reembolsable',
                ],
            },
            {
                'codigo': 'MH',
                'nombre': 'Material y Herramientas',
                'items': [
                    'Material obra',
                    'Material fungible',
                    'Insumos maquinaria y equipos',
                    'Activos fijos',
                    'Herramientas menores',
                    'Gastos reembolsables',
                ],
            },
            {
                'codigo': 'SC',
                'nombre': 'Subcontratistas',
                'items': [
                    'Subcontratistas',
                ],
            },
        ],
    },
    'costos_fijos': {
        'titulo': 'Costos Fijos',
        'categorias': [
            {
                'codigo': 'MO',
                'nombre': 'Mano de Obra Administración',
                'items': [
                    'Nómina administración',
                    'Tiempo extra administración',
                    'Tiempo festivo administración',
                    'Beneficios administración',
                    'Aportes y parafiscales administración',
                    'Prestaciones sociales administración',
                    'FIC administración',
                    'Seguro de vida administración',
                    'Viáticos reembolsables administración',
                    'Viáticos NO reembolsables administración',
                    'Viáticos descanso administración',
                    'Auxilio alimentación/alojamiento administración',
                    'Hidratación administración',
                    'Campamento administración',
                    'Alimentación administración',
                ],
            },
            {
                'codigo': 'SST',
                'nombre': 'SST Administración',
                'items': [
                    'Dotación administración',
                    'Examen médico administración',
                    'Certificación administración',
                ],
            },
            {
                'codigo': 'TA',
                'nombre': 'Transporte y Gastos Administración',
                'items': [
                    'Transporte administración',
                    'Transporte ocasional',
                    'Arrendamiento oficina',
                    'Arrendamiento patio',
                    'Servicios públicos',
                    'Celular / internet',
                    'Seguridad privada',
                    'Gastos menores',
                    'Insumos oficina',
                    'Ingreso sitio torre',
                    'Actividades por garantía / imprevistos',
                ],
            },
            {
                'codigo': 'PFS',
                'nombre': 'Garantías, Fianzas y Seguros',
                'items': [
                    'Garantías, fianzas y seguros',
                ],
            },
        ],
    },
}


def _build_empty_datos():
    """Return an empty datos structure with all zeros."""
    zero_months = {m: 0 for m in MESES}
    datos = {'ingreso_proyectado': dict(zero_months)}
    for seccion_key, seccion in ESTRUCTURA_COSTOS.items():
        datos[seccion_key] = {}
        for cat in seccion['categorias']:
            datos[seccion_key][cat['codigo']] = {}
            for item in cat['items']:
                datos[seccion_key][cat['codigo']][item] = dict(zero_months)
    return datos


def _build_display_rows(datos):
    """Build structured rows for template rendering from stored datos JSON."""
    rows = []

    # Ingreso proyectado
    ing = datos.get('ingreso_proyectado', {})
    valores_ing = [ing.get(m, 0) for m in MESES]
    rows.append({
        'tipo': 'ingreso',
        'nombre': 'INGRESO PROYECTADO',
        'codigo': '',
        'valores': valores_ing,
        'total': sum(valores_ing),
    })

    gran_total_variables = [0] * 12
    gran_total_fijos = [0] * 12

    for seccion_key, seccion in ESTRUCTURA_COSTOS.items():
        is_variable = seccion_key == 'costos_variables'
        # Section header
        rows.append({
            'tipo': 'seccion',
            'nombre': seccion['titulo'],
            'codigo': '',
            'valores': None,
            'total': None,
        })

        seccion_totales = [0] * 12

        for cat in seccion['categorias']:
            # Category header
            cat_totales = [0] * 12
            rows.append({
                'tipo': 'categoria',
                'nombre': cat['nombre'],
                'codigo': cat['codigo'],
                'valores': None,
                'total': None,
            })

            cat_data = datos.get(seccion_key, {}).get(cat['codigo'], {})
            for item_name in cat['items']:
                item_data = cat_data.get(item_name, {})
                valores = [item_data.get(m, 0) for m in MESES]
                for i in range(12):
                    cat_totales[i] += valores[i]
                rows.append({
                    'tipo': 'item',
                    'nombre': item_name,
                    'codigo': cat['codigo'],
                    'seccion': seccion_key,
                    'valores': valores,
                    'total': sum(valores),
                })

            # Category subtotal
            rows.append({
                'tipo': 'subtotal',
                'nombre': f"Subtotal {cat['nombre']}",
                'codigo': cat['codigo'],
                'valores': cat_totales,
                'total': sum(cat_totales),
            })
            for i in range(12):
                seccion_totales[i] += cat_totales[i]

        # Section total
        rows.append({
            'tipo': 'total_seccion',
            'nombre': f"Total {seccion['titulo']}",
            'codigo': '',
            'valores': seccion_totales,
            'total': sum(seccion_totales),
        })

        if is_variable:
            gran_total_variables = seccion_totales[:]
        else:
            gran_total_fijos = seccion_totales[:]

    # Summary rows
    ing_vals = [ing.get(m, 0) for m in MESES]
    total_gastos = [gran_total_variables[i] + gran_total_fijos[i] for i in range(12)]
    resultado = [ing_vals[i] - total_gastos[i] for i in range(12)]

    rows.append({
        'tipo': 'total_seccion',
        'nombre': 'TOTAL GASTOS',
        'codigo': '',
        'valores': total_gastos,
        'total': sum(total_gastos),
    })
    rows.append({
        'tipo': 'resultado',
        'nombre': 'RESULTADO (Ingreso - Gastos)',
        'codigo': '',
        'valores': resultado,
        'total': sum(resultado),
    })

    total_ing = sum(ing_vals)
    total_gasto = sum(total_gastos)
    utilidad_pct = ((total_ing - total_gasto) / total_ing * 100) if total_ing else 0

    return rows, {
        'total_ingreso': total_ing,
        'total_variables': sum(gran_total_variables),
        'total_fijos': sum(gran_total_fijos),
        'total_gastos': total_gasto,
        'resultado': total_ing - total_gasto,
        'utilidad_pct': utilidad_pct,
    }


class PresupuestoDetalladoBaseView(LoginRequiredMixin, RoleRequiredMixin, TemplateView):
    """Base view for detailed budget (Planeado / Real)."""
    template_name = 'financiero/presupuesto_detallado.html'
    allowed_roles = ['admin', 'director', 'coordinador']
    tipo_presupuesto = None  # Override in subclass

    def _get_contrato_filter(self):
        """Return contrato instance from GET params, or None."""
        from apps.contratos.models import Contrato
        contrato_id = self.request.GET.get('contrato') or self.request.POST.get('contrato')
        if contrato_id:
            try:
                return Contrato.objects.get(pk=contrato_id)
            except Contrato.DoesNotExist:
                pass
        return None

    def _get_or_create_presupuesto(self, anio, contrato):
        """Get or create PresupuestoDetallado with optional contrato."""
        return PresupuestoDetallado.objects.get_or_create(
            anio=anio,
            tipo=self.tipo_presupuesto,
            contrato=contrato,
            defaults={'datos': _build_empty_datos()},
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from django.utils import timezone
        from apps.contratos.models import Contrato

        anio = int(self.request.GET.get('anio', timezone.now().year))
        unidad_filter = self.request.GET.get('unidad', '')
        contrato = self._get_contrato_filter()

        obj, created = self._get_or_create_presupuesto(anio, contrato)

        rows, resumen = _build_display_rows(obj.datos)

        context['presupuesto_obj'] = obj
        context['tipo_presupuesto'] = self.tipo_presupuesto
        context['tipo_label'] = 'Planeado' if self.tipo_presupuesto == 'PLANEADO' else 'Real'
        context['anio'] = anio
        context['meses'] = MESES
        context['meses_cortos'] = [m[:3].title() for m in MESES]
        context['rows'] = rows
        context['resumen'] = resumen
        context['anios_disponibles'] = list(range(anio - 2, anio + 3))

        # Contract filters
        context['unidad_filter'] = unidad_filter
        context['contrato_seleccionado'] = contrato
        contratos_qs = Contrato.objects.all()
        if unidad_filter:
            contratos_qs = contratos_qs.filter(unidad_negocio=unidad_filter)
        context['contratos_disponibles'] = contratos_qs
        context['unidades_negocio'] = Contrato.UnidadNegocio.choices
        return context

    def post(self, request, *args, **kwargs):
        """Route POST actions: inline cell edit or Excel import."""
        action = request.POST.get('action', '')
        if action == 'import_excel':
            return self._handle_excel_import(request)
        return self._handle_cell_edit(request)

    def _handle_cell_edit(self, request):
        """Handle inline cell edits via HTMX."""
        from django.http import HttpResponse
        from django.utils import timezone

        anio = int(request.POST.get('anio', timezone.now().year))
        contrato = self._get_contrato_filter()
        obj, _ = self._get_or_create_presupuesto(anio, contrato)

        seccion = request.POST.get('seccion', '')
        categoria = request.POST.get('categoria', '')
        item = request.POST.get('item', '')
        mes = request.POST.get('mes', '')
        valor_raw = request.POST.get('valor', '0').replace(',', '').replace('.', '').strip()

        try:
            valor = int(valor_raw)
        except (ValueError, TypeError):
            valor = 0

        datos = obj.datos or _build_empty_datos()

        if seccion == 'ingreso_proyectado':
            if 'ingreso_proyectado' not in datos:
                datos['ingreso_proyectado'] = {m: 0 for m in MESES}
            datos['ingreso_proyectado'][mes] = valor
        elif seccion in ('costos_variables', 'costos_fijos') and categoria and item:
            if seccion not in datos:
                datos[seccion] = {}
            if categoria not in datos[seccion]:
                datos[seccion][categoria] = {}
            if item not in datos[seccion][categoria]:
                datos[seccion][categoria][item] = {m: 0 for m in MESES}
            datos[seccion][categoria][item][mes] = valor

        obj.datos = datos
        obj.save(update_fields=['datos', 'updated_at'])

        return HttpResponse(
            f'<span class="text-right">{valor:,.0f}</span>',
            content_type='text/html',
        )

    def _build_redirect_url(self, anio, contrato=None):
        """Build redirect URL preserving filters."""
        url = f'{self.request.path}?anio={anio}'
        if contrato:
            url += f'&contrato={contrato.pk}'
        return url

    def _handle_excel_import(self, request):
        """Handle Excel file upload and data import."""
        from django.contrib import messages
        from django.shortcuts import redirect
        from django.utils import timezone

        anio = int(request.POST.get('anio', timezone.now().year))
        contrato = self._get_contrato_filter()
        archivo = request.FILES.get('archivo')

        redirect_url = self._build_redirect_url(anio, contrato)

        if not archivo:
            messages.error(request, 'No se seleccionó ningún archivo.')
            return redirect(redirect_url)

        if not archivo.name.lower().endswith(('.xlsx', '.xls')):
            messages.error(request, 'Formato no soportado. Use archivos .xlsx o .xls')
            return redirect(redirect_url)

        if archivo.size > 10 * 1024 * 1024:
            messages.error(request, 'El archivo excede el tamaño máximo (10 MB).')
            return redirect(redirect_url)

        from .importers import (
            ContableExcelImporter,
            PresupuestoExcelImporter,
            detect_excel_format,
        )

        # Detect format and show recommendation
        deteccion = detect_excel_format(archivo)
        archivo.seek(0)

        formato = deteccion['formato']
        confianza = deteccion['confianza']

        if formato == 'desconocido':
            messages.warning(
                request,
                f'Formato no reconocido: {deteccion["descripcion"]} '
                f'Se intentará importar de todas formas.',
            )
        else:
            formato_labels = {
                'presupuesto': 'Presupuesto estándar',
                'contable_resumen': 'Contable (resumen)',
                'contable_transaccional': 'Contable (transaccional)',
            }
            messages.info(
                request,
                f'Formato detectado: {formato_labels.get(formato, formato)} '
                f'(confianza: {confianza}). {deteccion["descripcion"]}',
            )

        # Import using the detected format (or try both)
        if formato == 'presupuesto':
            importer = PresupuestoExcelImporter()
            result = importer.importar(archivo, ESTRUCTURA_COSTOS)
            if not result['exito'] or result.get('matched', 0) == 0:
                archivo.seek(0)
                contable_importer = ContableExcelImporter()
                contable_result = contable_importer.importar(archivo, ESTRUCTURA_COSTOS)
                if contable_result['exito'] and contable_result.get('matched', 0) > 0:
                    result = contable_result
        elif formato in ('contable_resumen', 'contable_transaccional'):
            contable_importer = ContableExcelImporter()
            result = contable_importer.importar(archivo, ESTRUCTURA_COSTOS)
            if not result['exito'] or result.get('matched', 0) == 0:
                archivo.seek(0)
                importer = PresupuestoExcelImporter()
                pres_result = importer.importar(archivo, ESTRUCTURA_COSTOS)
                if pres_result['exito'] and pres_result.get('matched', 0) > 0:
                    result = pres_result
        else:
            # Unknown format: try both
            importer = PresupuestoExcelImporter()
            result = importer.importar(archivo, ESTRUCTURA_COSTOS)
            if not result['exito'] or result.get('matched', 0) == 0:
                archivo.seek(0)
                contable_importer = ContableExcelImporter()
                contable_result = contable_importer.importar(archivo, ESTRUCTURA_COSTOS)
                if contable_result['exito'] and contable_result.get('matched', 0) > 0:
                    result = contable_result

        if not result['exito']:
            messages.error(request, f'Error al importar: {result.get("error", "Error desconocido")}')
            return redirect(redirect_url)

        obj, created = self._get_or_create_presupuesto(anio, contrato)
        if not created:
            obj.datos = result['datos']
            obj.save(update_fields=['datos', 'updated_at'])
        else:
            obj.datos = result['datos']
            obj.save(update_fields=['datos', 'updated_at'])

        matched = result['matched']
        unmatched = result['unmatched']
        msg = f'Importación exitosa: {matched} items importados.'
        if unmatched:
            msg += f' {unmatched} items no reconocidos.'
        messages.success(request, msg)

        for w in result.get('warnings', [])[:3]:
            messages.warning(request, w)
        for item in result.get('unmatched_items', [])[:5]:
            messages.info(request, f'No reconocido: {item}')

        return redirect(redirect_url)


class PresupuestoPlaneadoView(PresupuestoDetalladoBaseView):
    """View for the planned budget tab."""
    tipo_presupuesto = 'PLANEADO'


class PresupuestoRealView(PresupuestoDetalladoBaseView):
    """View for the actual/real budget tab."""
    tipo_presupuesto = 'REAL'


ROLES_OPERATIVOS = [
    'SUPERVISOR', 'LINIERO_I', 'LINIERO_II', 'AYUDANTE', 'CONDUCTOR',
    'SUPERVISOR_FOREST', 'ASISTENTE_FOREST',
]


class NominaView(LoginRequiredMixin, RoleRequiredMixin, TemplateView):
    """Payroll view: operative and administrative labor with personnel and budget items."""
    template_name = 'financiero/nomina.html'
    allowed_roles = ['admin', 'director', 'coordinador']

    def _get_filters(self, request):
        from django.utils import timezone
        from apps.contratos.models import Contrato

        anio = int(request.GET.get('anio') or request.POST.get('anio') or timezone.now().year)
        tipo = request.GET.get('tipo') or request.POST.get('tipo') or 'operativo'
        mes_idx = request.GET.get('mes') or request.POST.get('mes')
        if mes_idx is not None:
            try:
                mes_idx = int(mes_idx)
                if mes_idx < 0 or mes_idx > 11:
                    mes_idx = None
            except (ValueError, TypeError):
                mes_idx = None
        unidad_filter = request.GET.get('unidad', '')
        contrato = None
        contrato_id = request.GET.get('contrato') or request.POST.get('contrato')
        if contrato_id:
            try:
                contrato = Contrato.objects.get(pk=contrato_id)
            except Contrato.DoesNotExist:
                pass
        contratos_qs = Contrato.objects.all()
        if unidad_filter:
            contratos_qs = contratos_qs.filter(unidad_negocio=unidad_filter)
        return anio, tipo, unidad_filter, contrato, contratos_qs, mes_idx

    def _build_redirect(self, request, tipo, anio, contrato=None, mes_idx=None):
        url = f'{request.path}?tipo={tipo}&anio={anio}'
        if contrato:
            url += f'&contrato={contrato.pk}'
        if mes_idx is not None:
            url += f'&mes={mes_idx}'
        return url

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from apps.contratos.models import Contrato
        from apps.cuadrillas.models import CuadrillaMiembro

        from .models import PersonalAdministrativo

        anio, tipo, unidad_filter, contrato, contratos_qs, mes_idx = self._get_filters(self.request)
        mes_actual = MESES[mes_idx] if mes_idx is not None else None

        # --- Personnel ---
        if tipo == 'administrativo':
            # Use PersonalAdministrativo model
            personal_qs = PersonalAdministrativo.objects.filter(activo=True)
            if contrato:
                personal_qs = personal_qs.filter(contrato=contrato)
            personal_list = list(personal_qs)
            total_personal = len(personal_list)
            cuadrillas_list = []  # Not used for admin
        else:
            # Use CuadrillaMiembro for operative
            miembros_qs = CuadrillaMiembro.objects.filter(
                activo=True,
                rol_cuadrilla__in=ROLES_OPERATIVOS,
            ).select_related('usuario', 'cuadrilla', 'cuadrilla__linea_asignada').order_by(
                'cuadrilla__codigo', 'rol_cuadrilla', 'usuario__first_name'
            )
            personal_list = []  # Not used for operative
            total_personal = miembros_qs.count()
            # Group by cuadrilla
            cuadrillas_dict = {}
            for m in miembros_qs:
                cua = m.cuadrilla
                if cua.pk not in cuadrillas_dict:
                    cuadrillas_dict[cua.pk] = {'cuadrilla': cua, 'miembros': []}
                cuadrillas_dict[cua.pk]['miembros'].append(m)
            cuadrillas_list = list(cuadrillas_dict.values())

        # --- Budget data ---
        planeado_obj = PresupuestoDetallado.objects.filter(
            anio=anio, tipo='PLANEADO', contrato=contrato,
        ).first()
        real_obj = PresupuestoDetallado.objects.filter(
            anio=anio, tipo='REAL', contrato=contrato,
        ).first()
        datos_planeado = planeado_obj.datos if planeado_obj else _build_empty_datos()
        datos_real = real_obj.datos if real_obj else _build_empty_datos()

        if tipo == 'administrativo':
            seccion_key = 'costos_fijos'
            cat_codigo = 'MO'
            cat_nombre = 'Mano de Obra Administración'
        else:
            seccion_key = 'costos_variables'
            cat_codigo = 'MO'
            cat_nombre = 'Mano de Obra'

        estructura_cat = None
        for cat in ESTRUCTURA_COSTOS[seccion_key]['categorias']:
            if cat['codigo'] == cat_codigo and cat['nombre'] == cat_nombre:
                estructura_cat = cat
                break

        items_presupuesto = []
        if estructura_cat:
            cat_data_planeado = datos_planeado.get(seccion_key, {}).get(cat_codigo, {})
            cat_data_real = datos_real.get(seccion_key, {}).get(cat_codigo, {})
            for item_name in estructura_cat['items']:
                planeado_vals = cat_data_planeado.get(item_name, {})
                real_vals = cat_data_real.get(item_name, {})
                if mes_actual:
                    total_p = planeado_vals.get(mes_actual, 0)
                    total_r = real_vals.get(mes_actual, 0)
                else:
                    total_p = sum(planeado_vals.get(m, 0) for m in MESES)
                    total_r = sum(real_vals.get(m, 0) for m in MESES)
                pct = ((total_r - total_p) / total_p * 100) if total_p else 0
                items_presupuesto.append({
                    'nombre': item_name,
                    'total_planeado': total_p,
                    'total_real': total_r,
                    'diferencia': total_r - total_p,
                    'pct_desviacion': pct,
                    'seccion_key': seccion_key,
                    'cat_codigo': cat_codigo,
                })

        total_planeado = sum(i['total_planeado'] for i in items_presupuesto)
        total_real = sum(i['total_real'] for i in items_presupuesto)

        context.update({
            'anio': anio,
            'tipo': tipo,
            'unidad_filter': unidad_filter,
            'contrato_seleccionado': contrato,
            'contratos_disponibles': contratos_qs,
            'unidades_negocio': Contrato.UnidadNegocio.choices,
            'anios_disponibles': list(range(anio - 2, anio + 3)),
            'cuadrillas_list': cuadrillas_list,
            'personal_list': personal_list,
            'total_personal': total_personal,
            'items_presupuesto': items_presupuesto,
            'cat_nombre': cat_nombre,
            'seccion_key': seccion_key,
            'cat_codigo': cat_codigo,
            'total_planeado': total_planeado,
            'total_real': total_real,
            'diferencia_total': total_real - total_planeado,
            'meses_cortos': [m[:3].title() for m in MESES],
            'meses': MESES,
            'mes_idx': mes_idx,
            'mes_actual': mes_actual,
            'mes_nombre': MESES[mes_idx].title() if mes_idx is not None else None,
            'meses_opciones': [{'idx': i, 'nombre': m.title()} for i, m in enumerate(MESES)],
            'cargos_admin': PersonalAdministrativo.Cargo.choices,
        })
        return context

    def post(self, request, *args, **kwargs):
        from django.contrib import messages
        from django.shortcuts import redirect

        from .models import PersonalAdministrativo

        action = request.POST.get('action', '')
        anio, tipo, unidad_filter, contrato, _, mes_idx = self._get_filters(request)
        redirect_url = self._build_redirect(request, tipo, anio, contrato, mes_idx)

        if action == 'agregar_personal':
            nombre = request.POST.get('nombre', '').strip()
            documento = request.POST.get('documento', '').strip()
            cargo = request.POST.get('cargo', 'OTRO')
            salario = request.POST.get('salario_mensual', '0') or '0'

            if not nombre:
                messages.error(request, 'El nombre es obligatorio.')
                return redirect(redirect_url)

            PersonalAdministrativo.objects.create(
                nombre=nombre,
                documento=documento,
                cargo=cargo if cargo in dict(PersonalAdministrativo.Cargo.choices) else 'OTRO',
                salario_mensual=float(salario),
                contrato=contrato,
            )
            messages.success(request, f'{nombre} agregado al personal administrativo.')

        elif action == 'remover_personal':
            persona_id = request.POST.get('persona_id')
            try:
                persona = PersonalAdministrativo.objects.get(pk=persona_id, activo=True)
                persona.activo = False
                persona.save(update_fields=['activo', 'updated_at'])
                messages.success(request, f'{persona.nombre} removido.')
            except PersonalAdministrativo.DoesNotExist:
                messages.error(request, 'Persona no encontrada.')

        elif action == 'upload_personal':
            archivo = request.FILES.get('archivo')
            if not archivo:
                messages.error(request, 'No se seleccionó archivo.')
                return redirect(redirect_url)
            try:
                import openpyxl
                wb = openpyxl.load_workbook(archivo, read_only=True)
                ws = wb.active
                creados = 0
                actualizados = 0
                cargos_map = {v.lower(): k for k, v in PersonalAdministrativo.Cargo.choices}

                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row or not row[0]:
                        continue
                    nombre = str(row[0]).strip()
                    documento = str(row[1]).strip() if len(row) > 1 and row[1] else ''
                    cargo_raw = str(row[2]).strip().lower() if len(row) > 2 and row[2] else ''
                    salario_raw = row[3] if len(row) > 3 and row[3] else 0

                    cargo_code = cargos_map.get(cargo_raw, '')
                    if not cargo_code:
                        # Try matching by code directly
                        if cargo_raw.upper() in dict(PersonalAdministrativo.Cargo.choices):
                            cargo_code = cargo_raw.upper()
                        else:
                            cargo_code = 'OTRO'

                    try:
                        salario = float(salario_raw)
                    except (ValueError, TypeError):
                        salario = 0

                    if documento:
                        obj, created = PersonalAdministrativo.objects.update_or_create(
                            documento=documento,
                            defaults={
                                'nombre': nombre,
                                'cargo': cargo_code,
                                'salario_mensual': salario,
                                'contrato': contrato,
                                'activo': True,
                            },
                        )
                    else:
                        PersonalAdministrativo.objects.create(
                            nombre=nombre,
                            documento=documento,
                            cargo=cargo_code,
                            salario_mensual=salario,
                            contrato=contrato,
                        )
                        created = True

                    if created:
                        creados += 1
                    else:
                        actualizados += 1

                messages.success(request, f'Personal cargado: {creados} nuevos, {actualizados} actualizados.')
            except Exception as e:
                messages.error(request, f'Error al procesar archivo: {str(e)}')

        elif action == 'edit_presupuesto':
            # Inline edit of a budget item value (total annual)
            from django.http import HttpResponse

            seccion_key = request.POST.get('seccion_key', '')
            cat_codigo = request.POST.get('cat_codigo', '')
            item_name = request.POST.get('item_name', '')
            tipo_pres = request.POST.get('tipo_presupuesto', '')  # PLANEADO or REAL
            mes = request.POST.get('mes', '')
            valor_raw = request.POST.get('valor', '0').replace(',', '').replace('.', '').strip()

            try:
                valor = int(valor_raw)
            except (ValueError, TypeError):
                valor = 0

            obj, _ = PresupuestoDetallado.objects.get_or_create(
                anio=anio, tipo=tipo_pres, contrato=contrato,
                defaults={'datos': _build_empty_datos()},
            )

            datos = obj.datos or _build_empty_datos()
            if seccion_key not in datos:
                datos[seccion_key] = {}
            if cat_codigo not in datos[seccion_key]:
                datos[seccion_key][cat_codigo] = {}
            if item_name not in datos[seccion_key][cat_codigo]:
                datos[seccion_key][cat_codigo][item_name] = {m: 0 for m in MESES}
            datos[seccion_key][cat_codigo][item_name][mes] = valor

            obj.datos = datos
            obj.save(update_fields=['datos', 'updated_at'])

            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return HttpResponse('OK')
            return redirect(redirect_url)

        return redirect(redirect_url)


class DescargarPlantillaExcelView(LoginRequiredMixin, RoleRequiredMixin, TemplateView):
    """Generate and download an Excel template for budget data entry."""
    allowed_roles = ['admin', 'director', 'coordinador']
    template_name = ''  # Not used

    def get(self, request, *args, **kwargs):
        import io

        from django.http import HttpResponse

        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'PRESUPUESTO'

        anio = request.GET.get('anio', '')

        # Styles
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        seccion_font = Font(bold=True, color='FFFFFF', size=10)
        seccion_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
        cat_font = Font(bold=True, size=10)
        cat_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
        subtotal_font = Font(bold=True, italic=True, size=10)
        subtotal_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
        total_font = Font(bold=True, color='FFFFFF', size=10)
        total_fill = PatternFill(start_color='375623', end_color='375623', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )
        money_fmt = '#,##0'

        meses_header = ['Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun',
                        'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic']

        # Title row
        ws.merge_cells('A1:N1')
        title_cell = ws['A1']
        title_cell.value = f'Presupuesto {anio}' if anio else 'Presupuesto'
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center')

        # Header row (row 3)
        row = 3
        headers = ['Concepto'] + meses_header + ['Total']
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        # Column widths
        ws.column_dimensions['A'].width = 45
        for c in range(2, 15):
            ws.column_dimensions[get_column_letter(c)].width = 15

        row = 4

        # Helper to write a row
        def write_row(label, font_style, fill_style, is_formula_row=False, item_row_start=None, item_row_end=None):
            nonlocal row
            cell_a = ws.cell(row=row, column=1, value=label)
            cell_a.font = font_style
            cell_a.fill = fill_style
            cell_a.border = thin_border
            for col in range(2, 15):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                cell.number_format = money_fmt
                cell.fill = fill_style
                cell.font = font_style
                if is_formula_row and item_row_start and col <= 13:
                    col_letter = get_column_letter(col)
                    cell.value = f'=SUM({col_letter}{item_row_start}:{col_letter}{item_row_end})'
                elif col == 14:
                    col_b = get_column_letter(2)
                    col_m = get_column_letter(13)
                    cell.value = f'=SUM({col_b}{row}:{col_m}{row})'
                else:
                    cell.value = 0
            row += 1
            return row - 1

        # INGRESO PROYECTADO
        ingreso_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
        write_row('INGRESO PROYECTADO', Font(bold=True, size=10), ingreso_fill)

        row += 1  # blank row

        # Build sections
        for seccion_key, seccion in ESTRUCTURA_COSTOS.items():
            # Section header
            cell_a = ws.cell(row=row, column=1, value=seccion['titulo'].upper())
            cell_a.font = seccion_font
            cell_a.fill = seccion_fill
            cell_a.border = thin_border
            for col in range(2, 15):
                cell = ws.cell(row=row, column=col)
                cell.fill = seccion_fill
                cell.border = thin_border
            seccion_item_rows = []
            row += 1

            for cat in seccion['categorias']:
                # Category header
                cell_a = ws.cell(row=row, column=1, value=f"  {cat['nombre']}")
                cell_a.font = cat_font
                cell_a.fill = cat_fill
                cell_a.border = thin_border
                for col in range(2, 15):
                    cell = ws.cell(row=row, column=col)
                    cell.fill = cat_fill
                    cell.border = thin_border
                row += 1

                item_start = row
                for item_name in cat['items']:
                    write_row(f"    {item_name}", Font(size=10), PatternFill())
                item_end = row - 1

                # Subtotal with formulas
                st_row = write_row(
                    f"  Subtotal {cat['nombre']}",
                    subtotal_font, subtotal_fill,
                    is_formula_row=True, item_row_start=item_start, item_row_end=item_end,
                )
                seccion_item_rows.append(st_row)
                row += 1  # spacing

            # Section total (sum of subtotals)
            cell_a = ws.cell(row=row, column=1, value=f"TOTAL {seccion['titulo'].upper()}")
            cell_a.font = total_font
            cell_a.fill = total_fill
            cell_a.border = thin_border
            for col in range(2, 15):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                cell.number_format = money_fmt
                cell.fill = total_fill
                cell.font = total_font
                if col <= 13:
                    col_letter = get_column_letter(col)
                    formula_parts = [f'{col_letter}{r}' for r in seccion_item_rows]
                    cell.value = f'={"+".join(formula_parts)}'
                else:
                    col_b = get_column_letter(2)
                    col_m = get_column_letter(13)
                    cell.value = f'=SUM({col_b}{row}:{col_m}{row})'
            row += 2  # spacing

        # Instructions sheet
        ws2 = wb.create_sheet('Instrucciones')
        ws2['A1'] = 'Instrucciones para llenar la plantilla'
        ws2['A1'].font = Font(bold=True, size=14)
        instructions = [
            '',
            '1. Llene los valores mensuales (Ene-Dic) en la hoja PRESUPUESTO.',
            '2. Solo modifique las celdas de los items (filas blancas). Los subtotales y totales se calculan automaticamente.',
            '3. El INGRESO PROYECTADO es la facturacion esperada por mes.',
            '4. Los valores deben ser en pesos colombianos (COP), sin decimales.',
            '5. Cuando termine, guarde el archivo y suba desde el boton "Subir Excel" en la aplicacion.',
            '',
            'Nota: No cambie los nombres de los conceptos ni la estructura de la hoja.',
        ]
        for i, txt in enumerate(instructions, 2):
            ws2.cell(row=i, column=1, value=txt)
        ws2.column_dimensions['A'].width = 80

        # Save to response
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f'Plantilla_Presupuesto_{anio}.xlsx' if anio else 'Plantilla_Presupuesto.xlsx'
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


class CargarCostosCuadrillaView(LoginRequiredMixin, RoleRequiredMixin, TemplateView):
    """Calculate labor costs from cuadrilla attendance data and fill Presupuesto Real."""
    allowed_roles = ['admin', 'director', 'coordinador']
    template_name = 'financiero/presupuesto_detallado.html'

    # Colombian legal overtime multipliers over base hourly rate
    FACTOR_HE_DIURNA = Decimal('1.25')
    FACTOR_HE_NOCTURNA = Decimal('1.75')
    FACTOR_HE_DOMINICAL_DIURNA = Decimal('2.00')
    FACTOR_HE_DOMINICAL_NOCTURNA = Decimal('2.50')

    # Standard working hours per day (average)
    HORAS_JORNADA = Decimal('8')

    def post(self, request, *args, **kwargs):
        from calendar import monthrange

        from django.contrib import messages
        from django.db.models import Q, Sum
        from django.shortcuts import redirect
        from django.utils import timezone

        from apps.contratos.models import Contrato
        from apps.cuadrillas.models import Asistencia, CuadrillaMiembro

        anio = int(request.POST.get('anio', timezone.now().year))
        redirect_url = request.POST.get('redirect_url', reverse_lazy('financiero:presupuesto_real'))

        # Handle optional contrato filter
        contrato = None
        contrato_id = request.POST.get('contrato')
        if contrato_id:
            try:
                contrato = Contrato.objects.get(pk=contrato_id)
            except Contrato.DoesNotExist:
                pass

        obj, created = PresupuestoDetallado.objects.get_or_create(
            anio=anio,
            tipo='REAL',
            contrato=contrato,
            defaults={'datos': _build_empty_datos()},
        )
        datos = obj.datos or _build_empty_datos()

        total_nomina = 0
        total_he = 0
        total_viaticos = 0
        total_transporte = 0

        for mes_num in range(1, 13):
            mes_name = MESES[mes_num - 1]
            fecha_inicio = date(anio, mes_num, 1)
            _, last_day = monthrange(anio, mes_num)
            fecha_fin = date(anio, mes_num, last_day)

            # Get all attendance records for this month
            asistencias = Asistencia.objects.filter(
                fecha__gte=fecha_inicio,
                fecha__lte=fecha_fin,
            ).select_related('cuadrilla', 'usuario')

            # --- Nomina operacion ---
            # Sum of costo_dia for all present days
            nomina_mes = Decimal('0')
            for asist in asistencias.filter(tipo_novedad='PRESENTE'):
                # Get the member's daily cost
                miembro = CuadrillaMiembro.objects.filter(
                    cuadrilla=asist.cuadrilla,
                    usuario=asist.usuario,
                    activo=True,
                ).first()
                if miembro and miembro.costo_dia:
                    nomina_mes += miembro.costo_dia

            datos['costos_variables']['MO']['Nómina operación'][mes_name] = int(nomina_mes)
            total_nomina += nomina_mes

            # --- Tiempo extra operacion ---
            # Calculate overtime cost based on hourly rate and overtime multipliers
            he_costo_mes = Decimal('0')
            he_qs = asistencias.filter(
                Q(he_diurna__gt=0) | Q(he_nocturna__gt=0) |
                Q(he_dominical_diurna__gt=0) | Q(he_dominical_nocturna__gt=0)
            )
            for asist in he_qs:
                miembro = CuadrillaMiembro.objects.filter(
                    cuadrilla=asist.cuadrilla,
                    usuario=asist.usuario,
                    activo=True,
                ).first()
                if miembro and miembro.costo_dia:
                    tarifa_hora = miembro.costo_dia / self.HORAS_JORNADA
                    he_costo = (
                        (asist.he_diurna or 0) * tarifa_hora * self.FACTOR_HE_DIURNA +
                        (asist.he_nocturna or 0) * tarifa_hora * self.FACTOR_HE_NOCTURNA +
                        (asist.he_dominical_diurna or 0) * tarifa_hora * self.FACTOR_HE_DOMINICAL_DIURNA +
                        (asist.he_dominical_nocturna or 0) * tarifa_hora * self.FACTOR_HE_DOMINICAL_NOCTURNA
                    )
                    he_costo_mes += he_costo

            datos['costos_variables']['MO']['Tiempo extra operación'][mes_name] = int(he_costo_mes)
            total_he += he_costo_mes

            # --- Viaticos ---
            viaticos_reemb = asistencias.filter(
                viatico_aplica=True,
            ).aggregate(total=Sum('viaticos'))['total'] or 0
            datos['costos_variables']['MO']['Viáticos reembolsables operación'][mes_name] = int(viaticos_reemb)
            total_viaticos += viaticos_reemb

            # --- Transporte (vehicle costs for days with attendance) ---
            from apps.cuadrillas.models import Cuadrilla
            cuadrillas_activas = Cuadrilla.objects.filter(
                asistencias__fecha__gte=fecha_inicio,
                asistencias__fecha__lte=fecha_fin,
                vehiculo__isnull=False,
            ).distinct()
            transporte_mes = Decimal('0')
            for cuad in cuadrillas_activas:
                dias_activos = Asistencia.objects.filter(
                    cuadrilla=cuad,
                    fecha__gte=fecha_inicio,
                    fecha__lte=fecha_fin,
                    tipo_novedad='PRESENTE',
                ).values('fecha').distinct().count()
                if cuad.vehiculo and cuad.vehiculo.costo_dia:
                    transporte_mes += cuad.vehiculo.costo_dia * dias_activos

            datos['costos_variables']['TA']['Transporte operación'][mes_name] = int(transporte_mes)
            total_transporte += transporte_mes

        obj.datos = datos
        obj.save(update_fields=['datos', 'updated_at'])

        messages.success(
            request,
            f'Costos cargados desde cuadrillas para {anio}: '
            f'Nómina ${total_nomina:,.0f}, '
            f'Horas Extra ${total_he:,.0f}, '
            f'Viáticos ${total_viaticos:,.0f}, '
            f'Transporte ${total_transporte:,.0f}.',
        )

        url = f'{redirect_url}?anio={anio}'
        if contrato:
            url += f'&contrato={contrato.pk}'
        return redirect(url)
