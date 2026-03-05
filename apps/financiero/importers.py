"""
Excel importers for PresupuestoDetallado.
"""
import logging
import unicodedata
from datetime import datetime

from openpyxl import load_workbook

logger = logging.getLogger(__name__)

MESES = [
    'enero', 'febrero', 'marzo', 'abril', 'mayo', 'junio',
    'julio', 'agosto', 'septiembre', 'octubre', 'noviembre', 'diciembre',
]

# Map 1-based month number to month name
MES_NUM_TO_NAME = {i + 1: m for i, m in enumerate(MESES)}

# Items where multiple Excel rows map to one ESTRUCTURA item (values are SUMMED)
MERGE_ITEMS = {
    'Auxilio alimentación/alojamiento operación',
    'Auxilio alimentación/alojamiento administración',
    'Servicios públicos',
}

# Explicit mapping: normalized Excel name -> exact ESTRUCTURA_COSTOS item name.
# Keys are the output of _normalize_name() applied to the raw Excel cell.
EXCEL_TO_ESTRUCTURA = {
    # ── Costos Variables / MO ──
    'nomina operacion': 'Nómina operación',
    'tiempo extra operacion': 'Tiempo extra operación',
    'tiempo festivo operacion': 'Tiempo festivo operación',
    'beneficios operacion': 'Beneficios operación',
    'aportes parafiscales operacion': 'Aportes y parafiscales operación',
    'prestaciones sociales operacion': 'Prestaciones sociales operación',
    'fic operacion': 'FIC operación',
    'seguro de vida operacion': 'Seguro de vida operación',
    'viaticos reembolsables operacion': 'Viáticos reembolsables operación',
    'viaticos no reembolsables operacion': 'Viáticos NO reembolsables operación',
    'viaticos descanso operacion': 'Viáticos descanso operación',
    'auxilio de alimentacion operacion': 'Auxilio alimentación/alojamiento operación',
    'auxilio de alojamiento operacion': 'Auxilio alimentación/alojamiento operación',
    'hidratacion operacion': 'Hidratación operación',
    'campamento operacion': 'Campamento operación',
    'alimentacion operacion': 'Alimentación operación',
    'celaduria campo': 'Celaduría campo',
    # ── Costos Variables / SST ──
    'epp consumibles': 'EPP consumibles',
    'epp alturas': 'EPP alturas',
    'epp servidumbre': 'EPP servidumbre',
    'seguridad grupal': 'Seguridad grupal',
    'bioseguridad': 'Bioseguridad',
    'dotacion operacion': 'Dotación operación',
    'examenmedico operacion': 'Examen médico operación',
    'examen medico operacion': 'Examen médico operación',
    'certificacion operacion': 'Certificación operación',
    'capacitaciones': 'Capacitaciones',
    'ambiental': 'Ambiental',
    # ── Costos Variables / TA ──
    'transporte operacion': 'Transporte operación',
    'transporte de materiales y herramientas': 'Transporte de materiales y herramientas',
    'transporte reembolsable': 'Transporte reembolsable',
    # ── Costos Variables / MH ──
    'material obra': 'Material obra',
    'material fungible': 'Material fungible',
    'insumos maquinaria y equipos': 'Insumos maquinaria y equipos',
    'activos fijos': 'Activos fijos',
    'httas menores': 'Herramientas menores',
    'herramientas menores': 'Herramientas menores',
    'gastos reembolsables': 'Gastos reembolsables',
    # ── Costos Variables / SC ──
    'subcontratistas': 'Subcontratistas',
    # ── Costos Fijos / MO ──
    'nomina admi': 'Nómina administración',
    'nomina administracion': 'Nómina administración',
    'tiempo extra admi': 'Tiempo extra administración',
    'tiempo extra administracion': 'Tiempo extra administración',
    'tiempo festivo admi': 'Tiempo festivo administración',
    'tiempo festivo administracion': 'Tiempo festivo administración',
    'beneficios admi': 'Beneficios administración',
    'beneficios administracion': 'Beneficios administración',
    'aportes parafiscales admi': 'Aportes y parafiscales administración',
    'aportes parafiscales administracion': 'Aportes y parafiscales administración',
    'prestaciones sociales admi': 'Prestaciones sociales administración',
    'prestaciones sociales administracion': 'Prestaciones sociales administración',
    'fic administracion': 'FIC administración',
    'fic admi': 'FIC administración',
    'seguro de vida admi': 'Seguro de vida administración',
    'seguro de vida administracion': 'Seguro de vida administración',
    'viaticos reembolsables administracion': 'Viáticos reembolsables administración',
    'viaticos no reembolsables administracion': 'Viáticos NO reembolsables administración',
    'viaticos descanso administracion': 'Viáticos descanso administración',
    'auxilio de alimentacion administracion': 'Auxilio alimentación/alojamiento administración',
    'auxilio de alojamiento administracion': 'Auxilio alimentación/alojamiento administración',
    'hidratacion administracion': 'Hidratación administración',
    'campamento administracion': 'Campamento administración',
    'alimentacion administracion': 'Alimentación administración',
    # ── Costos Fijos / SST ──
    'dotacion administracion': 'Dotación administración',
    'examenmedico admi': 'Examen médico administración',
    'examen medico admi': 'Examen médico administración',
    'examen medico administracion': 'Examen médico administración',
    'certificacion admi': 'Certificación administración',
    'certificacion administracion': 'Certificación administración',
    # ── Costos Fijos / TA ──
    'transporte admi': 'Transporte administración',
    'transporte administracion': 'Transporte administración',
    'transporte ocasional': 'Transporte ocasional',
    'arrendamiento oficina bodega': 'Arrendamiento oficina',
    'arrendamiento oficina': 'Arrendamiento oficina',
    'arrendamiento patio': 'Arrendamiento patio',
    'servicos publicos oficina bodega': 'Servicios públicos',
    'servicios publicos oficina bodega': 'Servicios públicos',
    'servicios publicos patio': 'Servicios públicos',
    'servicios publicos': 'Servicios públicos',
    'celular internet': 'Celular / internet',
    'seguridad privada': 'Seguridad privada',
    'gastos menores': 'Gastos menores',
    'insumos y elementos de oficina': 'Insumos oficina',
    'insumos oficina': 'Insumos oficina',
    'ingreso a sitio de torre': 'Ingreso sitio torre',
    'ingreso sitio torre': 'Ingreso sitio torre',
    'actividades por garantia y o imprevistos': 'Actividades por garantía / imprevistos',
    'actividades por garantia imprevistos': 'Actividades por garantía / imprevistos',
    # ── Costos Fijos / PFS ──
    'garantias fianzas seguros': 'Garantías, fianzas y seguros',
    'garantias fianzas y seguros': 'Garantías, fianzas y seguros',
}


def _normalize_name(name: str) -> str:
    """Normalize an Excel item name for dictionary lookup."""
    s = name.strip().lower()
    s = s.replace('_', ' ').replace('-', ' ')
    # Strip accents
    s = unicodedata.normalize('NFD', s)
    s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
    # Remove punctuation
    s = s.replace('/', ' ').replace('(', '').replace(')', '').replace(',', '')
    s = ' '.join(s.split())
    return s


def _parse_value(val):
    """Convert a cell value to integer."""
    if val is None:
        return 0
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return 0


def _month_from_value(val):
    """Extract month name from YYYYMM int, datetime, or month name string."""
    if val is None:
        return None
    # Already a month name string
    if isinstance(val, str):
        s = val.strip().lower()
        if s in MESES:
            return s
        # Try YYYYMM string like '202503'
        if len(s) == 6 and s.isdigit():
            month_num = int(s[4:])
            return MES_NUM_TO_NAME.get(month_num)
        return None
    # YYYYMM integer (e.g. 202503)
    if isinstance(val, (int, float)):
        ival = int(val)
        if 100000 <= ival <= 999999:
            month_num = ival % 100
            return MES_NUM_TO_NAME.get(month_num)
        return None
    # datetime object
    if isinstance(val, datetime):
        return MES_NUM_TO_NAME.get(val.month)
    return None


# Mapping from contable/accounting category names (normalized) to
# ESTRUCTURA_COSTOS items: normalized_name -> (seccion_key, codigo, item_name).
# Used when importing files with "Cuenta Equiv" style categories.
CONTABLE_TO_ESTRUCTURA = {
    # ── Costos Variables / MO ──
    'salarios': ('costos_variables', 'MO', 'Nómina operación'),
    'prestaciones sociales': ('costos_variables', 'MO', 'Prestaciones sociales operación'),
    'aportes parafiscales': ('costos_variables', 'MO', 'Aportes y parafiscales operación'),
    'seguridad social': ('costos_variables', 'MO', 'Aportes y parafiscales operación'),
    'gastos de personal': ('costos_variables', 'MO', 'Beneficios operación'),
    'auxilio de transporte': ('costos_variables', 'MO', 'Beneficios operación'),
    'gastos de viaje': ('costos_variables', 'MO', 'Viáticos reembolsables operación'),
    'viaticos reembolsables': ('costos_variables', 'MO', 'Viáticos reembolsables operación'),
    'cif': ('costos_variables', 'MO', 'FIC operación'),
    # ── Costos Variables / SST ──
    'sst': ('costos_variables', 'SST', 'EPP consumibles'),
    'dotacion': ('costos_variables', 'SST', 'Dotación operación'),
    # ── Costos Variables / TA ──
    'conductores': ('costos_variables', 'TA', 'Transporte operación'),
    'combustible': ('costos_variables', 'TA', 'Transporte operación'),
    'transporte': ('costos_variables', 'TA', 'Transporte operación'),
    # ── Costos Variables / MH ──
    'materiales': ('costos_variables', 'MH', 'Material obra'),
    'equipos y herramientas': ('costos_variables', 'MH', 'Herramientas menores'),
    'reembolsable': ('costos_variables', 'MH', 'Gastos reembolsables'),
    # ── Costos Variables / SC ──
    'subcontratistas': ('costos_variables', 'SC', 'Subcontratistas'),
    # ── Costos Fijos / MO ──
    'administrativos': ('costos_fijos', 'MO', 'Nómina administración'),
    # ── Costos Fijos / TA ──
    'arrendamiento': ('costos_fijos', 'TA', 'Arrendamiento oficina'),
    'servicios publicos': ('costos_fijos', 'TA', 'Servicios públicos'),
    'vigilancia': ('costos_fijos', 'TA', 'Seguridad privada'),
    'financieros': ('costos_fijos', 'TA', 'Gastos menores'),
    'depreciacion': ('costos_fijos', 'TA', 'Gastos menores'),
    # ── Costos Fijos / PFS ──
    'seguros': ('costos_fijos', 'PFS', 'Garantías, fianzas y seguros'),
}


def detect_excel_format(archivo):
    """Analyze an Excel file and recommend the best import format.

    Returns a dict with:
        - formato: 'presupuesto' | 'contable_resumen' | 'contable_transaccional' | 'desconocido'
        - confianza: 'alta' | 'media' | 'baja'
        - descripcion: human-readable description
        - detalles: dict with detection details (meses found, items matched, etc.)
    """
    try:
        wb = load_workbook(archivo, read_only=True, data_only=True)
    except Exception as e:
        return {
            'formato': 'desconocido',
            'confianza': 'baja',
            'descripcion': f'No se pudo leer el archivo: {e}',
            'detalles': {},
        }

    result = {
        'formato': 'desconocido',
        'confianza': 'baja',
        'descripcion': '',
        'detalles': {'hojas': wb.sheetnames},
    }

    # --- Check for standard PRESUPUESTO format ---
    presupuesto_sheet = None
    for name in wb.sheetnames:
        if 'presupuesto' in name.lower():
            presupuesto_sheet = wb[name]
            break

    if presupuesto_sheet is None:
        presupuesto_sheet = wb.active

    # Scan for month columns (standard format)
    month_cols = {}
    for scan_row in [3, 12, 1, 2, 4]:
        for col in range(1, 30):
            val = presupuesto_sheet.cell(row=scan_row, column=col).value
            if val and str(val).strip().lower() in MESES:
                month_cols[str(val).strip().lower()] = col
        if len(month_cols) == 12:
            break

    # Scan for category codes (MO, SST, TA, etc.)
    category_codes_found = set()
    items_found = 0
    max_row = min(presupuesto_sheet.max_row or 50, 50)
    for row_num in range(1, max_row + 1):
        col_a = presupuesto_sheet.cell(row=row_num, column=1).value
        col_b = presupuesto_sheet.cell(row=row_num, column=2).value
        if col_a and str(col_a).strip() in ('MO', 'SST', 'TA', 'MH', 'SC', 'PFS'):
            category_codes_found.add(str(col_a).strip())
            if col_b:
                normalized = _normalize_name(str(col_b).strip())
                if normalized in EXCEL_TO_ESTRUCTURA:
                    items_found += 1

    if len(month_cols) >= 6 and len(category_codes_found) >= 2 and items_found >= 3:
        confianza = 'alta' if len(month_cols) == 12 and items_found >= 10 else 'media'
        result = {
            'formato': 'presupuesto',
            'confianza': confianza,
            'descripcion': (
                f'Formato Presupuesto estándar detectado. '
                f'{len(month_cols)} meses encontrados, '
                f'{items_found} ítems reconocidos en categorías '
                f'{", ".join(sorted(category_codes_found))}.'
            ),
            'detalles': {
                'hojas': wb.sheetnames,
                'meses_detectados': len(month_cols),
                'items_reconocidos': items_found,
                'categorias': sorted(category_codes_found),
            },
        }
        wb.close()
        return result

    # --- Check for contable/accounting format ---
    # Check for summary/pivot sheets
    for name in wb.sheetnames:
        if name.lower() in ('res', 'resumen', 'summary', 'pivot'):
            sheet = wb[name]
            period_cols = 0
            for scan_row in range(1, 15):
                for col in range(1, 30):
                    val = sheet.cell(row=scan_row, column=col).value
                    if _month_from_value(val):
                        period_cols += 1
                if period_cols >= 2:
                    break

            if period_cols >= 2:
                result = {
                    'formato': 'contable_resumen',
                    'confianza': 'alta' if period_cols >= 6 else 'media',
                    'descripcion': (
                        f'Formato Contable (resumen/pivot) detectado en hoja "{name}". '
                        f'{period_cols} períodos encontrados.'
                    ),
                    'detalles': {
                        'hojas': wb.sheetnames,
                        'hoja_detectada': name,
                        'periodos': period_cols,
                    },
                }
                wb.close()
                return result

    # Check for transactional format
    for name in wb.sheetnames:
        sheet = wb[name]
        headers = {}
        for col in range(1, 25):
            val = sheet.cell(row=1, column=col).value
            if val:
                headers[_normalize_name(str(val))] = col

        has_neto = 'neto' in headers
        has_periodo = 'periodo' in headers or 'fecha' in headers
        has_cuenta = 'cuenta equiv' in headers

        if has_neto and has_cuenta and has_periodo:
            result = {
                'formato': 'contable_transaccional',
                'confianza': 'alta',
                'descripcion': (
                    f'Formato Contable (transaccional) detectado en hoja "{name}". '
                    f'Columnas: Neto, Cuenta Equiv, '
                    f'{"Periodo" if "periodo" in headers else "Fecha"}.'
                ),
                'detalles': {
                    'hojas': wb.sheetnames,
                    'hoja_detectada': name,
                    'columnas': list(headers.keys()),
                },
            }
            wb.close()
            return result

    wb.close()

    # Could not determine format
    result['descripcion'] = (
        'No se reconoció el formato del archivo. '
        'Se espera un Excel con: (a) hoja PRESUPUESTO con meses en columnas y códigos '
        'MO/SST/TA/MH/SC/PFS en columna A, o (b) formato contable con columnas '
        'Neto/Cuenta Equiv/Periodo.'
    )
    return result


class PresupuestoExcelImporter:
    """Parses a presupuesto Excel and returns a datos dict."""

    def __init__(self):
        self.matched = 0
        self.unmatched_items = []
        self.warnings = []

    def importar(self, archivo, estructura_costos):
        """
        Parse the Excel file and return a datos dict compatible with
        PresupuestoDetallado.datos.
        """
        try:
            wb = load_workbook(archivo, read_only=True, data_only=True)
        except Exception as e:
            return {'exito': False, 'error': f'Error al cargar archivo: {e}'}

        # Find PRESUPUESTO sheet
        sheet = None
        for name in wb.sheetnames:
            if 'presupuesto' in name.lower():
                sheet = wb[name]
                break
        if sheet is None:
            sheet = wb.active
            self.warnings.append('No se encontró hoja PRESUPUESTO, usando hoja activa.')

        # Detect month columns by scanning header rows
        month_cols = {}
        for scan_row in [3, 12, 1, 2, 4]:
            for col in range(1, 30):
                val = sheet.cell(row=scan_row, column=col).value
                if val and str(val).strip().lower() in MESES:
                    month_cols[str(val).strip().lower()] = col
            if len(month_cols) == 12:
                break

        if len(month_cols) < 12:
            wb.close()
            return {
                'exito': False,
                'error': f'Solo se detectaron {len(month_cols)} de 12 meses en las cabeceras.',
            }

        # Build reverse lookup: item name -> (seccion_key, codigo)
        item_lookup = {}
        for seccion_key, seccion in estructura_costos.items():
            for cat in seccion['categorias']:
                for item_name in cat['items']:
                    item_lookup[item_name] = (seccion_key, cat['codigo'])

        # Initialize datos with zeros
        from apps.financiero.views import _build_empty_datos
        datos = _build_empty_datos()

        # Parse all rows
        max_row = min(sheet.max_row or 200, 200)
        for row_num in range(1, max_row + 1):
            col_b = sheet.cell(row=row_num, column=2).value
            if not col_b:
                continue
            col_b_str = str(col_b).strip()

            # Check if this is the INGRESO PROYECTADO row
            if 'ingreso proyectado' in col_b_str.lower():
                # Only use the first occurrence (row ~5, not the summary at row ~91)
                if any(datos['ingreso_proyectado'][m] != 0 for m in MESES):
                    continue
                for mes, col in month_cols.items():
                    datos['ingreso_proyectado'][mes] = _parse_value(
                        sheet.cell(row=row_num, column=col).value
                    )
                continue

            # For cost items, column A must have a category code
            col_a = sheet.cell(row=row_num, column=1).value
            if not col_a:
                continue
            col_a_str = str(col_a).strip()
            if col_a_str not in ('MO', 'SST', 'TA', 'MH', 'SC', 'PFS'):
                continue

            # Normalize and lookup
            normalized = _normalize_name(col_b_str)
            estructura_name = EXCEL_TO_ESTRUCTURA.get(normalized)

            if not estructura_name:
                self.unmatched_items.append(f'Fila {row_num}: [{col_a_str}] {col_b_str}')
                continue

            if estructura_name not in item_lookup:
                self.warnings.append(
                    f'Fila {row_num}: "{estructura_name}" no está en ESTRUCTURA_COSTOS'
                )
                continue

            seccion_key, codigo = item_lookup[estructura_name]
            self.matched += 1

            # Extract monthly values
            for mes, col in month_cols.items():
                val = _parse_value(sheet.cell(row=row_num, column=col).value)
                if estructura_name in MERGE_ITEMS:
                    datos[seccion_key][codigo][estructura_name][mes] += val
                else:
                    datos[seccion_key][codigo][estructura_name][mes] = val

        wb.close()

        return {
            'exito': True,
            'datos': datos,
            'matched': self.matched,
            'unmatched': len(self.unmatched_items),
            'unmatched_items': self.unmatched_items,
            'warnings': self.warnings,
        }


class ContableExcelImporter:
    """Parses an accounting/contable Excel (Res summary or BD transactions)."""

    SKIP_LABELS = frozenset(('total', 'etiquetas de fila', ''))

    def __init__(self):
        self.matched = 0
        self.unmatched_items = []
        self.warnings = []

    def importar(self, archivo, estructura_costos):
        try:
            wb = load_workbook(archivo, read_only=True, data_only=True)
        except Exception as e:
            return {'exito': False, 'error': f'Error al cargar archivo: {e}'}

        datos = None

        # Try Res/summary sheet first (already aggregated, faster)
        for name in wb.sheetnames:
            if name.lower() in ('res', 'resumen', 'summary', 'pivot'):
                datos = self._parse_resumen(wb[name])
                if datos:
                    break

        # Fall back to any transactional sheet
        if datos is None:
            for name in wb.sheetnames:
                datos = self._parse_transaccional(wb[name])
                if datos:
                    break

        wb.close()

        if datos is None:
            return {
                'exito': False,
                'error': 'No se pudo detectar formato contable en el archivo.',
            }

        return {
            'exito': True,
            'datos': datos,
            'matched': self.matched,
            'unmatched': len(self.unmatched_items),
            'unmatched_items': self.unmatched_items,
            'warnings': self.warnings,
        }

    def _parse_resumen(self, sheet):
        """Parse a summary/pivot sheet with YYYYMM period columns."""
        # Find header row with YYYYMM/datetime columns
        month_cols = {}
        header_row = None

        for scan_row in range(1, 15):
            row_months = {}
            for col in range(1, 30):
                val = sheet.cell(row=scan_row, column=col).value
                month = _month_from_value(val)
                if month and month not in row_months:
                    row_months[month] = col
            # Accept row if it has at least 2 distinct months
            if len(row_months) >= 2:
                month_cols = row_months
                header_row = scan_row
                break

        if not month_cols:
            return None

        self.warnings.append(
            f'Formato contable resumen detectado ({len(month_cols)} meses).'
        )

        from apps.financiero.views import _build_empty_datos
        datos = _build_empty_datos()

        max_row = min(sheet.max_row or 100, 100)
        for row_num in range(header_row + 1, max_row + 1):
            label = sheet.cell(row=row_num, column=1).value
            if not label:
                continue
            label_str = str(label).strip()
            normalized = _normalize_name(label_str)

            if normalized in self.SKIP_LABELS:
                continue

            # Income rows
            if 'ingreso' in normalized and 'operacional' in normalized:
                for mes, col in month_cols.items():
                    val = _parse_value(sheet.cell(row=row_num, column=col).value)
                    datos['ingreso_proyectado'][mes] += abs(val)
                self.matched += 1
                continue

            # Skip other income-type rows (e.g. "Otros Ingresos")
            if 'ingreso' in normalized:
                continue

            # Map contable category to ESTRUCTURA item
            mapping = CONTABLE_TO_ESTRUCTURA.get(normalized)
            if mapping:
                seccion_key, codigo, item_name = mapping
                for mes, col in month_cols.items():
                    val = _parse_value(sheet.cell(row=row_num, column=col).value)
                    datos[seccion_key][codigo][item_name][mes] += val
                self.matched += 1
            else:
                self.unmatched_items.append(f'Fila {row_num}: {label_str}')

        return datos

    def _parse_transaccional(self, sheet):
        """Parse a transactional sheet (individual entries with Periodo/Fecha)."""
        # Detect columns by header names in row 1
        headers = {}
        for col in range(1, 25):
            val = sheet.cell(row=1, column=col).value
            if val:
                headers[_normalize_name(str(val))] = col

        neto_col = headers.get('neto')
        periodo_col = headers.get('periodo')
        fecha_col = headers.get('fecha')
        cuenta_col = headers.get('cuenta equiv')

        if not neto_col or not cuenta_col:
            return None

        month_source = periodo_col or fecha_col
        if not month_source:
            return None

        self.warnings.append('Formato contable transaccional detectado.')

        from apps.financiero.views import _build_empty_datos
        datos = _build_empty_datos()

        max_col = max(neto_col, month_source, cuenta_col) + 1
        seen_unmatched = set()

        for row in sheet.iter_rows(
            min_row=2, max_row=10000, max_col=max_col, values_only=False,
        ):
            neto_raw = row[neto_col - 1].value
            if neto_raw is None:
                continue

            period_val = row[month_source - 1].value
            month = _month_from_value(period_val)
            if not month:
                continue

            cuenta_val = row[cuenta_col - 1].value
            if not cuenta_val:
                continue

            neto = _parse_value(neto_raw)
            normalized = _normalize_name(str(cuenta_val))

            # Income
            if 'ingreso' in normalized:
                datos['ingreso_proyectado'][month] += abs(neto)
                self.matched += 1
                continue

            mapping = CONTABLE_TO_ESTRUCTURA.get(normalized)
            if mapping:
                seccion_key, codigo, item_name = mapping
                datos[seccion_key][codigo][item_name][month] += abs(neto)
                self.matched += 1
            else:
                if normalized not in seen_unmatched:
                    seen_unmatched.add(normalized)
                    self.unmatched_items.append(f'Categoría: {cuenta_val}')

        return datos
