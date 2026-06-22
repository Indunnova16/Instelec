"""
Importador contable completo v2 — B1 (#120).

Lee la hoja 'BD' del archivo 'BASE DE DATOS.xlsx' (≈23k filas), agrupa por la
columna O ("Cta equivalente") sumando la columna C ("Neto"), y mapea cada
cuenta equivalente a un rubro presupuestal via ``MapeoCtaRubro`` (o el dict
semilla ``MAPEO_CTA_RUBRO_DEFAULT``). Las cuentas no mapeadas caen en
"Otros / No Clasificado".

Devuelve una estructura ``datos`` que se guarda en
``PresupuestoDetallado.datos`` bajo la llave ``finv2_bd`` (no colisiona con la
estructura legacy de ``_build_display_rows`` / ``_build_empty_datos``, que vive
en las otras llaves del mismo JSONField).

Reusa el patrón de detección de columnas y lectura openpyxl de
``apps.financiero.importers`` (``_normalize_name`` / ``load_workbook
read_only``).

A1 (#120) — Bucketing mensual
-----------------------------
Además del total anual por cuenta/rubro, el importer acumula el ``Neto`` por
**mes fiscal** (orden julio→junio, ``MESES_FISCALES``) leyendo la columna D
``Fecha`` (con fallback a la columna F ``Periodo`` YYYYMM). Las filas sin fecha
reconocible caen en el bucket ``sin_mes``: cuentan en el total ANUAL pero NO en
ninguna columna de la matriz mensual. ``build_rubro_matrix_rows`` arma la matriz
rubro × 12 meses + totales por columna; ``build_rubro_display_rows`` conserva la
vista plana (fallback). La paridad ``sum(meses) + sin_mes == total`` la garantiza
el acumulado por fila (un test la verifica).
"""
import datetime
import unicodedata

from openpyxl import load_workbook

from .models_finv2_mapeo import (
    MAPEO_CTA_RUBRO_DEFAULT,
    RUBRO_NO_CLASIFICADO,
)


# Tamaño máximo de archivo aceptado (20 MB, por requerimiento #120).
MAX_UPLOAD_BYTES = 20 * 1024 * 1024

# Nombres de hoja candidatos para la base de datos contable.
_BD_SHEET_NAMES = ('bd', 'base de datos', 'base datos')

# Encabezados (normalizados) esperados en la fila 1.
_HEADER_CTA = 'cta equivalente'   # columna O
_HEADER_NETO = 'neto'             # columna C
_HEADER_DESC = 'desc auxiliar'    # columna B
_HEADER_FECHA = 'fecha'           # columna D
_HEADER_PERIODO = 'periodo'       # columna F (YYYYMM)


# --------------------------------------------------------------------------- #
# Año fiscal — orden julio → junio (oráculo: hoja 'Presupuesto' del xlsx #120)
# --------------------------------------------------------------------------- #
# Cada item: (key estable, etiqueta humana, número de mes calendario 1..12).
# El orden de la lista ES el orden de columnas de la matriz (12 meses).
MESES_FISCALES = [
    ('julio', 'Julio', 7),
    ('agosto', 'Agosto', 8),
    ('septiembre', 'Septiembre', 9),
    ('octubre', 'Octubre', 10),
    ('noviembre', 'Noviembre', 11),
    ('diciembre', 'Diciembre', 12),
    ('enero', 'Enero', 1),
    ('febrero', 'Febrero', 2),
    ('marzo', 'Marzo', 3),
    ('abril', 'Abril', 4),
    ('mayo', 'Mayo', 5),
    ('junio', 'Junio', 6),
]

# Llaves fiscales en orden (julio..junio) — usado para iterar buckets.
MESES_FISCALES_KEYS = [m[0] for m in MESES_FISCALES]

# {numero_mes_calendario: key_fiscal}, p.ej. {7: 'julio', 1: 'enero'}.
_MES_NUM_A_KEY = {num: key for key, _label, num in MESES_FISCALES}

# Bucket para filas sin fecha reconocible — entra solo en el total ANUAL,
# nunca en una columna mensual de la matriz (requerimiento #120).
SIN_MES_KEY = 'sin_mes'


def _norm(value) -> str:
    """Normaliza un texto para comparación: minúsculas, sin acentos/puntuación."""
    if value is None:
        return ''
    s = str(value).strip().lower()
    s = s.replace('_', ' ').replace('-', ' ').replace('.', ' ')
    s = unicodedata.normalize('NFD', s)
    s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
    s = s.replace('/', ' ').replace('(', '').replace(')', '').replace(',', '')
    return ' '.join(s.split())


def _to_number(value):
    """Convierte un valor de celda a float; None/no-numérico → 0.0."""
    if value is None:
        return 0.0
    try:
        return float(value)
    except (ValueError, TypeError):
        return 0.0


def _mes_fiscal_de(fecha_val, periodo_val):
    """Determina la key fiscal (julio..junio) de una fila.

    Prioridad: columna D ``Fecha`` (datetime/date) → fallback columna F
    ``Periodo`` (YYYYMM). Si ninguna es interpretable, devuelve ``None``
    (la fila cuenta solo en el total anual, bucket ``sin_mes``).
    """
    # 1) Fecha (col D): datetime/date directo.
    if isinstance(fecha_val, (datetime.datetime, datetime.date)):
        return _MES_NUM_A_KEY.get(fecha_val.month)

    # 1b) Fecha como string (algunos export traen texto): intentar parsear.
    if isinstance(fecha_val, str):
        s = fecha_val.strip()
        for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%Y/%m/%d'):
            try:
                d = datetime.datetime.strptime(s[:10], fmt)
                return _MES_NUM_A_KEY.get(d.month)
            except (ValueError, TypeError):
                continue

    # 2) Fallback: Periodo YYYYMM (col F). Puede venir como int o str.
    if periodo_val is not None:
        p = str(periodo_val).strip()
        # Aceptar formas tipo "202407" o "202407.0".
        if '.' in p:
            p = p.split('.', 1)[0]
        digits = ''.join(ch for ch in p if ch.isdigit())
        if len(digits) >= 6:
            try:
                mes_num = int(digits[4:6])
                if 1 <= mes_num <= 12:
                    return _MES_NUM_A_KEY.get(mes_num)
            except (ValueError, TypeError):
                pass
    return None


class ContableCompleteImporter:
    """Procesa la BD contable completa y la agrupa en rubros presupuestales."""

    def __init__(self):
        self.warnings = []

    # ------------------------------------------------------------------ #
    # Mapeo cuenta → rubro
    # ------------------------------------------------------------------ #
    def _build_mapeo(self, mapeo=None):
        """Construye el dict {cta_equivalente_normalizada: rubro}.

        Prioridad: ``mapeo`` explícito > registros activos de ``MapeoCtaRubro``
        en BD > ``MAPEO_CTA_RUBRO_DEFAULT`` semilla.
        """
        if mapeo:
            return {_norm(k): v for k, v in mapeo.items()}

        tabla = dict(MAPEO_CTA_RUBRO_DEFAULT)
        try:
            from .models_finv2_mapeo import MapeoCtaRubro
            qs = MapeoCtaRubro.objects.filter(activo=True)
            for m in qs:
                tabla[_norm(m.cta_equivalente)] = m.rubro_presupuestal
        except Exception:
            # En contextos sin BD (p.ej. validación de sintaxis) seguimos con
            # el dict semilla. El importador real corre dentro de la vista.
            pass
        return {_norm(k): v for k, v in tabla.items()}

    def _rubro_para(self, cta_equivalente, mapeo):
        """Devuelve el rubro para una cuenta equivalente (o NO_CLASIFICADO)."""
        return mapeo.get(_norm(cta_equivalente), RUBRO_NO_CLASIFICADO)

    # ------------------------------------------------------------------ #
    # Detección de columnas
    # ------------------------------------------------------------------ #
    def _locate_sheet(self, wb):
        for name in wb.sheetnames:
            if _norm(name) in _BD_SHEET_NAMES:
                return wb[name]
        return None

    def _locate_columns(self, sheet):
        """Detecta los índices (1-based) de las columnas O, C, B, D y F.

        Si no encuentra por nombre, cae a los índices fijos del layout #120
        (C=3, B=2, D=4, F=6, O=15).
        """
        headers = {}
        for col in range(1, 30):
            val = sheet.cell(row=1, column=col).value
            if val:
                headers[_norm(val)] = col

        cta_col = headers.get(_HEADER_CTA)
        neto_col = headers.get(_HEADER_NETO)
        desc_col = headers.get(_HEADER_DESC) or headers.get('descripcion')
        fecha_col = headers.get(_HEADER_FECHA)
        periodo_col = headers.get(_HEADER_PERIODO)

        # Fallback a índices fijos del layout documentado en #120.
        if not cta_col:
            cta_col = 15  # O
        if not neto_col:
            neto_col = 3   # C
        if not desc_col:
            desc_col = 2   # B
        if not fecha_col:
            fecha_col = 4  # D
        if not periodo_col:
            periodo_col = 6  # F

        return cta_col, neto_col, desc_col, fecha_col, periodo_col

    # ------------------------------------------------------------------ #
    # Entry point
    # ------------------------------------------------------------------ #
    def procesar_bd_completa(self, archivo, mapeo=None):
        """Procesa el archivo y devuelve un dict de resultado.

        Estructura de retorno::

            {
              'exito': bool,
              'error': str | None,        # mensaje ❌ del issue
              'advertencia': str | None,  # mensaje ⚠️ del issue
              'mensaje': str | None,      # mensaje ✅ del issue
              'datos': {...} | None,      # estructura finv2_bd para .datos
              'cuentas': int,
              'total': float,
              'no_mapeadas': [str, ...],
            }
        """
        # --- Validación de archivo (tamaño + extensión) ---
        nombre = getattr(archivo, 'name', '') or ''
        if not nombre.lower().endswith('.xlsx'):
            return self._error(
                'Archivo inválido. Verifique que sea .xlsx con estructura correcta'
            )

        tamano = getattr(archivo, 'size', None)
        if tamano is not None and tamano > MAX_UPLOAD_BYTES:
            return self._error(
                'Archivo inválido. El archivo excede el tamaño máximo (20 MB).'
            )

        try:
            wb = load_workbook(archivo, read_only=True, data_only=True)
        except Exception as exc:  # noqa: BLE001
            return self._error(
                'Archivo inválido. Verifique que sea .xlsx con estructura '
                f'correcta ({exc}).'
            )

        # --- Edge case 1: hoja BD ausente ---
        sheet = self._locate_sheet(wb)
        if sheet is None:
            sheet = wb.active
            self.warnings.append(
                f"No se encontró hoja 'BD'; usando hoja activa '{sheet.title}'."
            )

        cta_col, neto_col, desc_col, fecha_col, periodo_col = \
            self._locate_columns(sheet)

        # --- Edge case 2: columna O (Cta equivalente) ausente / vacía ---
        # Validamos leyendo el encabezado real de la columna detectada.
        header_cta = _norm(sheet.cell(row=1, column=cta_col).value)
        if header_cta and header_cta != _HEADER_CTA and 'cta' not in header_cta:
            wb.close()
            return self._advertencia(
                'Archivo sin datos válidos en columnas O (Cta equivalente) o '
                'C (Neto)'
            )

        mapeo_tabla = self._build_mapeo(mapeo)

        # --- Agrupación por cuenta equivalente (col O) sumando Neto (col C),
        #     bucketeada por mes fiscal (col D Fecha / fallback col F Periodo) ---
        # grupos[cta] = {'descripcion': str, 'total': float,
        #                'meses': {key_fiscal|sin_mes: float}}
        grupos = {}
        filas_validas = 0
        filas_sin_mes = 0
        max_col = max(cta_col, neto_col, desc_col, fecha_col, periodo_col)

        for row in sheet.iter_rows(min_row=2, max_col=max_col, values_only=True):
            cta_val = row[cta_col - 1] if len(row) >= cta_col else None
            neto_val = row[neto_col - 1] if len(row) >= neto_col else None
            desc_val = row[desc_col - 1] if len(row) >= desc_col else None
            fecha_val = row[fecha_col - 1] if len(row) >= fecha_col else None
            periodo_val = row[periodo_col - 1] if len(row) >= periodo_col else None

            # Validar que la fila tenga O, C, B (requerimiento #120 paso 1).
            if cta_val is None or neto_val is None:
                continue
            cta_str = str(cta_val).strip()
            if not cta_str:
                continue

            neto = _to_number(neto_val)
            filas_validas += 1

            mes_key = _mes_fiscal_de(fecha_val, periodo_val)
            if mes_key is None:
                mes_key = SIN_MES_KEY
                filas_sin_mes += 1

            grupo = grupos.setdefault(
                cta_str, {'descripcion': '', 'total': 0.0, 'meses': {}}
            )
            grupo['total'] += neto
            grupo['meses'][mes_key] = grupo['meses'].get(mes_key, 0.0) + neto
            if not grupo['descripcion'] and desc_val:
                grupo['descripcion'] = str(desc_val).strip()

        wb.close()

        # --- Edge case 3: archivo sin filas válidas ---
        if filas_validas == 0 or not grupos:
            return self._advertencia(
                'Archivo sin datos válidos en columnas O (Cta equivalente) o '
                'C (Neto)'
            )

        # --- Mapeo cuenta → rubro y construcción de la estructura datos ---
        # rubro -> {'total': float, 'meses': {key: float}, 'cuentas': [ {...} ]}
        rubros = {}
        no_mapeadas = []

        def _round_meses(meses):
            """Redondea cada bucket mensual a 2 decimales (incluye sin_mes)."""
            return {k: round(v, 2) for k, v in meses.items()}

        for cta, info in sorted(grupos.items()):
            rubro = self._rubro_para(cta, mapeo_tabla)
            if rubro == RUBRO_NO_CLASIFICADO:
                no_mapeadas.append(cta)
            destino = rubros.setdefault(
                rubro, {'total': 0.0, 'meses': {}, 'cuentas': []}
            )
            destino['total'] += info['total']
            for mk, mv in info['meses'].items():
                destino['meses'][mk] = destino['meses'].get(mk, 0.0) + mv
            destino['cuentas'].append({
                'cta_equivalente': cta,
                'descripcion': info['descripcion'],
                'total': round(info['total'], 2),
                'meses': _round_meses(info['meses']),
            })

        total_general = round(sum(r['total'] for r in rubros.values()), 2)
        for r in rubros.values():
            r['total'] = round(r['total'], 2)
            r['meses'] = _round_meses(r['meses'])

        datos = {
            'finv2_bd': {
                'rubros': rubros,
                'total': total_general,
                'cuentas_count': len(grupos),
                'cuentas_no_mapeadas': no_mapeadas,
                'filas_sin_mes': filas_sin_mes,
            }
        }

        mensaje = (
            f'Importación completada. {len(grupos)} cuentas procesadas. '
            f'Total importado: ${total_general:,.0f}'
        )
        if no_mapeadas:
            self.warnings.append(
                f'{len(no_mapeadas)} cuentas sin mapeo se agruparon en '
                f'"{RUBRO_NO_CLASIFICADO}".'
            )
        if filas_sin_mes:
            self.warnings.append(
                f'{filas_sin_mes} movimientos sin fecha reconocible (col D / '
                f'col F) cuentan solo en el total anual, no en la matriz mensual.'
            )

        return {
            'exito': True,
            'error': None,
            'advertencia': None,
            'mensaje': mensaje,
            'datos': datos,
            'cuentas': len(grupos),
            'total': total_general,
            'no_mapeadas': no_mapeadas,
            'warnings': self.warnings,
        }

    # ------------------------------------------------------------------ #
    # Helpers de resultado
    # ------------------------------------------------------------------ #
    def _error(self, msg):
        return {
            'exito': False, 'error': msg, 'advertencia': None, 'mensaje': None,
            'datos': None, 'cuentas': 0, 'total': 0.0, 'no_mapeadas': [],
            'warnings': self.warnings,
        }

    def _advertencia(self, msg):
        return {
            'exito': False, 'error': None, 'advertencia': msg, 'mensaje': None,
            'datos': None, 'cuentas': 0, 'total': 0.0, 'no_mapeadas': [],
            'warnings': self.warnings,
        }


def build_rubro_display_rows(datos):
    """Construye filas para el template a partir de ``datos['finv2_bd']``.

    Cada fila: {'rubro': str, 'total': float, 'pct': float, 'cuentas': [...]}.
    Devuelve (rows, total_general). Tolerante a datos vacíos / legacy sin
    la llave finv2_bd (devuelve ([], 0)).
    """
    bloque = (datos or {}).get('finv2_bd') or {}
    rubros = bloque.get('rubros') or {}
    total_general = bloque.get('total') or 0.0

    rows = []
    for rubro, info in sorted(
        rubros.items(), key=lambda kv: kv[1].get('total', 0), reverse=True
    ):
        rubro_total = info.get('total', 0.0)
        pct = (rubro_total / total_general * 100) if total_general else 0.0
        rows.append({
            'rubro': rubro,
            'total': rubro_total,
            'pct': round(pct, 1),
            'cuentas': info.get('cuentas', []),
        })
    return rows, total_general


def build_rubro_matrix_rows(datos):
    """Construye la matriz rubro × 12 meses fiscales desde ``datos['finv2_bd']``.

    Devuelve ``(rows, totales_columna, meses_fiscales, total_general)`` donde:

    - ``rows``: lista de filas, una por rubro (orden desc por total), cada una::

          {'rubro': str,
           'meses': [float, ...],        # 12 valores, orden julio..junio
           'total': float,               # total ANUAL del rubro (incl. sin_mes)
           'pct': float}

    - ``totales_columna``: lista de 12 floats (suma por columna mensual).
    - ``meses_fiscales``: ``MESES_FISCALES`` (para render del encabezado).
    - ``total_general``: total anual de todos los rubros.

    Tolerante a datos vacíos / legacy sin la llave ``finv2_bd`` (→ filas vacías)
    y a rubros legacy sin la sub-llave ``meses`` (→ matriz de ceros, el total
    anual se respeta). Las filas con bucket ``sin_mes`` NO aparecen en ninguna
    columna mensual pero SÍ suman al total anual (paridad la valida el test).
    """
    bloque = (datos or {}).get('finv2_bd') or {}
    rubros = bloque.get('rubros') or {}
    total_general = bloque.get('total') or 0.0

    rows = []
    totales_columna = [0.0] * len(MESES_FISCALES_KEYS)

    for rubro, info in sorted(
        rubros.items(), key=lambda kv: kv[1].get('total', 0), reverse=True
    ):
        meses_dict = info.get('meses') or {}
        fila_meses = []
        for i, key in enumerate(MESES_FISCALES_KEYS):
            valor = round(meses_dict.get(key, 0.0) or 0.0, 2)
            fila_meses.append(valor)
            totales_columna[i] += valor

        rubro_total = info.get('total', 0.0)
        pct = (rubro_total / total_general * 100) if total_general else 0.0
        rows.append({
            'rubro': rubro,
            'meses': fila_meses,
            'total': rubro_total,
            'pct': round(pct, 1),
        })

    totales_columna = [round(v, 2) for v in totales_columna]
    return rows, totales_columna, MESES_FISCALES, total_general


def build_mes_filter_rows(datos, mes_key):
    """Filas de un único mes fiscal (vista 'Filtro Mes').

    Devuelve ``(rows, total_mes, mes_label)``: una fila por rubro con su valor
    en ``mes_key`` (orden desc), excluyendo rubros con 0 en ese mes. Si
    ``mes_key`` no es un mes fiscal válido, devuelve ``([], 0.0, '')``.
    """
    label = next((lbl for k, lbl, _n in MESES_FISCALES if k == mes_key), '')
    if not label:
        return [], 0.0, ''

    bloque = (datos or {}).get('finv2_bd') or {}
    rubros = bloque.get('rubros') or {}

    rows = []
    total_mes = 0.0
    for rubro, info in rubros.items():
        valor = round((info.get('meses') or {}).get(mes_key, 0.0) or 0.0, 2)
        if valor == 0.0:
            continue
        rows.append({'rubro': rubro, 'total': valor})
        total_mes += valor

    rows.sort(key=lambda r: r['total'], reverse=True)
    return rows, round(total_mes, 2), label
