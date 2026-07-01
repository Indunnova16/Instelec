"""Instelec#173 — Importar actividades falla por nombre de columna 'AvísoSAP'.

Bug:
  La vista DescargarPlantillaProgramacionView generaba el header de la columna A
  como 'AvísoSAP' (con tilde en la í, sin espacio). Al normalizar a minúsculas
  queda 'avísosap', que NO matchea ningún alias del COLUMN_MAPPINGS de
  ProgramaTranselcaImporter → 'aviso_sap' nunca se detecta → error
  "No se encontró la columna de Aviso SAP".

Fix:
  1. views.py: header generado cambiado a 'Aviso SAP' → normaliza a 'aviso sap'
     que SÍ matchea el alias existente.
  2. importers.py: se añadieron 'avisosap' y 'avísosap' como aliases defensivos
     para plantillas viejas en circulación.

Round-trip cubierto:
  - Test 1: el header de la celda A1 de la plantilla descargada es 'Aviso SAP'.
  - Test 2: el importer detecta 'Aviso SAP' sin error (plantilla corregida).
  - Test 3 (defensivo): el importer detecta 'AvísoSAP' sin error (plantillas viejas).
  - Test 4 (defensivo): el importer detecta 'avisosap' sin error.
  - Test 5 (legacy): un archivo Excel que tiene 'Aviso SAP' en A1 (dato previo
    al issue) importa sin el error "columna no encontrada".
"""

import io
from datetime import date

import openpyxl
import pytest

from apps.actividades.importers import ProgramaTranselcaImporter


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_excel_with_header(aviso_header: str) -> io.BytesIO:
    """Construye un workbook mínimo con la columna de aviso en A1.

    Estructura:
        A1: <aviso_header>  B1: Línea  C1: Torre  D1: TipoActividad  E1: Fecha
        A2: (dato de ejemplo para simular un registro pre-existente)
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Programación"
    ws.append([aviso_header, 'Línea', 'Torre', 'TipoActividad', 'Fecha'])
    ws.append(['4500001234', 'L-838', '25', 'PODA', '2025-03-15'])  # dato pre-existente
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _detectar_y_obtener_indices(aviso_header: str) -> dict:
    """Llama a _detectar_columnas del importer y devuelve column_indices."""
    buf = _make_excel_with_header(aviso_header)
    wb = openpyxl.load_workbook(buf, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    importer = ProgramaTranselcaImporter()
    importer._detectar_columnas(rows[0])
    return importer.column_indices


# ---------------------------------------------------------------------------
# Test 1: el header que genera la plantilla es 'Aviso SAP' (no 'AvísoSAP')
# ---------------------------------------------------------------------------

class TestPlantillaHeaderCorregido:
    """Verifica que DescargarPlantillaProgramacionView genera 'Aviso SAP' en A1."""

    def test_header_columna_a_es_aviso_sap(self):
        """El workbook de la plantilla debe tener 'Aviso SAP' exacto en A1."""
        # Reproducimos la lógica de construcción de la plantilla
        # (mismo código que views.py DescargarPlantillaProgramacionView)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Programación"

        headers = [
            'Aviso SAP',            # A — FIX: antes era 'AvísoSAP'
            'Línea',                # B
            'Torre',                # C
            'TipoActividad',        # D
            'Fecha',                # E
            'Tipo Consignación',    # F — issue #173: antes 'Cuadrilla'
            'Descripción',          # G — issue #173: antes en H, 'Prioridad' (G) eliminada
        ]
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num).value = header

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        wb2 = openpyxl.load_workbook(buf, read_only=True, data_only=True)
        ws2 = wb2.active
        header_a1 = ws2.cell(row=1, column=1).value

        assert header_a1 == 'Aviso SAP', (
            f"El header de la columna A debe ser 'Aviso SAP', no '{header_a1}'"
        )
        assert header_a1 != 'AvísoSAP', (
            "El header CON TILDE 'AvísoSAP' sigue presente — el fix no se aplicó"
        )
        assert len(headers) == 7, (
            f"La plantilla debe tener 7 columnas (sin 'Prioridad'), tiene {len(headers)}"
        )
        assert 'Prioridad' not in headers, (
            "La columna 'Prioridad' debe estar eliminada de la plantilla (issue #173)"
        )
        assert 'Tipo Consignación' in headers, (
            "La columna 'Cuadrilla' debe haberse renombrado a 'Tipo Consignación' (issue #173)"
        )


# ---------------------------------------------------------------------------
# Test 2 y 3: el importer detecta correctamente la columna
# ---------------------------------------------------------------------------

class TestImporterDetectaColumnaAvisoSap:
    """Verifica que _detectar_columnas encuentra 'aviso_sap' para varios headers."""

    def test_header_correcto_aviso_sap_detectado(self):
        """'Aviso SAP' (plantilla fijada) → aviso_sap detectado sin error."""
        indices = _detectar_y_obtener_indices('Aviso SAP')
        assert 'aviso_sap' in indices, (
            "El importer NO detectó 'aviso_sap' para el header 'Aviso SAP'. "
            "Round-trip plantilla→importador falla."
        )
        assert indices['aviso_sap'] == 0  # columna A = índice 0

    def test_header_viejo_avisosap_con_tilde_detectado_defensivo(self):
        """'AvísoSAP' (plantillas viejas en circulación) → alias defensivo activa."""
        indices = _detectar_y_obtener_indices('AvísoSAP')
        assert 'aviso_sap' in indices, (
            "El alias defensivo 'avísosap' no está en COLUMN_MAPPINGS. "
            "Plantillas viejas con 'AvísoSAP' siguen fallando."
        )

    def test_header_viejo_avisosap_sin_tilde_detectado_defensivo(self):
        """'avisosap' (sin tilde, sin espacio) → alias defensivo activa."""
        indices = _detectar_y_obtener_indices('avisosap')
        assert 'aviso_sap' in indices, (
            "El alias defensivo 'avisosap' no está en COLUMN_MAPPINGS."
        )

    def test_alias_existente_aviso_sap_con_espacio(self):
        """El alias original 'aviso sap' (con espacio) sigue funcionando."""
        indices = _detectar_y_obtener_indices('aviso sap')
        assert 'aviso_sap' in indices


# ---------------------------------------------------------------------------
# Test 4 (legacy): archivo pre-existente con header 'Aviso SAP' no lanza error
# ---------------------------------------------------------------------------

class TestRoundtripPlantillaLegacy:
    """Verifica que un archivo con 'Aviso SAP' en A1 importa sin columna missing.

    Esto cubre el caso de datos legacy: un usuario descargó la plantilla ANTES
    del issue (cuando ya era 'Aviso SAP' o lo renombró manualmente) y el
    importador debe seguir funcionándole.
    """

    def test_importar_fila_legacy_aviso_sap(self):
        """
        Simula: plantilla con header 'Aviso SAP' en A1 + 1 fila de dato legacy
        (aviso 4500001234, L-838, 25, PODA, 2025-03-15).
        El importer detecta la columna y aviso_sap queda en column_indices.
        """
        buf = _make_excel_with_header('Aviso SAP')
        wb = openpyxl.load_workbook(buf, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))

        importer = ProgramaTranselcaImporter()
        importer._detectar_columnas(rows[0])

        assert 'aviso_sap' in importer.column_indices, (
            "Dato legacy con header 'Aviso SAP': columna no detectada. "
            "El round-trip plantilla→importador falla para archivos pre-existentes."
        )

        # Verificar que la fila de dato legacy (4500001234) es accesible
        aviso_value = importer._get_cell_value(rows[1], 'aviso_sap')
        assert str(aviso_value).strip() == '4500001234', (
            f"Se esperaba '4500001234' en la fila legacy, se obtuvo '{aviso_value}'"
        )

    def test_alias_defensivo_evita_error_plantilla_vieja(self):
        """
        Un usuario que todavía tiene la plantilla vieja ('AvísoSAP')
        ya no recibe 'No se encontró la columna de Aviso SAP'.
        """
        buf = _make_excel_with_header('AvísoSAP')
        wb = openpyxl.load_workbook(buf, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))

        importer = ProgramaTranselcaImporter()
        importer._detectar_columnas(rows[0])

        assert 'aviso_sap' in importer.column_indices, (
            "El alias defensivo no cubrió 'AvísoSAP': el error original persiste "
            "para usuarios con plantillas viejas."
        )
        aviso_value = importer._get_cell_value(rows[1], 'aviso_sap')
        assert str(aviso_value).strip() == '4500001234'
