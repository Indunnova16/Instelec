from datetime import date
from io import BytesIO

import pytest
from django.urls import reverse
from openpyxl import load_workbook

from apps.contratos.models import Contrato
from apps.construccion.excel_psc import HEADERS, importar_programacion_semanal
from apps.construccion.models import (
    AsignacionPersonalProyectoConstruccion, ProgramacionSemanalConstruccion, ProyectoConstruccion,
)
from apps.cuadrillas.models import Cargo, PersonalCuadrilla, Vehiculo


@pytest.fixture
def excel_data(db):
    contrato = Contrato.objects.create(codigo='PSC-XLSX', nombre='Contrato XLSX', unidad_negocio='CONSTRUCCION')
    proyecto = ProyectoConstruccion.objects.create(contrato=contrato, nombre='Proyecto XLSX')
    cargo, _ = Cargo.objects.get_or_create(codigo='PSC-XLSX', defaults={'nombre': 'Operario XLSX'})
    persona = PersonalCuadrilla.objects.create(nombre='Ana XLSX', documento='PSC-XLSX-1', rol_cuadrilla=cargo)
    AsignacionPersonalProyectoConstruccion.objects.create(proyecto=proyecto, personal=persona, fecha_inicio=date(2026, 1, 1))
    vehiculo = Vehiculo.objects.create(placa='XLSX225')
    return proyecto, persona, vehiculo


def _file(rows):
    from openpyxl import Workbook
    book = Workbook()
    sheet = book.active
    sheet.append(HEADERS)
    for row in rows:
        sheet.append(row)
    output = BytesIO()
    book.save(output)
    output.seek(0)
    output.name = 'programacion.xlsx'
    return output


def _row(proyecto, persona='', vehiculo='', **overrides):
    values = [proyecto.nombre, 'Obra Civil', 'Excavación', '', persona, vehiculo,
              '2026-08-17', '2026-08-21', '07:00', '16:00', 'Importación test']
    mapping = dict(zip(HEADERS, values))
    mapping.update(overrides)
    return [mapping[key] for key in HEADERS]


@pytest.mark.django_db
def test_http_get_xlsx(admin_user, client):
    client.force_login(admin_user)
    response = client.get(reverse('construccion:psc_plantilla_excel'))
    assert response.status_code == 200
    assert response['Content-Type'].startswith('application/vnd.openxmlformats-officedocument')
    assert tuple(next(load_workbook(BytesIO(b''.join(response.streaming_content))).active.values)) == HEADERS


@pytest.mark.django_db
def test_exportar_xlsx_incluye_once_columnas(admin_user, client, excel_data):
    client.force_login(admin_user)
    response = client.get(reverse('construccion:psc_exportar_excel'))
    assert response.status_code == 200
    assert tuple(next(load_workbook(BytesIO(b''.join(response.streaming_content))).active.values)) == HEADERS


@pytest.mark.django_db
def test_upload_y_reporte(admin_user, client, excel_data):
    proyecto, persona, vehiculo = excel_data
    client.force_login(admin_user)
    response = client.post(reverse('construccion:psc_importar_excel'), {'archivo': _file([_row(proyecto, persona.documento, vehiculo.placa)])})
    assert response.status_code == 200
    assert ProgramacionSemanalConstruccion.objects.count() == 1
    assert 'Se importaron 1 programaciones' in response.content.decode()


@pytest.mark.django_db
def test_importacion_atomica_si_una_fila_es_invalida(excel_data):
    proyecto, persona, vehiculo = excel_data
    result = importar_programacion_semanal(_file([
        _row(proyecto, persona.documento, vehiculo.placa),
        _row(proyecto, 'NO-EXISTE', vehiculo.placa),
    ]))
    assert not result.ok
    assert result.errors[0]['row'] == 3
    assert ProgramacionSemanalConstruccion.objects.count() == 0


@pytest.mark.django_db
def test_rechaza_cruce_de_personal_en_archivo(excel_data):
    proyecto, persona, vehiculo = excel_data
    result = importar_programacion_semanal(_file([
        _row(proyecto, persona.documento, vehiculo.placa),
        _row(proyecto, persona.documento, vehiculo.placa, **{'Fecha Inicio': '2026-08-20'}),
    ]))
    assert not result.ok
    assert 'se cruza con la fila 2' in result.errors[0]['error']
